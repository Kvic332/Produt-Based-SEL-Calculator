# Copyright (c) 2026 Kenechukwu (Kvic7). All rights reserved.
# Proprietary and confidential — see LICENSE. No license granted.
from __future__ import annotations
import datetime
import gc
import json
import math
import os
import pathlib
import re
import uuid
import pandas as pd
import streamlit as st
from bank_parser import (
    monthly_analysis, parse_transactions, parse_firstcentral,
    ym_label, CreditAccount,
    extract_stated_totals, verify_extraction_accuracy,
    extract_account_no_excel,
    get_last_full_text,
)
from sel_rules import calculate_eligibility, get_interest_rate, get_dti, loan_limits
from report_generator import generate_pdf_report, generate_credit_memo
from tracker import (track, admin_stats, save_history, get_history, export_audit_csv,
                     check_blacklist, check_duplicate_application,
                     save_blacklist_entries, get_blacklist, delete_blacklist_entry, clear_blacklist,
                     get_applicant_assessments, save_officer_note, get_officer_notes)

# ── Page config ───────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="PARSIO",
    page_icon="▶",
    layout="wide",
)

# ── Nigeria time (WAT, UTC+1 — no DST, so a fixed offset is exact) ────────────
# Streamlit Cloud servers run UTC; every user-facing clock/date must use this.
_TZ_LAGOS = datetime.timezone(datetime.timedelta(hours=1), name="WAT")

def now_lagos() -> datetime.datetime:
    return datetime.datetime.now(_TZ_LAGOS)

# ── Dark theme CSS — PARSIO design system (portal-aligned, Phase 1) ──────────
st.markdown("""
<style>
  /* Import fonts — match the PARSIO portal */
  @import url('https://fonts.googleapis.com/css2?family=Plus+Jakarta+Sans:wght@400;500;600;700;800&family=Inter:wght@400;500;600;700&family=Space+Mono:wght@400;700&display=swap');

  /* Root variables — Emerald + Gold palette (refined surfaces + elevation) */
  :root {
    --bg: #0a0f0d; --surface: #101814; --surface2: #17211c;
    --border: #20392c; --border-soft: #1a2e24;
    --accent: #10b981; --accent2: #f59e0b;
    --green: #34d399; --red: #f87171; --text: #e6ebe8;
    --muted: #9fb0a7; --dim: #6b7c73; --gold: #fbbf24; --orange: #fb923c;
    --font-head: 'Plus Jakarta Sans', sans-serif;
    --font-body: 'Inter', sans-serif;
    --font-mono: 'Space Mono', monospace;
    --shadow-sm: 0 1px 3px rgba(0,0,0,.35);
    --shadow-md: 0 8px 24px -8px rgba(0,0,0,.45);
    --glow-accent: 0 0 24px rgba(16,185,129,.12);
    --radius: 12px; --radius-sm: 8px;
  }

  /* Global */
  .stApp { background: var(--bg) !important; color: var(--text) !important; font-family: var(--font-body) !important; }
  .block-container { padding: 2rem 2rem 4rem !important; max-width: 1040px !important; }
  .stApp p, .stApp li { font-family: var(--font-body); }

  /* Headers */
  h1 { font-family: var(--font-head) !important; font-weight: 800 !important;
       letter-spacing: -1px !important; color: #fff !important; }
  h1 em { color: var(--accent) !important; font-style: normal !important; }
  h2, h3 { font-family: var(--font-head) !important; color: var(--accent) !important;
           font-size: 13px !important; letter-spacing: 2.5px !important; text-transform: uppercase !important; font-weight: 700 !important; }

  /* Sections */
  .sel-section { background: var(--surface); border: 1px solid var(--border-soft);
                 border-radius: var(--radius); padding: 26px; margin-bottom: 22px;
                 box-shadow: var(--shadow-sm); }
  .sel-section-title { font-family: var(--font-head); font-size: 16px; letter-spacing: 1.2px; color: #fff;
                       text-transform: uppercase; border-bottom: 2px solid var(--accent);
                       padding-bottom: 12px; margin-bottom: 18px; font-weight: 800; }
  .sel-caption { font-size: 13.5px; color: #c4cfc9; font-weight: 500;
                 margin: -8px 0 14px 0; line-height: 1.65; font-family: var(--font-body); }

  /* Metric cards */
  .sel-card { background: var(--surface2); border: 1px solid var(--border-soft);
              border-radius: var(--radius-sm); padding: 16px 18px;
              transition: border-color .2s, box-shadow .2s; }
  .sel-card:hover { border-color: var(--border); }
  .sel-card.highlight { border-color: var(--accent); box-shadow: var(--glow-accent); }
  .sel-label { font-family: var(--font-head); font-size: 10.5px; letter-spacing: 1.8px; color: var(--muted);
               text-transform: uppercase; margin-bottom: 6px; font-weight: 700; }
  .sel-value { font-family: var(--font-head); font-size: 23px; font-weight: 800; color: var(--accent);
               font-variant-numeric: tabular-nums; letter-spacing: -.3px; }
  .sel-value.green  { color: var(--green) !important; }
  .sel-value.gold   { color: var(--gold) !important; }
  .sel-value.red    { color: var(--red) !important; }
  .sel-value.orange { color: var(--orange) !important; }

  /* Banners */
  .banner-approved { background: rgba(52,211,153,.09); border: 1px solid rgba(52,211,153,.35);
                     color: var(--green); padding: 16px 20px; border-radius: var(--radius-sm);
                     font-family: var(--font-head); font-size: 15px; font-weight: 700; letter-spacing: .5px; }
  .banner-rejected { background: rgba(248,113,113,.09); border: 1px solid rgba(248,113,113,.35);
                     color: var(--red); padding: 16px 20px; border-radius: var(--radius-sm);
                     font-family: var(--font-head); font-size: 15px; font-weight: 700; letter-spacing: .5px; }
  .banner-info     { background: rgba(16,185,129,.06); border: 1px solid rgba(16,185,129,.22);
                     color: var(--accent); padding: 13px 17px; border-radius: var(--radius-sm);
                     font-size: 13px; line-height: 1.6; }
  .banner-bad      { background: rgba(248,113,113,.08); border: 1px solid rgba(248,113,113,.28);
                     color: var(--red); padding: 13px 17px; border-radius: var(--radius-sm);
                     font-size: 13px; line-height: 1.6; }
  .banner-good     { background: rgba(52,211,153,.08); border: 1px solid rgba(52,211,153,.28);
                     color: var(--green); padding: 13px 17px; border-radius: var(--radius-sm);
                     font-size: 13px; line-height: 1.6; }

  /* ── Hero verdict card ──────────────────────────────────────────── */
  .verdict-hero { border-radius: 16px; padding: 28px 30px; margin: 4px 0 18px;
                  display: grid; grid-template-columns: 1.1fr 1fr; gap: 28px;
                  align-items: center; box-shadow: var(--shadow-md); }
  .vh-approved { background: linear-gradient(135deg, rgba(16,185,129,.16) 0%, rgba(6,78,54,.28) 100%);
                 border: 1px solid rgba(52,211,153,.4); }
  .vh-declined { background: linear-gradient(135deg, rgba(248,113,113,.12) 0%, rgba(127,29,29,.22) 100%);
                 border: 1px solid rgba(248,113,113,.38); }
  .vh-badge { display: inline-block; font-family: var(--font-head); font-weight: 800;
              font-size: 12px; letter-spacing: 2.5px; text-transform: uppercase;
              padding: 7px 18px; border-radius: 999px; margin-bottom: 16px; }
  .vh-approved .vh-badge { background: rgba(52,211,153,.18); color: var(--green);
                           border: 1px solid rgba(52,211,153,.45); }
  .vh-declined .vh-badge { background: rgba(248,113,113,.15); color: var(--red);
                           border: 1px solid rgba(248,113,113,.4); }
  .vh-label { font-family: var(--font-head); font-size: 11px; letter-spacing: 2.5px;
              color: var(--muted); text-transform: uppercase; font-weight: 700; margin-bottom: 4px; }
  .vh-amount { font-family: var(--font-head); font-size: 46px; font-weight: 800;
               letter-spacing: -1.5px; line-height: 1.05; font-variant-numeric: tabular-nums; }
  .vh-approved .vh-amount { color: #fff; text-shadow: 0 0 30px rgba(52,211,153,.3); }
  .vh-declined .vh-amount { color: #fecaca; }
  .vh-sub { font-size: 13px; color: var(--muted); margin-top: 8px; font-weight: 500; }
  .vh-grid { display: grid; grid-template-columns: 1fr 1fr; gap: 10px; }
  .vh-cell-gauge { background: rgba(0,0,0,.22); border: 1px solid rgba(255,255,255,.06);
                  border-radius: 10px; padding: 10px 8px 6px;
                  display: flex; align-items: center; justify-content: center; }
  .vh-cell { background: rgba(0,0,0,.22); border: 1px solid rgba(255,255,255,.06);
             border-radius: 10px; padding: 12px 14px; }
  .vh-cell-label { font-family: var(--font-head); font-size: 9.5px; letter-spacing: 1.6px;
                   color: var(--dim); text-transform: uppercase; font-weight: 700; margin-bottom: 3px; }
  .vh-cell-value { font-family: var(--font-head); font-size: 17px; font-weight: 800;
                   color: var(--text); font-variant-numeric: tabular-nums; }
  .vh-approved .vh-cell-value.em { color: var(--green); }
  .vh-declined .vh-cell-value.em { color: var(--red); }
  @media (max-width: 768px) {
    .verdict-hero { grid-template-columns: 1fr; padding: 22px; gap: 18px; }
    .vh-amount { font-size: 36px; }
  }

  /* ── Tabs — segmented control style ─────────────────────────────── */
  .stTabs [data-baseweb="tab-list"] { gap: 6px; background: var(--surface);
      border: 1px solid var(--border-soft); border-radius: 10px; padding: 5px;
      margin-bottom: 4px; }
  .stTabs [data-baseweb="tab"] { font-family: var(--font-head) !important;
      font-weight: 700 !important; font-size: 13px !important; letter-spacing: .4px;
      color: var(--muted) !important; background: transparent !important;
      border-radius: 7px !important; padding: 8px 16px !important; }
  .stTabs [data-baseweb="tab"]:hover { color: var(--text) !important; }
  .stTabs [aria-selected="true"] { background: rgba(16,185,129,.13) !important;
      color: var(--accent) !important; }
  .stTabs [data-baseweb="tab-highlight"], .stTabs [data-baseweb="tab-border"] { display: none !important; }

  /* Tags / badges */
  .badge { display: inline-block; padding: 3px 10px; border-radius: 999px;
           font-family: var(--font-head); font-weight: 700;
           font-size: 10px; letter-spacing: 1px; text-transform: uppercase; }
  .badge-blue   { background: rgba(16,185,129,.1); color: var(--accent); border: 1px solid rgba(16,185,129,.2); }
  .badge-red    { background: rgba(248,113,113,.1);  color: var(--red);    border: 1px solid rgba(248,113,113,.25); }
  .badge-orange { background: rgba(251,146,60,.1);  color: var(--orange); border: 1px solid rgba(251,146,60,.25); }
  .badge-green  { background: rgba(52,211,153,.1);  color: var(--green);  border: 1px solid rgba(52,211,153,.25); }

  /* Tables */
  .preview-table { width: 100%; border-collapse: collapse; font-size: 13px; margin-top: 8px;
                   font-variant-numeric: tabular-nums; }
  .preview-table th { font-family: var(--font-head); font-size: 11px; letter-spacing: 1.2px; color: var(--muted);
                      text-transform: uppercase; padding: 9px 12px; font-weight: 700;
                      border-bottom: 2px solid var(--accent); text-align: right; }
  .preview-table th:first-child { text-align: left; }
  .preview-table td { padding: 8px 12px; border-bottom: 1px solid var(--border-soft);
                      text-align: right; font-weight: 500; }
  .preview-table tr:hover td { background: rgba(16,185,129,.03); }
  .preview-table td:first-child { text-align: left; color: var(--accent); font-weight: 700;
                                  font-family: var(--font-head); }
  .col-gross  { color: var(--green); }
  .col-self   { color: var(--orange); }
  .col-rev    { color: #a78bfa; }
  .col-nonbiz { color: var(--muted); }
  .col-loan   { color: var(--gold); }
  .col-net    { color: var(--accent); font-weight: 700; }

  /* Credit table */
  .credit-table { width: 100%; border-collapse: collapse; font-size: 12.5px; margin-top: 10px;
                  font-variant-numeric: tabular-nums; }
  .credit-table th { font-family: var(--font-head); font-size: 10.5px; letter-spacing: 1.2px; color: var(--muted);
                     text-transform: uppercase; padding: 9px 8px; font-weight: 700;
                     border-bottom: 1px solid var(--border); text-align: left; }
  .credit-table td { padding: 9px 8px; border-bottom: 1px solid var(--border-soft); vertical-align: top; }
  .credit-table tr:hover td { background: rgba(16,185,129,.03); }
  .credit-table tfoot td { border-top: 1px solid var(--border); font-weight: 700; }

  /* Expanders */
  [data-testid="stExpander"] { background: var(--surface) !important;
      border: 1px solid var(--border-soft) !important;
      border-radius: var(--radius-sm) !important; box-shadow: var(--shadow-sm); }
  [data-testid="stExpander"] summary { font-family: var(--font-head) !important;
      font-weight: 600 !important; font-size: 14px !important; }
  [data-testid="stExpander"] summary:hover { color: var(--accent) !important; }

  /* Sidebar */
  [data-testid="stSidebar"] { background: var(--surface) !important;
                               border-right: 1px solid var(--border-soft) !important; }
  [data-testid="stSidebar"] [data-testid="stMetric"] { background: var(--surface2) !important; }
  [data-testid="stSidebar"] hr { margin: 16px 0 !important; }
  [data-testid="stSidebar"] label { color: var(--muted) !important; font-size: 11px !important;
                                    letter-spacing: 1px !important; text-transform: uppercase !important; font-weight: 600 !important; }

  /* Inputs */
  input, select, textarea, [data-testid="stTextInput"] input,
  [data-testid="stNumberInput"] input, [data-testid="stSelectbox"] select {
    background: var(--surface2) !important; border: 1px solid var(--border) !important;
    color: var(--text) !important; font-family: var(--font-body) !important;
    border-radius: var(--radius-sm) !important;
  }

  /* Buttons — default (secondary) = outline, primary = filled emerald */
  .stButton button { background: transparent !important; border: 1px solid var(--accent) !important;
                     color: var(--accent) !important; font-family: var(--font-head) !important;
                     font-weight: 700 !important; letter-spacing: 1.2px !important;
                     text-transform: uppercase !important; border-radius: var(--radius-sm) !important;
                     transition: background .2s, transform .15s, box-shadow .2s !important; }
  .stButton button:hover { background: rgba(16,185,129,.08) !important; }
  .stButton button[kind="primary"] { background: var(--accent) !important;
                     color: #04130c !important;
                     box-shadow: 0 4px 14px rgba(16,185,129,.25) !important; }
  .stButton button[kind="primary"]:hover { background: #0ea271 !important; transform: translateY(-1px);
                     box-shadow: 0 8px 20px rgba(16,185,129,.3) !important; }

  /* Selectbox (BaseWeb) — dark control + readable value */
  [data-testid="stSelectbox"] [data-baseweb="select"] > div {
    background: var(--surface2) !important; border-color: var(--border) !important;
    color: var(--text) !important; border-radius: var(--radius-sm) !important;
  }
  [data-testid="stSelectbox"] [data-baseweb="select"] * { color: var(--text) !important; }
  [data-baseweb="popover"] [data-baseweb="menu"], [data-baseweb="popover"] ul {
    background: var(--surface2) !important; border: 1px solid var(--border) !important;
  }
  [data-baseweb="popover"] li { color: var(--text) !important; }
  [data-baseweb="popover"] li:hover { background: rgba(16,185,129,.1) !important; }

  /* Number input steppers */
  [data-testid="stNumberInput"] button {
    background: var(--surface2) !important; color: var(--muted) !important;
    border: 1px solid var(--border) !important;
  }

  /* File uploader */
  [data-testid="stFileUploader"] { background: var(--surface2) !important;
                                    border: 2px dashed var(--border) !important;
                                    border-radius: var(--radius) !important;
                                    transition: border-color .2s !important; }
  [data-testid="stFileUploader"]:hover { border-color: var(--accent) !important; }
  [data-testid="stFileUploaderDropzone"] { background: var(--surface2) !important;
                                           color: var(--muted) !important; }
  [data-testid="stFileUploaderDropzone"] small,
  [data-testid="stFileUploaderDropzone"] span { color: var(--muted) !important; }
  [data-testid="stFileUploaderDropzone"] button {
    background: var(--surface) !important; color: var(--accent) !important;
    border: 1px solid var(--border) !important; border-radius: var(--radius-sm) !important;
    font-family: var(--font-head) !important; font-weight: 700 !important;
  }

  /* BaseWeb input wrapper — covers the adornment segment (password eye etc.) */
  [data-baseweb="input"], [data-baseweb="base-input"] {
    background: var(--surface2) !important; border-color: var(--border) !important;
    border-radius: var(--radius-sm) !important;
  }
  [data-testid="stTextInput"] button, [data-testid="stNumberInput"] svg {
    background: transparent !important; color: var(--muted) !important;
  }

  /* Dataframe */
  .stDataFrame { background: var(--surface2) !important; }

  /* Metric */
  [data-testid="stMetric"] { background: var(--surface2) !important;
                              border: 1px solid var(--border-soft) !important;
                              border-radius: var(--radius-sm) !important; padding: 14px 16px !important; }
  [data-testid="stMetricLabel"] { color: var(--muted) !important; font-size: 10.5px !important;
                                   letter-spacing: 1.8px !important; text-transform: uppercase !important; font-weight: 700 !important;
                                   font-family: var(--font-head) !important; }
  [data-testid="stMetricValue"] { color: var(--accent) !important; font-size: 23px !important;
                                   font-family: var(--font-head) !important; font-weight: 800 !important;
                                   font-variant-numeric: tabular-nums !important; }

  /* Divider */
  hr { border: none !important; border-top: 1px solid var(--border-soft) !important; margin: 28px 0 !important; }

  /* Caption */
  [data-testid="stCaptionContainer"] { color: var(--muted) !important; font-size: 12px !important; }

  /* Search input highlight */
  [data-testid="stTextInput"] input:focus {
    border-color: var(--accent) !important;
    box-shadow: 0 0 0 2px rgba(16,185,129,.15) !important;
  }

  /* Number input */
  [data-testid="stNumberInput"] input:focus {
    border-color: var(--accent) !important;
    box-shadow: 0 0 0 2px rgba(16,185,129,.1) !important;
  }

  /* Spinner */
  .stSpinner > div { border-top-color: var(--accent) !important; }

  /* Success / error messages */
  [data-testid="stAlert"] { font-family: var(--font-body) !important; font-size: 13px !important;
                            border-radius: var(--radius-sm) !important; }

  /* Widget labels — Location, Product Type, Tenor, Loan Amount, etc. */
  [data-testid="stWidgetLabel"] p,
  [data-testid="stWidgetLabel"] label,
  .stSelectbox label, .stNumberInput label,
  .stTextInput label, .stRadio label > div > p,
  div[data-testid="stSelectbox"] > label,
  div[data-testid="stNumberInput"] > label {
    font-size: 12px !important; font-weight: 700 !important;
    color: #e2e8f0 !important; letter-spacing: 0.5px !important;
  }

  /* Inflow grid header (legacy class — kept for compatibility) */
  .inflow-grid-header { font-size: 12px; letter-spacing: 1.5px; color: #e2e8f0;
                        text-transform: uppercase; font-weight: 800; padding-bottom: 4px; }

  /* Radio option labels (SEL / SME toggle) */
  [data-testid="stRadio"] label span,
  [data-testid="stRadio"] > div > label > div > p {
    font-size: 14px !important; font-weight: 700 !important;
    color: #e2e8f0 !important; letter-spacing: 1px !important;
  }
  /* Radio group title (hidden but keep consistent) */
  [data-testid="stRadio"] > label { display: none !important; }

  /* Selectbox option text */
  [data-testid="stSelectbox"] > div > div { color: #e2e8f0 !important; font-weight: 600 !important; }

  /* Download button */
  [data-testid="stDownloadButton"] button {
    background: rgba(16,185,129,.06) !important;
    border: 1px solid rgba(16,185,129,.35) !important;
    color: var(--green) !important;
    font-family: var(--font-head) !important;
    font-size: 12px !important;
    letter-spacing: 1px !important;
    font-weight: 700 !important;
    border-radius: var(--radius-sm) !important;
  }
  [data-testid="stDownloadButton"] button:hover {
    background: rgba(16,185,129,.12) !important;
    border-color: var(--accent) !important;
    color: var(--accent) !important;
  }

  /* ── Mobile / Tablet responsive layout ────────────────────────────── */
  @media (max-width: 768px) {
    /* Reduce container padding on small screens */
    .block-container { padding: 1rem 0.75rem 3rem !important; }

    /* Stack Streamlit columns vertically on mobile */
    [data-testid="stHorizontalBlock"] {
      flex-direction: column !important;
      gap: 0.5rem !important;
    }
    [data-testid="stHorizontalBlock"] > [data-testid="stVerticalBlock"] {
      width: 100% !important;
      min-width: 0 !important;
      flex: 1 1 100% !important;
    }

    /* Make metric cards more compact */
    [data-testid="stMetric"] { padding: 8px 10px !important; }
    [data-testid="stMetricValue"] { font-size: 20px !important; }
    [data-testid="stMetricLabel"] { font-size: 10px !important; }

    /* Larger touch targets for buttons */
    button, [data-testid="stDownloadButton"] button {
      min-height: 44px !important;
      font-size: 13px !important;
    }

    /* Full-width selectboxes and inputs */
    [data-testid="stSelectbox"], [data-testid="stTextInput"],
    [data-testid="stNumberInput"] {
      width: 100% !important;
    }

    /* Prevent table overflow — allow horizontal scroll */
    .preview-table { font-size: 10px !important; }
    div[data-testid="stMarkdownContainer"] { overflow-x: auto; }

    /* Shrink section titles on mobile */
    .sel-section-title { font-size: 13px !important; letter-spacing: 1px !important; }

    /* Upload areas — larger on touch screens */
    [data-testid="stFileUploaderDropzone"] {
      min-height: 80px !important;
      padding: 16px !important;
    }
  }

  @media (max-width: 480px) {
    .block-container { padding: 0.75rem 0.5rem 2rem !important; }
    .sel-section-title { font-size: 11px !important; }
    [data-testid="stMetricValue"] { font-size: 17px !important; }
  }
</style>
""", unsafe_allow_html=True)


# ── Helpers ───────────────────────────────────────────────────────────────────
def money(v: float) -> str:
    try:
        return f"₦{float(v):,.0f}"
    except (TypeError, ValueError):
        return "₦0"

def pct(v) -> str:
    try:
        return "--" if v is None else f"{float(v) * 100:.2f}%"
    except (TypeError, ValueError):
        return "--"

def render_parse_confidence(buckets: dict, pct_match, bank: str,
                            filename: str = "", session: str = "",
                            officer: str = "") -> None:
    """Parser self-audit — shown after every extraction.

    Components
    ----------
    accuracy   kobo-match vs the statement's own stated totals (when available)
    coverage   months with data vs the full span of months in the statement
               (a gap month usually means pages the parser silently skipped)
    density    average transactions per active month (a 6-month statement
               yielding 3 txns/month suggests dropped lines)

    Composite score drives a High / Medium / Low badge; Low statements are
    flagged for manual review and every report is tracked for the admin
    dashboard so weak parsers surface systematically instead of one
    complaint at a time.
    """
    import re as _re_pc
    months = sorted(k for k in buckets if _re_pc.match(r"^20\d{2}-\d{2}$", k))
    if not months:
        return
    active = [m for m in months if buckets[m].get("gross", 0) > 0]
    if not active:
        return

    # Month coverage across the statement's own span
    _y0, _m0 = int(months[0][:4]),  int(months[0][5:])
    _y1, _m1 = int(months[-1][:4]), int(months[-1][5:])
    span = (_y1 - _y0) * 12 + (_m1 - _m0) + 1
    coverage = len(active) / span if span else 1.0
    gap_months = span - len(active)

    # Transaction density
    txn_total = sum(buckets[m].get("count", 0) for m in active)
    avg_txn   = txn_total / len(active)
    density   = min(avg_txn / 10.0, 1.0)   # 10+ txns/month = fully dense

    # Accuracy (neutral when the statement states no totals to check against)
    has_acc = pct_match is not None
    acc     = min(float(pct_match) / 100.0, 1.0) if has_acc else 0.70

    score = acc * 0.55 + coverage * 0.30 + density * 0.15
    if score >= 0.90 and (not has_acc or pct_match >= 95) and gap_months == 0:
        level, col, icon = "High", "#34d399", "●"
    elif score >= 0.72:
        level, col, icon = "Medium", "#fbbf24", "●"
    else:
        level, col, icon = "Low", "#f87171", "●"

    _parts = []
    _parts.append(
        f'<span style="color:#9fb0a7">Accuracy '
        f'<strong style="color:{"#34d399" if has_acc and pct_match >= 95 else "#fbbf24" if has_acc else "#6b7c73"}">'
        f'{f"{pct_match:.0f}%" if has_acc else "n/a"}</strong></span>'
    )
    _parts.append(
        f'<span style="color:#9fb0a7">Coverage '
        f'<strong style="color:{"#34d399" if gap_months == 0 else "#f87171"}">'
        f'{len(active)}/{span} months</strong></span>'
    )
    _parts.append(
        f'<span style="color:#9fb0a7">Density '
        f'<strong style="color:{"#34d399" if avg_txn >= 10 else "#fbbf24"}">'
        f'{avg_txn:.0f} txn/mo</strong></span>'
    )
    _warn = ""
    if level == "Low":
        _warn = ('<div style="margin-top:6px;font-size:11.5px;color:#f87171;font-weight:600">'
                 '⚠ Low parser confidence — verify the monthly figures against the '
                 'original statement before relying on this assessment.</div>')
    elif gap_months > 0:
        _warn = (f'<div style="margin-top:6px;font-size:11.5px;color:#fbbf24">'
                 f'⚑ {gap_months} month(s) in the statement period have no extracted '
                 f'transactions — check for skipped pages.</div>')

    st.markdown(
        f'<div style="background:rgba(0,0,0,.2);border:1px solid {col}44;'
        f'border-left:4px solid {col};border-radius:8px;padding:10px 14px;'
        f'margin-top:8px;font-size:12px">'
        f'<span style="font-family:Plus Jakarta Sans,sans-serif;font-weight:800;'
        f'color:{col};letter-spacing:1px">{icon} PARSER CONFIDENCE: {level.upper()}'
        f'</span>&nbsp;&nbsp;<span style="color:#6b7c73">score {score:.0%}</span>'
        f'<div style="display:flex;gap:18px;flex-wrap:wrap;margin-top:5px">'
        + "".join(_parts) + '</div>' + _warn + '</div>',
        unsafe_allow_html=True,
    )

    try:
        track("parse_confidence", session=session, officer=officer, bank=bank,
              filename=filename, score=round(score, 3), level=level,
              accuracy_pct=(round(pct_match, 1) if has_acc else None),
              months_active=len(active), months_span=span,
              avg_txn_per_month=round(avg_txn, 1))
    except Exception:
        pass


def extract_account_no(raw_text: str) -> str:
    """Extract 10-digit Nigerian NUBAN account number from raw statement text.

    Handles multiple statement formats:
    - mybankStatement:  'Account No. 0036641218'
    - OPay / Carbon:    'Account Name  Account Number\\nNAME  7026155943'
    - Generic label:    'Account Number: 1234567890'
    """
    # 1. Explicit label on same line: "Account No. XXXXXXXXXX"
    m = re.search(r'Account\s*No\.?\s*[:\s]+(\d{10})\b', raw_text, re.I)
    if m:
        return m.group(1)

    # 2. OPay / Carbon column-header format:
    #    "Account Name  Account Number\n<NAME>  7026155943"
    m2 = re.search(
        r'Account\s*Number\s*[\r\n]+[^\r\n]{1,80}\s(\d{10})\b',
        raw_text, re.I,
    )
    if m2:
        return m2.group(1)

    # 3. "Account Number" label followed (same line) by 10-digit number
    m3 = re.search(r'Account\s*Number\s*[:\s]+(\d{10})\b', raw_text, re.I)
    if m3:
        return m3.group(1)

    # 4. Fallback: first standalone 10-digit number in header area (first 2000 chars)
    m4 = re.search(r'\b(\d{10})\b', raw_text[:2000])
    return m4.group(1) if m4 else ""

def html_bar_chart(labels, values, color: str = "#10b981", money_fmt: bool = False,
                   height: int = 120) -> str:
    """Render a dependency-free vertical bar chart as HTML.

    Used instead of st.bar_chart/st.line_chart, which import altair —
    altair's TypedDict(closed=True) schema crashes on Python 3.14.

    height: bar column height in px (default 120; pass 180+ for full-width charts).
    """
    vals  = [float(v or 0) for v in values]
    scale = max(vals) if vals and max(vals) > 0 else 1
    n     = len(vals)
    BAR_H = height

    # label density — show every Nth label to avoid crowding on dense charts
    _show_every = max(1, n // 20)

    def _fmt(v: float) -> str:
        if money_fmt:
            return f"₦{v/1_000_000:.1f}m" if v >= 1_000_000 else f"₦{v/1_000:.0f}k" if v >= 1000 else f"₦{v:.0f}"
        return f"{v:,.0f}"

    # detect which labels are month-name markers (e.g. "Jun 2026") vs day numbers
    def _is_month_label(lbl): return len(lbl) > 3  # day numbers are "01".."31"

    bars = ""
    for i, (lbl, v) in enumerate(zip(labels, vals)):
        h = int(v / scale * BAR_H) if v > 0 else 0
        _is_month = _is_month_label(lbl)

        # value label: show on every non-zero bar
        _val_lbl = (
            f'<div style="font-size:9px;font-weight:600;color:{color};margin-bottom:3px;'
            f'white-space:nowrap;min-height:14px">'
            + (_fmt(v) if v > 0 else "")
            + '</div>'
        )
        # x-axis label:
        #   - month names always shown (truncated to "M..." if too narrow)
        #   - day numbers shown only on active (non-zero) bars
        if _is_month:
            _show_lbl = lbl[:4] + "…" if len(lbl) > 6 else lbl
            _lbl_col  = "#e2e8f0"    # bright white for month markers
            _lbl_size = "9px"
        elif v > 0:
            _show_lbl = lbl
            _lbl_col  = "#94a3b8"
            _lbl_size = "9px"
        else:
            _show_lbl = ""
            _lbl_col  = "#94a3b8"
            _lbl_size = "9px"

        _x_lbl = (
            f'<div style="font-size:{_lbl_size};font-weight:{"700" if _is_month else "600"};'
            f'color:{_lbl_col};margin-top:5px;white-space:nowrap;'
            f'overflow:hidden;text-overflow:ellipsis;max-width:100%">'
            + _show_lbl
            + '</div>'
        )
        # bar colour: month-marker column gets a subtle highlight
        _bar_bg = (
            f'background:linear-gradient(180deg,{color} 0%,{color}88 100%)'
        )
        bars += (
            f'<div style="flex:1;display:flex;flex-direction:column;align-items:center;min-width:0">'
            + _val_lbl
            + f'<div style="width:100%;height:{BAR_H}px;display:flex;align-items:flex-end;justify-content:center">'
            f'<div style="width:80%;height:{h}px;{_bar_bg};'
            f'border-radius:3px 3px 0 0;min-height:{"3" if v > 0 else "0"}px"></div>'
            f'</div>'
            + _x_lbl
            + '</div>'
        )
    return (
        f'<div style="display:flex;align-items:flex-end;gap:3px;padding:16px 14px 10px;'
        f'background:rgba(0,0,0,.18);border:1px solid #1a3d2b;border-radius:6px">{bars}</div>'
    )


def html_line_chart(labels, values, color: str = "#10b981", money_fmt: bool = False,
                    height: int = 180) -> str:
    """Render a bold SVG line chart — no altair dependency.

    height: viewBox height in px (default 180; pass 240+ for full-width stacked charts).
    """
    vals  = [float(v or 0) for v in values]
    scale = max(vals) if vals and max(vals) > 0 else 1
    n     = len(vals)

    def _fmt(v: float) -> str:
        if money_fmt:
            return f"₦{v/1_000_000:.1f}m" if v >= 1_000_000 else f"₦{v/1_000:.0f}k" if v >= 1000 else f"₦{v:.0f}"
        return f"{v:,.0f}"

    SVG_W = 900          # wider viewBox for full-width use
    SVG_H = height
    PL = 14; PR = 14
    PT = 32              # top  — room for value labels
    PB = 40              # bottom — room for date labels
    pw = SVG_W - PL - PR
    ph = SVG_H - PT - PB

    def _x(i): return PL + (pw / (n - 1) * i if n > 1 else pw / 2)
    def _y(v): return PT + ph - (v / scale * ph)

    pts  = " ".join(f"{_x(i):.1f},{_y(v):.1f}" for i, v in enumerate(vals))
    area = f"{_x(0):.1f},{PT+ph:.1f} {pts} {_x(n-1):.1f},{PT+ph:.1f}"

    # grid lines at 25 / 50 / 75 %
    grid = "".join(
        f'<line x1="{PL}" y1="{PT + ph*(1-f):.1f}" x2="{SVG_W-PR}" y2="{PT + ph*(1-f):.1f}" '
        f'stroke="#1a3d2b" stroke-width="1" stroke-dasharray="5 4"/>'
        for f in [0.25, 0.5, 0.75]
    )

    # decide label density: show every Nth value label to avoid crowding
    _show_every = max(1, n // 20)   # ~20 value labels max regardless of density

    dots = val_labels = lbl_els = ""
    for i, (lbl, v) in enumerate(zip(labels, vals)):
        cx = _x(i); cy = _y(v)

        # dot size scales slightly with height
        r_outer = 6 if height >= 220 else 5
        r_inner = 3 if height >= 220 else 2.5
        sw      = 3  if height >= 220 else 2.5

        dots += (
            f'<circle cx="{cx:.1f}" cy="{cy:.1f}" r="{r_outer}" fill="#0d1f17" '
            f'stroke="{color}" stroke-width="{sw}"/>'
            f'<circle cx="{cx:.1f}" cy="{cy:.1f}" r="{r_inner}" fill="{color}"/>'
        )

        # value label — only on non-zero peaks and every Nth point to stay readable
        if (i % _show_every == 0 or i == n - 1) and v > 0:
            val_labels += (
                f'<text x="{cx:.1f}" y="{cy-11:.1f}" text-anchor="middle" '
                f'font-size="10" font-weight="600" fill="{color}" '
                f'font-family="Space Mono,monospace">{_fmt(v)}</text>'
            )

        # x-axis label — every Nth to avoid overlap on dense charts
        if i % _show_every == 0 or i == n - 1:
            lbl_els += (
                f'<text x="{cx:.1f}" y="{PT+ph+18}" text-anchor="middle" '
                f'font-size="10" font-weight="600" fill="#94a3b8" '
                f'font-family="Space Mono,monospace">{lbl}</text>'
            )

    svg = (
        f'<svg viewBox="0 0 {SVG_W} {SVG_H}" xmlns="http://www.w3.org/2000/svg" '
        f'style="width:100%;height:auto;display:block">'
        + grid
        + f'<polygon points="{area}" fill="{color}18"/>'
        + f'<polyline points="{pts}" fill="none" stroke="{color}" '
        f'stroke-width="4" stroke-linejoin="round" stroke-linecap="round"/>'
        + dots + val_labels + lbl_els
        + '</svg>'
    )
    return (
        f'<div style="padding:16px 14px 10px;background:rgba(0,0,0,.18);'
        f'border:1px solid #1a3d2b;border-radius:6px">{svg}</div>'
    )


def card(label: str, value: str, cls: str = "") -> str:
    return (f'<div class="sel-card{"highlight" if cls=="_h" else ""}" style="margin-bottom:8px">'
            f'<div class="sel-label">{label}</div>'
            f'<div class="sel-value {cls}">{value}</div></div>')

def section(title: str) -> str:
    return f'<div class="sel-section-title">{title}</div>'


# ── Excel Export Helper ────────────────────────────────────────────────────────
def generate_xlsx(rows: list[dict], result: dict | None = None,
                  account_name: str = "", bank: str = "",
                  params: dict | None = None,
                  officer: str = "") -> bytes:
    """Generate a formatted .xlsx with monthly breakdown + optional eligibility sheet."""
    from io import BytesIO
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()

    # ── Palette ──────────────────────────────────────────────────────────────
    DARK   = "0D2818"
    MID    = "0F1A15"
    LIGHT  = "162019"
    ACCENT = "10B981"
    GREEN  = "34D399"
    GOLD   = "F59E0B"
    ORANGE = "FB923C"
    PURPLE = "A78BFA"
    MUTED  = "64748B"
    WHITE  = "E2E8F0"
    RED    = "F87171"

    def hdr_font(color=WHITE):  return Font(name="Calibri", bold=True, color=color, size=10)
    def body_font(color=WHITE): return Font(name="Calibri", color=color, size=10)
    def bold_font(color=WHITE): return Font(name="Calibri", bold=True, color=color, size=10)
    def fill(hex_color):        return PatternFill("solid", fgColor=hex_color)
    def center():               return Alignment(horizontal="center", vertical="center")
    def right():                return Alignment(horizontal="right", vertical="center")
    def left():                 return Alignment(horizontal="left",  vertical="center")
    def thin_border():
        s = Side(style="thin", color="1A3D2B")
        return Border(bottom=s)

    today_ym = datetime.date.today().strftime("%Y-%m")
    display_rows = [r for r in rows if r.get("ym","") < today_ym and r.get("gross",0) > 0][-12:]

    # ── Sheet 1: Monthly Breakdown ────────────────────────────────────────────
    ws = wb.active
    ws.title = "Monthly Breakdown"
    ws.sheet_properties.tabColor = ACCENT

    # Info rows
    ws["A1"] = "PARSIO"
    ws["A1"].font = Font(name="Calibri", bold=True, color=ACCENT, size=13)
    _off_str = f"   |   Assessed by: {officer}" if officer else ""
    ws["A2"] = f"Account: {account_name}   |   Bank: {bank}   |   Generated: {datetime.date.today()}{_off_str}"
    ws["A2"].font = body_font(MUTED)
    ws.merge_cells("A1:H1")
    ws.merge_cells("A2:H2")

    # Column headers
    has_self  = any(r.get("self_transfer",0)  > 0 for r in display_rows)
    has_rev   = any(r.get("reversal",0)       > 0 for r in display_rows)
    has_nb    = any(r.get("non_business",0)   > 0 for r in display_rows)
    has_loan  = any(r.get("loan_disbursal",0) > 0 for r in display_rows)

    base_cols = ["Month", "Total Inflow (NGN)"]
    col_colors = [WHITE, GREEN]
    if has_self:  base_cols.append("Self Deposits (NGN)");  col_colors.append(ORANGE)
    if has_rev:   base_cols.append("Reversals (NGN)");      col_colors.append(PURPLE)
    if has_nb:    base_cols.append("Non-Business (NGN)");   col_colors.append(MUTED)
    if has_loan:  base_cols.append("Loan Disbursals (NGN)");col_colors.append(GOLD)
    base_cols.append("Eligible Income (NGN)")
    col_colors.append(ACCENT)

    HDR_ROW = 4
    for ci, (colname, colclr) in enumerate(zip(base_cols, col_colors), 1):
        cell = ws.cell(row=HDR_ROW, column=ci, value=colname)
        cell.fill = fill(DARK)
        cell.font = hdr_font(colclr)
        cell.alignment = right() if ci > 1 else left()
        cell.border = thin_border()
        ws.column_dimensions[get_column_letter(ci)].width = 22 if ci > 1 else 12

    # Data rows
    t_gross = t_self = t_rev = t_nb = t_loan = t_net = 0.0
    for ri, r in enumerate(display_rows):
        row_num = HDR_ROW + 1 + ri
        row_fill = fill(MID) if ri % 2 == 0 else fill(LIGHT)
        vals = [r.get("label", r.get("ym","")), r.get("gross",0)]
        if has_self:  vals.append(r.get("self_transfer",0))
        if has_rev:   vals.append(r.get("reversal",0))
        if has_nb:    vals.append(r.get("non_business",0))
        if has_loan:  vals.append(r.get("loan_disbursal",0))
        vals.append(r.get("eligible_income",0))
        t_gross += r.get("gross",0); t_self += r.get("self_transfer",0)
        t_rev   += r.get("reversal",0); t_nb += r.get("non_business",0)
        t_loan  += r.get("loan_disbursal",0); t_net += r.get("eligible_income",0)
        for ci, (val, clr) in enumerate(zip(vals, col_colors), 1):
            cell = ws.cell(row=row_num, column=ci, value=val)
            cell.fill = row_fill
            cell.font = body_font(clr)
            cell.alignment = right() if ci > 1 else left()
            if ci > 1 and isinstance(val, (int, float)):
                cell.number_format = '#,##0.00'

    # Totals row
    tot_row = HDR_ROW + 1 + len(display_rows)
    tot_vals = ["TOTAL", t_gross]
    if has_self:  tot_vals.append(t_self)
    if has_rev:   tot_vals.append(t_rev)
    if has_nb:    tot_vals.append(t_nb)
    if has_loan:  tot_vals.append(t_loan)
    tot_vals.append(t_net)
    for ci, (val, clr) in enumerate(zip(tot_vals, col_colors), 1):
        cell = ws.cell(row=tot_row, column=ci, value=val)
        cell.fill = fill("0A2E1F")
        cell.font = bold_font(ACCENT if ci == len(tot_vals) else clr)
        cell.alignment = right() if ci > 1 else left()
        if ci > 1 and isinstance(val, (int, float)):
            cell.number_format = '#,##0.00'

    # ── Sheet 2: Eligibility Summary (if result provided) ──────────────────
    if result:
        ws2 = wb.create_sheet("Eligibility Summary")
        ws2.sheet_properties.tabColor = GOLD
        _tenor_v = result.get("tenor")

        # ── Loan Parameters section ─────────────────────────────────────────
        ws2["A1"] = "Eligibility Summary"
        ws2["A1"].font = Font(name="Calibri", bold=True, color=ACCENT, size=13)
        ws2.column_dimensions["A"].width = 30
        ws2.column_dimensions["B"].width = 22

        _p = params or {}
        _lp_pairs = [
            ("LOAN PARAMETERS", ""),
            ("Assessed By",           officer or "—"),
            ("Location",              _p.get("location", "—")),
            ("Product Type",          _p.get("product_type", "—")),
            ("Tenor (Months)",        _p.get("tenor", "—")),
            ("Other Monthly Repayments", f"NGN {_p.get('other_loans', 0):,.2f}"),
        ]
        if _p.get("req_loan", 0) > 0:
            _lp_pairs.append(("Requested Loan Amount", f"NGN {_p['req_loan']:,.2f}"))
        if _p.get("manual_rate", 0) > 0:
            _lp_pairs.append(("Manual Interest Rate", f"{_p['manual_rate']:.2f}%"))

        _lp_pairs.append(("", ""))  # spacer

        _elig_pairs = [
            ("ELIGIBILITY RESULT", ""),
            ("Decision",             "Approved" if result.get("approved") else "Below Minimum"),
            ("Max Loan Amount",      result.get("max_loan", 0)),
            ("Applicable Turnover",  result.get("applicable_turnover", 0)),
            ("Total Eligible Net",   result.get("total_net", 0)),
            ("DTI",                  f"{result.get('dti',0)*100:.2f}%"),
            ("Interest Rate",        f"{(result.get('interest_rate') or 0)*100:.2f}%"),
            ("Repayment Frequency",  result.get("repayment_frequency","")),
            ("Max Repayment/Period", result.get("max_repayment_display", 0)),
            ("Max Total Repayment",  result.get("max_total_repayment", 0)),
        ]

        # Requested loan analysis rows (only when present)
        _req_pairs = []
        if _p.get("req_loan", 0) > 0 and "requested" in result:
            _rq = result["requested"]
            _rq_within = _rq.get("within_max", False)
            _rq_diff = abs(_p["req_loan"] - result.get("max_loan", 0))
            _rq_sign = "+" if _p["req_loan"] >= result.get("max_loan", 0) else "-"
            _req_pairs = [
                ("", ""),
                ("REQUESTED LOAN ANALYSIS", ""),
                ("Requested Amount",   f"NGN {_p['req_loan']:,.2f}"),
                ("Interest Rate",      f"{(_rq.get('rate') or 0)*100:.2f}%"),
                ("Repayment / Period", f"NGN {_rq.get('repayment', 0):,.2f}"),
                ("DTI for Requested",  f"{(_rq.get('dti') or 0)*100:.2f}%"),
                ("vs Max Loan",        f"{_rq_sign}NGN {_rq_diff:,.2f}"),
                ("Status",             "Below max — eligible" if _rq_within else "Above max — not eligible"),
            ]

        pairs = _lp_pairs + _elig_pairs + _req_pairs

        for ri2, (label, val) in enumerate(pairs, 3):
            lc = ws2.cell(row=ri2, column=1, value=label)
            vc = ws2.cell(row=ri2, column=2, value=val)
            # Section headers
            if label in ("LOAN PARAMETERS", "ELIGIBILITY RESULT", "REQUESTED LOAN ANALYSIS"):
                lc.font = Font(name="Calibri", bold=True, color=GOLD, size=10)
                lc.fill = fill(DARK)
                vc.fill = fill(DARK)
                ws2.merge_cells(f"A{ri2}:B{ri2}")
                continue
            if label == "" and val == "":
                continue  # spacer row
            lc.font = body_font(MUTED); lc.fill = fill(MID)
            row_clr = GREEN if (label == "Decision" and result.get("approved")) else (RED if label == "Decision" else WHITE)
            vc.font = bold_font(row_clr); vc.fill = fill(LIGHT)
            vc.alignment = right()
            if isinstance(val, (int, float)) and val > 1:
                vc.number_format = '#,##0.00'

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── What-if Reverse Calculator ────────────────────────────────────────────────
def required_income_for_loan(target_loan: float, tenor: int,
                              location: str, product_type: str,
                              other_loans: float = 0,
                              manual_rate_pct: float | None = None) -> dict:
    """Reverse-calculate the monthly income needed to qualify for target_loan."""
    rate = (manual_rate_pct / 100) if (manual_rate_pct and manual_rate_pct > 0) \
           else get_interest_rate(target_loan, location, product_type)
    if not rate:
        return {"ok": False, "reason": "Rate unavailable for this amount/product combo"}
    min_loan, max_loan = loan_limits(location, product_type)
    if target_loan < min_loan or target_loan > max_loan:
        return {"ok": False, "reason": f"Target ₦{target_loan:,.0f} outside product range {money(min_loan)}–{money(max_loan)}"}
    pmt = target_loan * rate / (1 - math.pow(1 + rate, -tenor))
    # Use a safe total_net=5M to get non-zero DTI (avoids 0.0 from RENEWAL <801k edge case)
    dti = get_dti(total_net=5_000_000, product_type=product_type, location=location)
    if not dti:
        return {"ok": False, "reason": "DTI is zero for this product — check product/location combo"}
    required_turnover = (pmt + other_loans) / dti
    freq = "Weekly (est.)" if required_turnover >= 200_000 else "Monthly"
    return {
        "ok": True, "rate": rate, "pmt": pmt, "dti": dti,
        "required_turnover": required_turnover,
        "repayment_frequency": freq,
    }


# ── Session state init ────────────────────────────────────────────────────────
for key in ["buckets_a","summary_a","bank_a","name_a","account_no_a",
            "buckets_b","summary_b","bank_b","name_b","account_no_b",
            "credit_data","rows_a","rows_b","txns_a","txns_b",
            "debit_txns_a","debit_txns_b",
            "last_calc_params", "batch_results", "batch_details", "last_share"]:
    if key not in st.session_state:
        st.session_state[key] = None

# ── Analytics session ID — unique per browser session ─────────────────────────
if "sel_session_id" not in st.session_state:
    st.session_state.sel_session_id = str(uuid.uuid4())[:12]
_SID = st.session_state.sel_session_id

if "assessment_count" not in st.session_state:
    st.session_state.assessment_count = 0

if "officer_name" not in st.session_state:
    # Restore from URL query param so browser refresh keeps the name
    st.session_state.officer_name = st.query_params.get("officer", "")

# ════════════════════════════════════════════════════════════════════════════
# DAILY SIGN-IN GATE
# Shows a full-screen welcome page once per day per browser session.
# The officer enters their name and clicks Sign In.  After that the full
# tool is revealed.  The signed-in state survives browser refreshes
# (stored in URL query params: ?officer=Name&signed=YYYY-MM-DD).
# ════════════════════════════════════════════════════════════════════════════
_today_iso   = now_lagos().date().isoformat()
_signed_date = st.query_params.get("signed", "")
_signed_name = st.query_params.get("officer", "").strip()
_is_signed_in = (_signed_date == _today_iso and bool(_signed_name))

if _is_signed_in:
    # Sync authoritative name into session state
    st.session_state.officer_name = _signed_name

if not _is_signed_in:
    # ── Build time-of-day greeting ─────────────────────────────────────────
    _sh = now_lagos().hour
    if   5  <= _sh < 12: _sw, _sc, _sicon = "Good morning",   "#34d399", "☀️"
    elif 12 <= _sh < 17: _sw, _sc, _sicon = "Good afternoon", "#fbbf24", "🌤"
    elif 17 <= _sh < 21: _sw, _sc, _sicon = "Good evening",   "#f59e0b", "🌆"
    else:                 _sw, _sc, _sicon = "Welcome",        "#a78bfa", "🌙"

    # ── Full-screen sign-in page (hides normal Streamlit chrome) ───────────
    st.markdown("""
    <style>
      header[data-testid="stHeader"]   { display:none !important; }
      [data-testid="stToolbar"]        { display:none !important; }
      .block-container {
        max-width: 540px !important;
        padding-top: 5vh !important;
        padding-bottom: 0 !important;
        margin: 0 auto !important;
      }
    </style>
    """, unsafe_allow_html=True)

    st.markdown(
        f"""
        <div style="text-align:center;margin-bottom:8px">
          <div style="font-size:16px;letter-spacing:4px;color:#10b981;font-weight:900;
                      text-transform:uppercase;margin-bottom:16px">
            ▶ PARSIO
          </div>
          <div style="font-size:52px;margin-bottom:4px">{_sicon}</div>
          <div style="font-size:38px;font-weight:900;color:{_sc};
                      font-family:'DM Serif Display',serif;margin-bottom:6px">
            {_sw}!
          </div>
          <div style="font-size:26px;font-weight:900;color:#fff;
                      font-family:'DM Serif Display',serif;margin-bottom:4px">
            Credit <em style="color:#10b981;font-style:italic">Intelligence</em> Engine
          </div>
          <div style="font-size:14px;font-weight:600;color:#94a3b8;margin-bottom:28px;letter-spacing:0.5px">
            Powered by Kenechukwu Kvic7™ &nbsp;·&nbsp; All Products &nbsp;·&nbsp; Auto-decisioning
          </div>
        </div>
        <div style="background:#0f1a15;border:1px solid #1a3d2b;border-top:3px solid {_sc};
                    border-radius:6px;padding:28px 32px 24px;box-shadow:0 8px 32px rgba(0,0,0,.4)">
          <div style="font-size:22px;font-weight:900;color:#e2e8f0;margin-bottom:10px;
                      letter-spacing:0.5px">
            PARSIO has a new home
          </div>
          <div style="font-size:15px;font-weight:600;color:#94a3b8;margin-bottom:18px;line-height:1.7">
            Sign-in now happens on the PARSIO portal. Click the button below
            to continue &mdash; you only sign in once per day.
          </div>
        </div>
        """,
        unsafe_allow_html=True,
    )

    # Sign-in now lives on the PARSIO portal (GitHub Pages): visitors land on
    # the landing page and click Sign in from there. The login page hands the
    # officer name back via ?officer=...&signed=... — override per deployment
    # with the PARSIO_LOGIN_URL env var / secret if needed.
    _LOGIN_URL = os.environ.get(
        "PARSIO_LOGIN_URL",
        "https://kvic332.github.io/Produt-Based-SEL-Calculator/",
    )

    # Tell the portal which deployment to return to after sign-in, so a
    # non-production deployment (e.g. the dev branch app) round-trips back to
    # ITSELF instead of production. login.html validates the host.
    try:
        from urllib.parse import urlsplit, quote as _q
        _self = urlsplit(st.context.url)
        if _self.hostname and _self.hostname.endswith(".streamlit.app"):
            _sep = "&" if "?" in _LOGIN_URL else "?"
            _LOGIN_URL = f"{_LOGIN_URL}{_sep}app={_q(f'{_self.scheme}://{_self.hostname}/', safe='')}"
    except Exception:
        pass  # older Streamlit without st.context.url — portal default applies

    # Streamlit Cloud hosts the app in a sandboxed iframe that blocks both
    # auto-redirects (meta refresh) and same-tab/_top navigation out of the
    # frame — but it does allow opening a new tab, so use target="_blank".
    st.markdown(
        f'<div style="text-align:center;margin-top:22px">'
        f'<a href="{_LOGIN_URL}" target="_blank" rel="noopener" '
        f'style="display:inline-block;'
        f'background:#10b981;color:#000;font-weight:900;font-size:15px;'
        f'letter-spacing:1px;padding:14px 36px;border-radius:6px;'
        f'text-decoration:none">GO TO PARSIO PORTAL →</a></div>',
        unsafe_allow_html=True,
    )

    # ── Direct sign-in (no portal round-trip) ─────────────────────────────
    # Keeps non-production deployments self-contained: the dev branch app
    # signs in here and never leaves its own URL. Also serves as the fallback
    # if GitHub Pages is unreachable. Production officers normally use the
    # portal button above.
    st.markdown(
        '<div style="display:flex;align-items:center;gap:12px;margin:24px 0 10px">'
        '<div style="flex:1;border-top:1px solid #1a3d2b"></div>'
        '<div style="font-size:10px;letter-spacing:2px;color:#64748b;text-transform:uppercase">'
        'or sign in directly</div>'
        '<div style="flex:1;border-top:1px solid #1a3d2b"></div></div>',
        unsafe_allow_html=True,
    )
    _si_col1, _si_col2 = st.columns([3, 1])
    with _si_col1:
        _si_name = st.text_input(
            "Officer Name / Staff ID",
            placeholder="e.g. Adaobi Nwosu  or  SEL-042",
            key="signin_name_input",
            label_visibility="collapsed",
        )
    with _si_col2:
        _si_btn = st.button("Sign In →", key="signin_btn", use_container_width=True)
    if _si_btn:
        if not _si_name.strip():
            st.error("Please enter your name or staff ID before signing in.")
        else:
            _clean = _si_name.strip()
            st.session_state.officer_name = _clean
            st.query_params["officer"] = _clean
            st.query_params["signed"]  = _today_iso
            track("signin", session=_SID, officer=_clean, bank="", filename="")
            st.rerun()

    st.markdown(
        f'<div style="margin-top:14px;font-size:16px;font-weight:700;color:#64748b;text-align:center">'
        f'{now_lagos().strftime("%A, %d %B %Y")}</div>',
        unsafe_allow_html=True,
    )
    st.stop()   # ← nothing else renders until the officer has signed in


# ════════════════════════════════════════════════════════════════════════════
# HEADER
# ════════════════════════════════════════════════════════════════════════════
_now_h = now_lagos().hour
if   5  <= _now_h < 12: _greet, _greet_col, _greet_sub = "Good morning ☀️",  "#34d399", "Ready for a productive day of assessments."
elif 12 <= _now_h < 17: _greet, _greet_col, _greet_sub = "Good afternoon 🌤", "#fbbf24", "Keep the momentum going."
elif 17 <= _now_h < 21: _greet, _greet_col, _greet_sub = "Good evening 🌆",   "#f59e0b", "Wrapping up for the day?"
else:                    _greet, _greet_col, _greet_sub = "Working late 🌙",   "#a78bfa", "The dedication doesn't go unnoticed."

st.markdown(f"""
<div style="border-bottom:1px solid #1a3d2b;padding-bottom:24px;margin-bottom:32px">
  <div style="display:flex;align-items:center;justify-content:space-between;margin-bottom:8px">
    <div style="font-size:10px;letter-spacing:4px;color:#10b981;text-transform:uppercase">▶ PARSIO</div>
    <a href="https://kvic332.github.io/Produt-Based-SEL-Calculator" target="_blank"
       style="font-size:10px;letter-spacing:2px;color:#10b981;text-transform:uppercase;
              text-decoration:none;border:1px solid rgba(16,185,129,.35);
              padding:5px 14px;border-radius:3px;
              background:rgba(16,185,129,.06);
              transition:background .2s"
       onmouseover="this.style.background='rgba(16,185,129,.14)'"
       onmouseout="this.style.background='rgba(16,185,129,.06)'">
      About PARSIO ↗
    </a>
  </div>
  <h1 style="font-family:DM Serif Display,serif;font-size:clamp(28px,4vw,44px);color:#fff;line-height:1.1">
    PARSIO <em style="color:#10b981;font-style:italic">Credit</em><br>Intelligence
  </h1>
  <div style="font-size:13px;color:#cbd5e1;margin-top:6px;font-weight:700;letter-spacing:0.5px">
    All Products &nbsp;|&nbsp; Auto-computes DTI, Repayment, Turnover &amp; Loan Amount &nbsp;|&nbsp; Recycling Detection
  </div>
  <div style="margin-top:16px;display:inline-flex;align-items:center;gap:14px;
              padding:10px 22px;border-radius:4px;
              background:rgba(255,255,255,.04);
              border:1px solid {_greet_col}55;
              border-left:4px solid {_greet_col}">
    <span style="font-size:22px;line-height:1">{_greet.split()[2]}</span>
    <div>
      <div style="font-size:24px;font-weight:900;color:{_greet_col};letter-spacing:0.5px;line-height:1.2">
        {" ".join(_greet.split()[:2])}
      </div>
      <div style="font-size:13px;color:#94a3b8;margin-top:3px;font-weight:600">{_greet_sub}</div>
    </div>
    <div style="border-left:1px solid #1a3d2b;padding-left:14px;text-align:center">
      <div style="font-family:'Space Mono',monospace;font-size:28px;font-weight:700;
                  color:{_greet_col};letter-spacing:3px;line-height:1">
        {now_lagos().strftime("%H:%M:%S")}
      </div>
      <div style="font-size:9px;letter-spacing:2px;color:#64748b;text-transform:uppercase;margin-top:4px">
        Session Time
      </div>
    </div>
  </div>
</div>
""", unsafe_allow_html=True)


# ── HTML component wrapper ────────────────────────────────────────────────────
# st.components.v1.html was removed after 2026-06-01. Use st.html() instead.
# st.html() does not execute scripts, so JS decorations (clock, badge, quotes,
# confetti) become no-ops — that's fine; they're non-essential.
# For the share-buttons component (height > 0), the HTML renders but onclick
# handlers are inert; users can still download via the Streamlit download button.
def _html(html: str, height: int = 0) -> None:
    try:
        if height > 0:
            st.html(f'<div style="min-height:{height}px">{html}</div>')
        # height==0 → pure JS decoration, skip silently
    except Exception:
        pass

# ── JS decorations removed ────────────────────────────────────────────────────
# st.components.v1.html removed after 2026-06-01; st.html() does not execute
# scripts. Clock is now Python-rendered in the header above. Badge and quote
# toast are rendered as static HTML below.

# ── Static quote (replaces JS toast — st.components.v1.html removed 2026-06-01) ─
import random as _random
_QUOTES = [
    ("Risk comes from not knowing what you are doing.", "Warren Buffett"),
    ("An investment in knowledge pays the best interest.", "Benjamin Franklin"),
    ("Every naira lent wisely builds a stronger Nigeria.", "PARSIO"),
    ("Credit is not given. It is earned through trust and consistency.", "PARSIO"),
    ("The goal of a good credit officer is not to say no — it is to find the right yes.", "PARSIO"),
    ("Data tells the story. Your judgment writes the ending.", "PARSIO"),
    ("Diligence is the mother of good fortune.", "Miguel de Cervantes"),
    ("Small daily improvements are the key to staggering long-term results.", "Robin Sharma"),
    ("Every loan decision shapes a family's future. Make it count.", "PARSIO"),
    ("Price is what you pay. Value is what you get.", "Warren Buffett"),
    ("In credit, character is the first C for a reason.", "PARSIO"),
    ("The secret of getting ahead is getting started.", "Mark Twain"),
    ("Africa is not poor. It is poorly managed.", "Fela Durotoye"),
    ("The more you learn, the more you earn.", "Warren Buffett"),
    ("A good decision is based on knowledge, not on numbers.", "Plato"),
    ("The best time to assess a loan was yesterday. The second best time is now.", "PARSIO"),
    ("Opportunities do not go away. They pass to someone else.", "PARSIO"),
    ("Champions keep playing until they get it right.", "Billie Jean King"),
    ("Do not save what is left after spending — spend what is left after saving.", "Warren Buffett"),
    ("Precision in underwriting protects both lender and borrower.", "PARSIO"),
]
_q, _a = _QUOTES[hash(str(datetime.date.today())) % len(_QUOTES)]
st.markdown(
    f'<div style="background:#0f1a15;border-left:4px solid #10b981;border-radius:4px;'
    f'padding:10px 14px;margin-bottom:8px">'
    f'<div style="font-size:9px;letter-spacing:3px;color:#10b981;text-transform:uppercase;margin-bottom:4px">✦ Thought for the day</div>'
    f'<div style="font-size:12px;color:#e2e8f0;font-style:italic;line-height:1.5">"{_q}"</div>'
    f'<div style="font-size:10px;color:#64748b;text-align:right;margin-top:4px">— {_a}</div>'
    f'<div style="font-size:9px;letter-spacing:2px;color:#374151;margin-top:6px;text-transform:uppercase">'
    f'Powered by Kenechukwu Kvic7 ™</div>'
    f'</div>',
    unsafe_allow_html=True,
)


# ── Officer status bar (name is locked in from sign-in gate) ─────────────────
_OFFICER = st.session_state.officer_name or "Unknown Officer"
_off_sb1, _off_sb2 = st.columns([4, 1])
with _off_sb1:
    st.markdown(
        f'<div style="display:flex;align-items:center;gap:10px;margin-bottom:4px">'
        f'<div style="font-size:11px;letter-spacing:2px;color:#64748b;text-transform:uppercase">Signed in as</div>'
        f'<div style="font-size:13px;font-weight:800;color:#10b981;letter-spacing:0.5px">'
        f'👤 {_OFFICER}</div>'
        f'<div style="font-size:10px;color:#374151">'
        f'· {now_lagos().strftime("%A, %d %b %Y")}</div>'
        f'</div>',
        unsafe_allow_html=True,
    )
with _off_sb2:
    if st.button("🔄 Switch Officer", key="switch_officer", use_container_width=True):
        # Clear sign-in state → sign-in page will show on next rerun
        st.query_params.pop("signed", None)
        st.query_params.pop("officer", None)
        st.session_state.officer_name = ""
        st.rerun()

st.markdown("---")

# ════════════════════════════════════════════════════════════════════════════
# PRODUCT SELECTOR
# ════════════════════════════════════════════════════════════════════════════
st.markdown(
    '<div style="font-size:15px;font-weight:800;color:#e2e8f0;letter-spacing:1px;'
    'text-transform:uppercase;margin-bottom:6px">Product</div>',
    unsafe_allow_html=True,
)
_product = st.radio(
    "Product",
    options=["SEL", "SME"],
    horizontal=True,
    key="product",
    help="SEL uses 6 months of bank statement data. SME uses 12 months.",
)
N_MONTHS = 6 if _product == "SEL" else 12
st.markdown(
    f'<div style="font-size:13px;font-weight:700;color:#34d399;margin:-8px 0 24px 0;">'
    f'{"▶ 6-month analysis window" if _product == "SEL" else "▶ 12-month analysis window"}'
    f'</div>',
    unsafe_allow_html=True,
)


# ════════════════════════════════════════════════════════════════════════════
# BATCH PROCESSING MODE (Feature 7)
# ════════════════════════════════════════════════════════════════════════════
with st.expander("⚡  Batch Processing — Assess multiple applicants at once", expanded=False):
    st.markdown(
        '<div style="font-size:12px;color:#cbd5e1;margin-bottom:12px;line-height:1.7">'
        'Upload up to 10 PDF bank statements at once. Each will be parsed and assessed '
        'using the shared loan parameters below. Results export as a single Excel sheet.</div>',
        unsafe_allow_html=True,
    )
    _bp_files = st.file_uploader(
        "Upload statements (PDF or Excel)",
        type=["pdf","xlsx","xls"],
        accept_multiple_files=True,
        key="batch_upload",
    )
    _bp_pw = st.text_input(
        "Shared PDF Password (leave blank if none)", type="password", key="batch_pw"
    )
    _bc1, _bc2, _bc3 = st.columns(3)
    with _bc1: _bp_loc  = st.selectbox("Location",     ["Lagos","Outside Lagos","Expansion"], key="batch_loc")
    with _bc2: _bp_prod = st.selectbox("Product Type", ["NTB","RENEWAL","TOP-UP"],            key="batch_prod")
    with _bc3: _bp_ten  = st.selectbox("Tenor",        list(range(2, 13)), index=4,           key="batch_tenor")

    # ── Live queue display ────────────────────────────────────────────────
    if _bp_files:
        _q_slots = []
        for _qf in _bp_files[:10]:
            _q_slots.append(st.empty())

        def _render_queue(statuses: list) -> None:
            for _qi, (_qf, _qs) in enumerate(zip(_bp_files[:10], statuses)):
                _qicon = ("⏳" if _qs == "queued" else "🔄" if _qs == "processing"
                          else "✅" if _qs == "approved" else "❌" if _qs == "below_min"
                          else "⚠" if _qs == "no_data" else "🔴")
                _qcol  = ("#94a3b8" if _qs == "queued" else "#fbbf24" if _qs == "processing"
                          else "#34d399" if _qs == "approved" else "#f87171" if _qs == "below_min"
                          else "#fb923c" if _qs == "no_data" else "#ef4444")
                _q_slots[_qi].markdown(
                    f'<div style="display:flex;align-items:center;gap:10px;padding:4px 0;'
                    f'font-size:11px;border-bottom:1px solid rgba(255,255,255,.04)">'
                    f'<span style="color:{_qcol}">{_qicon}</span>'
                    f'<span style="color:#94a3b8;flex:1;overflow:hidden;text-overflow:ellipsis;'
                    f'white-space:nowrap">{_qf.name}</span>'
                    f'<span style="color:{_qcol};font-weight:700;font-size:10px;'
                    f'text-transform:uppercase">{_qs.replace("_"," ")}</span>'
                    f'</div>',
                    unsafe_allow_html=True,
                )

        _statuses = ["queued"] * len(_bp_files[:10])
        _render_queue(_statuses)

    if st.button("▶  Run Batch Assessment", key="btn_batch", use_container_width=True):
        if not _bp_files:
            st.error("Please upload at least one statement.")
        else:
            _bp_rows = []
            _bp_details = []
            _bp_bar  = st.progress(0, text="Processing…")
            _statuses = ["queued"] * len(_bp_files[:10])
            for _bfi, _bpf in enumerate(_bp_files[:10]):
                _statuses[_bfi] = "processing"
                _render_queue(_statuses)
                _bp_bar.progress(_bfi / len(_bp_files[:10]), text=f"Processing {_bpf.name} ({_bfi+1}/{len(_bp_files[:10])})…")
                try:
                    _bk, _bsumm, _bbank, _bname, _btxns, _bdebits = parse_transactions(
                        _bpf.getvalue(), _bp_pw, filename=_bpf.name
                    )
                    _brows = monthly_analysis(_bk, _bsumm)

                    # ── Per-applicant insight detail (credit/debit/verification) ──
                    _bd_total   = sum(t["amount"] for t in _bdebits)
                    _bd_pri     = {"loan_repayment", "credit_card", "rent"}
                    _bd_flagged = [t for t in _bdebits if t["category"] in _bd_pri]
                    _bd_loans   = [t for t in _bdebits if t["category"] == "loan_repayment"]
                    try:
                        _b_stated  = extract_stated_totals(get_last_full_text())
                        _b_verif   = verify_extraction_accuracy(_bk, _b_stated)
                    except Exception:
                        _b_verif   = {"matched": None, "message": ""}
                    _bp_details.append({
                        "file":        _bpf.name,
                        "name":        _bname or _bpf.name,
                        "bank":        _bbank or "—",
                        "months":      _brows,
                        "total_credits":  sum(r["parsed_gross"] for r in _brows),
                        "total_eligible": sum(r["eligible_income"] for r in _brows),
                        "total_debits":   _bd_total,
                        "debit_count":    len(_bdebits),
                        "loan_repay_sum": sum(t["amount"] for t in _bd_loans),
                        "loan_repay_ct":  len(_bd_loans),
                        "rent_sum":       sum(t["amount"] for t in _bdebits if t["category"] == "rent"),
                        "flagged":        sorted(_bd_flagged, key=lambda t: -t["amount"])[:15],
                        "verified":       _b_verif.get("matched"),
                        "verify_msg":     _b_verif.get("message", ""),
                    })
                    _today_ym_b = datetime.date.today().strftime("%Y-%m")
                    _b_valid = [r for r in _brows if r["ym"] < _today_ym_b and r["gross"] > 0][-N_MONTHS:]
                    if _b_valid:
                        _b_nets   = [r["eligible_income"] for r in _b_valid]
                        _b_counts = [r["count"] for r in _b_valid]
                        _b_res    = calculate_eligibility(
                            nets=_b_nets, counts=_b_counts,
                            location=_bp_loc, product_type=_bp_prod, tenor=_bp_ten,
                            sel_mode=(_product == "SEL"),
                        )
                        _b_avg    = sum(_b_nets) / len(_b_nets)
                        _b_appr   = _b_res.get("approved", False)
                        # Blacklist check for batch
                        _b_acct   = extract_account_no(_bpf.getvalue().decode("latin-1", errors="ignore")[:5000]) if not _bpf.name.lower().endswith((".xlsx",".xls")) else ""
                        _b_bl     = check_blacklist(_bname or "", _b_acct)
                        _b_flag   = f"⚠ WATCHLIST: {_b_bl[0].get('value','')}" if _b_bl else ""
                        _statuses[_bfi] = "approved" if _b_appr else "below_min"
                        _bp_rows.append({
                            "Name":          _bname or "—",
                            "Bank":          _bbank or "—",
                            "Months":        len(_b_valid),
                            "Avg Income":    round(_b_avg),
                            "Max Loan":      _b_res.get("max_loan", 0),
                            "Rate":          f"{(_b_res.get('interest_rate') or 0)*100:.2f}%",
                            "Tenor":         f"{_bp_ten} mo",
                            "Repayment":     round(_b_res.get("max_repayment_display", 0)),
                            "Frequency":     _b_res.get("repayment_frequency", "—"),
                            "Decision":      "Approved" if _b_appr else "Below Min",
                            "Watchlist":     _b_flag,
                            "Total Credits": round(_bp_details[-1]["total_credits"]),
                            "Total Debits":  round(_bp_details[-1]["total_debits"]),
                            "Loan Repayments": round(_bp_details[-1]["loan_repay_sum"]),
                            "Verified":      ("✓" if _bp_details[-1]["verified"]
                                              else "✗" if _bp_details[-1]["verified"] is False else "—"),
                            "File":          _bpf.name,
                        })
                    else:
                        _statuses[_bfi] = "no_data"
                        _bp_rows.append({
                            "Name": _bname or _bpf.name, "Bank": _bbank or "—",
                            "Months": 0, "Avg Income": 0, "Max Loan": 0,
                            "Rate": "—", "Tenor": f"{_bp_ten} mo", "Repayment": 0,
                            "Frequency": "—", "Decision": "No data",
                            "Watchlist": "",
                            "Total Credits": round(_bp_details[-1]["total_credits"]),
                            "Total Debits":  round(_bp_details[-1]["total_debits"]),
                            "Loan Repayments": round(_bp_details[-1]["loan_repay_sum"]),
                            "Verified": ("✓" if _bp_details[-1]["verified"]
                                         else "✗" if _bp_details[-1]["verified"] is False else "—"),
                            "File": _bpf.name,
                        })
                except Exception as _be:
                    _statuses[_bfi] = "error"
                    _bp_rows.append({
                        "Name": _bpf.name, "Bank": "—", "Months": 0, "Avg Income": 0,
                        "Max Loan": 0, "Rate": "—", "Tenor": f"{_bp_ten} mo",
                        "Repayment": 0, "Frequency": "—",
                        "Decision": f"Error: {str(_be)[:40]}",
                        "Watchlist": "",
                        "Total Credits": 0, "Total Debits": 0,
                        "Loan Repayments": 0, "Verified": "—",
                        "File": _bpf.name,
                    })
                _render_queue(_statuses)
            _bp_bar.progress(1.0, text=f"Done — {len(_bp_files[:10])} files processed.")
            st.session_state.batch_results = _bp_rows
            st.session_state.batch_details = _bp_details

    if st.session_state.batch_results:
        _bdf = pd.DataFrame(st.session_state.batch_results)
        _approved_ct = sum(1 for r in st.session_state.batch_results if r["Decision"] == "Approved")
        st.markdown(
            f'<div style="font-size:11px;color:#64748b;margin:8px 0">'
            f'<span style="color:#34d399;font-weight:700">{_approved_ct} approved</span>'
            f' / {len(st.session_state.batch_results)} assessed</div>',
            unsafe_allow_html=True,
        )
        st.dataframe(
            _bdf.style.apply(
                lambda col: ["color:#34d399" if v == "Approved" else "color:#f87171" if "Error" in str(v) else "" for v in col],
                subset=["Decision"],
            ),
            hide_index=True,
            use_container_width=True,
            column_config={
                "Avg Income":  st.column_config.NumberColumn("Avg Income", format="₦%d"),
                "Max Loan":    st.column_config.NumberColumn("Max Loan",   format="₦%d"),
                "Repayment":   st.column_config.NumberColumn("Repayment",  format="₦%d"),
            },
        )
        # Excel export
        import io as _bio
        _bxl = _bio.BytesIO()
        _bdf.to_excel(_bxl, index=False, sheet_name="Batch Results")
        _bxl.seek(0)
        st.download_button(
            "⬇  Download Batch Results (Excel)",
            _bxl.getvalue(),
            file_name=f"PARSIO_Batch_{datetime.date.today():%Y%m%d}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="dl_batch",
            use_container_width=True,
        )

        # ── Per-applicant credit/debit insight (officer visibility) ──────────
        if st.session_state.get("batch_details"):
            st.markdown(
                '<div style="margin-top:18px;margin-bottom:4px;font-size:12px;font-weight:700;'
                'letter-spacing:2px;color:#10b981;text-transform:uppercase">'
                '🔍 Applicant Insight — Credits &amp; Debits'
                '<span style="font-size:10px;font-weight:400;color:#94a3b8;margin-left:10px;'
                'text-transform:none;letter-spacing:0">Officer view only · '
                'Upload individually for full transaction detail</span></div>',
                unsafe_allow_html=True,
            )
            for _bdi, _bd in enumerate(st.session_state.batch_details):
                _bd_ver_icon = ("✅ Verified" if _bd["verified"]
                                else "⚠ Mismatch" if _bd["verified"] is False
                                else "— No stated total")
                with st.expander(f"💳  {_bd['name']}  ·  {_bd['bank']}  ·  {_bd_ver_icon}",
                                 expanded=False):
                    # Summary cards: credits | eligible | debits | loan repayments
                    _bic1, _bic2, _bic3, _bic4 = st.columns(4)
                    with _bic1:
                        st.markdown(
                            f'<div style="background:rgba(52,211,153,.08);border:1px solid rgba(52,211,153,.3);'
                            f'border-radius:10px;padding:14px 16px">'
                            f'<div style="font-size:9px;letter-spacing:2px;color:#6ee7b7;text-transform:uppercase;'
                            f'font-weight:600;margin-bottom:6px">Total Credits</div>'
                            f'<div style="font-size:16px;font-weight:800;color:#34d399">{money(_bd["total_credits"])}</div></div>',
                            unsafe_allow_html=True,
                        )
                    with _bic2:
                        st.markdown(
                            f'<div style="background:rgba(16,185,129,.08);border:1px solid rgba(16,185,129,.3);'
                            f'border-radius:10px;padding:14px 16px">'
                            f'<div style="font-size:9px;letter-spacing:2px;color:#6ee7b7;text-transform:uppercase;'
                            f'font-weight:600;margin-bottom:6px">Eligible Income</div>'
                            f'<div style="font-size:16px;font-weight:800;color:#10b981">{money(_bd["total_eligible"])}</div></div>',
                            unsafe_allow_html=True,
                        )
                    with _bic3:
                        st.markdown(
                            f'<div style="background:rgba(248,113,113,.10);border:1px solid rgba(248,113,113,.35);'
                            f'border-radius:10px;padding:14px 16px">'
                            f'<div style="font-family:Plus Jakarta Sans,sans-serif;font-size:10px;letter-spacing:1.8px;color:#fca5a5;text-transform:uppercase;'
                            f'font-weight:600;margin-bottom:6px">Total Debits</div>'
                            f'<div style="font-size:16px;font-weight:800;color:#f87171">{money(_bd["total_debits"])}</div>'
                            f'<div style="font-size:9px;color:#94a3b8;margin-top:2px">{_bd["debit_count"]} transaction(s)</div></div>',
                            unsafe_allow_html=True,
                        )
                    with _bic4:
                        _bd_lr_col = "#f87171" if _bd["loan_repay_ct"] else "#94a3b8"
                        st.markdown(
                            f'<div style="background:rgba(248,113,113,.10);border:1px solid rgba(248,113,113,.35);'
                            f'border-radius:10px;padding:14px 16px">'
                            f'<div style="font-family:Plus Jakarta Sans,sans-serif;font-size:10px;letter-spacing:1.8px;color:#fca5a5;text-transform:uppercase;'
                            f'font-weight:600;margin-bottom:6px">Loan Repayments</div>'
                            f'<div style="font-size:16px;font-weight:800;color:{_bd_lr_col}">{money(_bd["loan_repay_sum"])}</div>'
                            f'<div style="font-size:9px;color:#94a3b8;margin-top:2px">{_bd["loan_repay_ct"]} flagged</div></div>',
                            unsafe_allow_html=True,
                        )

                    # Monthly credit table: gross / deductions / eligible
                    if _bd["months"]:
                        _bd_rows_html = "".join(
                            f'<tr><td>{r["label"]}</td>'
                            f'<td class="col-gross">{money(r["gross"])}</td>'
                            f'<td class="col-self">{money(r["deductions"])}</td>'
                            f'<td class="col-net">{money(r["eligible_income"])}</td>'
                            f'<td style="color:#94a3b8">{r["count"]}</td></tr>'
                            for r in _bd["months"]
                        )
                        st.markdown(
                            f'<table class="preview-table" style="margin-top:12px">'
                            f'<thead><tr><th>Month</th><th>Gross Credits</th>'
                            f'<th>Deductions</th><th>Eligible Income</th><th>Txns</th></tr></thead>'
                            f'<tbody>{_bd_rows_html}</tbody></table>',
                            unsafe_allow_html=True,
                        )

                    # Flagged debits: loan repayments / credit cards / rent
                    if _bd["flagged"]:
                        _bd_fl_html = "".join(
                            f'<tr><td>{t["ym"]}</td>'
                            f'<td style="color:#e2e8f0;text-align:left">{t["narration"][:60]}</td>'
                            f'<td style="color:#fbbf24">{t["label"]}</td>'
                            f'<td style="color:#f87171;font-weight:700">{money(t["amount"])}</td></tr>'
                            for t in _bd["flagged"]
                        )
                        st.markdown(
                            f'<div style="margin-top:14px;font-size:10px;letter-spacing:2px;color:#f87171;'
                            f'text-transform:uppercase;font-weight:700">🔴 Flagged Debits '
                            f'(loan repay · credit card · rent) — top {len(_bd["flagged"])}</div>'
                            f'<table class="preview-table" style="margin-top:6px">'
                            f'<thead><tr><th>Month</th><th style="text-align:left">Narration</th>'
                            f'<th>Type</th><th>Amount</th></tr></thead>'
                            f'<tbody>{_bd_fl_html}</tbody></table>',
                            unsafe_allow_html=True,
                        )
                    elif _bd["debit_count"]:
                        st.markdown(
                            '<div style="margin-top:12px;font-size:11px;color:#34d399">'
                            '✓ No loan repayments, credit cards or rent detected in debits.</div>',
                            unsafe_allow_html=True,
                        )

                    if _bd["verified"] is False and _bd["verify_msg"]:
                        st.markdown(
                            f'<div style="margin-top:10px;font-size:11px;color:#fbbf24">⚠ {_bd["verify_msg"]}</div>',
                            unsafe_allow_html=True,
                        )

st.markdown("---")

# ════════════════════════════════════════════════════════════════════════════
# SECTION 00 — FIRST BANK STATEMENT
# ════════════════════════════════════════════════════════════════════════════
st.markdown('<div class="sel-section-title">00 — Bank Statement Auto-Fill &nbsp;<span style="color:#94a3b8;font-size:11px">— Optional</span></div>', unsafe_allow_html=True)
st.markdown('<div class="sel-caption">Upload PDF bank statement. Credits are automatically classified into real income vs recycled amounts.</div>', unsafe_allow_html=True)

col1, col2 = st.columns(2)
with col1:
    file_a = st.file_uploader("Upload Bank Statement (PDF or Excel)", type=["pdf","xlsx","xls"], key="upload_a")
with col2:
    pw_a   = st.text_input("PDF Password", type="password", key="pw_a", placeholder="Leave blank if not encrypted")
    if st.button("Extract Monthly Credits", key="btn_extract_a"):
        if not file_a:
            st.error("Please select a PDF file first.")
        else:
            with st.spinner("Extracting..."):
                # ── Read file bytes ONCE — prevents 4-5 redundant BytesIO copies ──
                _pdf_bytes_a = file_a.getvalue()
                _size_mb_a   = len(_pdf_bytes_a) / 1_048_576

                # ── File size guard ──────────────────────────────────────────
                # Hard limit raised to 30 MB — chunked page extraction in
                # extract_pdf_text() now processes 50 pages at a time, which
                # keeps PyPDF2 peak memory manageable for 200–300 page PDFs.
                if _size_mb_a > 30:
                    st.error(
                        f"⚠️ This PDF is {_size_mb_a:.1f} MB — too large to process safely on "
                        f"Streamlit Cloud (limit: 30 MB). Please export a shorter date range "
                        f"(6 months) from your bank portal and upload again."
                    )
                    del _pdf_bytes_a
                    gc.collect()
                else:
                    if _size_mb_a > 10:
                        st.warning(
                            f"Large file ({_size_mb_a:.1f} MB, ~{round(_size_mb_a * 10):.0f} pages). "
                            f"Processing in chunks — this may take up to 60 seconds…"
                        )

                    track("upload", session=_SID, officer=_OFFICER, filename=file_a.name,
                          size_kb=round(_size_mb_a * 1024, 1))
                    try:
                        _pt0 = datetime.datetime.now()
                        buckets, summary, bank, name, txns, debit_txns = parse_transactions(
                            _pdf_bytes_a, pw_a, filename=file_a.name
                        )
                        _parse_secs = (datetime.datetime.now() - _pt0).total_seconds()
                        # Heavy-statement telemetry: lets the admin dashboard
                        # identify statements that strain memory/CPU (the ones
                        # that can OOM the shared Streamlit Cloud worker).
                        track("parsed", session=_SID, officer=_OFFICER, bank=bank,
                              filename=file_a.name,
                              size_kb=round(_size_mb_a * 1024, 1),
                              duration_s=round(_parse_secs, 1),
                              txns=len(txns), debits=len(debit_txns),
                              heavy=bool(_parse_secs > 60 or len(txns) > 5000))
                        rows = monthly_analysis(buckets, summary)
                        st.session_state.buckets_a       = buckets
                        st.session_state.summary_a       = summary
                        st.session_state.bank_a          = bank
                        st.session_state.bank_override_a = bank   # seed manual override
                        st.session_state.name_a          = name
                        st.session_state.rows_a          = rows
                        st.session_state.txns_a          = txns
                        st.session_state.debit_txns_a    = debit_txns

                        # ── Reuse text already extracted by parse_transactions ──
                        # Eliminates the previous pdfplumber + PyPDF2 re-parses
                        # (which each used 100–200 MB for a large PDF).
                        _reused_text_a = get_last_full_text()

                        # Extract account number — Excel: file bytes; PDF: reuse text
                        try:
                            if file_a.name.lower().endswith((".xlsx", ".xls")):
                                st.session_state.account_no_a = extract_account_no_excel(_pdf_bytes_a)
                            else:
                                st.session_state.account_no_a = extract_account_no(_reused_text_a)
                        except Exception:
                            st.session_state.account_no_a = ""

                        st.success(f"Extracted from {bank} statement — {name or 'account holder'}")
                        if _parse_secs > 60 or len(txns) > 5000:
                            st.warning(
                                f"⚠ Heavy statement: {len(txns):,} transactions, parsed in "
                                f"{_parse_secs:.0f}s. Logged for tracking — avoid re-extracting "
                                f"it repeatedly in one session."
                            )

                        # ── Blacklist / Watchlist Check ───────────────────────
                        try:
                            _bl_hits = check_blacklist(
                                name or "",
                                st.session_state.get("account_no_a", ""),
                            )
                            for _bh in _bl_hits:
                                _bh_type = "Account No" if _bh.get("entry_type") == "account_no" else "Name"
                                st.markdown(
                                    f'<div style="background:rgba(239,68,68,.1);border:2px solid #ef4444;'
                                    f'border-radius:4px;padding:10px 16px;margin-top:8px">'
                                    f'<div style="color:#ef4444;font-weight:900;font-size:13px;'
                                    f'letter-spacing:1px">🚫 WATCHLIST MATCH — {_bh_type}</div>'
                                    f'<div style="color:#fca5a5;margin-top:4px;font-size:12px">'
                                    f'<strong>Flagged value:</strong> {_bh.get("value","")}'
                                    + (f'&nbsp;&nbsp;|&nbsp;&nbsp;<strong>Reason:</strong> {_bh.get("reason","")}' if _bh.get("reason") else "")
                                    + f'</div></div>',
                                    unsafe_allow_html=True,
                                )
                        except Exception:
                            pass

                        _txn_count = sum(b.get("count", 0) for b in buckets.values())
                        _gross_tot = sum(b.get("gross", 0) for b in buckets.values())
                        track("parse_success", session=_SID, officer=_OFFICER, bank=bank,
                              filename=file_a.name, txn_count=_txn_count,
                              gross_total=round(_gross_tot, 2),
                              months=len([r for r in rows if r["gross"] > 0]))

                        # ── Accuracy Verification — reuses cached text, no re-parse ──
                        _conf_pct_a = None   # feeds the parser-confidence report below
                        is_pdf = not file_a.name.lower().endswith((".xlsx", ".xls"))
                        if is_pdf and buckets and _reused_text_a:
                            try:
                                stated  = extract_stated_totals(_reused_text_a)
                                verdict = verify_extraction_accuracy(buckets, stated)
                                if verdict["pct_match"] is not None:
                                    _pct_match = verdict["pct_match"]
                                    _conf_pct_a = _pct_match
                                    ext = verdict["extracted"]
                                    stl = verdict["stated_total"]
                                    colour = ("#34d399" if _pct_match >= 95
                                              else "#fb923c" if _pct_match >= 90
                                              else "#f87171")
                                    st.markdown(
                                        f'<div style="background:rgba(0,0,0,.2);border:1px solid {colour}33;'
                                        f'border-radius:3px;padding:10px 14px;margin-top:8px;font-size:12px;">'
                                        f'<span style="color:{colour};font-weight:700">▶ Accuracy Check — {_pct_match}% match</span>'
                                        f'&nbsp;&nbsp;<span style="color:#64748b">Extracted ₦{ext:,.0f} vs '
                                        f'stated ₦{stl:,.0f}</span><br>'
                                        f'<span style="color:#94a3b8;font-size:11px">{verdict["message"]}</span>'
                                        f'</div>',
                                        unsafe_allow_html=True,
                                    )
                            except Exception:
                                pass  # Accuracy check is best-effort; never block the main flow

                        # ── Parser Confidence Report ──────────────────────────
                        if buckets:
                            try:
                                render_parse_confidence(
                                    buckets, _conf_pct_a, bank,
                                    filename=file_a.name, session=_SID,
                                    officer=_OFFICER,
                                )
                            except Exception:
                                pass  # never block the main flow

                        # ── Statement Freshness Warning ───────────────────────
                        if buckets:
                            try:
                                import datetime as _dt_fw
                                _latest_ym = max(buckets.keys())
                                _fw_ly, _fw_lm = int(_latest_ym[:4]), int(_latest_ym[5:])
                                _fw_today = _dt_fw.date.today()
                                _fw_months_old = (_fw_today.year - _fw_ly) * 12 + (_fw_today.month - _fw_lm)
                                if _fw_months_old > 3:
                                    _fw_col  = "#f87171" if _fw_months_old > 6 else "#fbbf24"
                                    _fw_icon = "🔴" if _fw_months_old > 6 else "🟡"
                                    _fw_note = (
                                        f"Statement is <strong>{_fw_months_old} months old</strong> "
                                        f"(latest data: {ym_label(_latest_ym)}). "
                                        + ("Income figures may not reflect current repayment capacity — request a fresh statement."
                                           if _fw_months_old > 6
                                           else "Consider requesting a more recent statement before disbursement.")
                                    )
                                    st.markdown(
                                        f'<div style="background:rgba(0,0,0,.2);border:1px solid {_fw_col}55;'
                                        f'border-radius:3px;padding:8px 14px;margin-top:8px;font-size:12px;">'
                                        f'{_fw_icon}&nbsp;<span style="color:{_fw_col};font-weight:700">'
                                        f'Statement Freshness Warning</span>'
                                        f'&nbsp;&nbsp;<span style="color:#94a3b8">{_fw_note}</span>'
                                        f'</div>',
                                        unsafe_allow_html=True,
                                    )
                            except Exception:
                                pass

                    except Exception as e:
                        track("parse_error", session=_SID, officer=_OFFICER, filename=file_a.name,
                              error=str(e), error_type=type(e).__name__)
                        if "EOF marker not found" in str(e) or "Unexpected EOF" in str(e):
                            st.error(
                                "This PDF appears to be corrupted or incomplete. "
                                "Please download the bank statement again from your bank app/portal."
                            )
                        else:
                            st.error(f"Error: {e}")
                    finally:
                        # Free raw bytes + cached text immediately — a 10 MB PDF
                        # leaves ~100-200 MB of PyPDF2 residuals if not freed.
                        del _pdf_bytes_a
                        try:
                            del _reused_text_a
                        except NameError:
                            pass
                        gc.collect()

# Show breakdown table for statement A
# ── Supported bank options for the manual override dropdown ──────────────────
_BANK_OVERRIDE_OPTIONS = [
    "GTBank", "Access", "UBA", "Zenith", "Ecobank", "First Bank", "Fidelity",
    "Union", "Stanbic", "FCMB", "Wema", "Sterling", "OPay", "PalmPay", "Kuda",
    "Moniepoint", "Carbon", "FairMoney", "Providus", "Jaiz", "Parallex",
    "Lotus", "Renmoney", "mybankStatement", "Other",
]

def _bank_override_selector(slot: str) -> None:
    """Render an editable 'Detected Bank' selector for statement A or B.

    Some statements (Zenith, UBA, Ecobank, etc.) carry the bank name only as a
    logo image, so it can't be auto-detected from text. This lets the officer
    set the correct bank; the choice flows into reports, tracking and exports.
    """
    bank_key  = f"bank_{slot}"
    over_key  = f"bank_override_{slot}"
    detected  = st.session_state.get(bank_key) or "mybankStatement"
    opts = _BANK_OVERRIDE_OPTIONS if detected in _BANK_OVERRIDE_OPTIONS else [detected] + _BANK_OVERRIDE_OPTIONS
    if over_key not in st.session_state or st.session_state[over_key] not in opts:
        st.session_state[over_key] = detected
    _oc1, _oc2 = st.columns([2, 3])
    with _oc1:
        chosen = st.selectbox(
            "🏦 Detected Bank (correct it if the logo wasn't read)",
            opts, key=over_key,
        )
    # Sync the (possibly corrected) bank back into the canonical slot
    st.session_state[bank_key] = chosen
    if chosen != detected:
        with _oc2:
            st.markdown(
                f'<div style="margin-top:30px;font-size:11px;color:#fbbf24">'
                f'⚑ Bank set to <strong>{chosen}</strong> (auto-detected: {detected})</div>',
                unsafe_allow_html=True,
            )

if st.session_state.rows_a:
    _bank_override_selector("a")
    rows_a = [r for r in st.session_state.rows_a
              if r["ym"] < __import__("datetime").date.today().strftime("%Y-%m")
              and r["gross"] > 0][-N_MONTHS:]
    if rows_a:
        has_self  = any(r["self_transfer"]  > 0 for r in rows_a)
        has_rev   = any(r["reversal"]       > 0 for r in rows_a)
        has_nb    = any(r["non_business"]   > 0 for r in rows_a)
        has_loan  = any(r["loan_disbursal"] > 0 for r in rows_a)

        # ── Business Type Inference ───────────────────────────────────────
        if st.session_state.txns_a:
            _narr_lc = [t["narration"].lower() for t in st.session_state.txns_a]
            _biz_scores: dict[str, tuple] = {
                "Salary Earner":       (["salary","payroll","wage","payslip","staff pay"],              "💼"),
                "E-commerce":          (["flutterwave","paystack","stripe","shopify","jumia","konga","selar","paypal","remita"], "🛒"),
                "Food & Beverage":     (["food","restaurant","kitchen","catering","eatery","canteen","bakery","snack","meal","chop"], "🍽"),
                "Logistics":           (["delivery","dispatch","logistics","courier","shipping","transport","freight"], "🚚"),
                "POS / Agent Banking": (["pos ","mpos","terminal","agent banking"],                     "🏧"),
                "Professional Svcs":   (["consulting","invoice","professional","retainer","service fee","legal fee"], "📋"),
                "Real Estate":         (["rent","property","estate","tenancy","lease","landlord"],      "🏢"),
                "Trader / Market":     (["market","goods","supply","wholesale","retail","merchandise","dealer"], "🛍"),
            }
            _scored = {
                label: (sum(1 for n in _narr_lc if any(k in n for k in kws)), icon)
                for label, (kws, icon) in _biz_scores.items()
            }
            _best_biz  = max(_scored, key=lambda x: _scored[x][0])
            _best_cnt, _best_icon = _scored[_best_biz]
            if _best_cnt >= 3:
                st.markdown(
                    f'<div style="display:flex;align-items:center;gap:10px;margin-bottom:10px">'
                    f'<div style="font-size:9px;letter-spacing:2px;color:#64748b;text-transform:uppercase">Business Type</div>'
                    f'<div style="background:rgba(16,185,129,.12);border:1px solid rgba(16,185,129,.3);'
                    f'border-radius:20px;padding:3px 14px;font-size:12px;font-weight:700;color:#10b981">'
                    f'{_best_icon} {_best_biz}</div>'
                    f'<div style="font-size:10px;color:#64748b">{_best_cnt} matching narrations</div>'
                    f'</div>',
                    unsafe_allow_html=True,
                )

        # ── Statement analysis tabs ───────────────────────────────────────
        # Manual __enter__/__exit__ keeps the existing blocks at their
        # original indentation — zero logic changes, pure layout regrouping.
        _t_break, _t_trend, _t_insight = st.tabs([
            "📊  Monthly Breakdown", "📈  Trend & Forecast", "🔍  Income Insights",
        ])
        _t_break.__enter__()

        hdr = ('<tr>'
               '<th class="col-gross" style="text-align:left">Month</th>'
               '<th class="col-gross">Total Inflow</th>')
        if has_self:  hdr += '<th class="col-self">Self Deposits</th>'
        if has_rev:   hdr += '<th class="col-rev">Reversals</th>'
        if has_nb:    hdr += '<th class="col-nonbiz">Non-Business</th>'
        if has_loan:  hdr += '<th class="col-loan">Loan Disbursals</th>'
        hdr += '<th class="col-net">Eligible Income</th></tr>'

        body = ""
        t_gross = t_self = t_rev = t_nb = t_loan = t_net = 0.0
        for r in rows_a:
            t_gross += r["gross"]
            t_self  += r["self_transfer"]
            t_rev   += r["reversal"]
            t_nb    += r["non_business"]
            t_loan  += r["loan_disbursal"]
            t_net   += r["eligible_income"]
            body += (f'<tr><td>{r["label"]}</td>'
                     f'<td class="col-gross">{money(r["gross"])}</td>')
            if has_self:  body += f'<td class="col-self" style="color:var(--orange);font-size:11px">{money(r["self_transfer"]) if r["self_transfer"] > 0 else "—"}</td>'
            if has_rev:   body += f'<td class="col-rev">{("-"+money(r["reversal"])) if r["reversal"] > 0 else "—"}</td>'
            if has_nb:    body += f'<td class="col-nonbiz">{("-"+money(r["non_business"])) if r["non_business"] > 0 else "—"}</td>'
            if has_loan:  body += f'<td class="col-loan">{("-"+money(r["loan_disbursal"])) if r["loan_disbursal"] > 0 else "—"}</td>'
            body += f'<td class="col-net">{money(r["eligible_income"])}</td></tr>'

        # ── Totals footer ─────────────────────────────────────────────────
        foot = (f'<tfoot><tr>'
                f'<td style="color:#64748b;font-size:10px;text-transform:uppercase;letter-spacing:1px">Total</td>'
                f'<td class="col-gross" style="border-top:1px solid #1a3d2b">{money(t_gross)}</td>')
        if has_self:  foot += f'<td class="col-self" style="border-top:1px solid #1a3d2b">{money(t_self) if t_self > 0 else "—"}</td>'
        if has_rev:   foot += f'<td class="col-rev"  style="border-top:1px solid #1a3d2b">{("-"+money(t_rev))  if t_rev  > 0 else "—"}</td>'
        if has_nb:    foot += f'<td class="col-nonbiz" style="border-top:1px solid #1a3d2b">{("-"+money(t_nb))   if t_nb   > 0 else "—"}</td>'
        if has_loan:  foot += f'<td class="col-loan" style="border-top:1px solid #1a3d2b">{("-"+money(t_loan)) if t_loan > 0 else "—"}</td>'
        foot += (f'<td class="col-net" style="border-top:1px solid #1a3d2b;font-size:14px">{money(t_net)}</td>'
                 f'</tr></tfoot>')

        st.markdown(
            f'<table class="preview-table"><thead>{hdr}</thead><tbody>{body}</tbody>{foot}</table>',
            unsafe_allow_html=True,
        )
        if has_self:
            st.markdown(
                '<div style="font-size:10px;color:#fb923c;margin-top:4px">'
                '⚑ Self Deposits (OWealth, Renflex, Renvault, savings round-trips) are <strong>deducted</strong> from eligible income as they are not business income.</div>',
                unsafe_allow_html=True,
            )

        _t_break.__exit__(None, None, None)
        _t_trend.__enter__()

        # ── Income Trend Chart ────────────────────────────────────────────
        _scale = max(
            max(r["gross"]            for r in rows_a),
            max(r["eligible_income"]  for r in rows_a),
        ) or 1
        BAR_H = 110  # max bar height in px

        def _fmt_v(v: float) -> str:
            return f"₦{v/1_000_000:.1f}m" if v >= 1_000_000 else f"₦{v/1_000:.0f}k"

        avg6 = sum(r["eligible_income"] for r in rows_a) / len(rows_a)
        avg3 = sum(r["eligible_income"] for r in rows_a[-3:]) / min(3, len(rows_a))

        # Trend: compare last month to first month
        trend_delta = rows_a[-1]["eligible_income"] - rows_a[0]["eligible_income"]
        trend_icon  = "▲" if trend_delta > 0 else "▼" if trend_delta < 0 else "▬"
        trend_col   = "#34d399" if trend_delta > 0 else "#f87171" if trend_delta < 0 else "#6b7f74"

        # ── SVG line graph ────────────────────────────────────────────────
        _SVG_W  = 600   # viewBox width
        _SVG_H  = 160   # viewBox height
        _PAD_L  = 10    # left padding
        _PAD_R  = 10    # right padding
        _PAD_T  = 28    # top padding (room for value labels)
        _PAD_B  = 36    # bottom padding (room for month labels + MoM %)
        _plot_w = _SVG_W - _PAD_L - _PAD_R
        _plot_h = _SVG_H - _PAD_T - _PAD_B
        _n      = len(rows_a)
        _scale_svg = max(
            max(r["gross"]           for r in rows_a),
            max(r["eligible_income"] for r in rows_a),
        ) or 1

        def _x(i):
            return _PAD_L + (_plot_w / (_n - 1) * i if _n > 1 else _plot_w / 2)

        def _y(v):
            return _PAD_T + _plot_h - (v / _scale_svg * _plot_h)

        # build polyline point strings
        _pts_gross   = " ".join(f"{_x(i):.1f},{_y(r['gross']):.1f}"           for i, r in enumerate(rows_a))
        _pts_elig    = " ".join(f"{_x(i):.1f},{_y(r['eligible_income']):.1f}" for i, r in enumerate(rows_a))

        # filled area under eligible line
        _first_x = f"{_x(0):.1f}"
        _last_x  = f"{_x(_n-1):.1f}"
        _base_y  = f"{_PAD_T + _plot_h:.1f}"
        _area_pts = f"{_first_x},{_base_y} {_pts_elig} {_last_x},{_base_y}"

        # dots, labels, month names
        _dots_gross = ""
        _dots_elig  = ""
        _val_labels = ""
        _mon_labels = ""
        _mom_labels = ""

        for _bi, r in enumerate(rows_a):
            cx = _x(_bi)
            cy_g = _y(r["gross"])
            cy_e = _y(r["eligible_income"])

            # gross dot (faint)
            _dots_gross += f'<circle cx="{cx:.1f}" cy="{cy_g:.1f}" r="4" fill="rgba(16,185,129,.25)" stroke="#10b981" stroke-width="1.5"/>'
            # eligible dot (bright, white centre)
            _dots_elig  += (
                f'<circle cx="{cx:.1f}" cy="{cy_e:.1f}" r="5.5" fill="#0f1a15" stroke="#34d399" stroke-width="2.5"/>'
                f'<circle cx="{cx:.1f}" cy="{cy_e:.1f}" r="2.5" fill="#34d399"/>'
            )
            # value label above eligible dot
            _val_labels += (
                f'<text x="{cx:.1f}" y="{cy_e - 9:.1f}" text-anchor="middle" '
                f'font-size="9" fill="#34d399" font-family="Space Mono,monospace">'
                f'{_fmt_v(r["eligible_income"])}</text>'
            )
            # month label below
            _mon_y = _PAD_T + _plot_h + 14
            _mon_labels += (
                f'<text x="{cx:.1f}" y="{_mon_y}" text-anchor="middle" '
                f'font-size="9" fill="#6b7f74" font-family="Space Mono,monospace">'
                f'{r["label"]}</text>'
            )
            # MoM % label
            if _bi > 0:
                _prev_ei = rows_a[_bi - 1]["eligible_income"]
                if _prev_ei > 0:
                    _mom_pct = (r["eligible_income"] - _prev_ei) / _prev_ei * 100
                    _mom_col = "#34d399" if _mom_pct > 0 else "#f87171"
                    _mom_arr = "▲" if _mom_pct > 0 else "▼"
                    _mom_labels += (
                        f'<text x="{cx:.1f}" y="{_mon_y + 13}" text-anchor="middle" '
                        f'font-size="8" fill="{_mom_col}" font-family="Space Mono,monospace">'
                        f'{_mom_arr}{abs(_mom_pct):.0f}%</text>'
                    )

        # horizontal grid lines (3 levels)
        _grid = ""
        for _frac in [0.25, 0.5, 0.75]:
            _gy = _PAD_T + _plot_h * (1 - _frac)
            _grid += (
                f'<line x1="{_PAD_L}" y1="{_gy:.1f}" x2="{_SVG_W - _PAD_R}" y2="{_gy:.1f}" '
                f'stroke="#1a3d2b" stroke-width="1" stroke-dasharray="4 4"/>'
            )

        _svg = (
            f'<svg viewBox="0 0 {_SVG_W} {_SVG_H}" xmlns="http://www.w3.org/2000/svg" '
            f'style="width:100%;height:auto;display:block">'
            # grid
            f'{_grid}'
            # filled area under eligible line
            f'<polygon points="{_area_pts}" fill="rgba(16,185,129,.07)"/>'
            # gross line (dashed, faint)
            f'<polyline points="{_pts_gross}" fill="none" stroke="rgba(16,185,129,.30)" '
            f'stroke-width="2" stroke-dasharray="5 4" stroke-linejoin="round" stroke-linecap="round"/>'
            # eligible line (bold, solid)
            f'<polyline points="{_pts_elig}" fill="none" stroke="#34d399" '
            f'stroke-width="3.5" stroke-linejoin="round" stroke-linecap="round"/>'
            # dots
            f'{_dots_gross}{_dots_elig}'
            # labels
            f'{_val_labels}{_mon_labels}{_mom_labels}'
            f'</svg>'
        )

        st.markdown(
            f'<div style="margin-top:20px;padding:16px 14px 14px;background:rgba(0,0,0,.18);'
            f'border:1px solid #1a3d2b;border-radius:4px">'
            # header
            f'<div style="display:flex;justify-content:space-between;align-items:baseline;margin-bottom:14px">'
            f'<div style="font-size:9px;letter-spacing:2px;color:#6b7f74;text-transform:uppercase">Income Trend</div>'
            f'<div style="font-size:10px;color:{trend_col};letter-spacing:1px">'
            f'{trend_icon} {_fmt_v(abs(trend_delta))} vs first month</div>'
            f'</div>'
            # SVG line graph
            + _svg +
            # averages + legend footer
            f'<div style="display:flex;gap:24px;margin-top:12px;padding-top:10px;border-top:1px solid #1a3d2b;align-items:center">'
            f'<div><div style="font-size:8px;letter-spacing:2px;color:#6b7f74;text-transform:uppercase;margin-bottom:2px">6-mo avg</div>'
            f'<div style="font-size:13px;font-weight:700;color:#10b981">{_fmt_v(avg6)}</div></div>'
            f'<div><div style="font-size:8px;letter-spacing:2px;color:#6b7f74;text-transform:uppercase;margin-bottom:2px">3-mo avg</div>'
            f'<div style="font-size:13px;font-weight:700;color:#fbbf24">{_fmt_v(avg3)}</div></div>'
            f'<div style="margin-left:auto;display:flex;gap:12px;align-items:center">'
            f'<span style="font-size:9px;color:#6b7f74">'
            f'<span style="display:inline-block;width:24px;height:3px;'
            f'background:#34d399;border-radius:2px;margin-right:4px;vertical-align:middle"></span>Eligible</span>'
            f'<span style="font-size:9px;color:#6b7f74">'
            f'<span style="display:inline-block;width:24px;height:2px;'
            f'background:rgba(16,185,129,.35);border-radius:2px;margin-right:4px;vertical-align:middle;'
            f'border-top:2px dashed rgba(16,185,129,.5)"></span>Gross</span>'
            f'</div></div>'
            f'</div>',
            unsafe_allow_html=True,
        )

        _t_trend.__exit__(None, None, None)
        _t_insight.__enter__()

        # ── Recurring Income Detection ────────────────────────────────────
        if st.session_state.txns_a:
            _real_txns = [t for t in st.session_state.txns_a if t.get("category") == "real_credit"]
            if _real_txns:
                import re as _re_rid
                from collections import defaultdict as _dd_rid

                def _sender_key(narr: str) -> str:
                    n = narr.upper()
                    n = _re_rid.sub(r'\b\d{6,}\b', '', n)
                    n = _re_rid.sub(r'\d{2}/\d{2}/\d{4}', '', n)
                    n = _re_rid.sub(r'[/|\\-]+', ' ', n)
                    n = _re_rid.sub(r'\s{2,}', ' ', n).strip()
                    _skip = {'FROM','TO','BY','OF','THE','A','AN','AND','FOR','AT','TRANSFER','TRF',
                             'NIP','CR','DR','CREDIT','DEBIT','PAYMENT','INWARD','OUTWARD','FT','WT',
                             'INT','VIA','REF','MON','TUE','WED','THU','FRI','SAT','SUN'}
                    parts = [p for p in n.split() if len(p) >= 3 and p not in _skip]
                    return ' '.join(parts[:2]) if parts else narr[:20].upper()

                _sender_months: dict = _dd_rid(set)
                _sender_total:  dict = _dd_rid(float)
                for _t in _real_txns:
                    _k = _sender_key(_t["narration"])
                    _sender_months[_k].add(_t["ym"])
                    _sender_total[_k] += _t["amount"]

                _recurring = sorted(
                    [(_k, len(_mv), _sender_total[_k]) for _k, _mv in _sender_months.items() if len(_mv) >= 2],
                    key=lambda x: (-x[1], -x[2])
                )[:15]

                if _recurring:
                    with st.expander("🔁  Recurring Income Sources", expanded=False):
                        _rid_rows = ""
                        _all_real_total_rid = sum(_sender_total.values()) or 1
                        for _rk, _rm, _rt in _recurring:
                            _tag    = "🟢 Recurring" if _rm >= 3 else "🔵 Returning"
                            _tcolor = "#34d399" if _rm >= 3 else "#60a5fa"
                            _avg_m  = _rt / _rm
                            _rid_rows += (
                                f'<tr>'
                                f'<td style="font-size:11px;max-width:200px;overflow:hidden;'
                                f'text-overflow:ellipsis;white-space:nowrap">{_rk.title()}</td>'
                                f'<td style="text-align:center;font-size:11px">{_rm}</td>'
                                f'<td style="text-align:right;font-size:11px">{money(_rt)}</td>'
                                f'<td style="text-align:right;font-size:11px">{money(_avg_m)}</td>'
                                f'<td style="text-align:center">'
                                f'<span style="color:{_tcolor};font-size:10px;font-weight:700">{_tag}</span>'
                                f'</td></tr>'
                            )
                        st.markdown(
                            f'<table class="preview-table"><thead><tr>'
                            f'<th style="text-align:left">Sender / Source</th>'
                            f'<th style="text-align:center">Months Active</th>'
                            f'<th style="text-align:right">Total Received</th>'
                            f'<th style="text-align:right">Avg / Month</th>'
                            f'<th style="text-align:center">Consistency</th>'
                            f'</tr></thead><tbody>{_rid_rows}</tbody></table>'
                            f'<div style="font-size:10px;color:#64748b;margin-top:6px">'
                            f'Senders appearing in 2+ months. '
                            f'🟢 Recurring = 3+ months &nbsp;|&nbsp; 🔵 Returning = 2 months.</div>',
                            unsafe_allow_html=True,
                        )

        # ── Suspicious Pattern Flags ──────────────────────────────────────
        if st.session_state.txns_a:
            _real_txns2 = [t for t in st.session_state.txns_a if t.get("category") == "real_credit"]
            if _real_txns2:
                import re as _re_spf
                from collections import defaultdict as _dd_spf
                _spf_flags = []
                _all_real_total2 = sum(t["amount"] for t in _real_txns2) or 1

                # Flag 1 — Round-number concentration
                _round_txns = [t for t in _real_txns2
                               if t["amount"] > 0 and (t["amount"] % 100_000 == 0 or t["amount"] % 50_000 == 0)]
                _round_total = sum(t["amount"] for t in _round_txns)
                _round_pct   = _round_total / _all_real_total2 * 100
                if _round_pct >= 40 and len(_round_txns) >= 3:
                    _spf_flags.append((
                        "⚠ High Round-Number Credits",
                        f"{len(_round_txns)} credits ({_round_pct:.0f}% of income) are exact multiples of "
                        f"₦50,000 or ₦100,000 — may indicate structured or artificial deposits.",
                        "#fbbf24",
                    ))

                # Flag 2 — Concentration risk
                _spf_sender: dict = _dd_spf(float)
                for _t in _real_txns2:
                    _spf_n = _re_spf.sub(r'[^A-Z ]', ' ', _t["narration"].upper()).split()
                    _spf_k = ' '.join([w for w in _spf_n if len(w) >= 4][:2]) or "OTHER"
                    _spf_sender[_spf_k] += _t["amount"]
                if _spf_sender:
                    _top_s = max(_spf_sender, key=lambda x: _spf_sender[x])
                    _top_pct = _spf_sender[_top_s] / _all_real_total2 * 100
                    if _top_pct >= 50:
                        _spf_flags.append((
                            "⚠ Income Concentration Risk",
                            f'"{_top_s.title()}" accounts for {_top_pct:.0f}% of eligible income — '
                            f'heavy reliance on a single payer increases cash-flow risk.',
                            "#fb923c",
                        ))

                if _spf_flags:
                    with st.expander(f"🚩  Suspicious Pattern Flags ({len(_spf_flags)} found)", expanded=True):
                        for _ft, _fm, _fc in _spf_flags:
                            st.markdown(
                                f'<div style="border-left:3px solid {_fc};'
                                f'border-radius:0 3px 3px 0;padding:8px 14px;'
                                f'background:rgba(0,0,0,.18);margin-bottom:8px;font-size:12px">'
                                f'<div style="color:{_fc};font-weight:700;margin-bottom:3px">{_ft}</div>'
                                f'<div style="color:#94a3b8">{_fm}</div>'
                                f'</div>',
                                unsafe_allow_html=True,
                            )

        _t_insight.__exit__(None, None, None)
        _t_trend.__enter__()

        # ── Cash Flow Forecast (Feature 10) ──────────────────────────────
        _fcast_vals = [r["eligible_income"] for r in rows_a if r["eligible_income"] > 0]
        _n_forecast = 3 if _product == "SEL" else 6
        if len(_fcast_vals) >= 3:
            _fn = len(_fcast_vals)
            _fx_mean = (_fn - 1) / 2
            _fy_mean = sum(_fcast_vals) / _fn
            _ss_xy = sum((i - _fx_mean) * (_fcast_vals[i] - _fy_mean) for i in range(_fn))
            _ss_xx = sum((i - _fx_mean) ** 2 for i in range(_fn))
            if _ss_xx > 0:
                _fc_slope = _ss_xy / _ss_xx
                _fc_icept = _fy_mean - _fc_slope * _fx_mean
                _y_pred   = [_fc_icept + _fc_slope * i for i in range(_fn)]
                _ss_res   = sum((_fcast_vals[i] - _y_pred[i]) ** 2 for i in range(_fn))
                _ss_tot   = sum((_fcast_vals[i] - _fy_mean) ** 2 for i in range(_fn))
                _fc_r2    = max(0.0, 1 - _ss_res / _ss_tot) if _ss_tot > 0 else 0.0
                _forecast = [max(_fc_icept + _fc_slope * (_fn + i), 0) for i in range(_n_forecast)]

                # Build future month labels
                _last_ym  = rows_a[-1]["ym"]
                _ly, _lm  = int(_last_ym[:4]), int(_last_ym[5:])
                _fc_labels = []
                for _fi in range(1, _n_forecast + 1):
                    _fm, _fy = _lm + _fi, _ly
                    while _fm > 12: _fm -= 12; _fy += 1
                    _fc_labels.append(ym_label(f"{_fy}-{str(_fm).zfill(2)}"))

                _slope_pct     = (_fc_slope / _fy_mean * 100) if _fy_mean else 0
                _fc_reliable   = _fc_r2 >= 0.60
                _fc_col        = "#34d399" if _fc_slope > 0 else "#f87171" if _fc_slope < 0 else "#64748b"
                _fc_trend_lbl  = "Growing" if _fc_slope > 0 else "Declining" if _fc_slope < 0 else "Flat"
                _fc_scale = max(_forecast) if max(_forecast) > 0 else 1

                # ── SVG line graph for forecast ───────────────────────────
                _FC_W   = 600
                _FC_H   = 140
                _FC_PL  = 10
                _FC_PR  = 10
                _FC_PT  = 28    # top padding for value labels
                _FC_PB  = 32    # bottom for month labels
                _fc_pw  = _FC_W - _FC_PL - _FC_PR
                _fc_ph  = _FC_H - _FC_PT - _FC_PB
                _fc_n   = len(_forecast)
                _fc_area_col = "rgba(52,211,153,.08)" if _fc_slope >= 0 else "rgba(248,113,113,.08)"
                _fc_dot_col  = _fc_col

                def _fcx(i):
                    return _FC_PL + (_fc_pw / (_fc_n - 1) * i if _fc_n > 1 else _fc_pw / 2)

                def _fcy(v):
                    return _FC_PT + _fc_ph - (v / _fc_scale * _fc_ph)

                _fc_pts = " ".join(f"{_fcx(i):.1f},{_fcy(v):.1f}" for i, v in enumerate(_forecast))
                _fc_area_pts = (
                    f"{_fcx(0):.1f},{_FC_PT + _fc_ph:.1f} "
                    + _fc_pts +
                    f" {_fcx(_fc_n - 1):.1f},{_FC_PT + _fc_ph:.1f}"
                )

                # grid lines
                _fc_grid = ""
                for _fg in [0.33, 0.66]:
                    _fgy = _FC_PT + _fc_ph * (1 - _fg)
                    _fc_grid += (
                        f'<line x1="{_FC_PL}" y1="{_fgy:.1f}" x2="{_FC_W - _FC_PR}" y2="{_fgy:.1f}" '
                        f'stroke="#1a3d2b" stroke-width="1" stroke-dasharray="4 4"/>'
                    )

                _fc_dots_svg  = ""
                _fc_vlabels   = ""
                _fc_mlabels   = ""
                for _fi, (_fl, _fv) in enumerate(zip(_fc_labels, _forecast)):
                    _cx = _fcx(_fi)
                    _cy = _fcy(_fv)
                    # dot: outer ring + inner fill
                    _fc_dots_svg += (
                        f'<circle cx="{_cx:.1f}" cy="{_cy:.1f}" r="5.5" fill="#0f1a15" '
                        f'stroke="{_fc_dot_col}" stroke-width="2.5"/>'
                        f'<circle cx="{_cx:.1f}" cy="{_cy:.1f}" r="2.5" fill="{_fc_dot_col}"/>'
                    )
                    # value label above dot
                    _fc_vlabels += (
                        f'<text x="{_cx:.1f}" y="{_cy - 9:.1f}" text-anchor="middle" '
                        f'font-size="9" fill="{_fc_dot_col}" font-family="Space Mono,monospace">'
                        f'{_fmt_v(_fv)}</text>'
                    )
                    # month label below
                    _fc_mlabels += (
                        f'<text x="{_cx:.1f}" y="{_FC_PT + _fc_ph + 14}" text-anchor="middle" '
                        f'font-size="9" fill="#4a6a58" font-family="Space Mono,monospace">'
                        f'{_fl}</text>'
                    )

                _fc_svg = (
                    f'<svg viewBox="0 0 {_FC_W} {_FC_H}" xmlns="http://www.w3.org/2000/svg" '
                    f'style="width:100%;height:auto;display:block">'
                    + _fc_grid
                    + f'<polygon points="{_fc_area_pts}" fill="{_fc_area_col}"/>'
                    + f'<polyline points="{_fc_pts}" fill="none" stroke="{_fc_col}" '
                    f'stroke-width="3.5" stroke-dasharray="7 4" '
                    f'stroke-linejoin="round" stroke-linecap="round"/>'
                    + _fc_dots_svg
                    + _fc_vlabels
                    + _fc_mlabels
                    + f'</svg>'
                )

                _grow_note = ""
                if _fc_slope > 0 and _fc_reliable:
                    _grow_note = (
                        f'<div style="margin-top:10px;padding:6px 10px;'
                        f'background:rgba(52,211,153,.06);border-left:2px solid #34d399;'
                        f'border-radius:3px;font-size:10px;color:#34d399">'
                        f'✦ Consistent upward trend — projected growth of {_slope_pct:+.1f}%/month '
                        f'may support higher assessments in future cycles.'
                        f'</div>'
                    )
                elif _fc_slope < 0 and _fc_reliable:
                    _grow_note = (
                        f'<div style="margin-top:10px;padding:6px 10px;'
                        f'background:rgba(248,113,113,.06);border-left:2px solid #f87171;'
                        f'border-radius:3px;font-size:10px;color:#f87171">'
                        f'⚠ Declining income trend — consider a shorter tenor to reduce repayment risk.'
                        f'</div>'
                    )

                st.markdown(
                    f'<div style="margin-top:12px;padding:14px;background:rgba(0,0,0,.15);'
                    f'border:1px dashed {_fc_col}44;border-radius:4px">'
                    f'<div style="display:flex;justify-content:space-between;align-items:baseline;margin-bottom:10px">'
                    f'<div style="font-size:9px;letter-spacing:2px;color:#6b7f74;text-transform:uppercase">'
                    f'Cash Flow Forecast — Next {_n_forecast} Months</div>'
                    f'<div style="font-size:10px;color:{_fc_col}">'
                    f'{_fc_trend_lbl} &nbsp;{_slope_pct:+.1f}%/mo'
                    f'{"  ·  R²="+str(round(_fc_r2,2)) if _fc_reliable else "  ·  low confidence"}'
                    f'</div></div>'
                    + _fc_svg
                    + _grow_note
                    + f'</div>',
                    unsafe_allow_html=True,
                )

        _t_trend.__exit__(None, None, None)

        # ── Download buttons (always visible, below the tabs) ─────────────
        _pdf_stmt = generate_pdf_report(
            account_name = st.session_state.name_a or "Account Holder",
            bank         = st.session_state.bank_a or "Bank",
            rows         = rows_a,
        )
        _xlsx_stmt = generate_xlsx(
            rows         = rows_a,
            account_name = st.session_state.name_a or "Account Holder",
            bank         = st.session_state.bank_a or "Bank",
        )
        _safe_stmt = (st.session_state.name_a or "statement").replace(" ", "_").lower()
        _dl1, _dl2 = st.columns(2)
        with _dl1:
            st.download_button(
                label     = "⬇  Download Statement Analysis (PDF)",
                data      = _pdf_stmt,
                file_name = f"PARSIO_Statement_{_safe_stmt}_{datetime.date.today():%Y%m%d}.pdf",
                mime      = "application/pdf",
                key       = "dl_statement_pdf",
                use_container_width=True,
            )
        with _dl2:
            st.download_button(
                label     = "⬇  Download Monthly Breakdown (Excel)",
                data      = _xlsx_stmt,
                file_name = f"PARSIO_Breakdown_{_safe_stmt}_{datetime.date.today():%Y%m%d}.xlsx",
                mime      = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key       = "dl_statement_xlsx",
                use_container_width=True,
            )

        # ── Income Consistency Score (Income Insights tab) ────────────────
        _t_insight.__enter__()
        import statistics as _stat_mod
        _ei_vals = [r["eligible_income"] for r in rows_a if r["eligible_income"] > 0]
        if len(_ei_vals) >= 2:
            _ei_mean = sum(_ei_vals) / len(_ei_vals)
            _ei_std  = _stat_mod.stdev(_ei_vals)
            _cv      = _ei_std / _ei_mean if _ei_mean else 0
            _peak    = max(_ei_vals)
            _trough  = min(_ei_vals)
            _cliff   = _peak / _trough if _trough > 0 else 99
            if _cv < 0.30:
                _cs_label, _cs_col = "Stable",   "#34d399"
            elif _cv < 0.60:
                _cs_label, _cs_col = "Moderate", "#f59e0b"
            else:
                _cs_label, _cs_col = "Volatile", "#f87171"
            _cliff_html = ""
            if _cliff > 5 and len(_ei_vals) >= 3:
                _cliff_html = (
                    f'<span style="margin-left:8px;background:rgba(248,113,113,.12);'
                    f'border:1px solid rgba(248,113,113,.3);border-radius:4px;'
                    f'padding:2px 8px;font-size:10px;color:#f87171">⚠ Cliff {_cliff:.1f}× peak vs trough</span>'
                )
            st.markdown(
                f'<div style="display:flex;align-items:center;flex-wrap:wrap;gap:12px;margin:10px 0;'
                f'padding:10px 14px;background:rgba(255,255,255,.03);border-radius:6px;'
                f'border-left:3px solid {_cs_col}">'
                f'<div style="font-size:9px;letter-spacing:2px;color:#64748b;text-transform:uppercase">Income Consistency</div>'
                f'<div style="font-size:13px;font-weight:700;color:{_cs_col}">● {_cs_label}</div>'
                f'<div style="font-size:11px;color:#6b7f74">CV {_cv:.0%}</div>'
                f'<div style="font-size:11px;color:#6b7f74">Peak/Trough {_cliff:.1f}×</div>'
                f'{_cliff_html}'
                f'</div>',
                unsafe_allow_html=True,
            )
        _t_insight.__exit__(None, None, None)

# ── Debit Transaction Visibility Panel ───────────────────────────────────────
# For officer visibility only — NOT used in credit decisioning.
def _render_debit_panel(debit_txns: list, label: str = "") -> None:
    if not debit_txns:
        return

    _PRIORITY = {"loan_repayment", "credit_card", "rent"}

    # Summary metrics
    _total_debit  = sum(t["amount"] for t in debit_txns)
    _flagged      = [t for t in debit_txns if t["category"] in _PRIORITY]
    _loan_repays  = [t for t in debit_txns if t["category"] == "loan_repayment"]
    _total_loan_r = sum(t["amount"] for t in _loan_repays)
    _total_rent   = sum(t["amount"] for t in debit_txns if t["category"] == "rent")

    # Monthly debit totals
    from collections import defaultdict as _dd_deb
    _monthly_deb: dict = _dd_deb(float)
    for _t in debit_txns:
        _monthly_deb[_t["ym"]] += _t["amount"]

    # Bold section header outside expander so it's always visible
    st.markdown(
        f'<div style="margin-top:20px;margin-bottom:6px;padding:10px 16px;'
        f'background:rgba(248,113,113,.10);border-left:5px solid #f87171;border-radius:4px;'
        f'font-size:13px;font-weight:700;color:#fca5a5;letter-spacing:1px">'
        f'💳 &nbsp;DEBIT VISIBILITY PANEL{(" — " + label.upper()) if label else ""}'
        f'<span style="font-size:10px;font-weight:400;color:#fda4af;margin-left:12px">'
        f'Officer view only · Not used in decisioning</span></div>',
        unsafe_allow_html=True,
    )

    with st.expander(f"Show / Hide Debit Detail — {label or 'All'}", expanded=True):
        st.markdown(
            f'<div style="background:rgba(248,113,113,.10);border:1px solid rgba(248,113,113,.4);'
            f'border-left:4px solid #f87171;border-radius:4px;padding:10px 14px;margin-bottom:14px;'
            f'font-size:11px;color:#fca5a5;font-weight:500">'
            f'ℹ️ &nbsp;Debit data is shown for <strong style="color:#fda4af">officer visibility only</strong> — '
            f'it is <strong style="color:#fda4af">not used in credit decisioning</strong> or loan calculations.'
            f'</div>',
            unsafe_allow_html=True,
        )

        # ── Summary cards ──────────────────────────────────────────────────
        _dc1, _dc2, _dc3, _dc4 = st.columns(4)
        with _dc1:
            st.markdown(
                f'<div style="background:rgba(248,113,113,.12);border:1px solid rgba(248,113,113,.4);'
                f'border-radius:10px;padding:14px 16px">'
                f'<div style="font-family:Plus Jakarta Sans,sans-serif;font-size:10px;letter-spacing:1.8px;color:#fca5a5;text-transform:uppercase;'
                f'font-weight:600;margin-bottom:6px">Total Debits</div>'
                f'<div style="font-family:Plus Jakarta Sans,sans-serif;font-variant-numeric:tabular-nums;font-size:22px;font-weight:800;color:#f87171">{money(_total_debit)}</div></div>',
                unsafe_allow_html=True,
            )
        with _dc2:
            st.markdown(
                f'<div style="background:rgba(248,113,113,.12);border:1px solid rgba(248,113,113,.4);'
                f'border-radius:10px;padding:14px 16px">'
                f'<div style="font-family:Plus Jakarta Sans,sans-serif;font-size:10px;letter-spacing:1.8px;color:#fca5a5;text-transform:uppercase;'
                f'font-weight:600;margin-bottom:6px">Loan Repayments</div>'
                f'<div style="font-family:Plus Jakarta Sans,sans-serif;font-variant-numeric:tabular-nums;font-size:22px;font-weight:800;color:#f87171">{money(_total_loan_r)}</div>'
                f'<div style="font-size:10px;color:#fca5a5;margin-top:3px">{len(_loan_repays)} transaction(s)</div></div>',
                unsafe_allow_html=True,
            )
        with _dc3:
            st.markdown(
                f'<div style="background:rgba(245,158,11,.10);border:1px solid rgba(245,158,11,.4);'
                f'border-radius:10px;padding:14px 16px">'
                f'<div style="font-size:9px;letter-spacing:2px;color:#fcd34d;text-transform:uppercase;'
                f'font-weight:600;margin-bottom:6px">Rent / Property</div>'
                f'<div style="font-family:Plus Jakarta Sans,sans-serif;font-variant-numeric:tabular-nums;font-size:22px;font-weight:800;color:#f59e0b">{money(_total_rent)}</div></div>',
                unsafe_allow_html=True,
            )
        with _dc4:
            st.markdown(
                f'<div style="background:rgba(248,113,113,.12);border:1px solid rgba(248,113,113,.4);'
                f'border-radius:10px;padding:14px 16px">'
                f'<div style="font-family:Plus Jakarta Sans,sans-serif;font-size:10px;letter-spacing:1.8px;color:#fca5a5;text-transform:uppercase;'
                f'font-weight:600;margin-bottom:6px">Priority Flags</div>'
                f'<div style="font-family:Plus Jakarta Sans,sans-serif;font-variant-numeric:tabular-nums;font-size:22px;font-weight:800;color:#f87171">{len(_flagged)}</div>'
                f'<div style="font-size:10px;color:#fca5a5;margin-top:3px">loan repay · credit card · rent</div></div>',
                unsafe_allow_html=True,
            )

        st.markdown("<div style='margin-top:16px'></div>", unsafe_allow_html=True)

        # ── Category breakdown table ───────────────────────────────────────
        _cat_totals: dict = _dd_deb(float)
        _cat_counts: dict = _dd_deb(int)
        for _t in debit_txns:
            _cat_totals[_t["label"]] += _t["amount"]
            _cat_counts[_t["label"]] += 1

        _cat_rows = sorted(_cat_totals.items(), key=lambda x: -x[1])
        _cat_html = "".join(
            f'<tr style="border-bottom:1px solid rgba(248,113,113,.15)">'
            f'<td style="padding:7px 10px;font-size:12px;color:#e2e8f0;font-weight:500">{lbl}</td>'
            f'<td style="padding:7px 10px;font-size:12px;color:#cbd5e1;text-align:right">{_cat_counts[lbl]}</td>'
            f'<td style="padding:7px 10px;font-size:12px;color:#f87171;text-align:right;font-weight:700">{money(_cat_totals[lbl])}</td>'
            f'</tr>'
            for lbl, _ in _cat_rows
        )
        st.markdown(
            f'<div style="font-size:10px;letter-spacing:2px;color:#fca5a5;text-transform:uppercase;'
            f'font-weight:600;margin-bottom:8px">📊 Debit Category Breakdown</div>'
            f'<div style="overflow-x:auto"><table style="width:100%;border-collapse:collapse;'
            f'background:rgba(248,113,113,.06);border:1px solid rgba(248,113,113,.3);border-radius:6px">'
            f'<thead><tr style="border-bottom:1px solid rgba(248,113,113,.3);background:rgba(248,113,113,.12)">'
            f'<th style="padding:8px 10px;font-size:10px;color:#fca5a5;text-align:left;text-transform:uppercase;'
            f'letter-spacing:1px;font-weight:700">Category</th>'
            f'<th style="padding:8px 10px;font-size:10px;color:#fca5a5;text-align:right;text-transform:uppercase;'
            f'letter-spacing:1px;font-weight:700">Count</th>'
            f'<th style="padding:8px 10px;font-size:10px;color:#fca5a5;text-align:right;text-transform:uppercase;'
            f'letter-spacing:1px;font-weight:700">Total</th>'
            f'</tr></thead><tbody>{_cat_html}</tbody></table></div>',
            unsafe_allow_html=True,
        )

        st.markdown("<div style='margin-top:16px'></div>", unsafe_allow_html=True)

        # ── Flagged transactions (loan repayment, credit card, rent) ──────
        if _flagged:
            st.markdown(
                f'<div style="font-size:10px;letter-spacing:2px;color:#f87171;text-transform:uppercase;'
                f'font-weight:700;margin-bottom:8px">'
                f'🔴 &nbsp;Flagged Transactions ({len(_flagged)}) — Loan Repayments · Credit Cards · Rent</div>',
                unsafe_allow_html=True,
            )
            _flag_rows_html = "".join(
                f'<tr style="border-bottom:1px solid rgba(248,113,113,.15)">'
                f'<td style="padding:6px 10px;font-size:11px;color:#cbd5e1;font-weight:500">{t.get("ym","")}</td>'
                f'<td style="padding:6px 10px;font-size:11px;color:#e2e8f0;max-width:260px;word-break:break-word">{t["narration"][:80]}</td>'
                f'<td style="padding:6px 10px;font-size:11px;color:#f87171;text-align:right;font-weight:700">{money(t["amount"])}</td>'
                f'<td style="padding:6px 10px;font-size:11px;color:#fca5a5;font-weight:500">{t["label"]}</td>'
                f'</tr>'
                for t in sorted(_flagged, key=lambda x: -x["amount"])[:50]
            )
            st.markdown(
                f'<div style="overflow-x:auto"><table style="width:100%;border-collapse:collapse;'
                f'background:rgba(248,113,113,.06);border:1px solid rgba(248,113,113,.3);border-radius:6px">'
                f'<thead><tr style="border-bottom:1px solid rgba(248,113,113,.3);background:rgba(248,113,113,.12)">'
                f'<th style="padding:7px 10px;font-size:10px;color:#fca5a5;text-align:left;font-weight:700">Month</th>'
                f'<th style="padding:7px 10px;font-size:10px;color:#fca5a5;text-align:left;font-weight:700">Narration</th>'
                f'<th style="padding:7px 10px;font-size:10px;color:#fca5a5;text-align:right;font-weight:700">Amount</th>'
                f'<th style="padding:7px 10px;font-size:10px;color:#fca5a5;text-align:left;font-weight:700">Flag</th>'
                f'</tr></thead><tbody>{_flag_rows_html}</tbody></table></div>',
                unsafe_allow_html=True,
            )

        st.markdown("<div style='margin-top:16px'></div>", unsafe_allow_html=True)

        # ── All debit transactions (collapsible) ──────────────────────────
        with st.expander("📋  All Debit Transactions (full list)", expanded=False):
            _all_rows_html = "".join(
                f'<tr style="border-bottom:1px solid rgba(248,113,113,.10)">'
                f'<td style="padding:5px 10px;font-size:11px;color:#cbd5e1">{t.get("ym","")}</td>'
                f'<td style="padding:5px 10px;font-size:11px;color:#e2e8f0;max-width:300px;word-break:break-word">{t["narration"][:90]}</td>'
                f'<td style="padding:5px 10px;font-size:11px;color:#f87171;text-align:right;font-weight:600">{money(t["amount"])}</td>'
                f'<td style="padding:5px 10px;font-size:11px;color:#fca5a5">{t["label"]}</td>'
                f'</tr>'
                for t in sorted(debit_txns, key=lambda x: (-x["amount"]))
            )
            st.markdown(
                f'<div style="overflow-x:auto"><table style="width:100%;border-collapse:collapse;'
                f'background:rgba(248,113,113,.05);border:1px solid rgba(248,113,113,.2)">'
                f'<thead><tr style="border-bottom:1px solid rgba(248,113,113,.25);background:rgba(248,113,113,.10)">'
                f'<th style="padding:6px 10px;font-size:10px;color:#fca5a5;text-align:left;font-weight:700">Month</th>'
                f'<th style="padding:6px 10px;font-size:10px;color:#fca5a5;text-align:left;font-weight:700">Narration</th>'
                f'<th style="padding:6px 10px;font-size:10px;color:#fca5a5;text-align:right;font-weight:700">Amount</th>'
                f'<th style="padding:6px 10px;font-size:10px;color:#fca5a5;text-align:left;font-weight:700">Category</th>'
                f'</tr></thead><tbody>{_all_rows_html}</tbody></table></div>',
                unsafe_allow_html=True,
            )


# ── Debit Visibility Panel — Statement A ─────────────────────────────────────
if st.session_state.get("debit_txns_a"):
    _render_debit_panel(st.session_state.debit_txns_a, label="Statement 1")

# ── Transaction Search — Statement A ─────────────────────────────────────────
if st.session_state.txns_a:
    st.markdown(
        '<div style="margin-top:16px;font-size:10px;letter-spacing:2px;color:#10b981;'
        'text-transform:uppercase;margin-bottom:6px">Transaction Search</div>',
        unsafe_allow_html=True,
    )
    search_a = st.text_input(
        "Search keyword",
        key="search_a",
        placeholder="e.g. salary, transfer, POS, Flutterwave...",
        label_visibility="collapsed",
    )
    if search_a and search_a.strip():
        kw = search_a.strip().lower()
        matches = [t for t in st.session_state.txns_a if kw in t["narration"].lower()]
        if matches:
            total_match = sum(t["amount"] for t in matches)
            st.markdown(
                f'<div style="font-size:11px;color:#64748b;margin-bottom:6px">'
                f'Found <span style="color:#10b981;font-weight:700">{len(matches)}</span> credit(s) '
                f'matching <em>"{search_a}"</em> — '
                f'Total: <span style="color:#34d399;font-weight:700">{money(total_match)}</span></div>',
                unsafe_allow_html=True,
            )
            _CAT_COLOUR = {
                "real_credit": "#34d399", "self_transfer": "#fb923c",
                "reversal": "#a78bfa", "non_business": "#64748b", "loan_disbursal": "#fbbf24",
            }
            rows_html = ""
            for t in matches[:100]:  # cap at 100 rows for performance
                clr = _CAT_COLOUR.get(t["category"], "#e2e8f0")
                cat_label = t["category"].replace("_", " ").title()
                rows_html += (
                    f'<tr><td style="color:#10b981">{t["ym"]}</td>'
                    f'<td style="color:#34d399;text-align:right">{money(t["amount"])}</td>'
                    f'<td><span style="background:rgba(255,255,255,.05);padding:1px 6px;'
                    f'border-radius:3px;font-size:9px;color:{clr}">{cat_label}</span></td>'
                    f'<td style="color:#94a3b8;font-size:11px">{t["narration"][:80]}</td></tr>'
                )
            if len(matches) > 100:
                rows_html += f'<tr><td colspan="4" style="color:#64748b;font-size:10px;padding-top:6px">... and {len(matches)-100} more</td></tr>'
            st.markdown(
                f'<table class="preview-table" style="margin-top:4px">'
                f'<thead><tr>'
                f'<th style="text-align:left">Month</th>'
                f'<th style="text-align:right">Amount</th>'
                f'<th style="text-align:left">Category</th>'
                f'<th style="text-align:left">Narration</th>'
                f'</tr></thead><tbody>{rows_html}</tbody></table>',
                unsafe_allow_html=True,
            )
        else:
            st.markdown(
                f'<div style="color:#64748b;font-size:12px;padding:8px 0">No credit transactions found matching <em>"{search_a}"</em>.</div>',
                unsafe_allow_html=True,
            )


# ── Statement Intelligence Panel ─────────────────────────────────────────────
if st.session_state.txns_a:
    import re as _re_intel, statistics as _stat_intel
    from collections import defaultdict as _ddict_intel

    _txns_i  = st.session_state.txns_a
    _intel   = []   # each item is an HTML string for a panel card

    # ── Income Source Breakdown ───────────────────────────────────────────
    _tot_amt = sum(t["amount"] for t in _txns_i)
    if _tot_amt > 0:
        _src = [
            ("Bank Transfers",     ["transfer","trf/"," nip "," neft ","instant payment","mobile transfer"], "#34d399"),
            ("Aggregator/Fintech", ["settlement","flutterwave","paystack","remita","interswitch","nibss","squad","stripe","selar"], "#a78bfa"),
            ("Salary / Payroll",   ["salary","payroll","wage","payslip"], "#fbbf24"),
            ("POS / Terminal",     ["pos ","mpos","terminal","agent banking"], "#fb923c"),
        ]
        _src_amts = []
        _accounted = 0.0
        for _lbl, _kws, _clr in _src:
            _a = sum(t["amount"] for t in _txns_i if any(k in t["narration"].lower() for k in _kws))
            _src_amts.append((_lbl, _a, _clr))
            _accounted += _a
        _src_amts.append(("Other", max(_tot_amt - _accounted, 0), "#64748b"))
        _src_amts = [x for x in _src_amts if x[1] > 0]

        _pos_pct = next((a / _tot_amt for l, a, c in _src_amts if l == "POS / Terminal"), 0)
        _bars = ""
        for _lbl, _amt, _clr in _src_amts:
            _p = _amt / _tot_amt
            _bars += (
                f'<div style="display:flex;align-items:center;gap:8px;margin-bottom:5px">'
                f'<div style="width:120px;font-size:10px;color:#94a3b8;text-align:right">{_lbl}</div>'
                f'<div style="flex:1;background:rgba(255,255,255,.06);border-radius:3px;height:9px">'
                f'<div style="width:{_p:.0%};background:{_clr};border-radius:3px;height:9px"></div></div>'
                f'<div style="width:34px;font-size:10px;color:{_clr};font-weight:700">{_p:.0%}</div>'
                f'</div>'
            )
        _pos_flag = ""
        if _pos_pct >= 0.5:
            _pos_flag = (
                f'<div style="margin-top:8px;padding:6px 10px;background:rgba(251,146,60,.08);'
                f'border-left:3px solid #fb923c;border-radius:3px;font-size:10px;color:#fb923c">'
                f'⚠ POS-heavy — {_pos_pct:.0%} from terminal settlements. '
                f'Confirm business operations drive these credits, not float cycling.</div>'
            )
        _intel.append(
            f'<div style="flex:1;min-width:240px">'
            f'<div style="font-size:9px;letter-spacing:2px;color:#64748b;text-transform:uppercase;margin-bottom:8px">Income Source Breakdown</div>'
            f'{_bars}{_pos_flag}</div>'
        )

    # ── Recurring Income Detection ────────────────────────────────────────
    def _nkey(n: str) -> str:
        n = _re_intel.sub(r"[\d/\-:.,]", "", n.lower()).strip()
        return _re_intel.sub(r"\s+", " ", n)[:35]

    _grp_i: dict = _ddict_intel(lambda: _ddict_intel(float))
    for _t in _txns_i:
        _k = _nkey(_t["narration"])
        if len(_k) >= 6:
            _grp_i[_k][_t["ym"]] += _t["amount"]

    _recur = []
    for _k, _ym_map in _grp_i.items():
        if len(_ym_map) >= 3:
            _amts = list(_ym_map.values())
            _mean_a = sum(_amts) / len(_amts)
            if _mean_a > 0 and all(abs(a - _mean_a) / _mean_a <= 0.45 for a in _amts):
                _rep = next((t["narration"] for t in _txns_i if _nkey(t["narration"]) == _k), _k)
                _recur.append({"narr": _rep[:45], "months": len(_ym_map), "avg": _mean_a})
    _recur.sort(key=lambda x: -x["avg"])

    if _recur:
        _rrows = "".join(
            f'<tr>'
            f'<td style="color:#94a3b8;font-size:11px;padding:3px 4px">{rx["narr"]}</td>'
            f'<td style="color:#34d399;text-align:right;font-weight:700;padding:3px 4px;white-space:nowrap">{money(rx["avg"])}/mo</td>'
            f'<td style="color:#10b981;text-align:center;padding:3px 4px">{rx["months"]}mo</td>'
            f'</tr>'
            for rx in _recur[:8]
        )
        _intel.append(
            f'<div style="flex:1;min-width:240px">'
            f'<div style="font-size:9px;letter-spacing:2px;color:#34d399;text-transform:uppercase;margin-bottom:8px">Recurring Income Detected</div>'
            f'<table style="width:100%;border-collapse:collapse">'
            f'<thead><tr>'
            f'<th style="font-size:9px;color:#64748b;text-align:left;padding-bottom:4px">Narration</th>'
            f'<th style="font-size:9px;color:#64748b;text-align:right;padding-bottom:4px">Avg/Month</th>'
            f'<th style="font-size:9px;color:#64748b;text-align:center;padding-bottom:4px">Seen</th>'
            f'</tr></thead><tbody>{_rrows}</tbody></table>'
            f'</div>'
        )

    # ── Loan Cycling Flag ─────────────────────────────────────────────────
    # Signal 1: narration keywords in credit transactions
    _loan_kw = ["loan","disburs","credit facility","lending","overdraft","cash advance","borrow","float"]
    _loan_txns = [t for t in _txns_i if any(k in t["narration"].lower() for k in _loan_kw)]
    _loan_mos: dict = _ddict_intel(float)
    for _lt in _loan_txns:
        _loan_mos[_lt["ym"]] += _lt["amount"]

    # Signal 2: parser already flagged loan_disbursal category in rows
    _parser_loan_months = {
        r["ym"]: r["loan_disbursal"]
        for r in (st.session_state.rows_a or [])
        if r.get("loan_disbursal", 0) > 0
    }
    _all_loan_mos = set(_loan_mos) | set(_parser_loan_months)
    _loan_total = sum(_loan_mos.values()) + sum(_parser_loan_months.values())

    if len(_all_loan_mos) >= 2:
        # Show per-month detail
        _detail_rows = ""
        for _ym in sorted(_all_loan_mos):
            _narr_amt  = _loan_mos.get(_ym, 0)
            _parse_amt = _parser_loan_months.get(_ym, 0)
            _mo_total  = max(_narr_amt, _parse_amt)  # avoid double-counting same txns
            _lbl = ym_label(_ym)
            _detail_rows += (
                f'<tr>'
                f'<td style="color:#94a3b8;font-size:11px;padding:2px 4px">{_lbl}</td>'
                f'<td style="color:#f87171;font-weight:700;text-align:right;padding:2px 4px;white-space:nowrap">{money(_mo_total)}</td>'
                f'</tr>'
            )
        _intel.append(
            f'<div style="flex:1;min-width:240px">'
            f'<div style="padding:10px 14px;background:rgba(248,113,113,.07);'
            f'border-left:3px solid #f87171;border-radius:4px">'
            f'<div style="font-size:9px;letter-spacing:2px;color:#f87171;text-transform:uppercase;margin-bottom:6px">⚠ Loan Cycling Signal</div>'
            f'<div style="font-size:11px;color:#94a3b8;margin-bottom:8px">'
            f'Loan-like credits detected in <span style="color:#f87171;font-weight:700">{len(_all_loan_mos)} months</span>. '
            f'Verify disbursements are not recycling through the account.</div>'
            f'<table style="width:100%;border-collapse:collapse">{_detail_rows}</table>'
            f'</div></div>'
        )

    if _intel:
        st.markdown("---")
        st.markdown(
            '<div style="font-size:10px;letter-spacing:2px;color:#f59e0b;'
            'text-transform:uppercase;margin-bottom:14px">Statement Intelligence</div>',
            unsafe_allow_html=True,
        )
        st.markdown(
            f'<div style="display:flex;flex-wrap:wrap;gap:24px">{"".join(_intel)}</div>',
            unsafe_allow_html=True,
        )

    # ── Feature 12: Statement Integrity Check ────────────────────────────
    _integrity_flags: list[dict] = []   # {level: "high"|"medium", msg: str}

    # 1. Round-number concentration — natural transactions rarely land on ₦X00,000 exactly
    _round_100k = sum(1 for t in _txns_i if t["amount"] >= 100_000 and t["amount"] % 100_000 == 0)
    _big_txns   = sum(1 for t in _txns_i if t["amount"] >= 100_000)
    if _big_txns >= 5 and _round_100k / _big_txns > 0.55:
        _integrity_flags.append({
            "level": "high",
            "msg":   f"{_round_100k}/{_big_txns} large credits ({_round_100k/_big_txns:.0%}) are exact multiples "
                     f"of ₦100,000 — genuine business inflows rarely cluster on perfectly round figures.",
        })

    # 2. Duplicate transactions — identical amount + narration key in same month
    from collections import defaultdict as _dd_int
    _dup_map: dict = _dd_int(int)
    for _t in _txns_i:
        _dup_map[(_t["ym"], _t["narration"][:30].lower().strip(), int(_t["amount"]))] += 1
    _dups = {k: v for k, v in _dup_map.items() if v >= 3}
    if _dups:
        _dup_total = sum(_dups.values())
        _integrity_flags.append({
            "level": "high",
            "msg":   f"{len(_dups)} narration+amount combination(s) appear 3+ times in the same month "
                     f"({_dup_total} duplicate entries total) — may indicate copy-pasted transactions.",
        })

    # 3. Narration monotony — if >60% of credits share the same narration key
    _narr_counts: dict = _dd_int(int)
    for _t in _txns_i:
        _narr_counts[_nkey(_t["narration"])] += 1
    if _txns_i:
        _top_narr_cnt = max(_narr_counts.values())
        _mono_pct = _top_narr_cnt / len(_txns_i)
        if _mono_pct > 0.60 and len(_txns_i) >= 10:
            _top_narr_key = max(_narr_counts, key=lambda k: _narr_counts[k])
            _integrity_flags.append({
                "level": "medium",
                "msg":   f"{_mono_pct:.0%} of credit transactions share the same narration pattern "
                         f'("{_top_narr_key[:40]}") — unusual for a genuine business account.',
            })

    # 4. Month-over-month income cliff — last month > 3× previous average
    _ei_vals_chk = [r["eligible_income"] for r in (st.session_state.rows_a or []) if r["eligible_income"] > 0]
    if len(_ei_vals_chk) >= 3:
        _prev_avg_chk = sum(_ei_vals_chk[:-1]) / (len(_ei_vals_chk) - 1)
        _last_chk     = _ei_vals_chk[-1]
        if _prev_avg_chk > 0 and _last_chk / _prev_avg_chk > 3.0:
            _integrity_flags.append({
                "level": "high",
                "msg":   f"Last month eligible income ({money(_last_chk)}) is "
                         f"{_last_chk/_prev_avg_chk:.1f}× the prior average ({money(_prev_avg_chk)}) — "
                         f"sudden late-period spike is a common indicator of statement manipulation.",
            })

    # 5. Very few transactions but very large amounts (thin activity profile)
    _real_txns = [t for t in _txns_i if t["category"] == "real_credit"]
    _months_w_data = len(set(t["ym"] for t in _real_txns))
    if _months_w_data >= 2:
        _avg_txn_per_mo = len(_real_txns) / _months_w_data
        _total_real = sum(t["amount"] for t in _real_txns)
        _avg_txn_size = _total_real / len(_real_txns) if _real_txns else 0
        if _avg_txn_per_mo < 3 and _avg_txn_size > 500_000:
            _integrity_flags.append({
                "level": "medium",
                "msg":   f"Very thin activity: avg {_avg_txn_per_mo:.1f} real credit(s)/month, "
                         f"avg size {money(_avg_txn_size)}. "
                         f"Legitimate SME accounts typically show higher transaction frequency.",
            })

    if _integrity_flags:
        _high_cnt = sum(1 for f in _integrity_flags if f["level"] == "high")
        _hdr_col  = "#f87171" if _high_cnt else "#fb923c"
        _hdr_lbl  = f"{'🔴' if _high_cnt else '🟠'} Statement Integrity — {len(_integrity_flags)} signal{'s' if len(_integrity_flags)>1 else ''} detected"
        st.markdown("---")
        with st.expander(_hdr_lbl, expanded=_high_cnt > 0):
            st.markdown(
                f'<div style="font-size:11px;color:#94a3b8;margin-bottom:12px;line-height:1.7">'
                f'These checks flag statistical anomalies. They are indicators, not proof — '
                f'always cross-reference with physical documents before making a credit decision.</div>',
                unsafe_allow_html=True,
            )
            for _flg in _integrity_flags:
                _fc2 = "#f87171" if _flg["level"] == "high" else "#fb923c"
                st.markdown(
                    f'<div style="display:flex;gap:10px;margin-bottom:8px;padding:8px 12px;'
                    f'background:rgba({"248,113,113" if _flg["level"]=="high" else "251,146,60"},.06);'
                    f'border-left:3px solid {_fc2};border-radius:3px">'
                    f'<span style="color:{_fc2};font-size:14px;line-height:1.4">{"🔴" if _flg["level"]=="high" else "🟠"}</span>'
                    f'<span style="font-size:11px;color:#e2e8f0;line-height:1.6">{_flg["msg"]}</span>'
                    f'</div>',
                    unsafe_allow_html=True,
                )

# ════════════════════════════════════════════════════════════════════════════
# SECTION 00B — SECOND BANK STATEMENT
# ════════════════════════════════════════════════════════════════════════════
st.markdown("---")
st.markdown('<div class="sel-section-title">00B — Second Bank Statement &nbsp;<span style="color:#94a3b8;font-size:11px">— Optional</span></div>', unsafe_allow_html=True)
st.markdown('<div class="sel-caption">Upload a second bank statement. Nets are merged month-by-month across all available months from either statement.</div>', unsafe_allow_html=True)

col3, col4 = st.columns(2)
with col3:
    file_b = st.file_uploader("Upload Second Bank Statement (PDF or Excel)", type=["pdf","xlsx","xls"], key="upload_b")
with col4:
    pw_b   = st.text_input("PDF Password", type="password", key="pw_b", placeholder="Leave blank if not encrypted")
    if st.button("Extract & Merge with First Statement", key="btn_extract_b"):
        if not file_b:
            st.error("Please select a second PDF file first.")
        elif not st.session_state.buckets_a:
            st.error("Please extract the first statement first.")
        else:
            with st.spinner("Extracting second statement..."):
                # ── Read file bytes ONCE ──────────────────────────────────────
                _pdf_bytes_b = file_b.getvalue()
                _size_mb_b   = len(_pdf_bytes_b) / 1_048_576

                if _size_mb_b > 30:
                    st.error(
                        f"⚠️ This PDF is {_size_mb_b:.1f} MB — too large to process safely on "
                        f"Streamlit Cloud (limit: 30 MB). Please export a shorter date range "
                        f"(6 months) from your bank portal and upload again."
                    )
                    del _pdf_bytes_b
                    gc.collect()
                else:
                    if _size_mb_b > 10:
                        st.warning(
                            f"Large file ({_size_mb_b:.1f} MB, ~{round(_size_mb_b * 10):.0f} pages). "
                            f"Processing in chunks — this may take up to 60 seconds…"
                        )

                    track("upload", session=_SID, officer=_OFFICER, filename=file_b.name,
                          size_kb=round(_size_mb_b * 1024, 1), slot="B")
                    try:
                        buckets_b, summary_b, bank_b, name_b, txns_b, debit_txns_b = parse_transactions(
                            _pdf_bytes_b, pw_b, filename=file_b.name
                        )
                        rows_b = monthly_analysis(buckets_b, summary_b)
                        st.session_state.buckets_b       = buckets_b
                        st.session_state.summary_b       = summary_b
                        st.session_state.bank_b          = bank_b
                        st.session_state.bank_override_b = bank_b   # seed manual override
                        st.session_state.name_b          = name_b
                        st.session_state.rows_b          = rows_b
                        st.session_state.txns_b          = txns_b
                        st.session_state.debit_txns_b    = debit_txns_b

                        _reused_text_b = get_last_full_text()

                        try:
                            if file_b.name.lower().endswith((".xlsx", ".xls")):
                                st.session_state.account_no_b = extract_account_no_excel(_pdf_bytes_b)
                            else:
                                st.session_state.account_no_b = extract_account_no(_reused_text_b)
                        except Exception:
                            st.session_state.account_no_b = ""

                        st.success(f"Second statement extracted: {bank_b} — {name_b or 'account holder'}")
                        _txn_count_b = sum(b.get("count", 0) for b in buckets_b.values())
                        _gross_tot_b = sum(b.get("gross", 0) for b in buckets_b.values())
                        track("parse_success", session=_SID, officer=_OFFICER, bank=bank_b,
                              filename=file_b.name, txn_count=_txn_count_b,
                              gross_total=round(_gross_tot_b, 2),
                              months=len([r for r in rows_b if r["gross"] > 0]), slot="B")

                        # ── Accuracy Verification — reuses cached text, no re-parse ──
                        _conf_pct_b = None
                        is_pdf_b = not file_b.name.lower().endswith((".xlsx", ".xls"))
                        if is_pdf_b and buckets_b and _reused_text_b:
                            try:
                                stated_b  = extract_stated_totals(_reused_text_b)
                                verdict_b = verify_extraction_accuracy(buckets_b, stated_b)
                                if verdict_b["pct_match"] is not None:
                                    _pct_match = verdict_b["pct_match"]
                                    _conf_pct_b = _pct_match
                                    ext = verdict_b["extracted"]
                                    stl = verdict_b["stated_total"]
                                    colour = ("#34d399" if _pct_match >= 95
                                              else "#fb923c" if _pct_match >= 90
                                              else "#f87171")
                                    st.markdown(
                                        f'<div style="background:rgba(0,0,0,.2);border:1px solid {colour}33;'
                                        f'border-radius:3px;padding:10px 14px;margin-top:8px;font-size:12px;">'
                                        f'<span style="color:{colour};font-weight:700">▶ Accuracy Check — {_pct_match}% match</span>'
                                        f'&nbsp;&nbsp;<span style="color:#64748b">Extracted ₦{ext:,.0f} vs '
                                        f'stated ₦{stl:,.0f}</span><br>'
                                        f'<span style="color:#94a3b8;font-size:11px">{verdict_b["message"]}</span>'
                                        f'</div>',
                                        unsafe_allow_html=True,
                                    )
                            except Exception:
                                pass

                        # ── Parser Confidence Report (Statement B) ────────────
                        if buckets_b:
                            try:
                                render_parse_confidence(
                                    buckets_b, _conf_pct_b, bank_b,
                                    filename=file_b.name, session=_SID,
                                    officer=_OFFICER,
                                )
                            except Exception:
                                pass

                    except Exception as e:
                        track("parse_error", session=_SID, officer=_OFFICER, filename=file_b.name,
                              error=str(e), error_type=type(e).__name__, slot="B")
                        if "EOF marker not found" in str(e) or "Unexpected EOF" in str(e):
                            st.error(
                                "This PDF appears to be corrupted or incomplete. "
                                "Please download the bank statement again from your bank app/portal."
                            )
                        else:
                            st.error(f"Error: {e}")
                    finally:
                        del _pdf_bytes_b
                        try:
                            del _reused_text_b
                        except NameError:
                            pass
                        gc.collect()

# Show Statement B own analysis
if st.session_state.rows_b:
    _bank_override_selector("b")
    import datetime as _dt2
    _today_b = _dt2.date.today().strftime("%Y-%m")
    rows_b_own = [r for r in st.session_state.rows_b if r["ym"] < _today_b and r["gross"] > 0][-N_MONTHS:]
    if rows_b_own:
        _bname = st.session_state.bank_b or "Statement 2"
        _bacco = st.session_state.name_b or ""
        st.markdown(
            f'<div style="font-size:10px;letter-spacing:2px;color:#f59e0b;text-transform:uppercase;'
            f'margin:16px 0 6px">▷ Statement 2 Breakdown — {_bname}'
            f'{(" — " + _bacco) if _bacco else ""}</div>',
            unsafe_allow_html=True,
        )
        _b_has_rev  = any(r["reversal"]       > 0 for r in rows_b_own)
        _b_has_nb   = any(r["non_business"]   > 0 for r in rows_b_own)
        _b_has_loan = any(r["loan_disbursal"] > 0 for r in rows_b_own)
        _b_hdr = ('<tr>'
                  '<th class="col-gross" style="text-align:left">Month</th>'
                  '<th class="col-gross">Total Inflow</th>')
        if _b_has_rev:   _b_hdr += '<th class="col-rev">Reversals</th>'
        if _b_has_nb:    _b_hdr += '<th class="col-nonbiz">Non-Business</th>'
        if _b_has_loan:  _b_hdr += '<th class="col-loan">Loan Disbursals</th>'
        _b_hdr += '<th class="col-net">Eligible Income</th></tr>'
        _b_body = ""
        _btg = _btr = _btn = _btl = _bte = 0.0
        for r in rows_b_own:
            _btg += r["gross"]; _btr += r["reversal"]; _btn += r["non_business"]
            _btl += r["loan_disbursal"]; _bte += r["eligible_income"]
            _b_body += f'<tr><td>{r["label"]}</td><td class="col-gross">{money(r["gross"])}</td>'
            if _b_has_rev:   _b_body += f'<td class="col-rev">{("-"+money(r["reversal"])) if r["reversal"] > 0 else "—"}</td>'
            if _b_has_nb:    _b_body += f'<td class="col-nonbiz">{("-"+money(r["non_business"])) if r["non_business"] > 0 else "—"}</td>'
            if _b_has_loan:  _b_body += f'<td class="col-loan">{("-"+money(r["loan_disbursal"])) if r["loan_disbursal"] > 0 else "—"}</td>'
            _b_body += f'<td class="col-net">{money(r["eligible_income"])}</td></tr>'
        _b_foot = (f'<tfoot><tr>'
                   f'<td style="color:#64748b;font-size:10px;text-transform:uppercase">Total</td>'
                   f'<td class="col-gross" style="border-top:1px solid #1a3d2b">{money(_btg)}</td>')
        if _b_has_rev:   _b_foot += f'<td class="col-rev" style="border-top:1px solid #1a3d2b">{("-"+money(_btr)) if _btr > 0 else "—"}</td>'
        if _b_has_nb:    _b_foot += f'<td class="col-nonbiz" style="border-top:1px solid #1a3d2b">{("-"+money(_btn)) if _btn > 0 else "—"}</td>'
        if _b_has_loan:  _b_foot += f'<td class="col-loan" style="border-top:1px solid #1a3d2b">{("-"+money(_btl)) if _btl > 0 else "—"}</td>'
        _b_foot += (f'<td class="col-net" style="border-top:1px solid #1a3d2b;font-size:14px">{money(_bte)}</td>'
                    f'</tr></tfoot>')
        st.markdown(
            f'<table class="preview-table"><thead>{_b_hdr}</thead><tbody>{_b_body}</tbody>{_b_foot}</table>',
            unsafe_allow_html=True,
        )

# Show merged preview
if st.session_state.rows_a and st.session_state.rows_b:
    import datetime
    today = datetime.date.today().strftime("%Y-%m")
    rows_a_map = {r["ym"]: r for r in st.session_state.rows_a if r["ym"] < today and r["gross"] > 0}
    rows_b_map = {r["ym"]: r for r in st.session_state.rows_b if r["ym"] < today and r["gross"] > 0}
    # Union — keep all months from either statement
    all_months = sorted(set(rows_a_map) | set(rows_b_map))[-N_MONTHS:]

    if all_months:
        # ── Consolidated View header ──────────────────────────────────────
        _overlap_months = sorted(set(rows_a_map) & set(rows_b_map))
        _name_a_disp = st.session_state.get("name_a") or "Statement 1"
        _name_b_disp = st.session_state.get("name_b") or "Statement 2"
        _bank_a_disp = st.session_state.get("bank_a") or ""
        _bank_b_disp = st.session_state.get("bank_b") or ""
        _label_a = f"{_name_a_disp}" + (f" ({_bank_a_disp})" if _bank_a_disp else "")
        _label_b = f"{_name_b_disp}" + (f" ({_bank_b_disp})" if _bank_b_disp else "")

        st.markdown(
            '<div style="font-size:10px;letter-spacing:2px;color:#34d399;'
            'text-transform:uppercase;margin:16px 0 4px">'
            '▷ Consolidated View — Both Accounts Combined</div>',
            unsafe_allow_html=True,
        )

        html = ('<table class="preview-table"><thead><tr>'
                f'<th style="text-align:left">Month</th>'
                f'<th class="col-gross" title="{_label_a}">Acct 1 Income</th>'
                f'<th style="text-align:right;color:#f59e0b" title="{_label_b}">Acct 2 Income</th>'
                f'<th class="col-net">Combined Net</th>'
                f'<th style="text-align:center;font-size:9px;color:#64748b">Source</th>'
                '</tr></thead><tbody>')
        tA = tB = tC = 0
        for ym in all_months:
            rA = rows_a_map.get(ym)
            rB = rows_b_map.get(ym)
            netA = rA["eligible_income"] if rA else 0
            netB = rB["eligible_income"] if rB else 0
            combined = netA + netB
            tA += netA; tB += netB; tC += combined
            if rA and rB:
                _src_label = '<span style="font-size:9px;color:#a78bfa">Both</span>'
            elif rA:
                _src_label = '<span style="font-size:9px;color:#64748b">Acct 1</span>'
            else:
                _src_label = '<span style="font-size:9px;color:#f59e0b">Acct 2</span>'
            html += (f'<tr><td>{ym_label(ym)}</td>'
                     f'<td class="col-gross">{money(netA) if netA else "—"}</td>'
                     f'<td style="text-align:right;color:#f59e0b">{money(netB) if netB else "—"}</td>'
                     f'<td class="col-net">{money(combined)}</td>'
                     f'<td style="text-align:center">{_src_label}</td></tr>')

        # Contribution % per account
        _pct_a = round(tA / tC * 100) if tC else 0
        _pct_b = round(tB / tC * 100) if tC else 0
        html += (f'</tbody><tfoot><tr>'
                 f'<td style="color:#64748b;font-size:10px;text-transform:uppercase;letter-spacing:1px">Total</td>'
                 f'<td class="col-gross" style="border-top:1px solid #1a3d2b">'
                 f'{money(tA)}<br><span style="font-size:9px;color:#64748b">{_pct_a}% of combined</span></td>'
                 f'<td style="text-align:right;color:#f59e0b;border-top:1px solid #1a3d2b">'
                 f'{money(tB)}<br><span style="font-size:9px;color:#a16207">{_pct_b}% of combined</span></td>'
                 f'<td class="col-net" style="border-top:1px solid #1a3d2b;font-size:14px">{money(tC)}</td>'
                 f'<td style="border-top:1px solid #1a3d2b"></td>'
                 f'</tr></tfoot></table>')

        # Footer notes
        _footer_notes = [f'{len(all_months)} months shown — union of both accounts']
        if _overlap_months:
            _footer_notes.append(
                f'⚑ {len(_overlap_months)} overlapping month(s) ({", ".join(ym_label(m) for m in _overlap_months)}) '
                f'— both accounts active; totals are additive (dedup handled above if same-owner transfers detected)'
            )
        html += ''.join(
            f'<div style="font-size:10px;color:#64748b;margin-top:5px">{n}</div>'
            for n in _footer_notes
        )
        st.markdown(html, unsafe_allow_html=True)

# ── Feature 6: Multi-Account Transfer Deduplication ──────────────────────────
if st.session_state.txns_a and st.session_state.txns_b:
    import re as _re_dd
    from collections import defaultdict as _dd_dd

    _name_a_lc = (st.session_state.name_a or "").lower().strip()
    _name_b_lc = (st.session_state.name_b or "").lower().strip()

    # Build (ym, rounded_amount) lookup for each statement
    def _dd_key(amount: float) -> int:
        return round(amount / 1000) * 1000   # bucket to nearest ₦1,000

    _grp_a: dict = _dd_dd(list)
    for _t in st.session_state.txns_a:
        _grp_a[(_t["ym"], _dd_key(_t["amount"]))].append(_t)

    _grp_b: dict = _dd_dd(list)
    for _t in st.session_state.txns_b:
        _grp_b[(_t["ym"], _dd_key(_t["amount"]))].append(_t)

    # Find overlapping keys (same month + ~same amount in both statements)
    _overlap_keys = set(_grp_a) & set(_grp_b)

    _dedup_flags = []
    for _ok in sorted(_overlap_keys):
        _ym, _amt = _ok
        for _ta in _grp_a[_ok]:
            for _tb in _grp_b[_ok]:
                _narr_a = _ta["narration"].lower()
                _narr_b = _tb["narration"].lower()
                # Stronger signal: cross-narration name match
                _cross = (
                    (_name_b_lc and len(_name_b_lc) > 3 and _name_b_lc in _narr_a) or
                    (_name_a_lc and len(_name_a_lc) > 3 and _name_a_lc in _narr_b)
                )
                # Weaker signal: transfer keywords on both sides
                _kw_a = any(k in _narr_a for k in ["transfer","trf","self","own"])
                _kw_b = any(k in _narr_b for k in ["transfer","trf","self","own"])
                if _cross or (_kw_a and _kw_b):
                    _dedup_flags.append({
                        "ym": _ym, "amount": _ta["amount"],
                        "narr_a": _ta["narration"][:50],
                        "narr_b": _tb["narration"][:50],
                        "confidence": "High" if _cross else "Medium",
                    })

    if _dedup_flags:
        _dd_total = sum(d["amount"] for d in _dedup_flags)
        _dd_rows_html = "".join(
            f'<tr>'
            f'<td style="color:#10b981">{d["ym"]}</td>'
            f'<td style="color:#f87171;text-align:right;font-weight:700">{money(d["amount"])}</td>'
            f'<td style="color:#94a3b8;font-size:11px">{d["narr_a"]}</td>'
            f'<td style="color:#f59e0b;font-size:11px">{d["narr_b"]}</td>'
            f'<td style="text-align:center"><span style="background:rgba({"248,113,113" if d["confidence"]=="High" else "251,146,60"},.12);'
            f'border-radius:3px;padding:1px 6px;font-size:9px;color:{"#f87171" if d["confidence"]=="High" else "#fb923c"}">'
            f'{d["confidence"]}</span></td>'
            f'</tr>'
            for d in _dedup_flags[:20]
        )
        st.markdown("---")
        st.markdown(
            f'<div style="font-size:10px;letter-spacing:2px;color:#f87171;'
            f'text-transform:uppercase;margin-bottom:8px">⚠ Potential Inter-Account Transfers Detected</div>',
            unsafe_allow_html=True,
        )
        st.markdown(
            f'<div style="font-size:11px;color:#94a3b8;margin-bottom:10px;line-height:1.7">'
            f'<span style="color:#f87171;font-weight:700">{len(_dedup_flags)} credit(s)</span> '
            f'totalling <span style="color:#f87171;font-weight:700">{money(_dd_total)}</span> '
            f'may be the same money appearing in both statements. '
            f'Review and manually deduct from the inflow grid if confirmed.</div>',
            unsafe_allow_html=True,
        )
        st.markdown(
            f'<table class="preview-table">'
            f'<thead><tr>'
            f'<th style="text-align:left">Month</th>'
            f'<th style="text-align:right;color:#f87171">Amount</th>'
            f'<th style="text-align:left">Stmt 1 Narration</th>'
            f'<th style="text-align:left;color:#f59e0b">Stmt 2 Narration</th>'
            f'<th style="text-align:center">Confidence</th>'
            f'</tr></thead><tbody>{_dd_rows_html}</tbody></table>',
            unsafe_allow_html=True,
        )
        if len(_dedup_flags) > 20:
            st.markdown(
                f'<div style="font-size:10px;color:#64748b;margin-top:4px">'
                f'… and {len(_dedup_flags)-20} more. Use Transaction Search above to investigate.</div>',
                unsafe_allow_html=True,
            )

# ── Debit Visibility Panel — Statement B ─────────────────────────────────────
if st.session_state.get("debit_txns_b"):
    _render_debit_panel(st.session_state.debit_txns_b, label="Statement 2")

# ── Transaction Search — Statement B ─────────────────────────────────────────
if st.session_state.txns_b:
    st.markdown(
        '<div style="margin-top:16px;font-size:10px;letter-spacing:2px;color:#f59e0b;'
        'text-transform:uppercase;margin-bottom:6px">Transaction Search — Statement 2</div>',
        unsafe_allow_html=True,
    )
    search_b = st.text_input(
        "Search keyword (statement 2)",
        key="search_b",
        placeholder="e.g. salary, transfer, POS...",
        label_visibility="collapsed",
    )
    if search_b and search_b.strip():
        kw_b = search_b.strip().lower()
        matches_b = [t for t in st.session_state.txns_b if kw_b in t["narration"].lower()]
        if matches_b:
            total_b = sum(t["amount"] for t in matches_b)
            st.markdown(
                f'<div style="font-size:11px;color:#64748b;margin-bottom:6px">'
                f'Found <span style="color:#f59e0b;font-weight:700">{len(matches_b)}</span> credit(s) '
                f'matching <em>"{search_b}"</em> — '
                f'Total: <span style="color:#34d399;font-weight:700">{money(total_b)}</span></div>',
                unsafe_allow_html=True,
            )
            _CAT_COLOUR = {
                "real_credit": "#34d399", "self_transfer": "#fb923c",
                "reversal": "#a78bfa", "non_business": "#64748b", "loan_disbursal": "#fbbf24",
            }
            rows_b_html = ""
            for t in matches_b[:100]:
                clr = _CAT_COLOUR.get(t["category"], "#e2e8f0")
                cat_label = t["category"].replace("_", " ").title()
                rows_b_html += (
                    f'<tr><td style="color:#f59e0b">{t["ym"]}</td>'
                    f'<td style="color:#34d399;text-align:right">{money(t["amount"])}</td>'
                    f'<td><span style="background:rgba(255,255,255,.05);padding:1px 6px;'
                    f'border-radius:3px;font-size:9px;color:{clr}">{cat_label}</span></td>'
                    f'<td style="color:#94a3b8;font-size:11px">{t["narration"][:80]}</td></tr>'
                )
            if len(matches_b) > 100:
                rows_b_html += f'<tr><td colspan="4" style="color:#64748b;font-size:10px;padding-top:6px">... and {len(matches_b)-100} more</td></tr>'
            st.markdown(
                f'<table class="preview-table" style="margin-top:4px">'
                f'<thead><tr>'
                f'<th style="text-align:left">Month</th>'
                f'<th style="text-align:right">Amount</th>'
                f'<th style="text-align:left">Category</th>'
                f'<th style="text-align:left">Narration</th>'
                f'</tr></thead><tbody>{rows_b_html}</tbody></table>',
                unsafe_allow_html=True,
            )
        else:
            st.markdown(
                f'<div style="color:#64748b;font-size:12px;padding:8px 0">No credit transactions found matching <em>"{search_b}"</em>.</div>',
                unsafe_allow_html=True,
            )


# ════════════════════════════════════════════════════════════════════════════
# SECTION 01 — FIRSTCENTRAL CREDIT REPORT
# ════════════════════════════════════════════════════════════════════════════
st.markdown("---")
st.markdown('<div class="sel-section-title">01 — FirstCentral Credit Report &nbsp;<span style="color:#94a3b8;font-size:11px">— External obligations</span></div>', unsafe_allow_html=True)
st.markdown('<div class="sel-caption">Upload the FirstCentral consumer credit report. Closed accounts are ignored. Open performing accounts feed into monthly obligations.</div>', unsafe_allow_html=True)

col5, col6 = st.columns(2)
with col5:
    credit_file = st.file_uploader("Upload FirstCentral Report (PDF)", type=["pdf"], key="credit_upload")
with col6:
    credit_pw = st.text_input("Credit Report Password", type="password", key="credit_pw", placeholder="Leave blank if not encrypted")
    if st.button("Extract External Obligations", key="btn_credit"):
        if not credit_file:
            st.error("Please select a FirstCentral PDF first.")
        else:
            with st.spinner("Parsing credit report..."):
                try:
                    data = parse_firstcentral(credit_file.getvalue(), credit_pw)
                    st.session_state.credit_data = data
                    total = data["total_monthly_obligation"]
                    bad   = len(data["bad_credit_accounts"])
                    st.success(f"Extracted {len(data['records'])} active accounts — Monthly obligations: {money(total)}")
                    if bad > 0:
                        st.markdown(f'<div class="banner-bad">⚠ Bad credit flag: {bad} account{"s" if bad>1 else ""} marked lost, derogatory, or delinquent with outstanding balance above ₦10,000.</div>', unsafe_allow_html=True)
                    else:
                        st.markdown('<div class="banner-good">✓ No bad-credit triggers found in the active FirstCentral accounts.</div>', unsafe_allow_html=True)
                except Exception as e:
                    st.error(f"Error: {e}")

# Show credit table
if st.session_state.credit_data:
    data = st.session_state.credit_data
    records = data["records"]
    total = data["total_monthly_obligation"]
    if records:
        col_m1, col_m2, col_m3 = st.columns(3)
        col_m1.metric("External Monthly Obligations", money(total))
        col_m2.metric("Active Accounts", len(records))
        col_m3.metric("Bad Credit Flags", len(data["bad_credit_accounts"]))

        html = """<table class="credit-table"><thead><tr>
            <th>Subscriber</th><th>Account Status</th><th>Facility Classification</th>
            <th>Instalment</th><th>Outstanding</th><th>Tenor</th>
            <th>Monthly Obligation</th><th>Rule</th>
        </tr></thead><tbody>"""
        for r in records:
            sts_cls  = "badge-red"    if r.is_bad_credit else "badge-blue"
            cls_cls  = "badge-orange" if r.facility_classification.lower() != "performing" else "badge-green"
            rule = ("Flagged as bad credit" if r.is_bad_credit
                    else "Derived from balance / tenor" if r.derived_from_balance
                    else "Instalment amount used" if r.include_in_obligation
                    else "Monitoring only")
            mo_str = money(r.monthly_obligation) if r.include_in_obligation else "—"
            tenor_str = f"{r.tenor_months} mo" if r.tenor_months else r.loan_duration_days
            html += (f'<tr>'
                     f'<td>{r.subscriber_name}<div style="color:#64748b;margin-top:3px;font-size:10px">{r.account_number}</div></td>'
                     f'<td><span class="badge {sts_cls}">{r.account_status}</span></td>'
                     f'<td><span class="badge {cls_cls}">{r.facility_classification}</span></td>'
                     f'<td>{money(r.instalment_amount)}</td>'
                     f'<td>{money(r.outstanding_balance)}</td>'
                     f'<td>{tenor_str}</td>'
                     f'<td style="color:#fbbf24;font-weight:700">{mo_str}</td>'
                     f'<td style="color:#64748b;font-size:10px">{rule}</td>'
                     f'</tr>')
        html += (f'</tbody><tfoot><tr>'
                 f'<td colspan="6" style="color:#64748b">Total monthly external obligations</td>'
                 f'<td style="color:#fbbf24;font-weight:700">{money(total)}</td>'
                 f'<td>{len(data["bad_credit_accounts"])} bad flag(s)' if data["bad_credit_accounts"] else '<td>Clear'
                 + '</td></tr></tfoot></table>')
        st.markdown(html, unsafe_allow_html=True)


# ════════════════════════════════════════════════════════════════════════════
# SECTION 02 — MONTHLY INFLOWS (editable)
# ════════════════════════════════════════════════════════════════════════════
st.markdown("---")
st.markdown(f'<div class="sel-section-title">02 — Monthly Inflows (Last {N_MONTHS} Months)</div>', unsafe_allow_html=True)
st.markdown('<div class="sel-caption">Gross credits auto-filled from bank statement. Adjust deductions or add extra manual deductions below.</div>', unsafe_allow_html=True)

import datetime
today = datetime.date.today()

# Determine which rows to pre-fill
def get_prefill_rows():
    """Get last N_MONTHS completed months from merged or single statement.
    Uses UNION of both statements so that months present in only one
    statement are still included (not dropped by an intersection filter).
    """
    today_ym = today.strftime("%Y-%m")
    if st.session_state.rows_a and st.session_state.rows_b:
        rows_a_map = {r["ym"]: r for r in st.session_state.rows_a if r["ym"] < today_ym and r["gross"] > 0}
        rows_b_map = {r["ym"]: r for r in st.session_state.rows_b if r["ym"] < today_ym and r["gross"] > 0}
        # Union — include every month present in either statement
        all_months = sorted(set(rows_a_map) | set(rows_b_map))[-N_MONTHS:]
        if all_months:
            merged = []
            for ym in all_months:
                rA = rows_a_map.get(ym)
                rB = rows_b_map.get(ym)
                eiA = rA["eligible_income"] if rA else 0
                eiB = rB["eligible_income"] if rB else 0
                merged.append({
                    "ym": ym, "label": ym_label(ym),
                    "gross": eiA + eiB,
                    "deductions": 0,
                    "count": max(rA["count"] if rA else 0, rB["count"] if rB else 0),
                })
            return merged
    elif st.session_state.rows_a:
        return [r for r in st.session_state.rows_a if r["ym"] < today_ym and r["gross"] > 0][-N_MONTHS:]
    return None

prefill = get_prefill_rows()

# ── Sync prefill into session state so Streamlit widget cache is updated ──
# This fixes the bug where Oct 25 shows ₦0 because Streamlit cached the
# previous widget value. We push fresh values into st.session_state BEFORE
# the widgets render so they pick up the correct numbers.
if prefill:
    for i, r in enumerate(prefill[:6]):
        gross_key  = f"gross_{i}"
        deduct_key = f"deduct_{i}"
        count_key  = f"count_{i}"
        # Only overwrite if the extraction just happened (rows_a was just set)
        # Use a fingerprint to detect when new data arrives
        fp = f"{r['ym']}_{r.get('gross',0):.0f}"
        if st.session_state.get(f"fp_{i}") != fp:
            st.session_state[gross_key]  = float(r.get("gross", 0))
            st.session_state[deduct_key] = float(r.get("deductions", 0))
            st.session_state[count_key]  = max(1, int(r.get("count", 12)))
            st.session_state[f"fp_{i}"]  = fp

# Build default month labels (N_MONTHS entries, oldest first)
def default_months():
    months = []
    for i in range(1, N_MONTHS + 1):
        month = today.month - (N_MONTHS + 1 - i)
        year  = today.year
        while month <= 0:
            month += 12
            year  -= 1
        months.append((f"{year}-{str(month).zfill(2)}", ym_label(f"{year}-{str(month).zfill(2)}")))
    return months

default_m = default_months()
inflow_data = []

# ── Column header row ─────────────────────────────────────────────────────────
h1, h2, h3, h4, h5, h6 = st.columns([1.2, 1.8, 1.5, 1.5, 1.5, 0.8])
with h1: st.markdown('<div style="font-size:12px;letter-spacing:1.5px;color:#e2e8f0;text-transform:uppercase;font-weight:800;padding-bottom:4px;border-bottom:2px solid #10b981">Month</div>', unsafe_allow_html=True)
with h2: st.markdown('<div style="font-size:12px;letter-spacing:1.5px;color:#e2e8f0;text-transform:uppercase;font-weight:800;padding-bottom:4px;border-bottom:2px solid #10b981">Gross Credit ₦</div>', unsafe_allow_html=True)
with h3: st.markdown('<div style="font-size:12px;letter-spacing:1.5px;color:#e2e8f0;text-transform:uppercase;font-weight:800;padding-bottom:4px;border-bottom:2px solid #10b981">Deductions ₦</div>', unsafe_allow_html=True)
with h4: st.markdown('<div style="font-size:12px;letter-spacing:1.5px;color:#fb923c;text-transform:uppercase;font-weight:800;padding-bottom:4px;border-bottom:2px solid #fb923c">Extra Deduction ₦</div>', unsafe_allow_html=True)
with h5: st.markdown('<div style="font-size:12px;letter-spacing:1.5px;color:#34d399;text-transform:uppercase;font-weight:800;padding-bottom:4px;border-bottom:2px solid #34d399">Net Inflow ₦</div>', unsafe_allow_html=True)
with h6: st.markdown('<div style="font-size:12px;letter-spacing:1.5px;color:#e2e8f0;text-transform:uppercase;font-weight:800;padding-bottom:4px;border-bottom:2px solid #10b981">Count</div>', unsafe_allow_html=True)

for i in range(N_MONTHS):
    if prefill and i < len(prefill):
        r     = prefill[i]
        label = r["label"]
    else:
        label = default_m[i][1]

    c1, c2, c3, c4, c5, c6 = st.columns([1.2, 1.8, 1.5, 1.5, 1.5, 0.8])
    with c1:
        st.markdown(f'<div style="padding-top:8px;font-weight:700;color:#10b981">{label}</div>',
                    unsafe_allow_html=True)
    with c2:
        g = st.number_input("Gross Credit (₦)", min_value=0.0, step=1000.0,
                            key=f"gross_{i}", label_visibility="collapsed")
    with c3:
        d = st.number_input("Deductions (₦)", min_value=0.0, step=1000.0,
                            key=f"deduct_{i}", label_visibility="collapsed")
    with c4:
        x = st.number_input("Extra Deduction (₦)", min_value=0.0, step=1000.0,
                            key=f"extra_{i}", label_visibility="collapsed")
    with c5:
        net = max(g - d - x, 0)
        st.markdown(
            f'<div style="color:#34d399;font-weight:700;padding-top:8px;font-size:13px">{money(net)}</div>',
            unsafe_allow_html=True,
        )
    with c6:
        cnt = st.number_input("Count", min_value=1, max_value=9999,
                              key=f"count_{i}", label_visibility="collapsed")
    inflow_data.append({"label": label, "gross": g, "deduct": d, "extra": x, "net": net, "count": cnt})

st.markdown(
    '<div style="font-size:12px;color:#cbd5e1;font-weight:600;margin-top:8px;padding:10px 14px;'
    'background:rgba(0,0,0,.25);border-left:3px solid #fb923c;border-radius:3px;line-height:1.6">'
    '<strong style="color:#fb923c;font-size:13px">Extra Deduction</strong> — use this to manually subtract any amount you\'ve identified '
    'from the search above (e.g. a recurring transfer you want excluded from income). '
    'Auto Deductions are pre-filled from the bank statement parser (reversals, loan disbursals, non-business).'
    '</div>',
    unsafe_allow_html=True,
)

# ── Applicant History Search ──────────────────────────────────────────────────
with st.expander("🔍  Applicant History Search — look up any previous assessment", expanded=False):
    _hs_col1, _hs_col2 = st.columns([3, 1])
    with _hs_col1:
        _hs_query = st.text_input(
            "Search by applicant name",
            placeholder="e.g. JOHN DOE or CHUKWU ENTERPRISE",
            key="history_search_query",
            label_visibility="collapsed",
        )
    with _hs_col2:
        _hs_go = st.button("Search", key="history_search_btn", use_container_width=True)

    if _hs_go and _hs_query.strip():
        _hs_results = get_history(_hs_query.strip())
        if not _hs_results:
            st.markdown(
                f'<div style="font-size:12px;color:#64748b;padding:8px 0">'
                f'No past assessments found for "{_hs_query.strip()}".</div>',
                unsafe_allow_html=True,
            )
        else:
            _hs_rows_html = ""
            for _hr in _hs_results:
                _hr_col = "#34d399" if _hr.get("approved") else "#f87171"
                _hr_dec = "✅ Approved" if _hr.get("approved") else "❌ Below Min"
                _hs_rows_html += (
                    f'<tr>'
                    f'<td style="font-size:11px;color:#94a3b8">{(_hr.get("ts") or "")[:10]}</td>'
                    f'<td style="font-size:11px">{_hr.get("bank","—")}</td>'
                    f'<td style="font-size:11px">{_hr.get("location","—")} / {_hr.get("product","—")}</td>'
                    f'<td style="text-align:right;font-size:11px">{money(_hr.get("avg_income",0))}</td>'
                    f'<td style="text-align:right;font-size:11px">{money(_hr.get("max_loan",0))}</td>'
                    f'<td style="text-align:center;font-size:11px">{_hr.get("tenor","—")} mo</td>'
                    f'<td style="text-align:center;color:{_hr_col};font-size:11px;font-weight:700">{_hr_dec}</td>'
                    f'</tr>'
                )
            st.markdown(
                f'<div style="font-size:10px;color:#64748b;margin-bottom:6px">'
                f'{len(_hs_results)} assessment(s) found for <strong style="color:#e2e8f0">'
                f'{_hs_query.strip()}</strong></div>'
                f'<table class="preview-table"><thead><tr>'
                f'<th style="text-align:left">Date</th>'
                f'<th>Bank</th>'
                f'<th>Location / Product</th>'
                f'<th style="text-align:right">Avg Income</th>'
                f'<th style="text-align:right">Max Loan</th>'
                f'<th style="text-align:center">Tenor</th>'
                f'<th style="text-align:center">Decision</th>'
                f'</tr></thead><tbody>{_hs_rows_html}</tbody></table>',
                unsafe_allow_html=True,
            )

# ════════════════════════════════════════════════════════════════════════════
# SECTION 03 — LOAN PARAMETERS
# ════════════════════════════════════════════════════════════════════════════
st.markdown("---")
st.markdown('<div class="sel-section-title">03 — Loan Parameters</div>', unsafe_allow_html=True)

p1, p2, p3, p4 = st.columns(4)
with p1: location  = st.selectbox("Location", ["Lagos","Outside Lagos","Expansion"])
with p2: prod_type = st.selectbox("Product Type", ["NTB","RENEWAL","TOP-UP"])
with p3: tenor     = st.selectbox("Tenor (Months)", list(range(2,13)), index=4)
with p4:
    default_other = 0.0
    if st.session_state.credit_data:
        default_other = st.session_state.credit_data["total_monthly_obligation"]
    other_loans = st.number_input("Other Monthly Loan Repayments (₦)", min_value=0.0,
                                   value=float(default_other), step=1000.0)

r1, r2 = st.columns(2)
with r1: req_loan   = st.number_input("Requested Loan Amount (₦) — Optional", min_value=0.0, value=0.0, step=10000.0)
with r2: manual_rate= st.number_input("Manual Interest Rate (%) — Optional Override", min_value=0.0, value=0.0, step=0.01,
                                       help="If entered, overrides the rate grid. Leave at 0 to use rate grid.")

# ── What-if Reverse Calculator ────────────────────────────────────────────────
with st.expander("🔁  What-if: How much income is needed to qualify for ₦X?", expanded=False):
    st.markdown(
        '<div style="font-size:12px;color:#cbd5e1;font-weight:600;margin-bottom:12px;line-height:1.6">'
        'Enter a target loan amount below. The table shows the minimum monthly income '
        'required to qualify, across all tenors — using the product/location settings above.</div>',
        unsafe_allow_html=True,
    )
    wi_target = st.number_input("Target Loan Amount (₦)", min_value=0.0, value=1_000_000.0,
                                 step=50_000.0, key="wi_target")
    if wi_target > 0:
        wi_rows = []
        for _wt in range(2, 13):
            _wr = required_income_for_loan(
                target_loan=wi_target, tenor=_wt,
                location=location, product_type=prod_type,
                other_loans=other_loans,
                manual_rate_pct=manual_rate if manual_rate > 0 else None,
            )
            if _wr["ok"]:
                wi_rows.append({
                    "Tenor": f"{'▶ ' if _wt == tenor else ''}{_wt} mo{'  ◀' if _wt == tenor else ''}",
                    "Rate":             f"{_wr['rate']*100:.2f}%",
                    "Monthly PMT":      money(_wr["pmt"]),
                    "Min Monthly Income": money(_wr["required_turnover"]),
                    "DTI Used":         f"{_wr['dti']*100:.0f}%",
                })
            else:
                wi_rows.append({
                    "Tenor":              f"{_wt} mo",
                    "Rate":               "—",
                    "Monthly PMT":        "—",
                    "Min Monthly Income": _wr["reason"],
                    "DTI Used":           "—",
                })
        st.dataframe(pd.DataFrame(wi_rows), hide_index=True, use_container_width=True)
        st.markdown(
            f'<div class="sel-caption">Min Monthly Income = (PMT + other loans) ÷ DTI &nbsp;|&nbsp; '
            f'NTB note: applicable turnover uses trimmed mean — actual income may need to be higher.</div>',
            unsafe_allow_html=True,
        )

calc_btn = st.button("▶   Calculate Eligibility", key="calc", use_container_width=True, type="primary")


# ════════════════════════════════════════════════════════════════════════════
# SECTION 04 — RESULTS
# ════════════════════════════════════════════════════════════════════════════
if calc_btn:
    # Exclude unfilled placeholder rows (gross=0) so that a 3-month statement
    # doesn't pad to 6 months with zeros, which would cause RENEWAL nets[-3:]
    # to land on the zero rows instead of the real months.
    _active_rows = [r for r in inflow_data if r["gross"] > 0] or inflow_data
    nets   = [r["net"]   for r in _active_rows]
    counts = [r["count"] for r in _active_rows]

    if all(n == 0 for n in nets):
        st.error("Please enter monthly inflow data before calculating.")
    else:
        # ── Build merged report rows (used by ALL download generators) ──────
        # When two statements are present, merge their per-month data so that
        # PDF/Excel/CSV downloads reflect the combined analysis, not just Stmt A.
        _today_ym = today.strftime("%Y-%m")
        _has_b    = bool(st.session_state.get("rows_b"))

        if _has_b:
            _rA_map = {r["ym"]: r
                       for r in (st.session_state.rows_a or [])
                       if r["ym"] < _today_ym and r["gross"] > 0}
            _rB_map = {r["ym"]: r
                       for r in (st.session_state.rows_b or [])
                       if r["ym"] < _today_ym and r["gross"] > 0}
            # Union — keep all months from either statement
            _all_months = sorted(set(_rA_map) | set(_rB_map))[-N_MONTHS:]
            _report_rows = []
            for _ym in _all_months:
                rA = _rA_map.get(_ym)
                rB = _rB_map.get(_ym)
                eiA = rA["eligible_income"] if rA else 0
                eiB = rB["eligible_income"] if rB else 0
                _report_rows.append({
                    "ym":            _ym,
                    "label":         ym_label(_ym),
                    "gross":         eiA + eiB,
                    "parsed_gross":  (rA.get("parsed_gross", 0) if rA else 0) + (rB.get("parsed_gross", 0) if rB else 0),
                    "eligible_income": eiA + eiB,
                    "self_transfer": (rA.get("self_transfer", 0) if rA else 0) + (rB.get("self_transfer", 0) if rB else 0),
                    "reversal":      (rA.get("reversal", 0)      if rA else 0) + (rB.get("reversal", 0)      if rB else 0),
                    "non_business":  (rA.get("non_business", 0)  if rA else 0) + (rB.get("non_business", 0)  if rB else 0),
                    "loan_disbursal":(rA.get("loan_disbursal", 0)if rA else 0) + (rB.get("loan_disbursal", 0)if rB else 0),
                    "deductions":    (rA.get("deductions", 0)    if rA else 0) + (rB.get("deductions", 0)    if rB else 0),
                    "count":         (rA.get("count", 0)         if rA else 0) + (rB.get("count", 0)         if rB else 0),
                })
            _name_a = st.session_state.name_a or ""
            _name_b = st.session_state.name_b or ""
            _bank_a = st.session_state.bank_a or ""
            _bank_b = st.session_state.bank_b or ""
            _report_name = f"{_name_a} + {_name_b}".strip(" +")
            _report_bank = f"{_bank_a} + {_bank_b}".strip(" +")
        else:
            _report_rows = [r for r in (st.session_state.rows_a or [])
                            if r["ym"] < _today_ym and r["gross"] > 0][-N_MONTHS:]
            _report_name = st.session_state.name_a or "Account Holder"
            _report_bank = st.session_state.bank_a or "Bank"

        # ── Audit src rows: combine both statements ──────────────────────
        _audit_src = (st.session_state.rows_a or []) + (
            st.session_state.rows_b or [] if _has_b else []
        )

        result = calculate_eligibility(
            nets=nets, counts=counts,
            location=location, product_type=prod_type,
            tenor=tenor, other_loans=other_loans,
            requested_loan=req_loan if req_loan > 0 else 0,
            manual_rate_percent=manual_rate if manual_rate > 0 else None,
            sel_mode=(_product == "SEL"),
        )
        # Build combined applicant name + account number for tracking
        _acct_no = (st.session_state.account_no_a or "") or (st.session_state.account_no_b or "")

        # ── Cross-statement applicant memory ──────────────────────────────
        # Snapshot the borrower's PRIOR assessments before logging this one,
        # so the diff below compares against history, not against itself.
        try:
            _prior_assess = get_applicant_assessments(_acct_no, _report_name or "")
        except Exception:
            _prior_assess = []

        # Debit-side obligations for this assessment (visibility + history diff)
        _deb_all_mem = (st.session_state.get("debit_txns_a") or []) + \
                       (st.session_state.get("debit_txns_b") or [])
        _lr_mem      = [t for t in _deb_all_mem if t.get("category") == "loan_repayment"]
        _lr_months   = {t["ym"] for t in _lr_mem}
        _monthly_repay_mem = (sum(t["amount"] for t in _lr_mem) / len(_lr_months)) if _lr_months else 0.0
        _debit_total_mem   = sum(t["amount"] for t in _deb_all_mem)

        track("eligibility_result",
              session=_SID, officer=_OFFICER,
              bank=st.session_state.bank_a or "",
              approved=result.get("approved", False),
              max_loan=round(result.get("max_loan", 0), 2),
              tenor=tenor,
              dti=round((result.get("dti") or 0) * 100, 2),
              location=location,
              product=prod_type,
              total_net=round(result.get("total_net", 0), 2),
              applicant=_report_name or "",
              account_no=_acct_no,
              monthly_repay_obligation=round(_monthly_repay_mem, 2),
              debit_total=round(_debit_total_mem, 2))

        # ── Duplicate Application Detection ──────────────────────────────
        if _acct_no:
            try:
                _dup_hits = check_duplicate_application(_acct_no, _OFFICER, days=30)
                if _dup_hits:
                    _dup_rows_html = "".join(
                        f'<tr>'
                        f'<td style="font-size:11px;color:#94a3b8">{(h.get("ts") or "")[:10]}</td>'
                        f'<td style="font-size:11px;color:#fbbf24;font-weight:700">{h.get("officer","—")}</td>'
                        f'<td style="font-size:11px">{h.get("bank","—")}</td>'
                        f'<td style="text-align:right;font-size:11px">'
                        f'{money(float(h.get("max_loan") or 0))}</td>'
                        f'</tr>'
                        for h in _dup_hits
                    )
                    st.markdown(
                        f'<div style="background:rgba(251,191,36,.07);border:2px solid #f59e0b;'
                        f'border-radius:4px;padding:12px 16px;margin-bottom:12px">'
                        f'<div style="color:#f59e0b;font-weight:900;font-size:13px;letter-spacing:1px">'
                        f'⚠ DUPLICATE APPLICATION DETECTED</div>'
                        f'<div style="color:#94a3b8;font-size:11px;margin:6px 0">Account <strong style="color:#e2e8f0">'
                        f'{_acct_no}</strong> was assessed {len(_dup_hits)} time(s) by a different officer '
                        f'in the last 30 days — possible round-tripping.</div>'
                        f'<table class="preview-table" style="margin-top:6px"><thead><tr>'
                        f'<th style="text-align:left">Date</th><th>Officer</th>'
                        f'<th>Bank</th><th style="text-align:right">Max Loan</th>'
                        f'</tr></thead><tbody>{_dup_rows_html}</tbody></table>'
                        f'</div>',
                        unsafe_allow_html=True,
                    )
            except Exception:
                pass

        # ── Build loan params dict for all download generators ───────────
        _loan_params = {
            "location":     location,
            "product_type": prod_type,
            "tenor":        tenor,
            "other_loans":  other_loans,
            "req_loan":     req_loan,
            "manual_rate":  manual_rate,
        }

        # ── Save params for persistent What-If panel + Officer Notes ──────
        st.session_state.last_calc_params = {
            "nets": nets, "counts": counts, "location": location,
            "prod_type": prod_type, "other_loans": other_loans,
            "manual_rate": manual_rate, "result": result,
            "sel_mode": (_product == "SEL"),
            "report_name": _report_name, "acct_no": _acct_no,
        }
        st.session_state["_wi_tenor"] = tenor
        st.session_state["_wi_other"] = float(other_loans)

        # ── Save + show Applicant History (Feature 9) ─────────────────────
        _hist_name = _report_name or st.session_state.name_a or ""
        _hist_bank = _report_bank or st.session_state.bank_a or ""
        _avg_ei    = (result.get("applicable_turnover") or
                      sum(nets) / len(nets) if nets else 0)
        save_history(
            account_name = _hist_name,
            bank         = _hist_bank,
            avg_income   = round(_avg_ei),
            max_loan     = result.get("max_loan", 0),
            tenor        = tenor,
            location     = location,
            product      = prod_type,
            approved     = result.get("approved", False),
            session      = _SID,
        )
        # ── Borrower Memory — rich diff vs prior assessments ──────────────
        if _prior_assess:
            _pa  = _prior_assess[0]   # most recent prior assessment
            _f   = lambda v: float(v or 0)
            _cur_net   = result.get("total_net", 0)
            _cur_loan  = result.get("max_loan", 0)
            _cur_dti   = (result.get("dti") or 0) * 100
            _cur_obl   = _monthly_repay_mem
            _pr_net    = _f(_pa.get("total_net"))
            _pr_loan   = _f(_pa.get("max_loan"))
            _pr_dti    = _f(_pa.get("dti"))
            _pr_obl    = _f(_pa.get("monthly_repay_obligation"))
            _d_inc_pct = ((_cur_net - _pr_net) / _pr_net * 100) if _pr_net else 0
            _inc_col   = "#34d399" if _d_inc_pct >= 0 else "#f87171"
            _loan_col  = "#34d399" if _cur_loan >= _pr_loan else "#f87171"
            _dti_col   = "#34d399" if _cur_dti <= _pr_dti else "#fbbf24"

            _delta_bits = [
                f'<span style="color:#94a3b8">Eligible net: </span>'
                f'<span style="color:#e2e8f0;font-weight:700">{money(_pr_net)}</span> → '
                f'<span style="color:{_inc_col};font-weight:700">{money(_cur_net)} ({_d_inc_pct:+.1f}%)</span>',
                f'<span style="color:#94a3b8">Max loan: </span>'
                f'<span style="color:#e2e8f0;font-weight:700">{money(_pr_loan)}</span> → '
                f'<span style="color:{_loan_col};font-weight:700">{money(_cur_loan)}</span>',
                f'<span style="color:#94a3b8">DTI: </span>'
                f'<span style="color:#e2e8f0;font-weight:700">{_pr_dti:.1f}%</span> → '
                f'<span style="color:{_dti_col};font-weight:700">{_cur_dti:.1f}%</span>',
            ]
            # New / increased loan obligations — the classic repeat-borrower trap
            _obl_warn = ""
            if _cur_obl > 0 and _pr_obl == 0:
                _obl_warn = (f'<div style="margin-top:6px;color:#f87171;font-weight:700;font-size:12px">'
                             f'⚠ NEW loan repayment obligations: {money(_cur_obl)}/month '
                             f'(none in the previous assessment)</div>')
            elif _pr_obl > 0 and _cur_obl > _pr_obl * 1.5:
                _obl_warn = (f'<div style="margin-top:6px;color:#fbbf24;font-weight:700;font-size:12px">'
                             f'⚑ Loan repayment obligations up {(_cur_obl/_pr_obl-1)*100:.0f}%: '
                             f'{money(_pr_obl)}/mo → {money(_cur_obl)}/mo</div>')

            _hist_rows_html = "".join(
                f'<tr>'
                f'<td>{(h.get("ts") or "")[:10]}</td>'
                f'<td>{h.get("officer") or "—"}</td>'
                f'<td>{(h.get("bank") or "—").replace("_"," ")}</td>'
                f'<td style="text-align:right">{money(_f(h.get("total_net")))}</td>'
                f'<td style="text-align:right">{money(_f(h.get("max_loan")))}</td>'
                f'<td style="text-align:right">{_f(h.get("dti")):.1f}%</td>'
                f'<td style="color:{"#34d399" if h.get("approved") else "#f87171"};font-weight:700">'
                f'{"Approved" if h.get("approved") else "Declined"}</td>'
                f'</tr>'
                for h in _prior_assess[:5]
            )
            _notes_html = ""
            try:
                _prior_notes = get_officer_notes(_acct_no, _report_name or "")
            except Exception:
                _prior_notes = []
            if _prior_notes:
                _notes_html = '<div style="margin-top:12px;border-top:1px solid rgba(245,158,11,.25);padding-top:10px">'
                _notes_html += ('<div style="font-family:Plus Jakarta Sans,sans-serif;font-size:12px;'
                                'letter-spacing:1.8px;color:#fbbf24;text-transform:uppercase;'
                                'font-weight:800;margin-bottom:8px">📝 Officer Notes on File</div>')
                for _n in _prior_notes[:3]:
                    _notes_html += (
                        f'<div style="background:rgba(245,158,11,.10);border:1px solid rgba(245,158,11,.35);'
                        f'border-left:4px solid #f59e0b;border-radius:8px;'
                        f'padding:10px 14px;margin-bottom:8px">'
                        f'<div style="font-family:Plus Jakarta Sans,sans-serif;font-size:11px;'
                        f'font-weight:700;color:#fbbf24;margin-bottom:4px">'
                        f'{(_n.get("ts") or "")[:10]} &nbsp;·&nbsp; {_n.get("officer") or "—"}</div>'
                        f'<div style="font-size:14px;font-weight:700;color:#ffffff;line-height:1.6">'
                        f'{_n.get("note") or ""}</div></div>'
                    )
                _notes_html += '</div>'

            st.markdown(
                f'<div style="margin:6px 0 14px;padding:12px 16px;'
                f'background:rgba(245,158,11,.06);border:1px solid rgba(245,158,11,.22);'
                f'border-left:4px solid #f59e0b;border-radius:8px;font-size:12px">'
                f'<span style="font-family:Plus Jakarta Sans,sans-serif;color:#f59e0b;'
                f'font-weight:800;letter-spacing:1px">📋 RETURNING APPLICANT</span>'
                f'&nbsp;&nbsp;<span style="color:#64748b;font-size:11px">'
                f'{len(_prior_assess)} prior assessment(s) · last on {(_pa.get("ts") or "")[:10]} '
                f'by {_pa.get("officer") or "—"}</span>'
                f'<div style="display:flex;gap:18px;flex-wrap:wrap;margin-top:6px">'
                + "".join(_delta_bits) + '</div>'
                + _obl_warn +
                f'<table class="preview-table" style="margin-top:10px"><thead><tr>'
                f'<th style="text-align:left">Date</th><th style="text-align:left">Officer</th>'
                f'<th style="text-align:left">Bank</th><th>Eligible Net</th>'
                f'<th>Max Loan</th><th>DTI</th><th style="text-align:left">Decision</th>'
                f'</tr></thead><tbody>{_hist_rows_html}</tbody></table>'
                + _notes_html +
                f'</div>',
                unsafe_allow_html=True,
            )

        st.markdown("---")
        st.markdown('<div class="sel-section-title">04 — Results</div>', unsafe_allow_html=True)

        approved = result.get("approved", False)
        loan     = result["max_loan"]

        # ── Hero verdict card ─────────────────────────────────────────────
        def _dti_gauge_svg(dti_ratio: float) -> str:
            import math as _math
            p = min(max(float(dti_ratio or 0), 0.0), 1.0)
            if p <= 0.30:
                color, risk = "#34d399", "LOW RISK"
            elif p <= 0.45:
                color, risk = "#f59e0b", "MODERATE"
            else:
                color, risk = "#f87171", "HIGH RISK"
            cx, cy, r = 50, 50, 36
            bg = f"M {cx-r},{cy} A {r},{r} 0 0,1 {cx+r},{cy}"
            if p < 0.01:
                fg = ""
            elif p > 0.99:
                fg = bg
            else:
                angle = _math.pi * (1 - p)
                ex = cx + r * _math.cos(angle)
                ey = cy - r * _math.sin(angle)
                large = 1 if p > 0.5 else 0
                fg = f"M {cx-r},{cy} A {r},{r} 0 {large},1 {ex:.2f},{ey:.2f}"
            fg_path_html = (f'<path d="{fg}" fill="none" stroke="{color}" '
                            f'stroke-width="7" stroke-linecap="round"/>') if fg else ""
            return (
                f'<svg viewBox="0 0 100 60" xmlns="http://www.w3.org/2000/svg" '
                f'style="width:100%;max-width:130px;display:block;margin:0 auto 4px">'
                f'<path d="{bg}" fill="none" stroke="rgba(255,255,255,0.08)" '
                f'stroke-width="7" stroke-linecap="round"/>'
                f'{fg_path_html}'
                f'<text x="{cx}" y="{cy-2}" text-anchor="middle" '
                f'font-family="system-ui,sans-serif" font-size="15" font-weight="800" '
                f'fill="{color}">{p*100:.0f}%</text>'
                f'<text x="{cx}" y="{cy+12}" text-anchor="middle" '
                f'font-family="system-ui,sans-serif" font-size="7" font-weight="700" '
                f'letter-spacing="1.2" fill="{color}">{risk}</text>'
                f'<text x="{cx}" y="{cy+22}" text-anchor="middle" '
                f'font-family="system-ui,sans-serif" font-size="6" '
                f'fill="rgba(255,255,255,0.35)">DEBT-TO-INCOME</text>'
                f'</svg>'
            )

        _vh_cls   = "vh-approved" if approved else "vh-declined"
        _vh_badge = "✓ &nbsp;Approved" if approved else "✕ &nbsp;Below Product Minimum"
        _vh_rate  = (f"{pct(result['interest_rate'])} ★" if manual_rate > 0
                     else pct(result["interest_rate"]))
        _vh_freq  = result.get("repayment_frequency", "—")
        st.markdown(
            f'<div class="verdict-hero {_vh_cls}">'
            f'  <div class="vh-main">'
            f'    <div class="vh-badge">{_vh_badge}</div>'
            f'    <div class="vh-label">Maximum Loan</div>'
            f'    <div class="vh-amount">{money(loan)}</div>'
            f'    <div class="vh-sub">{prod_type} &nbsp;·&nbsp; {tenor} months &nbsp;·&nbsp; {location}</div>'
            f'  </div>'
            f'  <div class="vh-grid">'
            f'    <div class="vh-cell vh-cell-gauge">{_dti_gauge_svg(result["dti"])}</div>'
            f'    <div class="vh-cell"><div class="vh-cell-label">Interest Rate</div>'
            f'      <div class="vh-cell-value">{_vh_rate}</div></div>'
            f'    <div class="vh-cell"><div class="vh-cell-label">Repayment / Period</div>'
            f'      <div class="vh-cell-value">{money(result["max_repayment_display"])}</div></div>'
            f'    <div class="vh-cell"><div class="vh-cell-label">Frequency</div>'
            f'      <div class="vh-cell-value">{_vh_freq}</div></div>'
            f'  </div>'
            f'</div>',
            unsafe_allow_html=True,
        )

        # ── Turnover ceiling policy note ──────────────────────────────────
        if result.get("turnover_capped"):
            _cap_amt = result.get("turnover_cap_amount", 5_000_000)
            _cap_thr = result.get("turnover_cap_threshold", 20_000_000)
            st.markdown(
                f'<div style="margin-top:8px;padding:10px 14px;'
                f'background:rgba(245,158,11,.08);border:1px solid rgba(245,158,11,.3);'
                f'border-left:4px solid #f59e0b;border-radius:4px;font-size:12px;color:#fbbf24">'
                f'⚑ <strong>Policy cap applied:</strong> max loan limited to {money(_cap_amt)} '
                f'because applicable turnover ({money(result.get("applicable_turnover", 0))}) '
                f'is below {money(_cap_thr)}. Turnover of {money(_cap_thr)}+ is required for a '
                f'loan above {money(_cap_amt)}.</div>',
                unsafe_allow_html=True,
            )

        # ── Assessment streak badge ───────────────────────────────────────
        st.session_state.assessment_count += 1
        _ac = st.session_state.assessment_count
        _milestone = None
        if   _ac == 3:               _milestone = ("🔥", "3 assessments today — on a roll!")
        elif _ac == 5:               _milestone = ("⚡", "5 assessments — you're on fire!")
        elif _ac == 10:              _milestone = ("💪", "10 assessments — unstoppable!")
        elif _ac == 15:              _milestone = ("🏆", "15 assessments — legendary effort!")
        elif _ac > 10 and _ac % 5 == 0: _milestone = ("🚀", f"{_ac} assessments this session — keep pushing!")
        if _milestone:
            _ms_icon, _ms_text = _milestone
            st.markdown(
                f'<div style="margin-top:8px;padding:8px 14px;'
                f'background:rgba(245,158,11,.07);border:1px solid rgba(245,158,11,.25);'
                f'border-radius:4px;font-size:12px;color:#f59e0b;'
                f'text-align:center;font-weight:700;letter-spacing:1px">'
                f'{_ms_icon} &nbsp;{_ms_text}</div>',
                unsafe_allow_html=True,
            )

        # ── Confetti on approval ──────────────────────────────────────────
        if approved:
            _html(
                f'<script>'
                f'(function(){{ var p=window.parent;'
                f'if(p.__selConfetti) p.__selConfetti({loan});'
                f'}})();'
                f'/* loan={loan} */'   # unique comment forces re-render each assessment
                f'</script>',
                height=0,
            )

        st.markdown("")

        # Supporting result cards (hero covers loan, DTI, rate, repayment/period, frequency)
        m1, m2, m3 = st.columns(3)
        m1.metric("Applicable Turnover",    money(result["applicable_turnover"]))
        m2.metric("Total Eligible Net",     money(result["total_net"]))
        m3.metric("Max Total Repayment",    money(result["max_total_repayment"]))

        # ── Customer Risk Score ───────────────────────────────────────────
        try:
            _rs_rows = [r for r in rows_a if r["eligible_income"] > 0]
            if _rs_rows:
                import math as _rs_math
                import datetime as _rs_dt

                # Dimension 1 — Income Consistency (CV of eligible income)
                _rs_vals  = [r["eligible_income"] for r in _rs_rows]
                _rs_mean  = sum(_rs_vals) / len(_rs_vals)
                _rs_std   = (_rs_math.sqrt(sum((v - _rs_mean)**2 for v in _rs_vals) / len(_rs_vals))
                             if len(_rs_vals) > 1 else 0)
                _rs_cv    = _rs_std / _rs_mean if _rs_mean else 1
                # Low CV → consistent → good; high CV → erratic → bad
                _rs_cons  = 5 if _rs_cv < 0.15 else 4 if _rs_cv < 0.30 else 3 if _rs_cv < 0.50 else 2 if _rs_cv < 0.75 else 1

                # Dimension 2 — Deduction Ratio (total deductions / total gross)
                _rs_tgross = sum(r["gross"] for r in _rs_rows) or 1
                _rs_tdeduct = sum(
                    r.get("self_transfer",0) + r.get("reversal",0) +
                    r.get("non_business",0) + r.get("loan_disbursal",0)
                    for r in _rs_rows
                )
                _rs_ded_pct = _rs_tdeduct / _rs_tgross
                _rs_ded   = 5 if _rs_ded_pct < 0.10 else 4 if _rs_ded_pct < 0.20 else 3 if _rs_ded_pct < 0.35 else 2 if _rs_ded_pct < 0.55 else 1

                # Dimension 3 — Statement Freshness
                try:
                    _rs_latest = max(r["ym"] for r in _rs_rows)
                    _rs_ly, _rs_lm = int(_rs_latest[:4]), int(_rs_latest[5:])
                    _rs_today = _rs_dt.date.today()
                    _rs_age = (_rs_today.year - _rs_ly) * 12 + (_rs_today.month - _rs_lm)
                    _rs_fresh = 5 if _rs_age <= 1 else 4 if _rs_age <= 2 else 3 if _rs_age <= 3 else 2 if _rs_age <= 6 else 1
                except Exception:
                    _rs_fresh = 3

                # Dimension 4 — Recycling % (self_transfer + loan_disbursal share of gross)
                _rs_recyc_tot = sum(r.get("self_transfer",0) + r.get("loan_disbursal",0) for r in _rs_rows)
                _rs_recyc_pct = _rs_recyc_tot / _rs_tgross
                _rs_recyc = 5 if _rs_recyc_pct < 0.05 else 4 if _rs_recyc_pct < 0.15 else 3 if _rs_recyc_pct < 0.30 else 2 if _rs_recyc_pct < 0.50 else 1

                # Composite score (weighted average, round to nearest 0.5)
                _rs_raw = (
                    _rs_cons  * 0.35 +
                    _rs_ded   * 0.30 +
                    _rs_fresh * 0.20 +
                    _rs_recyc * 0.15
                )
                _rs_score = round(_rs_raw * 2) / 2   # nearest 0.5

                _rs_label = ("Very Low Risk" if _rs_score >= 4.5 else
                             "Low Risk"      if _rs_score >= 3.5 else
                             "Moderate Risk" if _rs_score >= 2.5 else
                             "High Risk"     if _rs_score >= 1.5 else
                             "Very High Risk")
                _rs_col   = ("#34d399" if _rs_score >= 4.5 else
                             "#10b981" if _rs_score >= 3.5 else
                             "#fbbf24" if _rs_score >= 2.5 else
                             "#fb923c" if _rs_score >= 1.5 else
                             "#f87171")

                # Render score card
                _rs_dims = [
                    ("Income Consistency", _rs_cons,  f"CV={_rs_cv:.2f}"),
                    ("Deduction Ratio",    _rs_ded,   f"{_rs_ded_pct*100:.0f}% of gross deducted"),
                    ("Statement Freshness",_rs_fresh, f"{_rs_age} month(s) old"),
                    ("Recycling Detected", _rs_recyc, f"{_rs_recyc_pct*100:.0f}% recycled credits"),
                ]
                _rs_dim_html = ""
                for _rd_label, _rd_score, _rd_note in _rs_dims:
                    _rd_col = ("#34d399" if _rd_score >= 4 else "#fbbf24" if _rd_score >= 3 else "#f87171")
                    _rs_dim_html += (
                        f'<div style="display:flex;align-items:center;gap:8px;margin-bottom:4px">'
                        f'<div style="width:120px;font-size:10px;color:#94a3b8">{_rd_label}</div>'
                        f'<div style="display:flex;gap:2px">'
                        + ''.join(
                            f'<div style="width:10px;height:10px;border-radius:2px;'
                            f'background:{"' + _rd_col + '" if _i < _rd_score else "rgba(255,255,255,.08)"}'
                            f'"></div>'
                            for _i in range(5)
                        )
                        + f'</div>'
                        f'<div style="font-size:9px;color:#64748b">{_rd_note}</div>'
                        f'</div>'
                    )

                st.markdown(
                    f'<div style="margin:14px 0;padding:14px 16px;'
                    f'background:rgba(0,0,0,.2);border:1px solid {_rs_col}44;border-radius:4px">'
                    f'<div style="display:flex;align-items:center;gap:14px;margin-bottom:10px">'
                    f'<div style="font-size:9px;letter-spacing:2px;color:#6b7f74;text-transform:uppercase">Customer Risk Score</div>'
                    f'<div style="font-size:22px;font-weight:900;color:{_rs_col}">{_rs_score:.1f}<span style="font-size:12px;color:#6b7f74">/5</span></div>'
                    f'<div style="background:{_rs_col}22;border:1px solid {_rs_col}55;border-radius:20px;'
                    f'padding:2px 12px;font-size:11px;font-weight:700;color:{_rs_col}">{_rs_label}</div>'
                    f'</div>'
                    f'{_rs_dim_html}'
                    f'</div>',
                    unsafe_allow_html=True,
                )
        except Exception:
            pass

        # ── Tenor Comparison Table ────────────────────────────────────────
        st.markdown("---")
        st.markdown(
            '<div style="font-size:10px;letter-spacing:2px;color:#f59e0b;'
            'text-transform:uppercase;margin-bottom:8px">Tenor Comparison — all tenors</div>',
            unsafe_allow_html=True,
        )
        _tenor_data = []
        for _t in range(2, 13):
            _tr = calculate_eligibility(
                nets=nets, counts=counts,
                location=location, product_type=prod_type,
                tenor=_t, other_loans=other_loans,
                manual_rate_percent=manual_rate if manual_rate > 0 else None,
                sel_mode=(_product == "SEL"),
            )
            _tenor_data.append({
                "Tenor":              f"{'▶ ' if _t == tenor else ''}{_t} mo{'  ◀' if _t == tenor else ''}",
                "Max Loan":           money(_tr["max_loan"]) if _tr["approved"] else "—",
                "Repayment / Period": money(_tr["max_repayment_display"]),
                "Rate":               pct(_tr["interest_rate"]) if _tr["interest_rate"] else "—",
                "Frequency":          _tr["repayment_frequency"],
                "Status":             "Approved" if _tr["approved"] else "Below min",
            })
        st.dataframe(
            pd.DataFrame(_tenor_data),
            hide_index=True,
            use_container_width=True,
        )

        # ── Repayment Schedule ────────────────────────────────────────────
        if result.get("approved") and result.get("interest_rate") and result.get("max_loan", 0) > 0:
            _loan_amt  = result["max_loan"]
            _m_rate    = result["interest_rate"]
            _m_pmt     = result["max_repayment_monthly"]
            _n_periods = result["tenor"]
            _freq      = result["repayment_frequency"]

            # Build monthly amortization schedule
            _sched = []
            _bal = float(_loan_amt)
            for _p in range(1, _n_periods + 1):
                _int   = _bal * _m_rate
                _prin  = _m_pmt - _int
                _close = max(_bal - _prin, 0.0)
                _sched.append({
                    "period":  _p,
                    "opening": _bal,
                    "payment": _m_pmt,
                    "interest": _int,
                    "principal": _prin,
                    "closing": _close,
                })
                _bal = _close

            _total_pmt  = sum(s["payment"]   for s in _sched)
            _total_int  = sum(s["interest"]  for s in _sched)
            _total_prin = sum(s["principal"] for s in _sched)
            _cost_pct   = _total_int / _loan_amt * 100 if _loan_amt else 0

            with st.expander(
                f"📅  Repayment Schedule — {_n_periods}-Month Amortization",
                expanded=False,
            ):
                _freq_note = "weekly" if _freq == "Weekly" else "monthly"
                st.markdown(
                    f'<div style="font-size:11px;color:#64748b;margin-bottom:10px;line-height:1.8">'
                    f'Loan: <span style="color:#10b981;font-weight:700">{money(_loan_amt)}</span>'
                    f'&nbsp;&nbsp;|&nbsp;&nbsp;'
                    f'Rate: <span style="color:#fbbf24;font-weight:700">{pct(_m_rate)}/month</span>'
                    f'&nbsp;&nbsp;|&nbsp;&nbsp;'
                    f'Payment: <span style="color:#34d399;font-weight:700">'
                    f'{money(result["max_repayment_display"])} {_freq_note}</span>'
                    f'{"&nbsp;&nbsp;|&nbsp;&nbsp;<em>Schedule shows monthly aggregates</em>" if _freq == "Weekly" else ""}'
                    f'</div>',
                    unsafe_allow_html=True,
                )

                _sched_body = ""
                for _s in _sched:
                    _sched_body += (
                        f'<tr>'
                        f'<td style="color:#94a3b8;text-align:center">{_s["period"]}</td>'
                        f'<td style="text-align:right">{money(_s["opening"])}</td>'
                        f'<td style="color:#34d399;font-weight:700;text-align:right">{money(_s["payment"])}</td>'
                        f'<td style="color:#f87171;text-align:right">{money(_s["interest"])}</td>'
                        f'<td style="color:#a78bfa;text-align:right">{money(_s["principal"])}</td>'
                        f'<td style="color:#10b981;font-weight:700;text-align:right">{money(_s["closing"])}</td>'
                        f'</tr>'
                    )
                _sched_body += (
                    f'<tr style="border-top:2px solid #1a3d2b">'
                    f'<td style="color:#64748b;font-size:10px;text-transform:uppercase;text-align:center">Total</td>'
                    f'<td></td>'
                    f'<td style="color:#34d399;font-weight:700;text-align:right">{money(_total_pmt)}</td>'
                    f'<td style="color:#f87171;font-weight:700;text-align:right">{money(_total_int)}</td>'
                    f'<td style="color:#a78bfa;font-weight:700;text-align:right">{money(_total_prin)}</td>'
                    f'<td></td>'
                    f'</tr>'
                )
                st.markdown(
                    f'<table class="preview-table">'
                    f'<thead><tr>'
                    f'<th style="text-align:center">Period</th>'
                    f'<th style="text-align:right">Opening Balance</th>'
                    f'<th style="text-align:right;color:#34d399">Payment</th>'
                    f'<th style="text-align:right;color:#f87171">Interest</th>'
                    f'<th style="text-align:right;color:#a78bfa">Principal</th>'
                    f'<th style="text-align:right;color:#10b981">Closing Balance</th>'
                    f'</tr></thead><tbody>{_sched_body}</tbody></table>',
                    unsafe_allow_html=True,
                )
                st.markdown(
                    f'<div style="margin-top:10px;padding:8px 12px;'
                    f'background:rgba(248,113,113,.07);border-left:3px solid #f87171;'
                    f'border-radius:3px;font-size:11px;color:#94a3b8">'
                    f'Total cost of credit: '
                    f'<span style="color:#f87171;font-weight:700">{money(_total_int)}</span>'
                    f' — <span style="color:#f87171">{_cost_pct:.1f}%</span>'
                    f' of loan principal repaid as interest over {_n_periods} months'
                    f'</div>',
                    unsafe_allow_html=True,
                )

        # ── Core Banking Export ───────────────────────────────────────────
        if approved:
            import datetime as _cbe_dt
            _cbe_name  = _report_name or st.session_state.get("name_a") or "UNKNOWN"
            _cbe_bank  = st.session_state.get("bank_a") or ""
            _cbe_acct  = st.session_state.get("account_no_a") or ""
            _cbe_loan  = result.get("max_loan", 0)
            _cbe_rate  = (result.get("interest_rate") or 0) * 100
            _cbe_pmt   = result.get("max_repayment_monthly", 0)
            _cbe_dti   = (result.get("dti") or 0) * 100
            _cbe_today = _cbe_dt.date.today().isoformat()
            _cbe_maturity = (_cbe_dt.date.today()
                             .replace(day=1)  # go to first of month
                             ).__class__(
                                 _cbe_dt.date.today().year + (tenor // 12),
                                 (_cbe_dt.date.today().month + tenor % 12 - 1) % 12 + 1,
                                 1,
                             ).isoformat()

            def _make_cbe_csv(fmt: str) -> bytes:
                import io as _io
                buf = _io.StringIO()
                if fmt == "flexcube":
                    # Oracle Flexcube OD/CL loan input format
                    hdr = ["LOAN_ACCOUNT_NO","CUSTOMER_NAME","CUSTOMER_ACCOUNT","PRODUCT_CODE",
                           "BRANCH_CODE","CURRENCY","LOAN_AMOUNT","TENOR_MONTHS",
                           "INTEREST_RATE_PCT","MONTHLY_REPAYMENT","DTI_PCT",
                           "DISBURSEMENT_DATE","MATURITY_DATE","SOURCE_BANK",
                           "ASSESSED_BY","PRODUCT_TYPE","LOCATION"]
                    row = ["", _cbe_name, _cbe_acct, prod_type,
                           location[:3].upper(), "NGN", f"{_cbe_loan:.2f}", str(tenor),
                           f"{_cbe_rate:.2f}", f"{_cbe_pmt:.2f}", f"{_cbe_dti:.2f}",
                           _cbe_today, _cbe_maturity, _cbe_bank,
                           _OFFICER, _product, location]
                else:  # temenos T24
                    hdr = ["ARRANGEMENT.ID","CUSTOMER","ACCOUNT.NO","PRODUCT",
                           "CURRENCY","AMOUNT","TERM","INT.RATE",
                           "INSTALMENT","DTI","START.DATE","MATURITY.DATE",
                           "BANK","OFFICER","TYPE","ZONE"]
                    row = ["", _cbe_name, _cbe_acct, f"SEL.{prod_type}",
                           "NGN", f"{_cbe_loan:.2f}", f"{tenor}M",
                           f"{_cbe_rate:.4f}", f"{_cbe_pmt:.2f}", f"{_cbe_dti:.2f}",
                           _cbe_today, _cbe_maturity,
                           _cbe_bank, _OFFICER, _product, location]
                buf.write(",".join(hdr) + "\r\n")
                buf.write(",".join(f'"{v}"' for v in row) + "\r\n")
                return buf.getvalue().encode()

            with st.expander("🏦  Export to Core Banking System", expanded=False):
                st.markdown(
                    '<div style="font-size:11px;color:#64748b;margin-bottom:10px">'
                    'Download a pre-filled CSV in the loan booking format for your core banking '
                    'system — paste directly into the import screen to eliminate re-entry.</div>',
                    unsafe_allow_html=True,
                )
                _cbe1, _cbe2 = st.columns(2)
                with _cbe1:
                    st.download_button(
                        "⬇  Flexcube Format (Oracle)",
                        data=_make_cbe_csv("flexcube"),
                        file_name=f"loan_booking_flexcube_{_cbe_name.replace(' ','_')}_{_cbe_today}.csv",
                        mime="text/csv",
                        key="dl_cbe_flexcube",
                        use_container_width=True,
                    )
                with _cbe2:
                    st.download_button(
                        "⬇  Temenos T24 Format",
                        data=_make_cbe_csv("temenos"),
                        file_name=f"loan_booking_t24_{_cbe_name.replace(' ','_')}_{_cbe_today}.csv",
                        mime="text/csv",
                        key="dl_cbe_t24",
                        use_container_width=True,
                    )
                st.markdown(
                    '<div style="font-size:10px;color:#64748b;margin-top:6px">'
                    '⚑ LOAN_ACCOUNT_NO / ARRANGEMENT.ID left blank — assign in your CBS before import. '
                    'Verify field mappings match your exact CBS version before bulk use.</div>',
                    unsafe_allow_html=True,
                )

        # Requested loan analysis
        if req_loan > 0 and "requested" in result:
            st.markdown("---")
            st.markdown('<div style="font-size:10px;letter-spacing:2px;color:#f59e0b;text-transform:uppercase;margin-bottom:8px">Requested Loan Analysis</div>', unsafe_allow_html=True)
            req = result["requested"]
            within = req["within_max"]
            r_col1, r_col2 = st.columns(2)
            with r_col1:
                st.markdown(f"""
                | | |
                |---|---|
                | Requested Amount | {money(req_loan)} |
                | Interest Rate | {pct(req.get("rate"))} |
                | Repayment / Period | {money(req["repayment"])} |
                | DTI for Requested | {pct(req["dti"])} |
                | vs Max Loan | {("+" if req_loan >= loan else "") + money(abs(req_loan - loan))} |
                | Status | {"✅ Below max — eligible" if within else "❌ Above max — not eligible"} |
                """)

        # Deduction audit table
        if st.session_state.rows_a or st.session_state.rows_b:
            st.markdown("---")
            st.markdown('<div style="font-size:10px;letter-spacing:2px;color:#64748b;text-transform:uppercase;margin-bottom:6px">Classification Audit — All Tagged Credits</div>', unsafe_allow_html=True)

            audit_rows = []
            today_ym = today.strftime("%Y-%m")
            for r in _audit_src:
                if r["ym"] >= today_ym or r["gross"] == 0:
                    continue
                # Self-transfers (savings round-trips) are deducted
                if r.get("self_transfer", 0) > 0:
                    audit_rows.append({
                        "Month": r["label"],
                        "Category": "Self Deposit / Savings Round-trip",
                        "Deducted": True,
                        "Amount": r["self_transfer"],
                    })
                for cat in ["reversal", "non_business", "loan_disbursal"]:
                    if r.get(cat, 0) > 0:
                        audit_rows.append({
                            "Month": r["label"],
                            "Category": cat.replace("_", " ").title(),
                            "Deducted": True,
                            "Amount": r[cat],
                        })
            # Extra manual deductions from grid
            for i, row in enumerate(inflow_data):
                if row.get("extra", 0) > 0:
                    audit_rows.append({
                        "Month": row["label"],
                        "Category": "Manual Deduction",
                        "Deducted": True,
                        "Amount": row["extra"],
                    })

            if audit_rows:
                df = pd.DataFrame(audit_rows)
                st.dataframe(
                    df, hide_index=True, use_container_width=True,
                    column_config={
                        "Amount": st.column_config.NumberColumn("Amount", format="₦%d"),
                        "Deducted": st.column_config.CheckboxColumn("Deducted from Eligible?"),
                    },
                )
                _xlsx_full = generate_xlsx(
                    rows         = _report_rows,
                    result       = result,
                    account_name = _report_name,
                    bank         = _report_bank,
                    params       = _loan_params,
                    officer      = _OFFICER,
                )
                _safe_xl = (_report_name or "report").replace(" ", "_").lower()
                _cav1, _cav2 = st.columns(2)
                with _cav1:
                    if st.download_button(
                        "⬇  Download Full Report (Excel)",
                        _xlsx_full,
                        file_name=f"PARSIO_Report_{_safe_xl}_{datetime.date.today():%Y%m%d}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        key="dl_audit_xlsx",
                        use_container_width=True,
                    ):
                        track("download", session=_SID, officer=_OFFICER,
                              bank=st.session_state.bank_a or "", fmt="excel")
                with _cav2:
                    # Build CSV with eligibility summary header + audit rows
                    import io as _io
                    _csv_buf = _io.StringIO()
                    # -- Summary section --
                    _csv_buf.write(f"Assessed By,{_OFFICER}\r\n")
                    _csv_buf.write(f"Date,{datetime.date.today().strftime('%d %b %Y')}\r\n")
                    _csv_buf.write("\r\n")
                    _csv_buf.write("LOAN PARAMETERS\r\n")
                    _csv_buf.write(f"Location,{location}\r\n")
                    _csv_buf.write(f"Product Type,{prod_type}\r\n")
                    _csv_buf.write(f"Tenor (Months),{tenor}\r\n")
                    _csv_buf.write(f"Other Monthly Repayments,{money(other_loans)}\r\n")
                    if req_loan > 0:
                        _csv_buf.write(f"Requested Loan Amount,{money(req_loan)}\r\n")
                    if manual_rate > 0:
                        _csv_buf.write(f"Manual Interest Rate,{manual_rate:.2f}%\r\n")
                    _csv_buf.write("\r\n")
                    _csv_buf.write("ELIGIBILITY SUMMARY\r\n")
                    _csv_buf.write(f"Decision,{'Approved' if result.get('approved') else 'Below Minimum'}\r\n")
                    _csv_buf.write(f"Max Loan Amount,{money(result.get('max_loan', 0))}\r\n")
                    _csv_buf.write(f"Tenor (Months),{result.get('tenor', '—')}\r\n")
                    _csv_buf.write(f"DTI,{pct(result.get('dti'))}\r\n")
                    _csv_buf.write(f"Interest Rate,{pct(result.get('interest_rate'))}\r\n")
                    _csv_buf.write(f"Repayment Frequency,{result.get('repayment_frequency', '')}\r\n")
                    _csv_buf.write(f"Repayment / Period,{money(result.get('max_repayment_display', 0))}\r\n")
                    _csv_buf.write(f"Max Total Repayment,{money(result.get('max_total_repayment', 0))}\r\n")
                    _csv_buf.write(f"Applicable Turnover,{money(result.get('applicable_turnover', 0))}\r\n")
                    _csv_buf.write(f"Total Net Income,{money(result.get('total_net', 0))}\r\n")
                    # -- Requested loan section (if applicable) --
                    if req_loan > 0 and "requested" in result:
                        _rq_csv = result["requested"]
                        _rq_diff = abs(req_loan - result.get("max_loan", 0))
                        _rq_sign = "+" if req_loan >= result.get("max_loan", 0) else "-"
                        _csv_buf.write("\r\n")
                        _csv_buf.write("REQUESTED LOAN ANALYSIS\r\n")
                        _csv_buf.write(f"Requested Amount,{money(req_loan)}\r\n")
                        _csv_buf.write(f"Interest Rate,{pct(_rq_csv.get('rate'))}\r\n")
                        _csv_buf.write(f"Repayment / Period,{money(_rq_csv.get('repayment', 0))}\r\n")
                        _csv_buf.write(f"DTI for Requested,{pct(_rq_csv.get('dti'))}\r\n")
                        _csv_buf.write(f"vs Max Loan,{_rq_sign}{money(_rq_diff)}\r\n")
                        _csv_buf.write(f"Status,{'Below max - eligible' if _rq_csv.get('within_max') else 'Above max - not eligible'}\r\n")
                    _csv_buf.write("\r\n")
                    # -- Audit rows section --
                    _csv_buf.write("CLASSIFICATION AUDIT\r\n")
                    _csv_buf.write(df.to_csv(index=False))
                    if st.download_button(
                        "⬇  Download Audit (CSV)",
                        _csv_buf.getvalue().encode("utf-8"),
                        file_name="sel_classification_audit.csv",
                        mime="text/csv",
                        key="dl_audit_csv",
                        use_container_width=True,
                    ):
                        track("download", session=_SID, officer=_OFFICER,
                              bank=st.session_state.bank_a or "", fmt="csv")

        # ── Download Full Eligibility Report PDF ──────────────────────────
        st.markdown("---")
        _pdf_full = generate_pdf_report(
            account_name = _report_name,
            bank         = _report_bank,
            rows         = _report_rows if _report_rows else [],
            result       = result,
            req_loan     = req_loan,
            params       = _loan_params,
            officer      = _OFFICER,
        )
        _safe_full = (_report_name or "report").replace(" ", "_").lower()

        # ── Credit Memo — one-page committee-ready decision document ──────
        try:
            _memo_risk = f"{_rs_label} ({_rs_score}/5)"
        except NameError:
            _memo_risk = ""
        try:
            _memo_flags = [f"{_ft} — {_fm}" for _ft, _fm, _fc in _spf_flags]
        except NameError:
            _memo_flags = []
        _pdf_memo = generate_credit_memo(
            account_name = _report_name,
            bank         = _report_bank,
            rows         = _report_rows if _report_rows else [],
            result       = result,
            params       = _loan_params,
            officer      = _OFFICER,
            account_no   = _acct_no,
            risk_label   = _memo_risk,
            flags        = _memo_flags,
        )
        _dlm1, _dlm2 = st.columns(2)
        with _dlm1:
            if st.download_button(
                label               = "📄  Credit Memo — one-page (PDF)",
                data                = _pdf_memo,
                file_name           = f"PARSIO_CreditMemo_{_safe_full}_{datetime.date.today():%Y%m%d}.pdf",
                mime                = "application/pdf",
                use_container_width = True,
                key                 = "dl_memo_pdf",
            ):
                track("download", session=_SID, officer=_OFFICER,
                      bank=st.session_state.bank_a or "", fmt="memo_pdf")
        with _dlm2:
            if st.download_button(
                label               = "⬇  Download Full Eligibility Report (PDF)",
                data                = _pdf_full,
                file_name           = f"PARSIO_Report_{_safe_full}_{datetime.date.today():%Y%m%d}.pdf",
                mime                = "application/pdf",
                use_container_width = True,
                key                 = "dl_full_pdf",
            ):
                track("download", session=_SID, officer=_OFFICER,
                      bank=st.session_state.bank_a or "", fmt="pdf")

        # ── Feature 8: WhatsApp / Email Share — PDF via Web Share API ───────
        import base64 as _b64s, urllib.parse as _uparse
        _share_name  = _report_name or "Applicant"
        _share_bank  = _report_bank or "Bank"
        _share_dec   = "APPROVED" if result.get("approved") else "BELOW MINIMUM"
        _share_msg   = (
            f"PARSIO Loan Assessment\n"
            f"Applicant: {_share_name} ({_share_bank})\n"
            f"Decision: {'✅' if result.get('approved') else '❌'} {_share_dec}\n"
            f"Max Loan: {money(result.get('max_loan', 0))}\n"
            f"Rate: {pct(result.get('interest_rate'))}/month\n"
            f"Repayment: {money(result.get('max_repayment_display', 0))} / "
            f"{result.get('repayment_frequency', 'Monthly').lower()[:3]}\n"
            f"Tenor: {result.get('tenor', '—')} months\n"
            f"Generated: {datetime.date.today().strftime('%d %b %Y')}"
        )
        _share_filename = f"PARSIO_Report_{_safe_full}_{datetime.date.today():%Y%m%d}.pdf"
        # Save for persistent render outside calc_btn block
        st.session_state.last_share = {
            "pdf_b64":  _b64s.b64encode(_pdf_full).decode(),
            "filename": _share_filename,
            "msg":      _share_msg,
            "name":     _share_name,
        }


# ════════════════════════════════════════════════════════════════════════════
# FEATURE 8 — PERSISTENT SHARE PANEL (Web Share API + PDF)
# ════════════════════════════════════════════════════════════════════════════
if st.session_state.last_share:
    _s = st.session_state.last_share
    # Escape values safe for JS template literals
    _js_msg      = _s["msg"].replace("\\", "\\\\").replace("`", "\\`").replace("${", "\\${")
    _js_name     = _s["name"].replace("\\", "\\\\").replace("`", "\\`")
    _js_filename = _s["filename"].replace("`", "\\`")
    _wa_fallback = "https://wa.me/?text=" + __import__("urllib.parse", fromlist=["quote"]).quote(_s["msg"])
    _ml_fallback = (
        "mailto:?subject=" + __import__("urllib.parse", fromlist=["quote"]).quote(f"PARSIO Result — {_s['name']}")
        + "&body=" + __import__("urllib.parse", fromlist=["quote"]).quote(_s["msg"])
    )

    st.markdown("---")
    st.markdown(
        '<div style="font-size:10px;letter-spacing:2px;color:#10b981;'
        'text-transform:uppercase;margin-bottom:10px">Share Result</div>',
        unsafe_allow_html=True,
    )
    # Web Share API component — shares the actual PDF on mobile, falls back to
    # download + text link on desktop browsers without file-share support.
    _html(f"""
<style>
  body{{margin:0;padding:0;background:transparent}}
  .sh-wrap{{display:flex;gap:12px}}
  .sh-btn{{
    flex:1;padding:10px 14px;border-radius:4px;font-size:12px;font-weight:700;
    cursor:pointer;letter-spacing:1px;font-family:"Space Mono",monospace;
    transition:opacity .2s;width:100%;
  }}
  .sh-btn:hover{{opacity:.75}}
  .sh-wa  {{background:rgba(52,211,153,.12);border:1px solid rgba(52,211,153,.35);color:#34d399}}
  .sh-mail{{background:rgba(245,158,11,.08); border:1px solid rgba(245,158,11,.30); color:#f59e0b}}
</style>
<div class="sh-wrap">
  <button class="sh-btn sh-wa"   onclick="doShare('wa')">📱  Share via WhatsApp</button>
  <button class="sh-btn sh-mail" onclick="doShare('email')">✉  Share via Email</button>
</div>
<script>
const PDF_B64  = `{_s["pdf_b64"]}`;
const FILENAME = `{_js_filename}`;
const MSG      = `{_js_msg}`;
const NAME     = `{_js_name}`;
const WA_FB    = `{_wa_fallback}`;
const ML_FB    = `{_ml_fallback}`;

function b64toBlob(b64){{
  const bin=atob(b64), buf=new Uint8Array(bin.length);
  for(let i=0;i<bin.length;i++) buf[i]=bin.charCodeAt(i);
  return new Blob([buf],{{type:"application/pdf"}});
}}

function triggerDownload(blob){{
  const url=URL.createObjectURL(blob);
  const a=document.createElement("a");
  a.href=url; a.download=FILENAME;
  document.body.appendChild(a); a.click();
  document.body.removeChild(a);
  setTimeout(()=>URL.revokeObjectURL(url),1000);
}}

async function doShare(via){{
  const blob=b64toBlob(PDF_B64);
  const file=new File([blob],FILENAME,{{type:"application/pdf"}});

  // Attempt native Web Share API (works on iOS/Android, modern desktop Chrome)
  if(navigator.canShare && navigator.canShare({{files:[file]}})){{
    try{{
      await navigator.share({{files:[file],title:"PARSIO Report — "+NAME,text:MSG}});
      return;
    }}catch(e){{ /* cancelled or blocked — fall through */ }}
  }}

  // Fallback: download PDF + open the channel with text message
  triggerDownload(blob);
  setTimeout(()=>{{
    if(via==="wa") window.open(WA_FB,"_blank");
    else window.location.href=ML_FB;
  }},400);
}}
</script>
""", height=52)
    with st.expander("📋  Copy message text", expanded=False):
        st.code(_s["msg"], language=None)


# ════════════════════════════════════════════════════════════════════════════
# OFFICER NOTES — persistent free-text note on the latest assessment
# Saved to the audit log; surfaces automatically when the same applicant
# returns (Borrower Memory panel) and in the admin data.
# ════════════════════════════════════════════════════════════════════════════
_lp_notes = st.session_state.last_calc_params
if isinstance(_lp_notes, dict) and isinstance(_lp_notes.get("result"), dict):
    st.markdown("---")
    with st.expander("📝  Officer Notes — attach a note to this assessment", expanded=False):
        _note_who = _lp_notes.get("report_name") or "this applicant"
        st.markdown(
            f'<div style="font-size:12px;color:#cbd5e1;margin-bottom:8px;line-height:1.7">'
            f'Notes are saved against <strong style="color:#e6ebe8">{_note_who}</strong> '
            f'and shown automatically the next time they are assessed.</div>',
            unsafe_allow_html=True,
        )
        _note_txt = st.text_area(
            "Note",
            key="officer_note_input",
            placeholder="e.g. Verified top counterparty by phone — genuine trade income. "
                        "Declined pending fresh statement (current one is 5 months old).",
            label_visibility="collapsed",
            height=90,
        )
        if st.button("💾  Save note", key="save_officer_note_btn"):
            if not (_note_txt or "").strip():
                st.error("Type a note before saving.")
            else:
                save_officer_note(
                    applicant  = _lp_notes.get("report_name") or "",
                    account_no = _lp_notes.get("acct_no") or "",
                    note       = _note_txt,
                    officer    = _OFFICER,
                    session    = _SID,
                )
                st.success("Note saved — it will appear whenever this applicant is assessed again.")
        # Existing notes for this applicant
        try:
            _exist_notes = get_officer_notes(
                _lp_notes.get("acct_no") or "", _lp_notes.get("report_name") or "")
        except Exception:
            _exist_notes = []
        if _exist_notes:
            st.markdown(
                '<div style="font-family:Plus Jakarta Sans,sans-serif;font-size:12px;'
                'letter-spacing:1.8px;color:#fbbf24;text-transform:uppercase;'
                'font-weight:800;margin:12px 0 8px">📝 Notes on file</div>'
                + "".join(
                    f'<div style="background:rgba(245,158,11,.10);border:1px solid rgba(245,158,11,.35);'
                    f'border-left:4px solid #f59e0b;border-radius:8px;'
                    f'padding:10px 14px;margin-bottom:8px">'
                    f'<div style="font-family:Plus Jakarta Sans,sans-serif;font-size:11px;'
                    f'font-weight:700;color:#fbbf24;margin-bottom:4px">'
                    f'{(n.get("ts") or "")[:10]} &nbsp;·&nbsp; {n.get("officer") or "—"}</div>'
                    f'<div style="font-size:14px;font-weight:700;color:#ffffff;line-height:1.6">'
                    f'{n.get("note") or ""}</div></div>'
                    for n in _exist_notes[:5]
                ),
                unsafe_allow_html=True,
            )


# ════════════════════════════════════════════════════════════════════════════
# FEATURE 5 — PERSISTENT WHAT-IF SCENARIOS PANEL
# Renders after any completed calculation, persists between rerenders.
# ════════════════════════════════════════════════════════════════════════════
_lp = st.session_state.last_calc_params
if isinstance(_lp, dict) and isinstance(_lp.get("result"), dict):
    _lp_result = _lp["result"]

    st.markdown("---")
    with st.expander("🔀  What-If Scenarios — adjust tenor or obligations instantly", expanded=False):
        st.markdown(
            '<div style="font-size:12px;color:#cbd5e1;margin-bottom:12px;line-height:1.7">'
            'Drag the sliders to instantly see how the result changes without re-uploading.'
            '</div>',
            unsafe_allow_html=True,
        )
        _wic1, _wic2 = st.columns(2)
        with _wic1:
            _wi_tenor = st.slider(
                "Tenor (months)",
                min_value=2, max_value=12,
                value=int(st.session_state.get("_wi_tenor", _lp["prod_type"] and 6)),
                key="_wi_tenor_sl",
            )
        with _wic2:
            _wi_other = st.slider(
                "Other monthly loan obligations (₦)",
                min_value=0,
                max_value=int(max(_lp["other_loans"] * 3, 200_000)),
                step=5_000,
                value=int(st.session_state.get("_wi_other", _lp["other_loans"])),
                key="_wi_other_sl",
                format="₦%d",
            )

        _wi_res = calculate_eligibility(
            nets    = _lp["nets"],
            counts  = _lp["counts"],
            location     = _lp["location"],
            product_type = _lp["prod_type"],
            tenor        = _wi_tenor,
            other_loans  = _wi_other,
            manual_rate_percent = _lp["manual_rate"] if _lp["manual_rate"] > 0 else None,
            sel_mode     = _lp.get("sel_mode", False),
        )

        # Side-by-side comparison
        _curr = _lp_result
        _chg_loan = _wi_res.get("max_loan", 0) - _curr.get("max_loan", 0)
        _chg_col  = "#34d399" if _chg_loan >= 0 else "#f87171"

        st.markdown(
            f'<div style="display:grid;grid-template-columns:1fr 1fr;gap:16px;margin-top:8px">'
            # Current
            f'<div style="padding:14px;background:rgba(0,0,0,.2);border:1px solid #1a3d2b;border-radius:4px">'
            f'<div style="font-size:9px;letter-spacing:2px;color:#64748b;text-transform:uppercase;margin-bottom:8px">Current</div>'
            f'<div style="font-size:11px;color:#94a3b8;margin-bottom:4px">Tenor: <span style="color:#e2e8f0">{_curr.get("tenor")} mo</span></div>'
            f'<div style="font-size:11px;color:#94a3b8;margin-bottom:4px">Max Loan: <span style="color:#10b981;font-weight:700">{money(_curr.get("max_loan",0))}</span></div>'
            f'<div style="font-size:11px;color:#94a3b8;margin-bottom:4px">Repayment: <span style="color:#e2e8f0">{money(_curr.get("max_repayment_display",0))}/{_curr.get("repayment_frequency","mo").lower()[:2]}</span></div>'
            f'<div style="font-size:11px;color:#94a3b8">Rate: <span style="color:#fbbf24">{pct(_curr.get("interest_rate"))}</span></div>'
            f'</div>'
            # What-If
            f'<div style="padding:14px;background:rgba(16,185,129,.04);border:1px solid rgba(16,185,129,.25);border-radius:4px">'
            f'<div style="font-size:9px;letter-spacing:2px;color:#10b981;text-transform:uppercase;margin-bottom:8px">What-If ({_wi_tenor} mo, ₦{_wi_other:,.0f} obligations)</div>'
            f'<div style="font-size:11px;color:#94a3b8;margin-bottom:4px">Tenor: <span style="color:#e2e8f0">{_wi_tenor} mo</span></div>'
            f'<div style="font-size:11px;color:#94a3b8;margin-bottom:4px">Max Loan: <span style="color:{_chg_col};font-weight:700">{money(_wi_res.get("max_loan",0))}</span>'
            f'&nbsp;<span style="font-size:10px;color:{_chg_col}">({_chg_loan:+,.0f})</span></div>'
            f'<div style="font-size:11px;color:#94a3b8;margin-bottom:4px">Repayment: <span style="color:#e2e8f0">{money(_wi_res.get("max_repayment_display",0))}/{_wi_res.get("repayment_frequency","mo").lower()[:2]}</span></div>'
            f'<div style="font-size:11px;color:#94a3b8">Rate: <span style="color:#fbbf24">{pct(_wi_res.get("interest_rate"))}</span></div>'
            f'</div>'
            f'</div>',
            unsafe_allow_html=True,
        )
        if not _wi_res.get("approved"):
            st.markdown(
                '<div style="margin-top:8px;font-size:11px;color:#f87171">'
                '⚠ This scenario falls below the product minimum — loan would not be approved.</div>',
                unsafe_allow_html=True,
            )


# ════════════════════════════════════════════════════════════════════════════
# ADMIN DASHBOARD — accessible only via ?admin=<PARSIO_ADMIN_KEY>
# Set the environment variable PARSIO_ADMIN_KEY to your chosen secret.
# Example URL: https://your-app.streamlit.app/?admin=mySecret99
# ════════════════════════════════════════════════════════════════════════════
_ADMIN_KEY = os.environ.get("PARSIO_ADMIN_KEY", os.environ.get("SEL_ADMIN_KEY", "kvic7admin"))   # ← change via env var
_qp = st.query_params
if _qp.get("admin") == _ADMIN_KEY:
    st.markdown("---")
    st.markdown(
        '<div style="font-size:10px;letter-spacing:4px;color:#f59e0b;'
        'text-transform:uppercase;margin-bottom:8px">⚙ Admin Only</div>',
        unsafe_allow_html=True,
    )
    st.markdown("## 📊 Usage Dashboard")

    _stats = admin_stats()

    # ── Storage backend indicator ──────────────────────────────────────────
    _backend = _stats.get("backend", "Unknown")
    _is_persistent = "PostgreSQL" in _backend or "Neon" in _backend
    _bk_col = "#34d399" if _is_persistent else "#f87171"
    _bk_note = ("✓ Data persists permanently across redeploys"
                if _is_persistent else
                "⚠ Ephemeral — data resets on redeploy. Set DATABASE_URL secret to persist.")
    st.markdown(
        f'<div style="display:inline-flex;align-items:center;gap:10px;margin-bottom:12px;'
        f'padding:6px 14px;background:rgba(255,255,255,.03);border:1px solid {_bk_col}44;'
        f'border-radius:4px;font-size:11px">'
        f'<span style="color:#64748b;letter-spacing:1px;text-transform:uppercase">Storage</span>'
        f'<span style="color:{_bk_col};font-weight:700">{_backend}</span>'
        f'<span style="color:#64748b">· {_bk_note}</span>'
        f'</div>',
        unsafe_allow_html=True,
    )

    if "_error" in _stats:
        st.error(f"Tracker DB error: {_stats['_error']}")
    else:
        # ── Top-line metrics ──────────────────────────────────────────────
        _sess   = _stats.get("sessions", {})
        _rate   = _stats.get("rate", {})
        _ok     = int(_rate.get("ok") or 0)
        _err    = int(_rate.get("err") or 0)
        _total  = _ok + _err
        _err_pct = round(_err / _total * 100, 1) if _total else 0

        mc1, mc2, mc3, mc4 = st.columns(4)
        mc1.metric("Total Sessions",   _sess.get("total_sessions", 0))
        mc2.metric("Statements Parsed", _ok)
        mc3.metric("Parse Errors",      _err,
                   delta=f"{_err_pct}% error rate",
                   delta_color="inverse")
        mc4.metric("Results Calculated", _sess.get("completed", 0))

        # ── Event summary ─────────────────────────────────────────────────
        _sum = _stats.get("summary", [])
        if _sum:
            st.markdown("#### Event Breakdown")
            st.dataframe(
                pd.DataFrame(_sum).rename(columns={"event": "Event", "total": "Count"}),
                hide_index=True, use_container_width=True,
            )

        # ── Daily activity ────────────────────────────────────────────────
        _daily = _stats.get("daily", [])
        if _daily:
            st.markdown("#### Daily Upload Activity (last 30 days)")
            _df_daily = pd.DataFrame(_daily).rename(
                columns={"day": "Date", "uploads": "Uploads"}
            ).sort_values("Date")
            st.markdown(
                html_line_chart(
                    [d[5:] for d in _df_daily["Date"].tolist()],  # MM-DD labels
                    _df_daily["Uploads"].tolist(),
                    color="#10b981",
                ),
                unsafe_allow_html=True,
            )

        # ── Bank distribution ─────────────────────────────────────────────
        _banks = _stats.get("banks", [])
        if _banks:
            st.markdown("#### Bank Distribution")
            st.dataframe(
                pd.DataFrame(_banks).rename(
                    columns={"bank": "Bank", "cnt": "Statements Parsed"}
                ),
                hide_index=True, use_container_width=True,
            )

        # ── Loan results ──────────────────────────────────────────────────
        _loans = _stats.get("loans", [])
        if _loans:
            st.markdown("#### Recent Eligibility Results")
            _loan_rows = []
            for _lr in _loans:
                try:
                    _d = json.loads(_lr["data"]) if isinstance(_lr["data"], str) else _lr["data"]
                except Exception:
                    _d = {}
                _loan_rows.append({
                    "Time":       _lr["ts"],
                    "Officer":    _d.get("officer",    "—") or "—",
                    "Applicant":  _d.get("applicant",  "—") or "—",
                    "Account No": _d.get("account_no", "—") or "—",
                    "Bank":       _lr.get("bank", ""),
                    "Decision":   "✅ Approved" if _d.get("approved") else "❌ Below Min",
                    "Max Loan":   f"NGN {_d.get('max_loan', 0):,.0f}",
                    "Tenor":      f"{_d.get('tenor', '—')} mo",
                    "DTI":        f"{_d.get('dti', 0):.1f}%",
                    "Product":    _d.get("product", ""),
                    "Location":   _d.get("location", ""),
                })
            st.dataframe(
                pd.DataFrame(_loan_rows),
                hide_index=True, use_container_width=True,
            )

        # ── Error log ─────────────────────────────────────────────────────
        _errors = _stats.get("errors", [])
        if _errors:
            st.markdown(
                f'<div style="color:#f87171;font-size:13px;font-weight:700;'
                f'margin-top:12px">⚠ Parse Errors ({len(_errors)} most recent)</div>',
                unsafe_allow_html=True,
            )
            for _e in _errors:
                try:
                    _ed = json.loads(_e["data"]) if isinstance(_e["data"], str) else _e["data"]
                except Exception:
                    _ed = {}
                with st.expander(
                    f"🔴  {_e['ts']}  |  {_e.get('bank') or 'Unknown bank'}"
                    f"  |  {_e.get('filename', '')}",
                    expanded=False,
                ):
                    st.code(
                        f"Error type : {_ed.get('error_type', '—')}\n"
                        f"Message    : {_ed.get('error', '—')}\n"
                        f"File       : {_e.get('filename', '—')}\n"
                        f"Session    : {_e.get('session', '—')}\n"
                        f"Slot       : {_ed.get('slot', 'A')}",
                        language="text",
                    )
        else:
            st.success("✅ No parse errors recorded.")

        # ── Heavy Statements (memory/CPU risk) ─────────────────────────────
        _heavy_all = _stats.get("heavy", [])
        _heavy_rows = []
        for _h in _heavy_all:
            try:
                _hd = json.loads(_h["data"]) if isinstance(_h["data"], str) else (_h["data"] or {})
            except Exception:
                _hd = {}
            _heavy_rows.append({
                "Time":       _h["ts"],
                "Officer":    _hd.get("officer", "—"),
                "Bank":       _h.get("bank") or "—",
                "File":       _h.get("filename", ""),
                "Size (KB)":  _hd.get("size_kb", 0),
                "Parse (s)":  _hd.get("duration_s", 0),
                "Txns":       _hd.get("txns", 0),
                "Heavy":      "🔴" if str(_hd.get("heavy")) == "True" else "",
            })
        if _heavy_rows:
            _n_heavy = sum(1 for r in _heavy_rows if r["Heavy"])
            st.markdown(
                f'<div style="color:#fbbf24;font-size:13px;font-weight:700;margin-top:18px">'
                f'🏋 Statement Load Tracker — {len(_heavy_rows)} parses logged, '
                f'<span style="color:#f87171">{_n_heavy} heavy</span> '
                f'(&gt;60s or &gt;5,000 txns — these can exhaust the server\'s memory)</div>',
                unsafe_allow_html=True,
            )
            st.dataframe(
                pd.DataFrame(_heavy_rows),
                hide_index=True,
                use_container_width=True,
            )

        # ── Officer Performance Dashboard ─────────────────────────────────
        _off_act = _stats.get("officer_activity", [])
        if _off_act:
            st.markdown("---")
            st.markdown("### 👤 Officer Performance Dashboard")

            # Compute fleet averages for outlier detection
            _oa_all_rates = [
                (int(_oa.get("approved") or 0) / int(_oa.get("assessments") or 1) * 100)
                for _oa in _off_act if int(_oa.get("assessments") or 0) >= 3
            ]
            _oa_fleet_rate = sum(_oa_all_rates) / len(_oa_all_rates) if _oa_all_rates else None
            _oa_all_loans  = [
                float(_oa.get("avg_loan_approved") or 0)
                for _oa in _off_act if (_oa.get("avg_loan_approved") or 0) > 0
            ]
            _oa_fleet_loan = sum(_oa_all_loans) / len(_oa_all_loans) if _oa_all_loans else None
            _oa_all_dtis   = [
                float(_oa.get("avg_dti") or 0)
                for _oa in _off_act if (_oa.get("avg_dti") or 0) > 0
            ]
            _oa_fleet_dti  = sum(_oa_all_dtis) / len(_oa_all_dtis) if _oa_all_dtis else None

            _oa_rows = []
            for _oa in _off_act:
                _oa_tot  = int(_oa.get("assessments") or 0)
                _oa_apr  = int(_oa.get("approved") or 0)
                _oa_rate = round(_oa_apr / _oa_tot * 100, 1) if _oa_tot else 0
                _oa_loan = float(_oa.get("avg_loan_approved") or 0)
                _oa_dti  = float(_oa.get("avg_dti") or 0)
                # Outlier flags (only when fleet has ≥2 officers with enough data)
                _flags = []
                if _oa_fleet_rate is not None and _oa_tot >= 3:
                    if _oa_rate > _oa_fleet_rate + 20:
                        _flags.append("⬆ approval")
                    elif _oa_rate < _oa_fleet_rate - 20:
                        _flags.append("⬇ approval")
                if _oa_fleet_loan and _oa_loan > 0:
                    if _oa_loan > _oa_fleet_loan * 1.4:
                        _flags.append("⬆ loan size")
                    elif _oa_loan < _oa_fleet_loan * 0.6:
                        _flags.append("⬇ loan size")
                if _oa_fleet_dti and _oa_dti > 0:
                    if _oa_dti > _oa_fleet_dti * 1.25:
                        _flags.append("⬆ DTI")
                _oa_rows.append({
                    "Officer":            _oa["officer"],
                    "Assessments":        _oa_tot,
                    "Approved":           _oa_apr,
                    "Below Min":          _oa_tot - _oa_apr,
                    "Approval Rate %":    _oa_rate,
                    "Avg Loan (Approved)":int(_oa_loan) if _oa_loan else 0,
                    "Avg DTI %":          _oa_dti,
                    "Last Active":        _oa.get("last_active", "—"),
                    "Outlier Flags":      ", ".join(_flags) if _flags else "—",
                })

            _oa_df = pd.DataFrame(_oa_rows)
            st.dataframe(
                _oa_df,
                hide_index=True,
                use_container_width=True,
                column_config={
                    "Approval Rate %": st.column_config.ProgressColumn(
                        "Approval Rate %", min_value=0, max_value=100, format="%.1f%%"
                    ),
                    "Avg Loan (Approved)": st.column_config.NumberColumn(
                        "Avg Loan (Approved)", format="₦%d"
                    ),
                    "Avg DTI %": st.column_config.NumberColumn(
                        "Avg DTI %", format="%.1f%%"
                    ),
                    "Outlier Flags": st.column_config.TextColumn(
                        "Outlier Flags", help="Officers whose metrics are >20% above/below fleet average"
                    ),
                },
            )

            # Fleet summary row
            if _oa_fleet_rate is not None:
                st.markdown(
                    f'<div style="font-size:10px;color:#64748b;margin-top:4px">'
                    f'Fleet avg (≥3 assessments): '
                    f'<span style="color:#10b981">approval {_oa_fleet_rate:.1f}%</span>'
                    + (f' &nbsp;|&nbsp; <span style="color:#fbbf24">avg loan ₦{_oa_fleet_loan:,.0f}</span>' if _oa_fleet_loan else '')
                    + (f' &nbsp;|&nbsp; avg DTI {_oa_fleet_dti:.1f}%' if _oa_fleet_dti else '')
                    + '</div>',
                    unsafe_allow_html=True,
                )

        # ── Sign-in Log ───────────────────────────────────────────────────
        # Shows every officer who signed in — even those who never ran a
        # calculation.  Useful for attendance / access-control monitoring.
        st.markdown("---")
        st.markdown("### 🔐 Sign-in Log")

        _signin_today  = _stats.get("signin_today", 0)
        _signin_log    = _stats.get("signin_log", [])
        _signin_summ   = _stats.get("signin_summary", [])

        _sl1, _sl2, _sl3 = st.columns(3)
        _sl1.metric("Sign-ins Today",   _signin_today)
        _sl2.metric("Total Sign-ins",   len(_signin_log))
        _sl3.metric("Unique Officers",  len(_signin_summ))

        if _signin_summ:
            st.markdown("#### Officers — All-time Sign-in Summary")
            # Flag officers who only signed in but never ran a calculation
            _calc_officers = {
                str(_oa.get("officer", "")).strip().lower()
                for _oa in _stats.get("officer_activity", [])
            }
            _ss_rows = []
            for _ss in _signin_summ:
                _off = str(_ss.get("officer") or "Unknown")
                _used_calc = _off.strip().lower() in _calc_officers
                _ss_rows.append({
                    "Officer":        _off,
                    "Total Sign-ins": int(_ss.get("total_signins") or 0),
                    "Last Seen":      _ss.get("last_seen") or "—",
                    "Used Calculator": "✅ Yes" if _used_calc else "🔲 Sign-in only",
                })
            st.dataframe(
                pd.DataFrame(_ss_rows),
                hide_index=True,
                use_container_width=True,
            )

        if _signin_log:
            st.markdown("#### Recent Sign-ins (newest first)")
            _sl_rows = []
            for _sl in _signin_log:
                _ts_raw = _sl.get("ts", "")
                # Format timestamp: "2025-06-03T14:22:11" → "03 Jun 2025  14:22"
                try:
                    _dt = datetime.datetime.strptime(_ts_raw[:16], "%Y-%m-%dT%H:%M")
                    _ts_fmt = _dt.strftime("%d %b %Y  %H:%M") + " UTC"
                except Exception:
                    _ts_fmt = _ts_raw
                _sl_rows.append({
                    "Date / Time (UTC)": _ts_fmt,
                    "Officer":           _sl.get("officer") or "—",
                    "Session ID":        (_sl.get("session") or "")[:16] + "…"
                                         if len(_sl.get("session") or "") > 16
                                         else (_sl.get("session") or "—"),
                })
            st.dataframe(
                pd.DataFrame(_sl_rows),
                hide_index=True,
                use_container_width=True,
            )
        else:
            st.info("No sign-in events recorded yet.")

        # ── Feature 11: Portfolio Analytics ───────────────────────────────
        st.markdown("---")
        st.markdown("### 📈 Portfolio Analytics")

        # Approval rate by bank
        _appr_bank = _stats.get("approval_by_bank", [])
        if _appr_bank:
            st.markdown("#### Approval Rate by Bank")
            _ab_rows = []
            for _ab in _appr_bank:
                _tot = int(_ab.get("total") or 0)
                _apr = int(_ab.get("approved") or 0)
                _rate_pct = round(_apr / _tot * 100, 1) if _tot else 0
                _ab_rows.append({
                    "Bank":           _ab["bank"],
                    "Submissions":    _tot,
                    "Approved":       _apr,
                    "Below Min":      _tot - _apr,
                    "Approval Rate %": _rate_pct,
                })
            _ab_df = pd.DataFrame(_ab_rows)
            st.dataframe(
                _ab_df,
                hide_index=True,
                use_container_width=True,
                column_config={
                    "Approval Rate %": st.column_config.ProgressColumn(
                        "Approval Rate %", min_value=0, max_value=100, format="%.1f%%"
                    ),
                },
            )

        # Loan volume & approval by month
        # ── Daily loan trend (current + previous month) ───────────────────
        _lbd = _stats.get("loans_by_day", [])
        _lbm = _stats.get("loans_by_month", [])
        if _lbd:
            # Build full date range: fill missing days with 0
            import datetime as _dt2
            _today      = _dt2.date.today()
            _prev_month = (_today.replace(day=1) - _dt2.timedelta(days=1)).replace(day=1)
            _day_map_count   = {r["day"]: int(r["count"]   or 0) for r in _lbd}
            _day_map_avgloan = {r["day"]: float(r["avg_loan"] or 0) for r in _lbd}
            # Generate all days from start of previous month to today
            _day_labels = []
            _day_counts = []
            _day_loans  = []
            _cur = _prev_month
            _month_markers = {}   # index → month label for divider
            while _cur <= _today:
                _ds = str(_cur)
                if _cur.day == 1:
                    _month_markers[len(_day_labels)] = _cur.strftime("%b %Y")
                _day_labels.append(_cur.strftime("%d"))   # just the day number
                _day_counts.append(_day_map_count.get(_ds, 0))
                _day_loans.append(_day_map_avgloan.get(_ds, 0))
                _cur += _dt2.timedelta(days=1)

            # Build x-axis labels: show month name at day=1, else day number
            _xlabels = []
            for _di, _lbl in enumerate(_day_labels):
                if _di in _month_markers:
                    _xlabels.append(_month_markers[_di])
                else:
                    _xlabels.append(_lbl)

            st.markdown("#### Daily Assessment Trend (previous + current month)")
            st.markdown("**Daily assessments**")
            st.markdown(
                html_bar_chart(
                    _xlabels,
                    _day_counts,
                    color="#10b981",
                    height=180,
                ),
                unsafe_allow_html=True,
            )
            st.markdown("<div style='margin-top:24px'></div>", unsafe_allow_html=True)
            st.markdown("**Daily avg max loan**")
            st.markdown(
                html_bar_chart(
                    _xlabels,
                    _day_loans,
                    color="#fbbf24",
                    money_fmt=True,
                    height=180,
                ),
                unsafe_allow_html=True,
            )
        elif _lbm:
            # Fallback: no daily data yet — show monthly aggregates
            st.markdown("#### Loan Volume by Month")
            _lbm_df = pd.DataFrame(_lbm).rename(columns={
                "month": "Month", "avg_loan": "Avg Max Loan (NGN)",
                "count": "Assessments", "approved": "Approved",
            }).sort_values("Month")
            _ac1, _ac2 = st.columns(2)
            with _ac1:
                st.markdown("**Assessment volume trend**")
                st.markdown(
                    html_line_chart(_lbm_df["Month"].tolist(), _lbm_df["Assessments"].tolist(), color="#10b981"),
                    unsafe_allow_html=True,
                )
            with _ac2:
                st.markdown("**Average max loan trend**")
                st.markdown(
                    html_line_chart(_lbm_df["Month"].tolist(), _lbm_df["Avg Max Loan (NGN)"].tolist(), color="#fbbf24", money_fmt=True),
                    unsafe_allow_html=True,
                )

        # Rejection breakdown
        _rej = _stats.get("rejection_reasons", [])
        if _rej:
            st.markdown("#### Top Rejection Combinations (Product × Location)")
            _rej_df = pd.DataFrame(_rej).rename(columns={
                "product": "Product", "location": "Location", "count": "Rejections"
            })
            st.dataframe(_rej_df, hide_index=True, use_container_width=True)

        # Download format breakdown
        _dl_fmt = _stats.get("download_formats", [])
        if _dl_fmt:
            st.markdown("#### Downloads by Format")
            _dl_df = pd.DataFrame(_dl_fmt).rename(columns={"fmt": "Format", "count": "Downloads"})
            st.dataframe(_dl_df, hide_index=True, use_container_width=True)

        # ── Audit log + raw DB downloads ──────────────────────────────────
        st.markdown("---")
        st.markdown("### 📥 Export")
        _ex1, _ex2 = st.columns(2)
        with _ex1:
            try:
                _audit_csv = export_audit_csv()
                st.download_button(
                    "⬇  Download Full Audit Log (CSV)",
                    _audit_csv.encode("utf-8-sig"),   # BOM → opens cleanly in Excel
                    file_name=f"PARSIO_Audit_Log_{datetime.date.today():%Y%m%d}.csv",
                    mime="text/csv",
                    key="dl_audit_log",
                    use_container_width=True,
                    help="Every assessment: officer, applicant, account no, decision, loan, etc.",
                )
            except Exception as _ex_err:
                st.caption(f"Audit export unavailable: {_ex_err}")
        with _ex2:
            _db_path = pathlib.Path(__file__).parent / "sel_analytics.db"
            if _db_path.exists():
                with open(_db_path, "rb") as _dbf:
                    st.download_button(
                        "⬇  Download Raw SQLite DB (local cache)",
                        _dbf.read(),
                        file_name="sel_analytics.db",
                        mime="application/octet-stream",
                        key="dl_admin_db",
                        use_container_width=True,
                        help="Local SQLite snapshot. On Neon, use the CSV export for live data.",
                    )

    # ── Blacklist / Watchlist Management ──────────────────────────────────────
    st.markdown("---")
    st.markdown("### 🚫 Blacklist / Watchlist")
    st.markdown(
        '<div style="font-size:11px;color:#64748b;margin-bottom:10px">'
        'Upload a CSV with flagged names or account numbers. '
        'The app will warn officers when a match is detected at parse time.</div>',
        unsafe_allow_html=True,
    )

    _bl_tab1, _bl_tab2 = st.tabs(["Upload / Add", "Current Entries"])

    with _bl_tab1:
        _bl_mode = st.radio("Add method", ["Upload CSV", "Enter manually"],
                            horizontal=True, key="bl_add_mode")

        if _bl_mode == "Upload CSV":
            st.markdown(
                '<div style="font-size:10px;color:#64748b;margin-bottom:6px">'
                'CSV format: <code>entry_type,value,reason</code> — '
                'entry_type = <code>name</code> or <code>account_no</code></div>',
                unsafe_allow_html=True,
            )
            _bl_file = st.file_uploader("Upload blacklist CSV", type=["csv"],
                                        key="bl_upload", label_visibility="collapsed")
            if _bl_file and st.button("Import Blacklist", key="bl_import_btn"):
                import csv as _csv_mod
                import io as _bl_io
                _bl_content = _bl_file.read().decode("utf-8-sig")
                _bl_reader  = _csv_mod.DictReader(_bl_io.StringIO(_bl_content))
                _bl_entries = []
                for _br in _bl_reader:
                    # Normalise headers — accept variations
                    _brow = {k.strip().lower().replace(" ", "_"): v.strip() for k, v in _br.items()}
                    _bl_entries.append({
                        "entry_type": _brow.get("entry_type", _brow.get("type", "name")),
                        "value":      _brow.get("value", _brow.get("name", _brow.get("account_no", ""))),
                        "reason":     _brow.get("reason", _brow.get("notes", "")),
                    })
                _bl_n = save_blacklist_entries(_bl_entries, added_by=_OFFICER)
                st.success(f"Imported {_bl_n} entries to the watchlist.")

        else:
            _bl_m1, _bl_m2 = st.columns([2, 3])
            with _bl_m1:
                _bl_etype = st.selectbox("Type", ["name", "account_no"], key="bl_etype_manual")
            with _bl_m2:
                _bl_val = st.text_input("Value (name or account number)", key="bl_val_manual",
                                        label_visibility="collapsed",
                                        placeholder="e.g. JOHN DOE or 0012345678")
            _bl_reason = st.text_input("Reason (optional)", key="bl_reason_manual",
                                       placeholder="e.g. Previous fraud, court order")
            if st.button("Add to Watchlist", key="bl_manual_add"):
                if _bl_val.strip():
                    save_blacklist_entries(
                        [{"entry_type": _bl_etype, "value": _bl_val.strip(), "reason": _bl_reason}],
                        added_by=_OFFICER,
                    )
                    st.success(f"Added '{_bl_val.strip()}' to watchlist.")
                else:
                    st.warning("Please enter a value before adding.")

    with _bl_tab2:
        _bl_current = get_blacklist()
        if not _bl_current:
            st.info("Watchlist is empty.")
        else:
            st.markdown(f'<div style="font-size:11px;color:#64748b;margin-bottom:6px">'
                        f'{len(_bl_current)} entr{"ies" if len(_bl_current) != 1 else "y"} on the watchlist</div>',
                        unsafe_allow_html=True)
            _bl_rows_html = ""
            for _be in _bl_current:
                _bl_rows_html += (
                    f'<tr>'
                    f'<td style="font-size:10px;color:#94a3b8">{(_be.get("ts") or "")[:10]}</td>'
                    f'<td style="font-size:11px;color:#fbbf24">{_be.get("entry_type","")}</td>'
                    f'<td style="font-size:11px;font-weight:700">{_be.get("value","")}</td>'
                    f'<td style="font-size:11px;color:#94a3b8">{_be.get("reason","")}</td>'
                    f'<td style="font-size:10px;color:#64748b">{_be.get("added_by","")}</td>'
                    f'</tr>'
                )
            st.markdown(
                f'<table class="preview-table"><thead><tr>'
                f'<th style="text-align:left">Added</th><th>Type</th>'
                f'<th>Value</th><th>Reason</th><th>Added By</th>'
                f'</tr></thead><tbody>{_bl_rows_html}</tbody></table>',
                unsafe_allow_html=True,
            )
            if st.button("🗑  Clear Entire Watchlist", key="bl_clear_all"):
                clear_blacklist()
                st.success("Watchlist cleared.")
                st.rerun()
