# NOTE: Do NOT add `from __future__ import annotations` here.
# This module is named `parser`, which collides with the removed CPython
# stdlib module of the same name. A dependency importing the old stdlib
# `parser` leaves sys.modules['parser'] = None; with stringized annotations,
# @dataclass then tries sys.modules['parser'].__dict__ and crashes on
# Python 3.12 ("'NoneType' object has no attribute '__dict__'"). Keeping
# real (non-stringized) annotations avoids that code path entirely.

import gc
import re
import subprocess
from collections import defaultdict
from dataclasses import dataclass
from decimal import Decimal
from io import BytesIO
from typing import Optional

from PyPDF2 import PdfReader

# ── Transaction log (reset on each parse_transactions call) ─────────────────
_txn_log: list[dict] = []

# ── Full-text cache — holds the PyPDF2 text from the most recent
#    parse_transactions call so the caller can reuse it for account-number
#    extraction and accuracy verification without re-parsing the PDF.
#    Reset to "" at the start of every parse_transactions call.
_last_full_text: str = ""


def get_last_full_text() -> str:
    """Return the raw extracted text from the most recent parse_transactions call.

    Allows the caller (app.py) to perform account-number extraction and accuracy
    verification using the text already produced during the main parse — avoiding
    two or three additional PyPDF2 / pdfplumber re-parses of the same PDF file.
    The returned string is freed by the caller after use; its lifetime ends when
    the caller does `del text; gc.collect()`.
    """
    return _last_full_text


# Excel support — openpyxl for .xlsx, xlrd for .xls
try:
    import openpyxl
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

# ── Shared constants ──────────────────────────────────────────────────────────
MONTH_ABBR = {1:"Jan",2:"Feb",3:"Mar",4:"Apr",5:"May",6:"Jun",
              7:"Jul",8:"Aug",9:"Sep",10:"Oct",11:"Nov",12:"Dec"}
MONTH_MAP  = {"jan":1,"feb":2,"mar":3,"apr":4,"may":5,"jun":6,
              "jul":7,"aug":8,"sep":9,"oct":10,"nov":11,"dec":12}
MONTH_NUM  = {v.lower(): str(k).zfill(2) for k, v in MONTH_ABBR.items()}

MONEY_RE   = re.compile(r"(?<!\d)(\d{1,3}(?:,\d{3})*|\d+)\.\d{2}(?!\d)")
ZENITH_ROW = re.compile(r"^(\d{2}/\d{2}/\d{4})\s+(\d{2}/\d{2}/\d{4})\s+(.*)$")
OPAY_ROW   = re.compile(
    r"^(\d{2})\s+(Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)"
    r"\s+(20\d{2})\s+\d{2}:\d{2}:\d{2}", re.I,
)

# ── FairMoney parser constants ────────────────────────────────────────────────
# Transaction anchor: DD/MM/YYYY  REFNUM  [+/-] ₦ AMOUNT  ₦ BALANCE[Narration]
_FM_TX = re.compile(
    r"^(\d{2}/\d{2}/\d{4})\s+\d+\s+"    # date  reference-number
    r"([+\-])\s*₦\s*([\d,]+\.\d{2})\s+" # sign  amount
    r"₦\s*[\d,]+\.\d{2}"                 # balance (consumed, not captured)
    r"(.*)"                               # narration start (may be empty or run-on)
)
# Lines that are page-header boilerplate to skip during narration accumulation
_FM_HEADER = re.compile(
    r"^(?:FairMoney MFB|Licensed by CBN|28 Pade Odanye|Phone:|Email:|"
    r"Account number|Date Reference|number|Transaction|details|"
    r"Credit Debit Account|balance|Opening Balance|Total Deposits|"
    r"Total Withdrawals|Closing Balance|Page \d+ of \d+)",
    re.I,
)
# Pure-money lines (e.g. "₦ 62.00") and date-range lines to skip
_FM_MONEY_ONLY = re.compile(r"^₦\s*[\d,]+\.\d{2}$")
_FM_DATE_RANGE = re.compile(r"^\d{2}/\d{2}/\d{4}\s+-\s+\d{2}/\d{2}/\d{4}$")


# ── Data classes ──────────────────────────────────────────────────────────────
@dataclass
class CreditAccount:
    subscriber_name: str
    account_number: str
    account_status: str
    facility_classification: str
    instalment_amount: float
    outstanding_balance: float
    loan_duration_days: str
    tenor_months: Optional[int]
    monthly_obligation: float
    derived_from_balance: bool
    include_in_obligation: bool
    is_closed: bool
    is_bad_credit: bool


# ════════════════════════════════════════════════════════════════════════════
# HELPERS
# ════════════════════════════════════════════════════════════════════════════
def ym_label(ym: str) -> str:
    year, month = ym.split("-")
    return f"{MONTH_ABBR[int(month)]} {year[-2:]}"


def _empty_bucket() -> dict:
    return {"gross": 0.0, "count": 0, "self_transfer": 0.0,
            "reversal": 0.0, "non_business": 0.0, "loan_disbursal": 0.0,
            "real_credit": 0.0}


def _parse_currency(v: str) -> float:
    return float(re.sub(r"[^\d.]", "", str(v or "")) or 0)


def _get_tenor_months(raw: str) -> Optional[int]:
    digits = int(re.sub(r"[^\d]", "", str(raw)) or 0)
    return max(1, -(-digits // 30)) if digits > 0 else None


# ════════════════════════════════════════════════════════════════════════════
# CLASSIFICATION  — self-transfer bug fix
# ════════════════════════════════════════════════════════════════════════════
def classify_credit(narration: str, account_name: str = "") -> tuple[str, str]:
    """
    Classify a credit transaction into one of:
      self_transfer  — OWealth, savings platforms, own-name transfers
      reversal       — RSVL, REV, refund, chargeback, dispute
      loan_disbursal — Loan apps, disbursement keywords
      non_business   — Gambling, betting, salary, allowance, support, contribution
      real_credit    — Genuine usable income

    SELF-TRANSFER FIX: Only flags self-transfer when the account holder name
    appears as the SENDER (60 chars after the transfer-from verb), NOT as the
    recipient. This prevents false flags like 'Transfer from X TO EMMANUEL'.

    CONTRIBUTION: ALL "contribution" credits are excluded (non_business) —
    whether from the account holder or from outsiders. Contributions (ajo,
    esusu, cooperative, group savings) are not verifiable business income.

    OVERDRAFT: Credits tagged as overdraft, OD credit, or facility drawdown
    are excluded as loan_disbursal — they are borrowed funds, not income.
    """
    text = narration.lower()

    # ── 1. OWealth / savings wallet round-trips ───────────────────────────
    owealth_kw = [
        "owealth withdrawal", "owealth deposit", "owealth interest",
        "auto-save to owealth", "savings withdrawal", "owealth balance",
        "owealth credit",
    ]
    if any(k in text for k in owealth_kw):
        return "self_transfer", "OWealth internal round-trip"

    # ── 1b. Renmoney internal savings products ────────────────────────────
    # Renflex = Renmoney flexible savings; RenVault = Renmoney vault savings;
    # RenSavings = any other Renmoney savings product.
    # These are own-fund round-trips — NOT business income.
    renmoney_savings_kw = ["renflex", "renvault", "rensavings"]
    if any(k in text for k in renmoney_savings_kw):
        return "self_transfer", "Renmoney internal savings round-trip"

    # ── 2. Savings platforms ──────────────────────────────────────────────
    savings_kw = [
        "piggyvest", "piggy vest", "piggy bank",
        "cowrywise", "cowry wise",
        "kuda save", "kuda vault",
    ]
    if any(k in text for k in savings_kw):
        return "self_transfer", "Savings platform round-trip"

    # ── 3. Overdraft / facility drawdown → loan_disbursal ────────────────
    # Check BEFORE reversals: some OD narrations contain "reversal" in them
    # (e.g. "overdraft credit limit reversal") but are still borrowed funds.
    if re.search(
        r"\boverdraft\b|\bod\s+credit\b|\bod\s+limit\b|\bod\s+utiliz",
        text,
    ):
        return "loan_disbursal", "Overdraft credit"
    if re.search(
        r"\bfacility\s+drawdown\b|\bcredit\s+facility\b|\bod\s+reversal\b",
        text,
    ):
        return "loan_disbursal", "Facility/OD drawdown"

    # ── 4. Reversals ──────────────────────────────────────────────────────
    if re.search(r"\*+rsvl|\brsvl\b|\brvsl\b|\brev\b|\brev-\b", text):
        return "reversal", "RSVL/RVSL/REV reversal marker"
    if any(k in text for k in [
        "reversal", "reversed", "refund", "chargeback", "chargbk",
        "dispute", "clawback", "returned funds", "charge back",
        "transaction reversed", "payment reversed",
    ]):
        return "reversal", "Reversal keyword"

    # ── 4. Loan disbursements ─────────────────────────────────────────────
    loan_exact = [
        "fairmoney", "carbon loan", "branch loan", "branch limit",
        "palmcredit", "palm credit", "aella credit", "aella",
        "kiakia", "kia kia", "quickcheck", "quick check",
        "migo ", "lidya", "zedvance", "creditwave", "creditmfb",
        "renmoney", "easemoni", "okash", "xcredit", "newcredit",
        "fast credit", "page financ", "lendigo", "specta",
        "loans by sterling", "credit direct", "gtbank loan",
        "access loan", "uba loan", "first bank loan",
        "zenith loan", "wema loan", "fcmb loan",
        # FairMoney internal credits (interest/wallet round-trips on FairMoney accounts)
        "incoming fairmoney",
    ]
    loan_regex = [
        r"\bloan\s+disburs", r"\bdisbursement\b", r"\bcredit\s+disburs",
        r"\bloan\s+credit\b", r"\bloan\s+repay",
    ]
    # "renmoney" in narration could mean the DESTINATION bank (e.g. "To RenMoney
    # Microfinance Bank | Payment for goods") — not a loan disbursement.
    # Only treat "renmoney" as a loan keyword when it is NOT the destination bank.
    _active_loan_exact = [
        k for k in loan_exact
        if k != "renmoney" or not re.search(r"\bto\s*:?\s*renmoney\b", text)
    ]
    if any(k in text for k in _active_loan_exact):
        return "loan_disbursal", "Loan app keyword"
    for pat in loan_regex:
        if re.search(pat, text):
            return "loan_disbursal", f"Loan pattern: {pat}"

    # ── 5. Self-transfer: own-name detection ─────────────────────────────
    if account_name and len(account_name) > 4:
        name_parts = [p for p in account_name.lower().split() if len(p) > 3]
        if name_parts:
            verb_m = re.search(
                r"\b(transfer from|transferred from|trf from|trf frm|"
                r"tfr from|transfer frm|trfr from)\b",
                text,
            )
            if verb_m:
                raw_window = text[verb_m.end(): verb_m.end() + 70]
                sender_window = re.split(r"\s+to\s+|\s*\|\s*", raw_window)[0]
                matched = sum(1 for p in name_parts if p in sender_window)
                if matched >= 2:
                    return "self_transfer", "Own-name sender detected"
            if "myself" in text or "| myself" in text or "/myself" in text:
                return "self_transfer", "Explicit self-transfer (myself)"

    # ── 5b. Contributions — ALL excluded regardless of sender ─────────────
    # Contributions (ajo, esusu, cooperative, group savings) are not
    # verifiable business income whether they come from the account holder
    # or from third parties. Classify as non_business so they are excluded.
    if re.search(r"\bcontribution\b|\bcontrib\b|\bajo\b|\besusu\b|\bcooperative\b", text):
        return "non_business", "Contribution/group-savings credit"

    # ── 6. Gambling / betting ─────────────────────────────────────────────
    betting_kw = [
        # ── Major Nigerian platforms ──────────────────────────────────────
        "sportybet", "sporty bet", "sporty|", "sporty internet",
        "bet9ja", "bet 9ja", "betnaija", "bet naija",
        "1xbet", "1x bet", "1x-bet",
        "betking", "bet king",
        "betway",
        "merrybet", "merry bet",
        "nairabet", "naira bet",
        "baba ijebu", "baba-ijebu",
        "naijabet",
        "lotto ", "lottomania", "premier lotto", "golden chance lotto",
        "supabets", "supa bets",
        "cloudbet",
        "msport", "m-sport",
        "bangbet", "bang bet",
        "parimatch", "pari match",
        "betboro",
        "betwinner", "bet winner",
        "22bet", "22 bet",
        "melbet", "mel bet",
        "betfair",
        "bet365",
        "frapapa",
        "linebet", "line bet",
        "xbet",
        "accessbet", "access bet",
        "wazobet", "wazobia bet",
        "betland",
        "surebet247", "sure bet",
        "betpawa", "bet pawa",
        "10bet", "ten bet",
        "nosporbet",
        "helabet", "hela bet",
        "premierbet", "premier bet",
        # ── Generic gambling signals ──────────────────────────────────────
        "casino ", "jackpot ", "betting winnings", "bet winnings",
        "winnings payout", "betting credit", "bet credit",
        "slot winnings", "poker winnings",
    ]
    if any(k in text for k in betting_kw):
        return "non_business", "Gambling/betting keyword"

    # ── 8. Non-business inflows ───────────────────────────────────────────
    non_biz_exact = [
        "salary", "salaries",
        "allowance",
        " support ",
        "monthly stipend", "stipend",
        "upkeep",
    ]
    if any(k in text for k in non_biz_exact):
        return "non_business", "Non-business keyword"

    return "real_credit", "Usable credit"


def add_credit(buckets: dict, ym: str, amount: float,
               narration: str, account_name: str = "") -> None:
    if not ym or not amount or amount <= 0:
        return
    if ym not in buckets:
        buckets[ym] = _empty_bucket()
    buckets[ym]["gross"]  += amount
    buckets[ym]["count"]  += 1
    cat, _ = classify_credit(narration, account_name)
    buckets[ym][cat] = buckets[ym].get(cat, 0.0) + amount
    # Append to global transaction log for search feature
    _txn_log.append({"ym": ym, "narration": narration, "amount": amount, "category": cat})


# ── Debit transaction log (visibility only — not used in decisioning) ─────────
_debit_log: list[dict] = []


def classify_debit(narration: str) -> tuple[str, str]:
    """Classify a debit transaction into a flagged category.

    Returns (category, label) — for display only, not used in credit decisioning.

    Categories
    ----------
    loan_repayment   Repayment of an existing loan / credit facility
    credit_card      Credit card payment
    rent             Rent or property-related payment
    utility          NEPA/electricity, water, cable TV, internet
    airtime_data     Airtime top-up or data purchase
    bank_charge      Bank fees, stamp duty, COT, VAT on transactions
    cash_withdrawal  ATM or counter cash withdrawal
    pos_purchase     POS / card purchase / online payment
    transfer_out     Regular outward transfer (default catch-all)
    """
    text = narration.lower()

    # ── Priority guard: "Charges -" prefix means bank transfer fee ──────────
    # Jaiz (and some other banks) prepend "Charges - " to the original transfer
    # narration when posting the ₦54 (or similar) inter-bank transfer fee.
    # The rest of the narration may mention "Loan Repayment" or other keywords
    # from the underlying transfer, but this row is ALWAYS a bank charge.
    if text.startswith("charges -") or text.startswith("charge -"):
        return "bank_charge", "⚪ Bank Charge / Fee"

    # Loan repayments — must contain explicit loan/repayment language
    # Use precise multi-word phrases or standalone words (with word boundaries
    # where needed) to avoid false positives like "ZEN" matching Zenith Bank.
    loan_kw = [
        "loan repay", "loan payment", "loan deduction", "loan instalment",
        "loan installment", "loan/repay", "loanrepay",
        " loan ",       # standalone word "loan" surrounded by spaces
        "/loan",        # e.g. "ref/loan123"
        "loan-",        # e.g. "loan-id"
        " laon ",       # common typo for loan (e.g. "isiaka laon")
        "equated monthly", "emi payment",
        "credit repay", "facility repay", "overdraft repay",
        "repayment",    # explicit word
        "renmoney repay", "fairmoney repay", "carbon repay",
        "migo repay", "branch repay", "aella repay", "lidya repay",
        "palmcredit repay", "creditwave", "page financials",
        "quick credit repay", "lendigo", "lendha repay",
        "flexpay repay", "creditcorp", "lapo repay",
    ]
    # Exclude "repayment for material/goods/supply" — these are vendor payments
    _loan_exclusions = [
        "repayment for material", "repayment for good", "repayment for supply",
        "repayment for item", "repayment for product",
    ]
    if any(k in text for k in loan_kw) and not any(e in text for e in _loan_exclusions):
        return "loan_repayment", "🔴 Loan Repayment"

    # Credit card payments
    cc_kw = ["credit card", "card payment", "creditcard", "card repay"]
    if any(k in text for k in cc_kw):
        return "credit_card", "🔴 Credit Card Payment"

    # Rent / property
    # NOTE: "service charge" is intentionally excluded here — in Nigerian
    # banking narrations "service charge" almost always means a vendor payment
    # for maintenance/repair work (e.g. "Generator Service Charge",
    # "Materials & Service Charge") which should NOT be flagged as rent.
    # True rent/property payments use "rent", "landlord", "estate" etc.
    rent_kw = ["rent", "landlord", "property", "estate agency", "caution fee",
               "agency fee", "ground rent"]
    if any(k in text for k in rent_kw):
        return "rent", "🟠 Rent / Property"

    # Utilities
    util_kw = ["nepa", "phcn", "ekedc", "ikedc", "phedc", "aedc", "kedco",
               "jed", "bedc", "electricity", "dstv", "gotv", "startimes",
               "multichoice", "lawma", "water board", "waterboard",
               "internet", "mtn", "airtel", "glo ", "9mobile", "spectranet",
               "swift network", "smile ", "ipnx"]
    if any(k in text for k in util_kw):
        return "utility", "🟡 Utility / Bill"

    # Airtime / data
    airtime_kw = ["airtime", "vtu", "data recharge", "data purchase",
                  "recharge", "topup", "top-up", "top up"]
    if any(k in text for k in airtime_kw):
        return "airtime_data", "⚪ Airtime / Data"

    # Bank charges
    charge_kw = ["bank charge", "stamp duty", "cot ", "vat ", "sms alert",
                 "maintenance fee", "card maintenance", "annual fee",
                 "account maintenance", "service fee", "commission",
                 "excise duty", "e-banking fee", "management fee"]
    if any(k in text for k in charge_kw):
        return "bank_charge", "⚪ Bank Charge / Fee"

    # Cash withdrawals
    cash_kw = ["atm ", "cash withdrawal", "counter withdrawal", "teller",
               "cash out", "cashout", "withdrawal"]
    if any(k in text for k in cash_kw):
        return "cash_withdrawal", "🟡 Cash Withdrawal"

    # POS / card purchase
    pos_kw = ["pos ", "purchase", "payment to", "pay to", "online payment",
              "web payment", "card transaction", "ussd purchase",
              "merchant", "supermarket", "restaurant", "hotel", "petrol",
              "fuel", "filling station"]
    if any(k in text for k in pos_kw):
        return "pos_purchase", "⚪ POS / Purchase"

    # Default: outward transfer
    return "transfer_out", "⚪ Transfer Out"


def add_debit(ym: str, narration: str, amount: float, date: str = "") -> None:
    """Log a debit transaction for UI display. Never affects credit decisioning."""
    if not ym or not amount or amount <= 0:
        return
    cat, label = classify_debit(narration)
    _debit_log.append({
        "ym":        ym,
        "date":      date,
        "narration": narration,
        "amount":    amount,
        "category":  cat,
        "label":     label,
    })


# ════════════════════════════════════════════════════════════════════════════
# PDF TEXT EXTRACTION
# ════════════════════════════════════════════════════════════════════════════
def extract_pdf_text(pdf_bytes: bytes, password: str = "") -> str:
    """Extract all text from a PDF using PyPDF2, processing pages in chunks.

    Processing in batches of CHUNK_SIZE pages (instead of all at once) lets
    Python's GC reclaim each batch's decompressed content streams before the
    next batch loads — critical for large PDFs (100+ pages) on memory-limited
    hosts (Streamlit Cloud free tier = 1 GB).

    PyPDF2 PdfReader reads the cross-reference table upfront but decompresses
    page content on demand via extract_text(). Explicit gc.collect() between
    chunks recovers the decompressed data from prior batches before the next
    batch inflates memory again.
    """
    CHUNK_SIZE = 50   # pages per batch — tune down to 30 if OOM persists
    buf = BytesIO(pdf_bytes)
    reader = PdfReader(buf)
    if reader.is_encrypted:
        if reader.decrypt(password or "") == 0:
            raise ValueError("Incorrect or missing PDF password.")

    pages  = reader.pages
    n      = len(pages)
    parts: list[str] = []

    for start in range(0, n, CHUNK_SIZE):
        end        = min(start + CHUNK_SIZE, n)
        chunk_text = "\n".join(pages[i].extract_text() or "" for i in range(start, end))
        parts.append(chunk_text)
        gc.collect()   # free decompressed page streams from this batch

    text = "\n".join(parts)
    del parts, pages, reader, buf
    gc.collect()
    return text


def extract_pdf_text_pdfplumber(pdf_bytes: bytes, password: str = "") -> str:
    """Extract text using pdfplumber — captures pages that PyPDF2 misses.
    Falls back to PyPDF2 if pdfplumber is not installed.
    """
    try:
        import pdfplumber
        buf = BytesIO(pdf_bytes)
        with pdfplumber.open(buf, password=password or "") as pdf:
            text = "\n".join(page.extract_text() or "" for page in pdf.pages)
        del buf
        gc.collect()
        return text
    except Exception:
        return extract_pdf_text(pdf_bytes, password)


def extract_pdf_text_layout(pdf_bytes: bytes) -> str:
    """
    Extract PDF text preserving spatial column layout using pdftotext -layout.
    Required for PDFs where amounts are in fixed-width columns (e.g. OPay v2).
    Returns empty string if pdftotext is unavailable.

    Uses a temp file because pdftotext v4.00 (Glyph & Cog) does not support
    stdin piping ('-' as input); only the open-source Poppler build does.
    """
    import tempfile, os
    tmp_path = None
    try:
        # Write to temp file — compatible with both Poppler and Glyph & Cog builds
        with tempfile.NamedTemporaryFile(suffix=".pdf", delete=False) as tmp:
            tmp.write(pdf_bytes)
            tmp_path = tmp.name
        result = subprocess.run(
            ["pdftotext", "-layout", tmp_path, "-"],
            capture_output=True,
            timeout=120,   # large files need more time
        )
        if result.returncode != 0:
            return ""
        return result.stdout.decode("utf-8", errors="replace")
    except Exception:
        return ""
    finally:
        if tmp_path and os.path.exists(tmp_path):
            try:
                os.unlink(tmp_path)
            except Exception:
                pass


# ════════════════════════════════════════════════════════════════════════════
# BANK DETECTION
# ════════════════════════════════════════════════════════════════════════════
def detect_bank(text: str) -> str:
    t = text.lower()

    # Header window = everything BEFORE the transaction table begins. The bank
    # name / account-summary area sits above the transactions, so cutting here
    # prevents narration text (e.g. "NIP/.../UBA /", "To Zenith Bank Plc",
    # "Transfer to STERLING BANK") from triggering a false bank match. Some
    # statements have a dense summary where narrations start well within the
    # first 1200 chars, which previously caused misdetection.
    _markers = [
        "transaction summary", "trandate valuedate", "tran date value date",
        "tran. date value", "tran.date value", "date narration",
        "transaction details", "transaction detail",
    ]
    _cuts = [t.find(m) for m in _markers]
    _cuts = [c for c in _cuts if c != -1]
    t_hdr = t[:min(_cuts)] if _cuts else t[:1200]

    # ── Renmoney MFB: MUST be before Jaiz ────────────────────────────────
    # Renmoney's column header "DateNarration Debit Credit Balance" contains
    # "datenarration", which would otherwise fire the Jaiz rule below.
    # "Product Name: Renmoney Account" appears in line 2 of every Renmoney
    # statement header — unique and always within t_hdr (first 1200 chars).
    if "product name: renmoney" in t_hdr or "renmoney account" in t_hdr:
        return "Renmoney"

    # ── Jaiz Bank: checked early — narrations mention "OPAY DIGITAL" ─────
    # PyPDF2 merges the adjacent column headers "DATE" + "NARRATION" into the
    # single token "DATENARRATION", which is unique to Jaiz's layout and never
    # appears in narration text of any other bank.  Must fire before OPay
    # because Jaiz narrations include "OPAY DIGITAL SERVICES LIMITED" credits.
    if "datenarration" in t:
        return "Jaiz"

    # ── Parallex Bank ────────────────────────────────────────────────────
    # "parallex savings" appears in the statement header (account type label).
    # "parallexbank" is from the footer email customercare@parallexbank.com.
    # Both are unique to Parallex and will not appear in the header area of
    # any other bank's statement.
    if "parallex savings" in t_hdr or "parallexbank" in t or "parallex bank" in t_hdr:
        return "Parallex"

    # ── OPay v2: "Account Statement" with Trans. Time column ─────────────
    # Checked FIRST: OPay narrations often mention "FairMoney MFB".
    if ("trans. time" in t and
            ("opay digital" in t or "wallet account" in t or "9payment service" in t)):
        return "OPay_v2"

    # ── Moniepoint Business v2: 19-column Settlement Debit/Credit layout ────
    # MUST fire BEFORE OPay checks: "OPay Digital" appears in the Source
    # Institution column of transactions, causing the generic OPay rule to
    # misfire.  This format is uniquely identified by "settlement" in the
    # column header area (first 1200 chars) alongside "business name".
    # The ISO timestamps are split across 4 lines by PyPDF2, so the
    # standard ISO-timestamp regex does NOT match — handled inside the parser.
    if "business name" in t_hdr and "settlement" in t_hdr:
        return "Moniepoint_Business_v2"

    # ── Carbon MFB: getcarbon.co domain or "Carbon MFB" in statement header ─
    # MUST fire before FairMoney and OPay checks: Carbon narrations can
    # contain "FairMoney MFB" (e.g. incoming transfers) or "OPay" merchant
    # names, both of which would cause earlier rules to mismatch.
    if "getcarbon" in t_hdr or "carbon mfb" in t_hdr:
        return "Carbon"

    # ── Moniepoint Business: ISO-timestamp layout ─────────────────────────
    # Checked BEFORE OPay_Business (same ISO layout).
    if ("moniepoint mfb" in t or "moniepoint microfinance" in t) and \
            re.search(r"\d{4}-\d{2}-\d{2}t\d{2}:", t):
        return "Moniepoint_Business"

    # ── OPay Business: ISO-timestamp "Account Statement" ─────────────────
    if "business name" in t and re.search(r"\d{4}-\d{2}-\d{2}t\d{2}:", t) and \
            ("2mpt" in t or "ap_trsf" in t or "business_credit" in t):
        return "OPay_Business"

    # ── FairMoney ─────────────────────────────────────────────────────────
    if ("fairmoney mfb" in t or "fairmoney microfinance" in t) and "mybankstatement" not in t:
        return "FairMoney"

    # ── OPay legacy ───────────────────────────────────────────────────────
    # Guard with mybankstatement: Access Bank narrations can contain
    # "OPAY DIGITAL SERVICES" as a POS merchant name — must not mismatch.
    if ("opay digital" in t or "wallet account" in t or "9payment service" in t) \
            and "mybankstatement" not in t:
        return "OPay"

    # ── PalmPay Business format ───────────────────────────────────────────
    # PyPDF2 splits "BUSINESS_ACCOUNT" across two table cells producing the
    # distinctive substring "BUSINESS_ACCOU" (followed by "NT" in next cell).
    # Combined with ISO-format dates (YYYY-MM-DD HH:MM:SS) this uniquely
    # identifies PalmPay Business Account statements.
    # Columns: Date+Time | Description | Debit | Credit | Balance | Extra | TxnID
    if ("business_accou" in t and
            re.search(r'\b20\d\d-\d\d-\d\d \d\d:\d\d:\d\d\b', t)):
        return "PalmPay_Business"

    # ── PalmPay NEW format ────────────────────────────────────────────────
    # Identified by its unique column headers: "Transaction Detail",
    # "Transaction ID", and "Money In (NGN)". The bank name "PalmPay" may
    # not appear in the PDF header text, so use format-based detection.
    if ("transaction detail" in t and "transaction id" in t and
            ("money in (ngn)" in t or "total money in" in t)):
        return "PalmPay_New"

    # ── Access Bank Oracle format ─────────────────────────────────────────
    # Unique column set: "Posted Date" (not "Date Posted") + "Value Date" +
    # "Debit (NGN)" + "Credit (NGN)". Must be checked BEFORE the generic
    # "access bank" name check because some Oracle statements omit the bank
    # name from the extracted text entirely.
    if ("posted date" in t and "value date" in t and
            "debit (ngn)" in t and "credit (ngn)" in t):
        return "Access_Oracle"

    # ── Zenith Corporate: MUST be before ALL named-bank checks ───────────
    # Zenith Corporate narrations contain "NIP/GTB/...", "NIP/FCMB/..." etc.
    # "date posted" + "value date" pair is unique to Zenith BOP format.
    if "date posted" in t and "value date" in t and \
            ("zenith" in t or "zenithdirect" in t):
        return "Zenith_Corporate"

    # ── Kuda: MUST be before GTBank ───────────────────────────────────────
    # Kuda statements contain "Gtbank Plc" in transaction narrations.
    if "kuda mf bank" in t or "kudabank" in t or "kuda technologies" in t:
        return "Kuda"

    # ── Providus Bank: MUST be before named-bank checks ──────────────────
    # Providus narrations reference other banks ("FROM STANBIC/...", "FROM
    # ACCESS/..."). The Stanbic check below uses t_hdr and "STANBIC IBTC"
    # appears inside the first 1200 chars of a Providus statement (transaction
    # narration), which would cause a mismatch. Detected by the unique
    # "CUST. NAME" + "TXN DATE" header combination.
    if "cust. name" in t_hdr and "txn date" in t_hdr:
        return "Providus"

    # ── mybankStatement-format banks ──────────────────────────────────────
    # IMPORTANT: a bank's own name only legitimately appears in the statement
    # HEADER (identity block). It must NEVER be detected from transaction
    # narrations ("NIP/FCMB/...", "To Zenith Bank Plc", "/UBA /"), which would
    # mislabel the statement. Therefore every bank-name check below matches
    # against t_hdr (the strict header, cut before the transaction table),
    # not the full text. Only structural markers (mybankstatement) use full t.
    _fidelity_url = "fidelitybank.ng"

    # Fidelity Direct: native Fidelity statement (fidelitybank.ng URL in the
    # page header, DD-Mon-YY dates — NOT a mybankStatement PDF).
    if _fidelity_url in t_hdr and "mybankstatement" not in t:
        return "Fidelity_Direct"

    if "guaranty trust" in t_hdr or "gtco" in t_hdr or \
            "gtbank" in t_hdr or "gt bank" in t_hdr:
        return "GTBank"
    if "access bank" in t_hdr:
        return "Access"
    if "first bank" in t_hdr or "firstbank" in t_hdr:
        return "FirstBank"
    if "united bank for africa" in t_hdr or re.search(r"\buba\b", t_hdr):
        return "UBA"
    if "fidelity bank" in t_hdr or "fidelity" in t_hdr:
        return "Fidelity"
    if "union bank" in t_hdr:
        return "Union"
    if "stanbic" in t_hdr:
        return "Stanbic"
    if "first city monument bank" in t_hdr or re.search(r"\bfcmb\b", t_hdr):
        return "FCMB"
    if "wema bank" in t_hdr or "wema" in t_hdr:
        return "Wema"
    if "sterling bank" in t_hdr or "sterling" in t_hdr:
        return "Sterling"
    # ── Zenith e-Statement (new 2025+ format) ────────────────────────────
    # Column header: "DATE DESCRIPTION DEBIT CREDIT VALUE DATE BALANCE"
    # Distinct from old Zenith ("Tran Date Value Date Narration").
    # Must fire BEFORE the generic "zenith" t_hdr check below.
    if "zenith" in t_hdr and "date description debit" in t:
        return "Zenith_New"
    if "zenith" in t_hdr:
        return "Zenith"

    # ── Ecobank: bank name is a logo IMAGE — never appears as text. ───────────
    # Detect via two independent signals that together are unambiguous:
    #   1. mybankStatement portal format (required)
    #   2. "ecobank" anywhere in text (e.g. NIP narrations reference the bank)
    #      OR account number starts with 2452 (Ecobank Nigeria NUBAN prefix)
    # No separate parser is needed — parse_gtbank handles all mybankStatement
    # banks identically regardless of bank name.
    if "mybankstatement" in t and (
        "ecobank" in t
        or re.search(r'\b2452\d{6}\b', t)
    ):
        return "Ecobank"

    # "tran date value date narration" is the native Zenith column header.
    # Guard with "mybankstatement not in t": PyPDF2 can tokenise the
    # mybankStatement portal column header "TranDate ValueDate Narration" with
    # extra spaces, producing the same lowercase string — that must NOT fire
    # the Zenith branch.
    # Plain "mybankstatement" without a named bank → unknown mybankStatement
    # bank; route to GTBank so parse_gtbank handles it (same parser for all
    # mybankStatement-portal banks).
    if "tran date value date narration" in t and "mybankstatement" not in t:
        return "Zenith"
    # Unidentified mybankStatement-portal bank (name is a logo image, not text).
    # Route through the shared parser but label neutrally so it is never
    # falsely shown as GTBank — the officer can set the correct bank via the
    # manual override in the UI.
    if "mybankstatement" in t:
        return "mybankStatement"
    if "moniepoint mfb" in t or "moniepoint microfinance" in t:
        return "Moniepoint"
    if "palmpay" in t:
        return "PalmPay"

    # ── Broad fallbacks (safe because more-specific checks fired first) ───
    if "access bank" in t:
        return "Access"
    if "united bank for africa" in t:
        return "UBA"
    if "first bank" in t or "firstbank" in t:
        return "FirstBank"
    if "zenith" in t:
        return "Zenith"
    if "sterling bank" in t:
        return "Sterling"
    if "fcmb" in t or "first city monument bank" in t:
        return "FCMB"
    if "wema bank" in t:
        return "Wema"
    if "fidelitybank.ng" in t and "mybankstatement" not in t:
        return "Fidelity_Direct"
    if "fidelity bank" in t:
        return "Fidelity"
    if "stanbic ibtc" in t or "stanbic" in t:
        return "Stanbic"
    if "union bank" in t:
        return "Union"
    return "Unknown"


# ════════════════════════════════════════════════════════════════════════════
# BANK PARSERS
# ════════════════════════════════════════════════════════════════════════════
def _extract_account_name(full_text: str) -> str:
    for pat in [
        r"Account\s*Name\s+([A-Z][A-Z .'\-]{4,})",   # allows hyphens, dots
        r"ACCOUNT\s*NAME[:\s]+([A-Z][A-Z .'\-]{4,})",
        r"AccountName\s+([A-Z][A-Z .'\-]{4,})",       # pdfplumber (no space)
        r"Name\s+([A-Z][A-Z .'\-]{4,})",
    ]:
        m = re.search(pat, full_text, re.I)
        if m:
            return m.group(1).strip().rstrip("-").strip()
    return ""


def parse_opay(full_text: str) -> tuple[dict, str]:
    buckets: dict = {}
    account_name = _extract_account_name(full_text)
    for line in full_text.splitlines():
        line = line.strip()
        m = OPAY_ROW.match(line)
        if not m:
            continue
        day, mon, year = m.group(1), m.group(2), m.group(3)
        ym = f"{year}-{MONTH_NUM[mon.lower()]}"
        rest = line[m.end():].strip()
        rest = re.sub(
            r"^\d{2}\s+(?:Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)\s+20\d{2}\s*",
            "", rest, flags=re.I,
        )
        money = list(MONEY_RE.finditer(rest))
        if len(money) < 2:
            continue
        amount_m   = money[-2]
        amount     = float(amount_m.group().replace(",", ""))
        before_amt = rest[:amount_m.start()]
        if not re.search(r"--\s*$", before_amt.rstrip()):
            continue
        narration = re.sub(r"--\s*$", "", before_amt).strip()
        narration = re.sub(r"\s+Mobile\S*$", "", narration).strip()
        add_credit(buckets, ym, amount, narration, account_name)
    return buckets, account_name


def _parse_opay_v2_pypdf2(full_text: str, account_name: str = "") -> tuple[dict, list]:
    """
    PyPDF2-based fallback for OPay v2 statements.

    Used when pdftotext is unavailable or produces misaligned output
    (e.g. pdftotext v4.00 Glyph & Cog on large PDFs).

    PyPDF2 output format:
        DD Mon YYYY HH:MM:SS  DD Mon YYYY  <description>  <debit|-->  <credit|-->  <balance>  <Channel>  <ref>
    Amounts may be glued to description text (no space). Multi-line
    descriptions accumulate until the next date-time anchor line.
    """
    buckets: dict = {}

    DT_RE = re.compile(
        r"^(\d{2})\s+(Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)\s+(20\d{2})\s+\d{2}:\d{2}:\d{2}",
        re.I,
    )
    # Three amounts + channel.  Use strict comma-formatted pattern only —
    # bare \d+ would swallow transaction reference numbers that PyPDF2
    # sometimes concatenates directly onto the next line's amount token.
    AMT_RE = re.compile(
        r"((?:\d{1,3}(?:,\d{3})*)\.\d{2}|--)\s+"
        r"((?:\d{1,3}(?:,\d{3})*)\.\d{2}|--)\s+"
        r"((?:\d{1,3}(?:,\d{3})*)\.\d{2})\s*"
        r"(?:Mobile|POS|Web|USSD|ATM|Agent)",
        re.I,
    )
    # Strip the two leading date stamps from a description block
    STRIP_DTS = re.compile(
        r"^\d{2}\s+\w{3}\s+20\d{2}\s+\d{2}:\d{2}:\d{2}\s*"
        r"(?:\d{2}\s+\w{3}\s+20\d{2}\s*)?",
        re.I,
    )

    in_wallet = False
    section_count = 0
    current_ym: str | None = None
    pending: list[str] = []
    debit_rows: list[tuple] = []   # (ym, desc, amount) — flushed to _debit_log by caller
    # Ref-number filter: bare 8+ digit strings are transaction references, not desc
    REF_RE = re.compile(r"^\d{8,}$")
    # Tokens to strip from tail: channel keywords, standalone digits, date parts
    _MONTH_NAMES = {"Jan","Feb","Mar","Apr","May","Jun",
                    "Jul","Aug","Sep","Oct","Nov","Dec"}
    _NOISE_WORDS  = {"Mobile","POS","Web","USSD","ATM","Agent","LANG","NG"}

    def _is_tail_noise(tok: str) -> bool:
        return (REF_RE.match(tok)
                or re.match(r"^\d{1,4}$", tok)            # short digit fragments
                or re.match(r"^\d{2}:\d{2}:\d{2}$", tok)  # time token HH:MM:SS
                or re.match(r"^20\d{2}$", tok)             # year token
                or tok.capitalize() in _MONTH_NAMES        # month name
                or tok in _NOISE_WORDS)

    DATE_NOISE = re.compile(
        r"\d{2}\s+(?:Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)\s+20\d{2}"
        r"(?:\s+\d{2}:\d{2}:\d{2})?",
        re.I,
    )

    def _flush() -> list[str]:
        """Process current pending block. Returns clean tail tokens (post-amounts
        text stripped of date/ref noise) as pre-description for the next txn."""
        nonlocal current_ym, pending
        if not pending or current_ym is None:
            return []
        block = " ".join(pending)
        m = AMT_RE.search(block)
        tail: list[str] = []
        if m:
            raw_desc = block[: m.start()]
            # Strip leading date stamps and any stray date fragments
            desc = DATE_NOISE.sub(" ", STRIP_DTS.sub("", raw_desc)).strip()
            desc = re.sub(r"\s{2,}", " ", desc).strip()
            debit_tok, credit_tok = m.group(1), m.group(2)
            if credit_tok != "--":
                credit = float(credit_tok.replace(",", ""))
                if credit > 0:
                    add_credit(buckets, current_ym, credit, desc, account_name)
            if debit_tok != "--":
                debit = float(debit_tok.replace(",", ""))
                if debit > 0:
                    debit_rows.append((current_ym, desc, debit))
            # Tail = post-amount tokens, scrubbed of refs, digits, and date parts.
            post = block[m.end():].strip()
            tail = [t for t in post.split() if not _is_tail_noise(t)]
        pending.clear()
        return tail

    for line in full_text.splitlines():
        if re.search(r"Trans\.\s*Time", line, re.I):
            section_count += 1
            if section_count > 1:
                _flush()
                break
            in_wallet = True
            pending.clear()
            current_ym = None
            continue

        if re.search(r"Savings Account", line, re.I) and section_count > 0:
            _flush()
            break

        if not in_wallet:
            continue

        dm = DT_RE.match(line.strip())
        if dm:
            # tail = post-amount text from the previous block = pre-description
            # for THIS transaction (lines that visually precede the date anchor).
            tail = _flush()
            current_ym = f"{dm.group(3)}-{MONTH_NUM[dm.group(2).lower()[:3]]}"
            pending = tail + [line.strip()]
            pre_desc = []
        elif current_ym is not None:
            s = line.strip()
            if s:
                pending.append(s)

    _flush()
    return buckets, debit_rows


def parse_opay_v2(pdf_bytes: bytes, full_text: str = "") -> tuple[dict, str]:
    """
    Parser for the OPay "Account Statement" format (2025+).

    Layout characteristics:
    - Header: Trans. Time | Value Date | Description | Debit(₦) | Credit(₦) | Balance After(₦) | Channel
    - Each transaction occupies one anchor line (date/time at the start) plus
      optional description fragment lines before and after it.
    - The debit / credit / balance / channel columns occupy fixed character
      positions starting around column 88 of the pdftotext -layout output.
    - Null values in numeric columns are rendered as "--".
    - Channel keyword ("Mobile", "POS", "Web", etc.) always follows the three
      numeric columns, making AMOUNTS_PAT reliable for anchor detection.
    - The PDF contains TWO account sections:
        1. Wallet Account  (the main transactional account)
        2. Savings Account (the OWealth sub-account — interest + round-trips)
      We parse only the first section, stopping when the second "Trans. Time"
      header (or "Savings Account" label) is encountered.

    Requires pdftotext (poppler-utils) installed on the host.
    Falls back to PyPDF2-based parsing when pdftotext is unavailable or
    produces misaligned output (e.g. pdftotext v4.00 Glyph & Cog).
    """
    # ── Layout-preserved text via pdftotext -layout ───────────────────────
    layout_text = extract_pdf_text_layout(pdf_bytes)

    # ── Patterns ──────────────────────────────────────────────────────────
    # Anchor line: starts with 0–4 spaces then DD Mon YYYY HH:MM:SS
    DATE_LINE = re.compile(
        r"^\s{0,4}(\d{2})\s+(Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)"
        r"\s+(20\d{2})\s+\d{2}:\d{2}:\d{2}",
        re.I,
    )
    # Three amount tokens (-- or N,NNN.NN) followed by a channel keyword.
    # This anchors on the channel word so it works regardless of how much
    # the description overflows into the amounts region.
    AMOUNTS_PAT = re.compile(
        r"([\d,]+\.\d{2}|--)\s+([\d,]+\.\d{2}|--)\s+([\d,]+\.\d{2})\s+"
        r"(?:Mobile|POS|Web|USSD|ATM|Agent)",
        re.I,
    )
    # Description continuation: deeply indented (36+ spaces), non-blank.
    DESC_CONT  = re.compile(r"^\s{36,}(\S.*)$")
    # Strip trailing transaction-reference digit runs from continuation lines.
    TRAIL_REF  = re.compile(r"\s+\d{10,}\s*$")
    # Section boundary markers.
    TRANS_TIME = re.compile(r"Trans\.\s*Time", re.I)
    SECTION_END = re.compile(r"Savings Account", re.I)

    # ── Account name ──────────────────────────────────────────────────────
    # OPay v2 pdftotext layout: "Account Name" and "Account Statement" land
    # on the SAME line (they're columns), the real name is the NEXT line.
    # PyPDF2 text has "Account Name\nCOMFORT BENARD WILLSON\n".
    _BOILERPLATE = {"account statement", "account number", "account name",
                    "generated on", "wallet account", "savings account"}
    account_name = ""

    # 1. Try layout text — scan for "Account Name" line, name is on next line
    if layout_text:
        layout_lines = layout_text.splitlines()
        for i, line in enumerate(layout_lines):
            if re.search(r"\bAccount\s+Name\b", line, re.I):
                for j in range(i + 1, min(i + 6, len(layout_lines))):
                    val = layout_lines[j].strip()[:60]
                    # Must be all-caps letters+spaces, not a boilerplate phrase
                    if (val and re.match(r"^[A-Z][A-Z ]{3,}$", val)
                            and val.lower() not in _BOILERPLATE):
                        account_name = val
                        break
                if account_name:
                    break

    # 2. PyPDF2 text — "Account Name\nNAME" pattern
    if not account_name and full_text:
        nm = re.search(r"Account Name\s*\n([A-Z][A-Z ]{3,})", full_text)
        if nm:
            candidate = nm.group(1).strip().split("\n")[0].strip()
            if candidate.lower() not in _BOILERPLATE:
                account_name = candidate

    # 3. Generic fallback (both texts)
    if not account_name:
        for txt in ([layout_text] if layout_text else []) + ([full_text] if full_text else []):
            candidate = _extract_account_name(txt)
            if candidate and candidate.lower() not in _BOILERPLATE:
                account_name = candidate
                break

    # ── Parse main Wallet section only (layout-based) ─────────────────────
    buckets: dict = {}
    layout_debit_rows: list[tuple] = []   # (ym, desc, amount)

    if layout_text:
        pre_desc: list[str] = []
        state = "pre"
        section_count = 0
        in_tx = False

        for line in layout_text.splitlines():
            if TRANS_TIME.search(line):
                section_count += 1
                if section_count > 1:
                    break
                in_tx = True
                pre_desc = []
                state = "pre"
                continue

            if SECTION_END.search(line):
                break

            if not in_tx:
                continue

            m = DATE_LINE.match(line)
            if m:
                mon  = m.group(2).lower()[:3]
                year = m.group(3)
                ym   = f"{year}-{MONTH_NUM[mon]}"

                amt_m = AMOUNTS_PAT.search(line)
                if amt_m:
                    debit_tok  = amt_m.group(1)
                    credit_tok = amt_m.group(2)
                    inline_desc = line[38:amt_m.start()].strip()
                    parts = pre_desc[:]
                    if inline_desc:
                        parts.append(inline_desc)
                    full_desc = " ".join(parts).strip()
                    if credit_tok != "--":
                        credit = float(credit_tok.replace(",", ""))
                        if credit > 0:
                            add_credit(buckets, ym, credit, full_desc, account_name)
                    if debit_tok != "--":
                        debit = float(debit_tok.replace(",", ""))
                        if debit > 0:
                            layout_debit_rows.append((ym, full_desc, debit))

                pre_desc = []
                state = "post"
                continue

            stripped = line.strip()
            if not stripped:
                pre_desc = []
                state = "pre"
                continue

            dm = DESC_CONT.match(line)
            if dm:
                text = TRAIL_REF.sub("", dm.group(1)).strip()
                if text and not re.match(r"^[\d\s]+$", text):
                    if state == "pre":
                        pre_desc.append(text)

    # ── PyPDF2 fallback: always run and use whichever parser found more data ───
    # Text-based fallback: use pdfplumber text (cleaner than PyPDF2 — no
    # reference-number concatenation).  Falls back to full_text (PyPDF2) if
    # pdfplumber is unavailable.  Pick whichever parser found more credit gross.
    # IMPORTANT: debits are NOT added to _debit_log until the winner is decided,
    # to avoid double-counting when both parsers run.
    winning_debit_rows = layout_debit_rows
    plumber_text = extract_pdf_text_pdfplumber(pdf_bytes) if pdf_bytes else full_text
    fallback_text = plumber_text or full_text
    if fallback_text:
        pypdf2_buckets, pypdf2_debit_rows = _parse_opay_v2_pypdf2(fallback_text, account_name)
        layout_gross  = sum(b.get("gross", 0) for b in buckets.values())
        pypdf2_gross  = sum(b.get("gross", 0) for b in pypdf2_buckets.values())
        if pypdf2_gross > layout_gross:
            buckets = pypdf2_buckets
            winning_debit_rows = pypdf2_debit_rows

    for ym, desc, amount in winning_debit_rows:
        add_debit(ym, desc, amount)

    return buckets, account_name


def parse_opay_business(pdf_bytes: bytes) -> tuple[dict, str]:
    """
    Parser for the OPay Business "Account Statement" format.

    Layout characteristics (pdftotext -layout):
    - Header: Business Name, Account Number, Currency, Date range, Address
    - Column header line: Date | Narration | Reference | Debit | Credit | Balance
    - Each transaction spans THREE physical lines:
        Line A (DATE):  YYYY-MM-DDThh:   [narration_part_1 at col ~17+]
        Line B (MID):   [narration_part_2 at col ~17–80]  [reference col ~82]
                        [debit]  [credit]  [balance]   ← at fixed end of line
        Line C (TIME):  mm:ss   [narration_part_3 at col ~17+]

    PAGE-BREAK MERGE: When a transaction straddles a page boundary, pdftotext
    concatenates the DATE line and the MID line (with reference + amounts) into
    a single long line prefixed with \x0c. In this case the amounts appear at
    the END of the DATE line itself and no separate MID line exists.

    Both cases are handled by first checking MONEY3 against the date line, then
    falling back to searching the subsequent middle line.

    Narration assembly: narr_part_1 (from DATE line after hh:) +
                        narr_part_2 (from MID line, cols 17–80) +
                        narr_part_3 (from TIME line after mm:ss)
    Reference tokens (≥15 chars, e.g. AP_TRSF|...) are stripped from narration.

    Requires pdftotext -layout (poppler-utils). Returns ({}, "") if unavailable.
    """
    layout_text = extract_pdf_text_layout(pdf_bytes)
    if not layout_text:
        return {}, ""

    DATE_LINE = re.compile(r"^\x0c?(\d{4})-(\d{2})-\d{2}T\d{2}:", re.I)
    MONEY3    = re.compile(
        r"([\d,]+\.\d{2})\s+([\d,]+\.\d{2})\s+([\d,]+\.\d{2})\s*$"
    )
    TIME_RE   = re.compile(r"^\d{2}:\d{2}\s*(.*)")
    # Strip trailing reference token (pipe-delimited IDs, ≥15 non-space chars)
    REF_RE    = re.compile(r"\s+\S{15,}\S*\s*$")

    # Account name from "Business Name" header (business account)
    # or "Account Name" (personal fallback)
    account_name = ""
    biz_m = re.search(r"Business Name\s{2,}([A-Z][A-Z ]{3,})", layout_text)
    if biz_m:
        account_name = biz_m.group(1).strip()
    if not account_name:
        account_name = _extract_account_name(layout_text)

    lines   = layout_text.splitlines()
    buckets: dict = {}
    i = 0

    while i < len(lines):
        l = lines[i].lstrip("\x0c")
        dm = DATE_LINE.match(lines[i])   # match on raw (preserves \x0c for page-break detection)
        if not dm:
            i += 1
            continue

        year, month = dm.group(1), dm.group(2)
        ym = f"{year}-{month}"

        # ── PAGE-BREAK MERGED: amounts are on the date line itself ────────
        date_am = MONEY3.search(l)
        if date_am:
            debit  = float(date_am.group(1).replace(",", ""))
            credit = float(date_am.group(2).replace(",", ""))
            if credit > 0 and debit == 0:
                # Narration: text after hh: up to where the reference starts
                after_hh = l[l.index("T") + 4:date_am.start()].strip()
                narr = REF_RE.sub("", after_hh).strip()
                add_credit(buckets, ym, credit, narr, account_name)
            # Skip to next date anchor
            j = i + 1
            while j < len(lines) and not DATE_LINE.match(lines[j]):
                j += 1
            i = j
            continue

        # ── NORMAL: narration on date line, amounts on mid line ───────────
        narr1 = l[l.index("T") + 4:].strip() if "T" in l else ""

        # Collect non-blank lines until next date anchor
        other_lines: list[str] = []
        j = i + 1
        while j < len(lines) and not DATE_LINE.match(lines[j]):
            if lines[j].strip():
                other_lines.append(lines[j].rstrip())
            j += 1

        for mid in other_lines:
            am = MONEY3.search(mid)
            if not am:
                continue
            debit  = float(am.group(1).replace(",", ""))
            credit = float(am.group(2).replace(",", ""))
            if credit > 0 and debit == 0:
                # Narration from mid line: cols 16–80 (before the reference)
                narr2 = mid[16:80].strip() if len(mid) > 16 else ""
                # Narration continuation from time line: after mm:ss
                time_cont = ""
                for ol in other_lines:
                    tm = TIME_RE.match(ol.strip())
                    if tm and tm.group(1).strip():
                        time_cont = tm.group(1).strip()
                        break
                narr = " ".join(p for p in [narr1, narr2, time_cont] if p)
                add_credit(buckets, ym, credit, narr, account_name)
            break  # only the first amount line per block

        i = j

    return buckets, account_name


# ════════════════════════════════════════════════════════════════════════════
# KUDA BANK PARSER
# ════════════════════════════════════════════════════════════════════════════
_KUDA_DATE  = re.compile(r"^(\d{2}/\d{2}/\d{2})$")
_KUDA_TIME  = re.compile(r"^\d{2}:\d{2}:\d{2}$")
_KUDA_MONEY = re.compile(r"^₦([\d,]+\.\d{2})$")
_KUDA_OUTWARD = re.compile(
    r"\boutward\b|\bpaid\b|\bpayment\b|\bbills?\b|\bairtime\b|\bdata\b|"
    r"\belectricity\b|\bcable\b|\bwithdrawal\b|\batm\b|\bpos\s+charge\b|"
    r"\bbank\s+charge\b|\bfee\b|\bstamp\s+duty\b|\bdebit\b|"
    r"\bspend\s+and\s+save\b|\bspend\s*\+\s*save\b",
    re.I,
)


def parse_kuda(full_text: str) -> tuple[dict, str]:
    """
    Parser for Kuda MF Bank PDF statements.

    PyPDF2 extracts Kuda columns as individual newline-separated tokens.
    Each transaction block follows this sequence:
      DD/MM/YY       ← date anchor
      HH:MM:SS       ← time
      ₦Amount        ← transaction amount (first ₦ token)
      category_word  ← direction indicator (inward/outward/local/house etc.)
      to/from info
      description
      ₦NewBalance    ← running balance (second ₦ token)

    Direction: any category token matching _KUDA_OUTWARD → debit (skip).
    All others (inward transfer, local funds transfer, house, etc.) → credit.

    Page-boundary dedup: Kuda PDFs repeat the last transaction on a page at
    the top of the next page. Dedup key = (date, time, first_amount).
    """
    lines = [ln.strip() for ln in full_text.split("\n")]

    # Account name: near top of statement as an ALL-CAPS multi-word line.
    # PyPDF2 may render tab characters between name parts; normalise before matching.
    account_name = ""
    for line in lines[:25]:
        normalised = re.sub(r"\t", " ", line).strip()
        if re.match(r"^[A-Z][A-Z ]{4,}$", normalised) and len(normalised.split()) >= 2:
            account_name = normalised
            break

    buckets: dict = {}
    seen_keys: set = set()
    i = 0

    while i < len(lines):
        dm = _KUDA_DATE.match(lines[i])
        if not dm:
            i += 1
            continue

        date_str = dm.group(1)  # DD/MM/YY

        # Collect lines until next date anchor or page-header keyword
        window: list[str] = []
        j = i + 1
        while j < len(lines):
            wl = lines[j].strip()
            if _KUDA_DATE.match(wl):
                break
            if re.match(r"^(?:Page \d+|All\s+Statements|Kuda\s+MF|Statement)", wl, re.I):
                j += 1
                continue
            window.append(wl)
            j += 1

        # Partition window into ₦amounts, time, and text tokens
        time_found: str | None = None
        amounts: list[str] = []
        text_tokens: list[str] = []

        for wl in window:
            if not wl:
                continue
            mm = _KUDA_MONEY.match(wl)
            if mm:
                amounts.append(mm.group(1))
                continue
            if _KUDA_TIME.match(wl):
                time_found = wl
                continue
            text_tokens.append(wl)

        if not amounts:
            i = j
            continue

        # Skip page-boundary duplicates
        dedup_key = (date_str, time_found or "notime", amounts[0])
        if dedup_key in seen_keys:
            i = j
            continue
        seen_keys.add(dedup_key)

        # Direction: check ONLY the first few tokens (the category cell).
        # Stop collecting at account-name tokens (contain digits or "/NNNN" pattern)
        # so that description text doesn't pollute the direction check.
        cat_tokens: list[str] = []
        for tok in text_tokens[:5]:
            if re.search(r"\d{6,}|/\d{4,}", tok):
                break
            cat_tokens.append(re.sub(r"\t", " ", tok).strip())
        cat_str = " ".join(cat_tokens).strip().lower()

        is_outward = bool(_KUDA_OUTWARD.search(cat_str))
        # Lone "spend" token = spend+save deduction → outward
        if not is_outward and cat_tokens and re.match(r"^spend$", cat_tokens[0], re.I):
            is_outward = True

        # Parse date: DD/MM/YY → YYYY-MM
        parts = date_str.split("/")
        if len(parts) != 3:
            i = j
            continue
        day, mon, yr = parts
        ym = f"{int(yr) + 2000}-{mon}"

        amount = float(amounts[0].replace(",", ""))
        narration = re.sub(r"\s+", " ", re.sub(r"\t", " ", " ".join(text_tokens))).strip()

        if is_outward:
            add_debit(ym, narration, amount)
            i = j
            continue

        # Skip stamp duty and government levies (these are bank charges, not income)
        if re.search(r"\bstamp\s+duty\b|\belectronic.*levy\b|\bfgn.*levy\b", narration, re.I):
            i = j
            continue

        add_credit(buckets, ym, amount, narration, account_name)
        i = j

    return buckets, account_name


def parse_zenith(full_text: str) -> tuple[dict, str]:
    buckets: dict = {}
    account_name = _extract_account_name(full_text)
    in_tx = False
    current: list[str] = []

    def process(lines: list[str]) -> None:
        full = " ".join(lines)
        m = ZENITH_ROW.match(full)
        if not m:
            return
        parts = m.group(1).split("/")
        ym    = f"{parts[2]}-{parts[1]}"
        rest  = m.group(3)
        money = list(re.finditer(r"[\d,]+\.\d{2}", rest))
        if len(money) < 2:
            return
        amount_m  = money[-2]
        amount    = float(amount_m.group().replace(",", ""))
        narration = rest[:amount_m.start()].strip()
        lower     = narration.lower()
        is_rev = bool(re.search(r"\*+rsvl\b|\brsvl\b", lower))
        is_cr  = bool(re.search(r"\b(?:nip\s*cr|cip\s*cr|inflow|etz\s+inflow)\b", lower)) or is_rev
        is_chg = bool(re.search(r"\b(?:stamp duty|sms alert|vat charge|charge\+vat|levy)\b", lower))
        if not is_cr or (is_chg and not is_rev):
            add_debit(ym, narration, amount)
            return
        add_credit(buckets, ym, amount, narration, account_name)

    for line in full_text.splitlines():
        line = line.strip()
        if "Tran Date Value Date Narration" in line:
            in_tx = True
            continue
        if not in_tx:
            continue
        if ZENITH_ROW.match(line):
            if current:
                process(current)
            current = [line]
        elif current and not line.startswith(("mybankStatement", "Tran Date")):
            current.append(line)
    if current:
        process(current)
    return buckets, account_name


def parse_gtbank(full_text: str) -> tuple[dict, str]:
    """
    Parser for all mybankStatement-engine banks: GTBank/GTCO, Access Bank,
    First Bank, UBA, Fidelity, Union Bank, Stanbic IBTC, FCMB, Wema, Sterling.

    pypdf renders the Debit/Credit columns in two ways depending on PDF vintage:

    Layout A — 3 numbers (debit=0.00 explicit on credit rows):
        10-01-2025  10-01-2025  TRANSFER FROM X  0.00  50,000.00  129,678.71

    Layout B — 2 numbers (debit column blank/omitted on credit rows):
        10-01-2025  10-01-2025  TRANSFER FROM X  50,000.00  129,678.71

    Layout A: debit==0 and credit>0 → confirmed credit.
    Layout B: balance delta > 0 → confirmed credit; delta < 0 → debit.

    The skip filter is anchored to line-start patterns so it never
    accidentally drops transaction narrations that mention a bank name.
    """
    buckets: dict     = {}

    # ── Wema Bank: PyPDF2 splits DD-MM-YYYY across three lines ───────────
    # Line 1: "02-10-"   Line 2: "202501-10-"   Line 3: "2025narration..."
    # Rejoin before parsing so GTB_DATE can see a full date.
    if re.search(r'^\d{2}-\d{2}-$', full_text, re.M):
        full_text = re.sub(
            r'(\d{2}-\d{2}-)\r?\n(\d{4})(\d{2}-\d{2}-)\r?\n(\d{4})',
            r'\1\2 \3\4 ',
            full_text,
        )

    # ── Sterling Bank mybankStatement (PyPDF2): M/D/YYYY split over lines ──
    # PyPDF2 wraps the year: "11/01/202" → newline → "511/03/202" → newline
    # → "5narration..." (the trailing "5" = last digit of year 2025).
    # After rejoining: "11/01/2025 11/03/2025 narration..."
    if re.search(r'^\d{1,2}/\d{1,2}/202$', full_text, re.M):
        full_text = re.sub(
            r'(\d{1,2}/\d{1,2}/202)\r?\n(\d)(\d{1,2}/\d{1,2}/202)\r?\n(\d)',
            r'\1\2 \3\4 ',
            full_text,
        )

    # ── Sterling Bank mybankStatement (pdfplumber): partial-year date rows ─
    # pdfplumber lays each row on ONE line:
    #   "M/D/202 M/D/202 narration debit credit balance"
    # with a continuation line starting "D D continuation..." where the
    # single digits complete the year (e.g. "5 5" → 202+5 = 2025).
    # Fix: append the year digit to both partial dates; strip "D D " prefix.
    if re.search(r'^\d{1,2}/\d{1,2}/202\s+\d{1,2}/\d{1,2}/202\s+', full_text, re.M):
        _PDFPL_DATE = re.compile(
            r'^(\d{1,2}/\d{1,2}/202)(\s+\d{1,2}/\d{1,2}/202\s+)',
        )
        _PDFPL_CONT = re.compile(r'^(\d)\s+(\d)\s*(.*)', re.DOTALL)
        _lines2, _j2 = full_text.splitlines(), 0
        _out2: list[str] = []
        while _j2 < len(_lines2):
            _dm2 = _PDFPL_DATE.match(_lines2[_j2])
            if _dm2 and _j2 + 1 < len(_lines2):
                _nm2 = _PDFPL_CONT.match(_lines2[_j2 + 1])
                if _nm2 and _nm2.group(1) == _nm2.group(2):
                    _yd = _nm2.group(1)           # year-end digit ("5")
                    # Complete every M/D/202 on this line with the year digit
                    _fixed2 = re.sub(
                        r'(\d{1,2}/\d{1,2}/202)(?=\s)',
                        rf'\g<1>{_yd}',
                        _lines2[_j2],
                    )
                    _out2.append(_fixed2)
                    _cont2 = _nm2.group(3).strip()
                    if _cont2:
                        _out2.append(_cont2)
                    _j2 += 2
                    continue
            _out2.append(_lines2[_j2])
            _j2 += 1
        full_text = '\n'.join(_out2)

    # ── mybankStatement split column header ───────────────────────────────
    # Some Sterling/Wema variants render the header as:
    #   "Tran. date Value\ndateTransaction details Debit Credit Balance"
    # Rejoin so the in_tx trigger ("Tran. date" + "Debit/Credit") fires.
    full_text = re.sub(
        r'(Tran\.\s*date\s+Value)\r?\n(date)',
        r'\1 \2',
        full_text,
        flags=re.I,
    )

    account_name: str = _extract_account_name(full_text)

    # Handles both numeric months (DD-MM-YYYY / M/D/YYYY) and alpha (DD-Mon-YYYY).
    # \d{1,2} accepts both zero-padded (DD/MM) and bare (M/D) single-digit components.
    GTB_DATE = re.compile(r"^(\d{1,2})[-/](\d{1,2}|[A-Za-z]{3})[-/](\d{4})", re.I)
    MONEY3   = re.compile(r"([\d,]+\.\d{2})\s+([\d,]+\.\d{2})\s+([\d,]+\.\d{2})\s*$")
    MONEY2   = re.compile(r"([\d,]+\.\d{2})\s+([\d,]+\.\d{2})\s*$")
    DATE_ORDER = "dmy"

    # ── MDY detection pass 1: "Period Mon D, YYYY" alpha-month format ─────
    period_m = re.search(
        r"\bPeriod\s+(Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)\s+\d{1,2},\s*20\d{2}",
        full_text,
        re.I,
    )
    if period_m:
        first_date_m = re.search(r"\n\s*(\d{1,2})[-/](\d{1,2})[-/](20\d{2})", full_text)
        if first_date_m and int(first_date_m.group(1)) == MONTH_MAP[period_m.group(1).lower()[:3]]:
            DATE_ORDER = "mdy"

    # ── MDY detection pass 2: numeric-Period format (e.g. UBA mybankStatement)
    # Scan the first 2000 chars for any slash-separated date where the SECOND
    # field (day position in M/D/YYYY) exceeds 12 — impossible as a month,
    # therefore confirming M/D/YYYY order.  Restricted to the header area so
    # that reference strings in narrations cannot trigger false positives.
    if DATE_ORDER == "dmy":
        for _dm in re.finditer(r'\b(\d{1,2})/(\d{1,2})/(20\d{2})\b', full_text[:2000]):
            if int(_dm.group(2)) > 12:
                DATE_ORDER = "mdy"
                break

    # ── MDY detection pass 3: period "DD-Mon-YYYY" consistency ──────────────
    # Catches statements whose header date is DMY (e.g. "29/05/2026") so pass 2
    # never fires, yet whose transactions are MDY (e.g. "11-03-2025" = Nov 3).
    # Logic: extract start month from "Period 01-Nov-2025…", compare to the
    # first transaction's first numeric field; if they match → MDY.
    if DATE_ORDER == "dmy":
        _pm3 = re.search(
            r'\bPeriod\s+\d{1,2}\s*[-/]\s*(jan|feb|mar|apr|may|jun|jul|aug|sep|oct|nov|dec)\s*[-/]\s*(\d{4})',
            full_text[:2000],
            re.I,
        )
        if _pm3:
            _pm3_month = MONTH_MAP.get(_pm3.group(1).lower()[:3], 0)
            _f3 = re.search(r'^(\d{1,2})-\d{1,2}-\d{4}', full_text, re.MULTILINE)
            if _f3 and _pm3_month and int(_f3.group(1)) == _pm3_month:
                DATE_ORDER = "mdy"

    # ── MDY sanity check: if MDY was triggered but the first transaction's
    # leading field is > 12 it MUST be a day (DMY).  This fixes Wema Bank
    # where the page-header "5/26/2026" triggers pass 2, but transactions are
    # "26-05-2026" (day 26 first).  Access Bank "11-01-2025" (month 11 first)
    # is ≤ 12 so MDY is correctly preserved.
    if DATE_ORDER == "mdy":
        _first_dash_tx = re.search(
            r'^(\d{1,2})-\d{1,2}-\d{4}\s+\d{1,2}-\d{1,2}-\d{4}',
            full_text, re.MULTILINE,
        )
        if _first_dash_tx and int(_first_dash_tx.group(1)) > 12:
            DATE_ORDER = "dmy"

    # Skip lines that ARE page headers — anchored so narrations mentioning
    # "GTBank" or "Access Bank" in the text are NOT accidentally dropped.
    SKIP_HDR = re.compile(
        r"^(?:mybankstatement|guaranty trust bank|first bank|access bank plc|"
        r"united bank for africa|fidelity bank|union bank|stanbic\s+(?:ibtc\s+)?bank|fcmb(?:\s|$)|"
        r"wema bank|sterling bank|ecobank|the pan african bank|gtco|rc no\.|"
        r"account\s+(?:name|no\.?|number|type|branch|currency|sort)|"
        r"available\s+balance|book\s+balance|total\s+(?:debit|credit)|"
        r"opening\s+balance|closing\s+balance|statement\s+period|"
        r"dear\s+customer|page\s+\d+)",
        re.I,
    )

    in_tx        = False
    prev_bal     : Optional[Decimal] = None
    pending_ym   = ""
    pending_narr = ""
    pending_cr   = 0.0   # confirmed credit amount (0 = nothing pending)

    def _flush() -> None:
        nonlocal pending_ym, pending_narr, pending_cr
        if pending_cr > 0:
            add_credit(buckets, pending_ym, pending_cr,
                       pending_narr.strip(), account_name)
        pending_ym = pending_narr = ""
        pending_cr = 0.0

    def _kind_from_delta(amount: Decimal, balance: Decimal, narr: str = "") -> tuple:
        """Return (kind, true_amount) for 2-column (Layout B) transaction rows.

        true_amount is the balance delta, not the raw amount column.
        Reference numbers are often concatenated to the amount by PyPDF2
        (e.g. "349347647473500.00" instead of "500.00"), so the amount column
        is unreliable.  We always use balance - prev_bal as the credit amount.

        When prev_bal is None (first transaction), delta is unknown so we
        return 0 for the credit amount — no credit is added.  The debit
        direction is still detected so prev_bal gets set for future rows.
        """
        nonlocal prev_bal
        if prev_bal is None:
            tl = narr.lower()
            if re.search(r"\bfrom\b|\bpayment\b|\bcredit\b|\bdeposit\b|\binflow\b|\breceived\b", tl):
                kind = "credit"
            elif re.search(r"\bto\b|\btransfer to\b|\bwithdraw\b|\bdebit\b|\boutflow\b|\bcharge\b|\bairtime\b|\bpurchase\b", tl):
                kind = "debit"
            else:
                kind = "credit"
            prev_bal = balance
            # Amount column may be contaminated — can't compute delta yet
            return kind, Decimal("0")
        else:
            delta = balance - prev_bal
            prev_bal = balance
            if   abs(delta - amount) <= Decimal("0.02"): return "credit", amount
            elif abs(delta + amount) <= Decimal("0.02"): return "debit",  amount
            elif delta > 0:  return "credit", delta   # delta is the reliable credit amount
            elif delta < 0:  return "debit",  -delta
            else:            return "credit", Decimal("0")

    def _process_amounts(m3, m2, narr_prefix: str) -> None:
        """Resolve amounts and set pending_cr / clear it."""
        nonlocal pending_cr, pending_narr, prev_bal
        if m3:
            debit   = float(m3.group(1).replace(",", ""))
            credit  = float(m3.group(2).replace(",", ""))
            balance = Decimal(m3.group(3).replace(",", ""))
            prev_bal = balance
            extra   = narr_prefix.strip()
            pending_narr = (pending_narr + " " + extra).strip() if extra else pending_narr
            if credit > 0 and debit == 0:
                pending_cr = credit
            else:
                # Debit transaction — log for officer visibility
                if debit > 0 and pending_ym:
                    add_debit(pending_ym, pending_narr.strip(), debit)
                pending_cr   = 0.0
                pending_narr = ""
        elif m2:
            amount  = Decimal(m2.group(1).replace(",", ""))
            balance = Decimal(m2.group(2).replace(",", ""))
            extra   = narr_prefix.strip()
            pending_narr = (pending_narr + " " + extra).strip() if extra else pending_narr
            kind, true_amt = _kind_from_delta(amount, balance, pending_narr)
            if kind == "credit":
                pending_cr = float(true_amt)
            else:
                # Debit transaction — log for officer visibility
                if float(true_amt) > 0 and pending_ym:
                    add_debit(pending_ym, pending_narr.strip(), float(true_amt))
                pending_cr   = 0.0
                pending_narr = ""

    for line in full_text.splitlines():
        stripped = line.strip()
        if not stripped:
            continue

        # ── Section-start detection ───────────────────────────────────────
        if re.search(r"tran\.?\s*date", stripped, re.I) and \
           re.search(r"debit|credit|balance", stripped, re.I):
            in_tx = True
            continue

        # ── Page-header skip (anchored — won't kill narration lines) ─────
        if SKIP_HDR.match(stripped):
            continue

        if not in_tx:
            continue

        date_m = GTB_DATE.match(stripped)

        if date_m:
            _flush()
            p1 = date_m.group(1)   # DD
            p2 = date_m.group(2)   # MM (numeric) or Mon (alpha e.g. "Oct")
            yy = date_m.group(3)   # YYYY
            if not p2.isdigit():
                # Alpha month: "Oct" → "10"
                mm = str(MONTH_MAP.get(p2.lower()[:3], 1)).zfill(2)
            elif DATE_ORDER == "mdy" or (int(p2) > 12 >= int(p1)):
                mm = p1.zfill(2)
            else:
                mm = p2.zfill(2)
            pending_ym = f"{yy}-{mm}"

            # Everything after the transaction date (strip optional value-date)
            # Value date may also use alpha months (e.g. "05-Nov-2025")
            rest = stripped[date_m.end():].strip()
            rest = re.sub(r"^\d{1,2}[-/](?:\d{1,2}|[A-Za-z]{3})[-/]\d{4}\s*", "", rest, flags=re.I).strip()

            m3 = MONEY3.search(rest)
            m2 = MONEY2.search(rest) if not m3 else None

            if m3 or m2:
                amt_match  = m3 or m2
                narr_chunk = rest[:rest.rfind(amt_match.group(0))].strip()
                pending_narr = narr_chunk
                _process_amounts(m3, m2, "")
            else:
                # Amounts not on this line yet — narration wraps
                pending_narr = rest
                pending_cr   = 0.0

        else:
            # ── Continuation line ─────────────────────────────────────────
            if not pending_ym:
                continue

            m3 = MONEY3.search(stripped)
            m2 = MONEY2.search(stripped) if not m3 else None

            if (m3 or m2) and pending_cr == 0.0:
                amt_match  = m3 or m2
                narr_chunk = stripped[:stripped.rfind(amt_match.group(0))].strip()
                _process_amounts(m3, m2, narr_chunk)
            elif not m3 and not m2:
                # Pure narration continuation
                pending_narr += " " + stripped

    _flush()
    return buckets, account_name


def parse_generic(full_text: str) -> tuple[dict, str]:
    """Balance-movement parser for Access, UBA, FirstBank, Sterling, etc."""
    buckets: dict = {}
    account_name  = _extract_account_name(full_text)
    prev_balance: Optional[Decimal] = None
    in_tx = False

    for line in full_text.splitlines():
        line = line.strip()
        if re.search(r"Tran Date|Value Date|Transaction Date", line, re.I):
            in_tx = True
            continue
        if not in_tx or not line:
            continue
        date_m = re.match(r"^(\d{2}/\d{2}/\d{4})", line)
        if not date_m:
            continue
        tran_date = date_m.group(1)
        parts = tran_date.split("/")
        ym    = f"{parts[2]}-{parts[1]}"
        money = list(MONEY_RE.finditer(line))
        if not money:
            continue
        balance = Decimal(money[-1].group().replace(",", ""))
        if len(money) < 2:
            prev_balance = balance
            continue
        amt_m     = money[-2]
        amount    = Decimal(amt_m.group().replace(",", ""))
        narration = line[date_m.end():amt_m.start()].strip()
        narration = re.sub(r"^\d{2}/\d{2}/\d{4}\s*", "", narration).strip()

        if prev_balance is None:
            kind = "credit" if balance >= amount else "unknown"
        else:
            delta = balance - prev_balance
            if   abs(delta - amount) <= Decimal("0.02"): kind = "credit"
            elif abs(delta + amount) <= Decimal("0.02"): kind = "debit"
            elif delta > 0:  kind = "credit";  amount = delta
            elif delta < 0:  kind = "debit";   amount = -delta
            else:            kind = "unknown"

        prev_balance = balance
        if kind == "credit" and amount > 0:
            add_credit(buckets, ym, float(amount), narration, account_name)

    return buckets, account_name


def parse_zenith_corporate(full_text: str) -> tuple[dict, str]:
    """
    Zenith Bank corporate/BOP statement format.
    Columns: DATE POSTED  VALUE DATE  DESCRIPTION  DEBIT  CREDIT  BALANCE
    """
    buckets: dict = {}
    account_name = _extract_account_name(full_text)
    if not account_name:
        name_m = re.search(r'^([A-Z][A-Z &]+(?:ENTERPRISES|LIMITED|LTD|COMPANY|CO\.?|PLC|NIG|NIGERIA))',
                            full_text, re.MULTILINE)
        account_name = name_m.group(1).strip() if name_m else "Unknown"
    account_name = re.sub(r'\s+(Account|Number|Currency|CA|NGN).*$', '', account_name, flags=re.I).strip()

    MONEY       = r'[\d,]+\.\d{2}'
    DATE_START  = re.compile(r'^\s*(\d{2}/\d{2}/\d{4})\s+(\d{2}/\d{2}/\d{4})\s*')
    THREE_NUMS  = re.compile(rf'({MONEY})\s+({MONEY})\s+({MONEY})\s*$')

    lines  = full_text.splitlines()
    in_tx  = False
    joined = []
    i      = 0

    def _fix_concat(line: str) -> str:
        line = re.sub(r'(\d{6,})(0\.\d{2})', r'\1 \2', line)
        line = re.sub(r'([A-Za-z/])(\d+\.\d{2})', r'\1 \2', line)
        line = re.sub(r'([A-Za-z/])(\d{1,3}(?:,\d{3})+\.\d{2})', r'\1 \2', line)
        return line

    while i < len(lines):
        line = lines[i]
        if 'DATE POSTED' in line and 'VALUE DATE' in line:
            in_tx = True
            i += 1
            continue
        if not in_tx:
            i += 1
            continue

        dm = DATE_START.match(line)
        if dm:
            if THREE_NUMS.search(_fix_concat(line)):
                joined.append(_fix_concat(line).strip())
                i += 1
            else:
                combined = line.rstrip()
                i += 1
                while i < len(lines):
                    nxt = lines[i].strip()
                    if not nxt:
                        i += 1
                        continue
                    if DATE_START.match(lines[i]):
                        break
                    combined = combined + ' ' + nxt
                    i += 1
                    if THREE_NUMS.search(_fix_concat(combined)):
                        break
                joined.append(_fix_concat(combined).strip())
        else:
            i += 1

    for row in joined:
        dm = DATE_START.match(row)
        if not dm:
            continue
        tdate = dm.group(1)
        parts = tdate.split('/')
        ym    = f"{parts[2]}-{parts[1]}"
        rest  = row[dm.end():]
        tm    = THREE_NUMS.search(rest)
        if not tm:
            continue
        debit     = float(tm.group(1).replace(',', ''))
        credit    = float(tm.group(2).replace(',', ''))
        narration = rest[:tm.start()].strip()

        lower = narration.lower()
        if any(k in lower for k in [
            'nip charge', 'stamp duty', 'sms charge', 'vat charge',
            'account maintenance', 'statement charge', 'airtime',
            'value added tax', 'charges on counter',
        ]):
            continue

        if credit > 0 and debit == 0:
            add_credit(buckets, ym, credit, narration, account_name)
        elif debit > 0 and credit == 0:
            add_debit(ym, narration, debit)

    return buckets, account_name


# ════════════════════════════════════════════════════════════════════════════
# FAIRMONEY PARSER
# ════════════════════════════════════════════════════════════════════════════
def parse_fairmoney(full_text: str) -> tuple[dict, str]:
    """
    FairMoney MFB statement parser.

    PyPDF2/pypdf renders each transaction as a single line in the form:
        DD/MM/YYYY  REFNUM  [+/-] ₦ AMOUNT  ₦ BALANCENarration start
    Narration continuation lines follow with no date or reference number.

    Page headers (bank name, address, account summary) repeat on every page
    and are filtered out during narration accumulation.

    Credits are identified by the '+' sign prefix on the amount.
    'Incoming FairMoney' credits (internal FairMoney wallet/interest credits)
    are classified as loan_disbursal and excluded from eligible income.
    """
    buckets: dict    = {}
    account_name: str = ""
    lines = full_text.splitlines()

    # ── Extract account holder name ───────────────────────────────────────
    # FairMoney puts the account name as a plain "Firstname Lastname" line
    # (Title-case, 2+ words) near the top of each page before the header
    # table. Grab the first such line from the full text.
    name_re = re.compile(r"^([A-Z][a-z]+(?:\s+[A-Z][a-z]+)+)$")
    skip_names = {"Date Reference", "Credit Debit Account"}
    for line in lines:
        stripped = line.strip()
        if stripped in skip_names:
            continue
        m = name_re.match(stripped)
        if m and len(stripped.split()) >= 2:
            account_name = stripped
            break

    # ── Parse transactions ────────────────────────────────────────────────
    pending_sign:   Optional[str]   = None
    pending_amount: float           = 0.0
    pending_ym:     str             = ""
    pending_narr:   str             = ""

    def _flush() -> None:
        nonlocal pending_sign, pending_amount, pending_ym, pending_narr
        if pending_sign == "+" and pending_amount > 0:
            add_credit(buckets, pending_ym, pending_amount,
                       pending_narr.strip(), account_name)
        elif pending_sign == "-" and pending_amount > 0:
            add_debit(pending_ym, pending_narr.strip(), pending_amount)
        pending_sign   = None
        pending_amount = 0.0
        pending_ym     = ""
        pending_narr   = ""

    for line in lines:
        stripped = line.strip()
        if not stripped:
            continue

        # Check for a new transaction anchor line first
        m = _FM_TX.match(stripped)
        if m:
            _flush()
            date_str, sign, amt_str, narr_start = m.groups()
            day, month, year = date_str.split("/")
            pending_ym     = f"{year}-{month}"
            pending_sign   = sign
            pending_amount = float(amt_str.replace(",", ""))
            pending_narr   = narr_start.strip()
            continue

        # Not a transaction anchor — handle as possible narration continuation
        if pending_sign is not None:
            # Skip page-header boilerplate lines
            if _FM_HEADER.match(stripped):
                continue
            if _FM_MONEY_ONLY.match(stripped):
                continue
            if _FM_DATE_RANGE.match(stripped):
                continue
            # Skip the account holder name line and account number that
            # repeat on each page header
            if stripped == account_name:
                continue
            if re.match(r"^\d{10}$", stripped):   # 10-digit account number
                continue
            # Append as narration continuation
            pending_narr += " " + stripped

    _flush()
    return buckets, account_name


# ════════════════════════════════════════════════════════════════════════════
# PALMPAY NEW-FORMAT PARSER
# ════════════════════════════════════════════════════════════════════════════
def parse_palmpay_new(full_text: str) -> tuple[dict, str]:
    """
    Parser for the PalmPay "Account Statement" format.

    Header: Name, Phone Number, Account Number (= phone number), Total Money In/Out
    Columns: Transaction Date | Transaction Detail | Money In (NGN) | Money Out (NGN) | Transaction ID
    Date format: MM/DD/YYYY HH:MM:SS AM/PM  (US-style date)
    Amount sign: immediately precedes the amount — '+' for Money In, '-' for Money Out.
    The sign and amount may be concatenated directly onto the narration end:
        "EMMANUEL ONAH+2000.00 6031thor1a907"
    Transaction ID is always the last token (8+ alphanumeric chars, no spaces).

    Bank name "PalmPay" may be absent from PyPDF2-extracted header text, so
    this parser is reached via format-based detection in detect_bank().
    """
    buckets: dict = {}
    account_name = ""

    # PalmPay header: "Account StatementName FIRSTNAME LASTNAME" on one line
    m = re.search(r'(?:Account\s*Statement\s*)?Name\s+([A-Z][A-Z ]{3,})', full_text)
    if m:
        account_name = m.group(1).strip()
    if not account_name:
        account_name = _extract_account_name(full_text)

    # Date anchor: MM/DD/YYYY HH:MM:SS [AP]M (AM/PM may be glued to narration)
    DATE_RE = re.compile(
        r'^(\d{2})/(\d{2})/(\d{4})\s+\d{2}:\d{2}:\d{2}\s*[AP]M\s*(.*)',
        re.I,
    )
    # Amount+sign with trailing Transaction ID at line end
    AMT_RE = re.compile(r'([+\-])([\d,]+\.\d{2})\s+\S{6,}\s*$')

    in_tx = False
    pending_ym = ""
    pending_narr = ""

    def _flush_pending(narr: str) -> None:
        nonlocal pending_ym, pending_narr
        am = AMT_RE.search(narr)
        if am and pending_ym:
            amount = float(am.group(2).replace(",", ""))
            clean_narr = narr[:am.start()].strip()
            if am.group(1) == "+":
                add_credit(buckets, pending_ym, amount, clean_narr, account_name)
            else:
                add_debit(pending_ym, clean_narr, amount)
        pending_ym = ""
        pending_narr = ""

    for line in full_text.splitlines():
        stripped = line.strip()
        if not stripped:
            continue

        if "transaction date" in stripped.lower() and "transaction detail" in stripped.lower():
            in_tx = True
            continue

        if not in_tx:
            continue

        dm = DATE_RE.match(stripped)
        if dm:
            # Flush any accumulated pending transaction
            if pending_ym:
                _flush_pending(pending_narr)

            month, day, year = dm.group(1), dm.group(2), dm.group(3)
            pending_ym = f"{year}-{month}"
            rest = dm.group(4).strip()

            am = AMT_RE.search(rest)
            if am:
                amount = float(am.group(2).replace(",", ""))
                narr_clean = rest[:am.start()].strip()
                if am.group(1) == "+":
                    add_credit(buckets, pending_ym, amount, narr_clean, account_name)
                else:
                    add_debit(pending_ym, narr_clean, amount)
                pending_ym = ""
                pending_narr = ""
            else:
                pending_narr = rest
        elif pending_ym:
            combined = (pending_narr + " " + stripped).strip()
            am = AMT_RE.search(combined)
            if am:
                amount = float(am.group(2).replace(",", ""))
                narr_clean = combined[:am.start()].strip()
                if am.group(1) == "+":
                    add_credit(buckets, pending_ym, amount, narr_clean, account_name)
                else:
                    add_debit(pending_ym, narr_clean, amount)
                pending_ym = ""
                pending_narr = ""
            else:
                pending_narr = combined

    if pending_ym:
        _flush_pending(pending_narr)

    return buckets, account_name


# ════════════════════════════════════════════════════════════════════════════
# PALMPAY BUSINESS ACCOUNT PARSER
# ════════════════════════════════════════════════════════════════════════════
def parse_palmpay_business(full_text: str) -> tuple[dict, str]:
    """
    Parser for PalmPay Business Account Statement (ISO-date format).

    Row layout as extracted by PyPDF2 (all tokens space-separated):
        YYYY-MM-DD HH:MM:SS  DESCRIPTION  DEBIT  CREDIT  BALANCE  [EXTRA...]  TXNID

    - Date: ISO format YYYY-MM-DD HH:MM:SS (not US-style like PalmPay_New)
    - DEBIT and CREDIT columns: one is 0.00, the other holds the amount
    - BALANCE: running account balance
    - EXTRA: optional — sender/receiver name, bank name, phone number
    - TXNID: last space-separated token on each row (alphanumeric, no spaces)

    Typical transaction types:
      Credits: "Pay With Transfer", "POS Card Purchase", "Withdraw from Agent"
      Debits:  "Send", "Stamp Duty", "Data bundle", "Top up Airtime"

    Detection signal: PyPDF2 splits "BUSINESS_ACCOUNT" (a cell in the Stamp Duty
    rows) into "BUSINESS_ACCOU" + "NT" because the word crosses a table-cell
    boundary.  Combined with ISO-format date stamps, this uniquely identifies
    this format.
    """
    from collections import Counter

    buckets: dict = {}
    account_name = ""

    # Account name: POS card purchases repeat the cardholder name before
    # "BALANCE_ACC". Count occurrences and pick the most common.
    name_cands = re.findall(
        r'([A-Z]{2,}(?:\s+[A-Z]{2,}){1,3})\s+BALANCE[_\s]ACC',
        full_text,
    )
    if name_cands:
        account_name = Counter(name_cands).most_common(1)[0][0]
    if not account_name:
        account_name = _extract_account_name(full_text)

    # Split the flat text at every ISO date-time stamp.
    # re.split() with a capturing group interleaves separators and segments:
    #   [preamble, date1, content1, date2, content2, ...]
    DATE_SPLIT_RE = re.compile(r'(\d{4}-\d{2}-\d{2} \d{2}:\d{2}:\d{2})')
    parts = DATE_SPLIT_RE.split(full_text)

    # Three consecutive decimal amounts: debit  credit  balance
    AMT3_RE = re.compile(r'([\d,]+\.\d{2})\s+([\d,]+\.\d{2})\s+([\d,]+\.\d{2})')

    # Walk pairs (date_str, row_content)
    idx = 1
    while idx + 1 < len(parts):
        date_str = parts[idx]      # "YYYY-MM-DD HH:MM:SS"
        row_text = parts[idx + 1]  # everything after the date until next date
        idx += 2

        ym = date_str[:7]  # "YYYY-MM"

        am = AMT3_RE.search(row_text)
        if not am:
            continue

        debit  = float(am.group(1).replace(',', ''))
        credit = float(am.group(2).replace(',', ''))
        # Narration: everything before the first amount trio, cleaned up
        narration = ' '.join(row_text[:am.start()].split()).strip()
        if not narration:
            continue

        if credit > 0 and debit == 0:
            add_credit(buckets, ym, credit, narration, account_name)
        elif debit > 0 and credit == 0:
            add_debit(ym, narration, debit)

    return buckets, account_name


# ════════════════════════════════════════════════════════════════════════════
# ACCESS BANK ORACLE-FORMAT PARSER
# ════════════════════════════════════════════════════════════════════════════
def parse_access_oracle(full_text: str) -> tuple[dict, str]:
    """
    Parser for Access Bank (Oracle-style) PDF statements.

    Column header: Posted Date  Value Date  Description  Debit (NGN)  Credit (NGN)  Balance (NGN)
    Date format:   DD-MON-YY  (e.g. 09-SEP-25)

    Amount layout per row (two possible patterns):
      Credit row: [text]-  CREDIT  BALANCE   ← debit column is "-" (may be glued to preceding text)
      Debit  row: DEBIT  -  BALANCE           ← credit column is "-" (standalone dash)

    PyPDF2 sometimes concatenates the value date onto the description without
    a space: "09-SEP-25 09-SEP-25NIP Transfer..." — handled by the DATE_RE
    which strips both date tokens then captures the rest.
    """
    buckets: dict = {}
    account_name = _extract_account_name(full_text)

    _MON3 = {
        "jan": 1, "feb": 2, "mar": 3, "apr": 4,
        "may": 5, "jun": 6, "jul": 7, "aug": 8,
        "sep": 9, "oct": 10, "nov": 11, "dec": 12,
    }

    # Date line: DD-MON-YY  DD-MON-YY  [Description...]
    # PyPDF2 may insert a stray space inside the date: "OCT -25" → tolerate
    # with \s* before each two-digit year component.
    DATE_RE = re.compile(
        r'^(\d{2})-([A-Za-z]{3})\s*-(\d{2})\s+\d{2}-[A-Za-z]{3}\s*-\d{2}\s*(.*)',
        re.I,
    )
    # Credit: leading dash (debit="-") then AMOUNT BALANCE at end of text
    # Balance uses [\d, ]+ to tolerate PyPDF2 stray spaces inside large numbers
    # e.g. "1,500,118.38" → "1,500,1 18.38" (space inserted mid-number)
    CREDIT_AMT = re.compile(r'-\s*([\d,]+\.\d{2})\s+([\d, ]+\.\d{2})\s*$')
    # Debit:  AMOUNT then dash then BALANCE at end of text
    # Same tolerance for stray spaces in balance column
    DEBIT_AMT  = re.compile(r'([\d,]+\.\d{2})\s+-\s*([\d, ]+\.\d{2})\s*$')
    # Explicit 3-column: DEBIT  CREDIT  BALANCE (rare in this format)
    THREE_AMT  = re.compile(
        r'([\d,]+\.\d{2})\s+([\d,]+\.\d{2})\s+([\d,]+\.\d{2})\s*$'
    )

    # Lines to skip regardless of in_tx state
    SKIP_RE = re.compile(
        r'^(?:posted\s+date|value\s+date|page\s+\d|generated\s+on|'
        r'account\s+details|financial\s+summary|transactions\s*$|'
        r'statement\s+period|currency\s*:|account\s+class)',
        re.I,
    )

    def _try_credit(text: str) -> tuple[float, str]:
        """Return (credit_amount, narration) if a credit row; else (0, '')."""
        m3 = THREE_AMT.search(text)
        if m3:
            debit  = float(m3.group(1).replace(",", ""))
            credit = float(m3.group(2).replace(",", ""))
            if credit > 0 and debit == 0:
                return credit, text[:m3.start()].strip()
            return 0, ""
        mc = CREDIT_AMT.search(text)
        if mc:
            credit = float(mc.group(1).replace(",", ""))
            return (credit, text[:mc.start()].strip()) if credit > 0 else (0, "")

    def _try_debit(text: str) -> tuple[float, str]:
        """Return (debit_amount, narration) if a debit row; else (0, '')."""
        m3 = THREE_AMT.search(text)
        if m3:
            debit  = float(m3.group(1).replace(",", ""))
            credit = float(m3.group(2).replace(",", ""))
            if debit > 0 and credit == 0:
                return debit, text[:m3.start()].strip()
            return 0, ""
        md = DEBIT_AMT.search(text)
        if md:
            debit = float(md.group(1).replace(",", ""))
            return (debit, text[:md.start()].strip()) if debit > 0 else (0, "")
        return 0, ""

    def _has_amounts(text: str) -> bool:
        return bool(
            THREE_AMT.search(text) or
            CREDIT_AMT.search(text) or
            DEBIT_AMT.search(text)
        )

    in_tx = False
    pending_ym = ""
    pending_narr = ""

    for line in full_text.splitlines():
        stripped = line.strip()
        if not stripped:
            continue

        # ── Section-start detection MUST fire before SKIP_RE ─────────────
        # "TRANSACTIONS" and "Posted Date" are both in SKIP_RE but they are
        # also the markers that enable transaction parsing. Check them first.
        if re.search(r'\bTRANSACTIONS\b', stripped) or \
           (re.search(r'\bPosted\s+Date\b', stripped, re.I) and
                re.search(r'\bValue\s+Date\b', stripped, re.I)):
            in_tx = True
            continue

        if SKIP_RE.match(stripped):
            continue

        if not in_tx:
            continue

        dm = DATE_RE.match(stripped)
        if dm:
            # Flush previous transaction
            if pending_ym and pending_narr:
                amt, narr = _try_credit(pending_narr)
                if amt > 0:
                    add_credit(buckets, pending_ym, amt, narr, account_name)

            mon_str = dm.group(2).lower()[:3]
            yr2     = int(dm.group(3))
            month_num = _MON3.get(mon_str, 0)
            if not month_num:
                pending_ym = ""
                pending_narr = ""
                continue

            pending_ym = f"{2000 + yr2}-{str(month_num).zfill(2)}"
            rest = dm.group(4).strip()

            amt, narr = _try_credit(rest)
            if amt > 0:
                add_credit(buckets, pending_ym, amt, narr, account_name)
                pending_ym = ""
                pending_narr = ""
            elif _has_amounts(rest):
                # Debit row — complete, log it
                damt, dnarr = _try_debit(rest)
                if damt > 0:
                    add_debit(pending_ym, dnarr, damt)
                pending_ym = ""
                pending_narr = ""
            else:
                pending_narr = rest

        elif pending_ym:
            combined = (pending_narr + " " + stripped).strip()
            amt, narr = _try_credit(combined)
            if amt > 0:
                add_credit(buckets, pending_ym, amt, narr, account_name)
                pending_ym = ""
                pending_narr = ""
            elif _has_amounts(combined):
                # Debit row complete — log it
                damt, dnarr = _try_debit(combined)
                if damt > 0:
                    add_debit(pending_ym, dnarr, damt)
                pending_ym = ""
                pending_narr = ""
            else:
                pending_narr = combined

    # Flush last pending
    if pending_ym and pending_narr:
        amt, narr = _try_credit(pending_narr)
        if amt > 0:
            add_credit(buckets, pending_ym, amt, narr, account_name)
        else:
            damt, dnarr = _try_debit(pending_narr)
            if damt > 0:
                add_debit(pending_ym, dnarr, damt)

    return buckets, account_name


# ════════════════════════════════════════════════════════════════════════════
# MONIEPOINT BUSINESS PARSER
# ════════════════════════════════════════════════════════════════════════════
def parse_moniepoint_business(full_text: str) -> tuple[dict, str]:
    """
    Parser for Moniepoint Business account statements (PDF via PyPDF2).

    Column layout: Date | Narration | Reference | Debit | Credit | Balance

    PyPDF2 splits the ISO timestamp across two lines:
      Line 1:  2025-11-02T11:                        ← YYYY-MM-DDThh: only
      Line 2:  28:40narration ref  0.00 32,400.00 40,726.03

    Occasionally the narration starts on the date line itself:
      2025-11-01T21: TRANSFER TO ALICE...
      01:50 /TRF|..._DEBIT_0  2,920.00  0.00  11,099.82

    Credit rows: Debit column = 0.00, Credit column > 0.
    """
    buckets: dict = {}

    # Business name: "Business Name Pearlsom Esthetic - Pearlsom Esthetic"
    name_m = re.search(
        r'Business\s+Name\s+(.*?)(?:\n|Account\s+Number)', full_text, re.I | re.S
    )
    if name_m:
        raw = name_m.group(1).strip()
        account_name = raw.split(' - ')[0].strip()
    else:
        account_name = _extract_account_name(full_text)

    # Date-start line: YYYY-MM-DDTHH: (with optional narration after)
    DATE_RE = re.compile(r'^(\d{4})-(\d{2})-\d{2}T\d{2}:', re.I)
    # Three amounts at end of line: DEBIT  CREDIT  BALANCE
    AMT3_RE = re.compile(
        r'([\d,]+\.\d{2})\s+([\d,]+\.\d{2})\s+([\d,]+\.\d{2})\s*$'
    )
    SKIP_RE = re.compile(
        r'^(?:account\s+statement|business\s+name|account\s+number|currency\s*$|'
        r'date\s+narration|opening\s+balance|total\s+debits|total\s+credits|'
        r'closing\s+balance|address)',
        re.I,
    )

    current_ym = ""

    for line in full_text.splitlines():
        stripped = line.strip()
        if not stripped or SKIP_RE.match(stripped):
            continue

        dm = DATE_RE.match(stripped)
        if dm:
            current_ym = f"{dm.group(1)}-{dm.group(2)}"
            # Edge case: date line already contains amounts (PyPDF2 merge)
            m3 = AMT3_RE.search(stripped)
            if m3 and current_ym:
                debit  = float(m3.group(1).replace(',', ''))
                credit = float(m3.group(2).replace(',', ''))
                narr   = stripped[:m3.start()].strip()
                if credit > 0:
                    add_credit(buckets, current_ym, credit, narr, account_name)
                elif debit > 0:
                    add_debit(current_ym, narr, debit)
            continue

        if not current_ym:
            continue

        m3 = AMT3_RE.search(stripped)
        if m3:
            debit  = float(m3.group(1).replace(',', ''))
            credit = float(m3.group(2).replace(',', ''))
            narr   = stripped[:m3.start()].strip()
            if credit > 0:
                add_credit(buckets, current_ym, credit, narr, account_name)
            elif debit > 0:
                add_debit(current_ym, narr, debit)

    return buckets, account_name


# ════════════════════════════════════════════════════════════════════════════
# MONIEPOINT BUSINESS v2 PARSER  (19-column Settlement format)
# ════════════════════════════════════════════════════════════════════════════
def parse_moniepoint_business_v2(full_text: str) -> tuple[dict, str]:
    """
    Parser for Moniepoint Business account statements with the 19-column
    "Settlement Debit / Settlement Credit" layout (v2 format).

    Column layout (PyPDF2 header text):
      Date | Account Name | Transaction Type | Transaction Status |
      Terminal ID | RRN | Transaction Ref | Reversal Status |
      Transaction Amount | Settlement Debit | Settlement Credit |
      Balance Before | Balance After | Charge |
      Beneficiary | Beneficiary Institution | Source | Source Institution |
      Narration

    PyPDF2 splits ISO timestamps across FOUR lines:
      Line 1: 2025-            ← year only
      Line 2: 12-[optional]   ← month (may have account-name text appended on
                                 page breaks)
      Line 3: 26T22:           ← day + T + hour (not needed)
      Line 4+: mm:ss<all remaining column data concatenated>

    Two credit sub-types:
      _CREDIT_0      — incoming transfer.
                       Amounts after ref: 0.00  CREDIT  BAL_BEFORE  BAL_AFTER
      _CBA_CREDIT_0  — counter bank account credit (inter-bank).
                       Amounts after ref: N/A  TA  0.00  CREDIT  BAL_BEFORE  BAL_AFTER  [CHARGE]

    In both sub-types, Settlement Debit = 0.00 immediately precedes the
    Settlement Credit amount.  The unified regex:
        (?<![0-9])0\\.00\\s+(amount)\\s+(bal_before)\\s+(bal_after)
    correctly extracts the credit amount for both patterns.

    Charge rows (_CREDIT_0_EMTL_DC_0) are excluded — they carry "_EMTL"
    in the reference field.

    Accuracy: ~99.8 % on tested statements.
    """
    buckets: dict = {}

    # Account name from "Business Name" header line
    name_m = re.search(
        r'Business\s+Name\s+(.*?)(?:\n|Account\s+Number)', full_text, re.I | re.S
    )
    if name_m:
        account_name = name_m.group(1).strip().split(" - ")[0].strip()
    else:
        account_name = _extract_account_name(full_text)

    # ── Patterns ──────────────────────────────────────────────────────────
    YEAR_RE = re.compile(r"^(20\d{2})-$")          # line = "YYYY-" only
    MONTH_RE = re.compile(r"^(\d{2})-")             # line starts with "MM-"

    # Settlement Debit=0.00 followed by Settlement Credit + 2 balance amounts.
    # Negative lookbehind prevents matching "500.00" (digit before the 0).
    CREDIT_RE = re.compile(
        r"(?<![0-9])0\.00\s+"             # Settlement Debit = 0.00
        r"([\d,]+\.\d{2})\s+"            # Settlement Credit (captured)
        r"[\d,]+\.\d{2}\s+"              # Balance Before
        r"[\d,]+\.\d{2}"                 # Balance After
    )

    lines = full_text.splitlines()
    n = len(lines)
    current_year: str | None = None
    current_month: str | None = None
    i = 0

    while i < n:
        stripped = lines[i].strip()

        # ── Year anchor ───────────────────────────────────────────────────
        ym_m = YEAR_RE.match(stripped)
        if ym_m:
            current_year = ym_m.group(1)
            # Look ahead for the month line (skip blanks)
            j = i + 1
            while j < n and not lines[j].strip():
                j += 1
            if j < n:
                mm_m = MONTH_RE.match(lines[j].strip())
                if mm_m:
                    current_month = mm_m.group(1)
                    i = j + 1
                    continue
            i += 1
            continue

        # ── Credit transaction line ───────────────────────────────────────
        if (current_year and current_month
                and "_CREDIT_" in stripped
                and "_EMTL" not in stripped):          # skip charge rows
            cm = CREDIT_RE.search(stripped)
            if cm:
                credit = float(cm.group(1).replace(",", ""))
                if credit > 0:
                    ym = f"{current_year}-{current_month}"
                    after = stripped[cm.end():].strip()
                    after = re.sub(r"^\d{10,}\s*", "", after).strip()
                    narration = after if after else "Credit transfer"
                    add_credit(buckets, ym, credit, narration, account_name)

        # ── Debit transaction line ────────────────────────────────────────
        elif (current_year and current_month
                and "_DEBIT_" in stripped
                and "_EMTL" not in stripped):          # skip charge rows
            # Debit pattern: Settlement Credit=0.00 immediately after debit amount
            # Extract: DEBIT_AMOUNT  0.00  BAL_BEFORE  BAL_AFTER
            DEBIT_V2 = re.compile(
                r'([\d,]+\.\d{2})\s+'       # Settlement Debit (captured)
                r'(?<![0-9])0\.00\s+'       # Settlement Credit = 0.00
                r'[\d,]+\.\d{2}\s+'         # Balance Before
                r'[\d,]+\.\d{2}'            # Balance After
            )
            dm2 = DEBIT_V2.search(stripped)
            if dm2:
                debit = float(dm2.group(1).replace(",", ""))
                if debit > 0:
                    ym = f"{current_year}-{current_month}"
                    after = stripped[dm2.end():].strip()
                    after = re.sub(r"^\d{10,}\s*", "", after).strip()
                    narration = after if after else "Debit transfer"
                    add_debit(ym, narration, debit)

        i += 1

    return buckets, account_name


# ════════════════════════════════════════════════════════════════════════════
# CARBON MFB PARSER
# ════════════════════════════════════════════════════════════════════════════
def parse_carbon(full_text: str) -> tuple[dict, str]:
    """
    Parser for Carbon MFB (getcarbon.co) PDF statements.

    PyPDF2 format per transaction block:
      DD/MM/YYYY                           ← date (standalone line)
      HH:MM AM/PM[Narration start]         ← time + narration start (concatenated)
      [narration continuation lines]       ← 0–N optional lines
      Carbon Account[CREDIT] [BALANCE]     ← credit row (no '-' prefix)
      Carbon Account- [DEBIT] [BALANCE]    ← debit row  ('-' prefix)
      Carbon Account- 10 -                 ← service charge (balance unchanged)

    Credits: rest after "Carbon Account" does NOT start with '-'.
    Amounts in credits: "amount balance" — amount may lack decimals (e.g. "13,000").
    Debits and service charges are skipped.

    Page headers (Page N of N, Licensed by CBN, DATETRANSACTION DETAILS, etc.)
    are skipped unconditionally so they never contaminate narration text.
    """
    buckets: dict = {}
    account_name = ""

    # ── Account name: line following "Currency:" ──────────────────────────
    lines = full_text.splitlines()
    for i, ln in enumerate(lines):
        if re.search(r'\bCurrency:\s*$', ln.strip(), re.I):
            for j in range(i + 1, min(i + 5, len(lines))):
                cand = lines[j].strip()
                # Title-case multi-word name: "Firstname Middlename Lastname"
                if cand and re.match(r'^[A-Z][a-zA-Z]+(?:\s+[A-Z][a-zA-Z]+)+$', cand):
                    account_name = cand
                    break
            if account_name:
                break
    # Fallback: find first Title-case multi-word line in header area
    if not account_name:
        _skip_exact = {
            'account statement', 'transaction summary', 'account balance',
            'opening balance', 'closing balance', 'total income', 'total expenses',
        }
        for ln in lines[:60]:
            cand = ln.strip()
            if (cand and re.match(r'^[A-Z][a-z]+(?:\s+[A-Z][a-z]+){1,4}$', cand)
                    and len(cand.split()) >= 2
                    and cand.lower() not in _skip_exact):
                account_name = cand
                break

    # ── Patterns ──────────────────────────────────────────────────────────
    DATE_RE    = re.compile(r'^(\d{2})/(\d{2})/(\d{4})$')
    TIME_RE    = re.compile(r'^\d{2}:\d{2}\s*[AP]M(.*)', re.I)
    CHANNEL_RE = re.compile(r'^Carbon\s+Account(.*)', re.I)
    # Credit: "AMOUNT BALANCE" — amount may have no decimal places; balance may
    # show only 1 decimal digit when the cents value ends in 0 (e.g. "562,336.2"
    # instead of "562,336.20") — accept 1–2 decimal places on both fields.
    CREDIT_RE  = re.compile(r'^([\d,]+(?:\.\d{1,2})?)\s+([\d,]+\.\d{1,2})$')
    # Page-header lines to skip unconditionally (never narration content)
    SKIP_RE    = re.compile(
        r'^(?:Page \d+ of \d+|Licensed by CBN|^Deposits$|Insured by NDIC|'
        r'Carbon is a financial|P: \d|DATETRANSACTION|ChannelDEBITS|'
        r'Account Statement|Here\'s the account|Disclaimer:|www\.getcarbon|'
        r'Transaction Summary|Account Balance|YOUR ACCOUNT STATEMENT|'
        r'Account No:|Period:|Client ID:|Currency:|Total Income|Total Expenses|'
        r'Opening balance|Closing balance|NGN[\d,])',
        re.I,
    )

    state        = 'idle'   # 'idle' | 'date' | 'narration'
    current_ym   = ""
    current_narr = ""

    # Debit amount+balance pattern: "- AMOUNT BALANCE"
    DEBIT_RE = re.compile(r'^-\s*([\d,]+(?:\.\d{1,2})?)\s+([\d,]+\.\d{1,2})$')

    def _flush_channel(rest: str) -> None:
        """Parse amounts from 'Carbon Account<rest>'; add credit/debit."""
        r = rest.strip()
        if not r:
            return
        if r.startswith('-'):
            # Debit row
            md = DEBIT_RE.match(r)
            if md and current_ym:
                amount = float(md.group(1).replace(',', ''))
                if amount > 0:
                    add_debit(current_ym, current_narr.strip(), amount)
            return
        m = CREDIT_RE.match(r)
        if m and current_ym:
            amount = float(m.group(1).replace(',', ''))
            if amount > 0:
                add_credit(buckets, current_ym, amount,
                           current_narr.strip(), account_name)

    for line in lines:
        stripped = line.strip()
        if not stripped:
            continue

        # Skip page-header boilerplate (safe regardless of state)
        if SKIP_RE.match(stripped):
            continue

        # Channel line → ends current transaction
        cm = CHANNEL_RE.match(stripped)
        if cm:
            _flush_channel(cm.group(1))
            state        = 'idle'
            current_narr = ""
            current_ym   = ""
            continue

        # Date line → starts new transaction
        dm = DATE_RE.match(stripped)
        if dm:
            state        = 'date'
            day, mon, yr = dm.group(1), dm.group(2), dm.group(3)
            current_ym   = f"{yr}-{mon}"
            current_narr = ""
            continue

        if state == 'idle':
            continue

        # Time line (may have narration text concatenated immediately after)
        tm = TIME_RE.match(stripped)
        if tm:
            current_narr = tm.group(1).strip()
            state        = 'narration'
            continue

        # Narration continuation
        if state in ('date', 'narration'):
            current_narr = (current_narr + ' ' + stripped).strip() if current_narr else stripped
            state        = 'narration'

    return buckets, account_name


# ════════════════════════════════════════════════════════════════════════════
# PROVIDUS BANK PARSER
# ════════════════════════════════════════════════════════════════════════════
def parse_providus(full_text: str) -> tuple[dict, str]:
    """
    Parser for Providus Bank PDF statements.

    Column header: TXN DATE  VAL DATE  REMARKS  DEBIT  CREDIT  BALANCE
    Date format:   DD-MM-YYYY

    PyPDF2 concatenates the value date directly to the narration start:
        DD-MM-YYYY DD-MM-YYYYNarration text...AMOUNT BALANCE

    Debit vs credit detection uses the balance-delta method (no separate debit/
    credit columns are reliably extractable from the collapsed PyPDF2 layout):
        delta > 0  → credit with amount = delta
        delta ≤ 0  → debit or charge, skip

    The opening balance from the PDF header initialises the tracking chain so
    the very first transaction (if a credit) is not missed.

    Multi-line transactions: narration + amount may span 2–3 physical lines.
    A row is considered complete when the combined text ends with an amount.
    """
    buckets: dict = {}

    # ── Account name ──────────────────────────────────────────────────────
    account_name = ""
    name_m = re.search(r'CUST\.?\s+NAME\s+([A-Z][A-Z .&\-]+)', full_text)
    if name_m:
        account_name = name_m.group(1).strip()
    if not account_name:
        account_name = _extract_account_name(full_text)

    # ── Opening balance (initialises delta tracking) ──────────────────────
    prev_balance: Optional[float] = None
    ob_m = re.search(r'OPENING\s+BAL\.\s+([\d,]+\.\d{2})', full_text, re.I)
    if ob_m:
        prev_balance = float(ob_m.group(1).replace(',', ''))

    # ── Patterns ──────────────────────────────────────────────────────────
    # Transaction anchor: DD-MM-YYYY followed by DD-MM-YYYY (value date)
    DATE_RE  = re.compile(r'^(\d{2})-(\d{2})-(\d{4})\s+\d{2}-\d{2}-\d{4}(.*)', re.S)
    # Balance = last money amount at end of line
    LAST_AMT = re.compile(r'([\d,]+\.\d{2})\s*$')
    # Lines to skip unconditionally
    SKIP_RE  = re.compile(
        r'^(?:STATEMENT\s+OF\s+ACCOUNT|CUST\.?\s+NAME|ADDRESS\s+|ACC\.?\s+NO|'
        r'ACC\.?\s+TYPE|CURRENCY\s+NGN|START\s+DATE|END\s+DATE|'
        r'OPENING\s+BAL|CLOSING\s+BAL|PRINTED\s+(?:ON|BY)|'
        r'TXN\s+DATE\s+VAL\s+DATE|Page\s+\d)',
        re.I,
    )
    # Explicit debit narration keywords — double-check to skip obvious charges
    # (balance delta is the primary filter; these are a safety net)
    DEBIT_NARR = re.compile(
        r'^(?:OUTWARD TRANSFER|COMMISSION TO|VAT TO|STAMP DUTY FROM|'
        r'BANK CHARGE|CHARGE RECOVERY|SMS ALERT|MAINTENANCE FEE|COT CHARGE)',
        re.I,
    )

    in_tx        = False
    pending_ym   = ""
    pending_narr = ""

    def _try_process(ym: str, text: str) -> bool:
        """
        Extract balance from text, compute delta, and add credit if applicable.
        Returns True if a balance was found (row is complete), False otherwise.
        """
        nonlocal prev_balance
        bal_m = LAST_AMT.search(text)
        if not bal_m:
            return False

        balance = float(bal_m.group(1).replace(',', ''))

        # Build clean narration: remove balance + trailing transaction amount
        raw_narr = text[:bal_m.start()].strip()
        # Strip the transaction amount (second-to-last money value)
        raw_narr = re.sub(r'[\d,]+\.\d{2}\s*$', '', raw_narr).strip()
        # Strip trailing reference codes: long digit strings, /XXXXX, \s+XXXXX
        raw_narr = re.sub(r'/\d{8,}\s*$', '', raw_narr).strip()
        raw_narr = re.sub(r'\s\d{10,}\s*$', '', raw_narr).strip()
        # Strip trailing pipe + digits (e.g. "| 0000232501...")
        raw_narr = re.sub(r'\s*\|\s*\d{6,}\s*$', '', raw_narr).strip()

        if prev_balance is not None:
            delta = balance - prev_balance
        else:
            delta = None

        prev_balance = balance   # always update chain

        if delta is not None and delta > 0.01:
            # Credit: use delta as the authoritative amount
            amount = delta
            if not DEBIT_NARR.match(raw_narr):
                add_credit(buckets, ym, amount, raw_narr, account_name)
        elif delta is not None and delta < -0.01:
            # Debit
            add_debit(ym, raw_narr, abs(delta))

        return True

    for line in full_text.splitlines():
        stripped = line.strip()
        if not stripped:
            continue

        # Enable transaction parsing once column header is seen
        if re.search(r'TXN\s+DATE\s+VAL\s+DATE', stripped, re.I):
            in_tx = True
            continue

        if SKIP_RE.match(stripped):
            continue

        if not in_tx:
            continue

        dm = DATE_RE.match(stripped)
        if dm:
            # Flush any incomplete previous accumulation (no balance found yet
            # on the previous row — treat as a discarded fragment)
            if pending_ym and pending_narr:
                _try_process(pending_ym, pending_narr)

            day, month, year = dm.group(1), dm.group(2), dm.group(3)
            pending_ym = f"{year}-{month}"
            rest = dm.group(4).strip()
            # Strip value date that PyPDF2 concatenates to narration start
            rest = re.sub(r'^\d{2}-\d{2}-\d{4}', '', rest).strip()

            if _try_process(pending_ym, rest):
                pending_ym   = ""
                pending_narr = ""
            else:
                pending_narr = rest

        elif pending_ym:
            combined = (pending_narr + ' ' + stripped).strip()
            if _try_process(pending_ym, combined):
                pending_ym   = ""
                pending_narr = ""
            else:
                pending_narr = combined

    # Final flush
    if pending_ym and pending_narr:
        _try_process(pending_ym, pending_narr)

    return buckets, account_name


# ════════════════════════════════════════════════════════════════════════════
# ACCURACY VERIFICATION — extract stated totals from PDF header
# ════════════════════════════════════════════════════════════════════════════
def extract_stated_totals(full_text: str) -> dict:
    """
    Parse the bank-stated summary figures from the top of a PDF statement.

    Returns a dict with any subset of:
      total_credits    — stated total inflows / deposits / money-in
      total_debits     — stated total outflows / withdrawals / money-out
      opening_balance  — opening balance
      closing_balance  — closing balance

    Only the first 3 000 characters are searched (the header/summary area)
    to avoid picking up transaction-line amounts.
    """
    header = full_text[:3000]
    result: dict = {}

    patterns = [
        # Total inflows
        (r'Total\s+(?:Money\s+In|Deposits?|Credits?|Inflow)\s*[:\s]*[₦\s]*([\d,]+\.\d{2})',
         "total_credits"),
        # Total outflows
        (r'Total\s+(?:Money\s+Out|Withdrawals?|Debits?|Outflow)\s*[:\s]*[₦\s]*([\d,]+\.\d{2})',
         "total_debits"),
        # Opening balance
        (r'Opening\s+Balance\s*[:\s]*[₦\s]*([\d,]+\.\d{2})',
         "opening_balance"),
        # Closing / cleared balance
        (r'(?:Closing|Cleared)\s+Balance\s*[:\s]*[₦\s]*([\d,]+\.\d{2})',
         "closing_balance"),
    ]
    for pat, key in patterns:
        m = re.search(pat, header, re.I)
        if m:
            result[key] = float(m.group(1).replace(",", ""))
    return result


def verify_extraction_accuracy(
    buckets: dict,
    stated: dict,
) -> dict:
    """
    Compare extracted gross credits against the bank-stated total credits.

    Returns:
      matched      — True if within tolerance
      extracted    — sum of extracted gross credits
      stated_total — bank-stated total (or None)
      pct_match    — 0–100 match percentage (or None)
      message      — human-readable summary
    """
    if not stated or "total_credits" not in stated:
        return {"matched": None, "extracted": None,
                "stated_total": None, "pct_match": None,
                "message": "No stated total found in header to verify against."}

    extracted = sum(b.get("gross", 0) for b in buckets.values())
    stated_total = stated["total_credits"]

    if stated_total == 0:
        return {"matched": None, "extracted": extracted,
                "stated_total": 0, "pct_match": None,
                "message": "Stated total is zero — cannot verify."}

    pct = min(extracted, stated_total) / max(extracted, stated_total) * 100
    matched = pct >= 90.0  # within 10% is considered a match
    diff = abs(extracted - stated_total)

    if pct >= 99:
        msg = f"Extraction matches stated total within 1% — ✓ High confidence"
    elif pct >= 95:
        msg = f"Extraction within 5% of stated total — ✓ Good match (₦{diff:,.0f} gap)"
    elif pct >= 90:
        msg = f"Extraction within 10% of stated total — ⚠ Acceptable (₦{diff:,.0f} gap)"
    else:
        msg = (f"Extracted ₦{extracted:,.0f} vs stated ₦{stated_total:,.0f} — "
               f"⚠ Large gap (₦{diff:,.0f}). Some transactions may not have been parsed.")

    return {
        "matched": matched,
        "extracted": extracted,
        "stated_total": stated_total,
        "pct_match": round(pct, 1),
        "message": msg,
    }


def parse_summary_credits(full_text: str) -> dict[str, float]:
    summary: dict[str, float] = {}
    pat = re.compile(
        r"\b(20\d{2})\s*(Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)"
        r"\s+([\d,]+\.\d{2})\s+([\d,]+\.\d{2})", re.I,
    )
    for m in pat.finditer(full_text.replace("\n", " ")):
        year  = m.group(1)
        month = str(MONTH_MAP[m.group(2).lower()[:3]]).zfill(2)
        credit = float(m.group(4).replace(",", ""))
        summary[f"{year}-{month}"] = credit
    return summary


# ════════════════════════════════════════════════════════════════════════════
# EXCEL / MONO STATEMENT PARSERS
# ════════════════════════════════════════════════════════════════════════════

def _excel_serial_to_ym(serial) -> Optional[str]:
    try:
        f = float(serial)
        if not (40000 < f < 70000):
            return None
        import datetime as _dt
        base = _dt.date(1899, 12, 30)
        d = base + _dt.timedelta(days=int(f))
        return f"{d.year}-{str(d.month).zfill(2)}"
    except Exception:
        return None


def _detect_excel_format(rows: list) -> Optional[dict]:
    for i, row in enumerate(rows[:30]):
        h = [str(c or "").lower().strip() for c in row]

        if ("date" in h and "narration" in h and "debit" in h and "credit" in h):
            if i + 1 < len(rows):
                try:
                    first_val = float(str(rows[i+1][0] or ""))
                    if 40000 < first_val < 70000:
                        return {
                            "type": "moniepoint_excel",
                            "hdr_idx": i,
                            "date_col": h.index("date"),
                            "narration_col": h.index("narration"),
                            "debit_col": h.index("debit"),
                            "credit_col": h.index("credit"),
                        }
                except (ValueError, IndexError):
                    pass

        if h and h[0] == "transaction date":
            credit_col = next((j for j, v in enumerate(h) if "credit" in v), None)
            debit_col  = next((j for j, v in enumerate(h) if "debit"  in v), None)
            narration_col = next((j for j, v in enumerate(h)
                                  if "narration" in v or "description" in v), None)
            if credit_col is not None:
                return {
                    "type": "mono_standard",
                    "hdr_idx": i,
                    "date_col": 0,
                    "narration_col": narration_col or 5,
                    "credit_col": credit_col,
                    "debit_col": debit_col,
                }

        has_date   = any("date" in v for v in h)
        has_credit = any("credit" in v for v in h)
        has_debit  = any("debit" in v for v in h)
        if has_date and has_credit and has_debit:
            date_col   = next((j for j, v in enumerate(h) if "date"   in v), 0)
            credit_col = next((j for j, v in enumerate(h) if "credit" in v), None)
            debit_col  = next((j for j, v in enumerate(h) if "debit"  in v), None)
            narr_col   = next((j for j, v in enumerate(h)
                               if "narration" in v or "description" in v or "details" in v), None)
            if credit_col is not None:
                return {
                    "type": "generic_excel",
                    "hdr_idx": i,
                    "date_col": date_col,
                    "narration_col": narr_col or 1,
                    "credit_col": credit_col,
                    "debit_col": debit_col,
                }
    return None


def _parse_excel_direct(file_bytes: bytes) -> tuple[dict, str]:
    """
    Fallback Excel parser using zipfile + lxml — bypasses openpyxl's stylesheet
    parser entirely. Handles files with invalid/missing stylesheet XML that
    openpyxl refuses to open (e.g. Moniepoint Business exports).

    Reads xl/sharedStrings.xml and xl/worksheets/sheet1.xml directly.
    Uses column letter references (A, B, J…) instead of positional indices
    so non-contiguous column layouts (common in Moniepoint Business) are
    handled correctly.
    """
    import zipfile
    import datetime as _dt
    from lxml import etree as _etree

    NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
    ns = {"s": NS}

    def _col_letter(cell_ref: str) -> str:
        return re.match(r"([A-Z]+)", cell_ref).group(1)

    def _serial_to_ym(val: str) -> Optional[str]:
        try:
            f = float(val)
            if not (40000 < f < 70000):
                return None
            base = _dt.date(1899, 12, 30)
            d = base + _dt.timedelta(days=int(f))
            return f"{d.year}-{str(d.month).zfill(2)}"
        except Exception:
            return None

    with zipfile.ZipFile(BytesIO(file_bytes)) as zf:
        # ── Shared strings ────────────────────────────────────────────────
        shared: list[str] = []
        with zf.open("xl/sharedStrings.xml") as f:
            tree = _etree.parse(f)
            for si in tree.findall(".//s:si", ns):
                texts = si.findall(".//s:t", ns)
                shared.append("".join(t.text or "" for t in texts))

        # ── Sheet data ────────────────────────────────────────────────────
        with zf.open("xl/worksheets/sheet1.xml") as f:
            tree = _etree.parse(f)
        xml_rows = tree.findall(".//s:row", ns)

    def _row_dict(xml_row) -> dict[str, str]:
        """Return {col_letter: value} for one XML row."""
        d: dict[str, str] = {}
        for c in xml_row.findall("s:c", ns):
            ref = c.get("r", "")
            if not ref:
                continue
            t_attr = c.get("t", "n")
            v_el = c.find("s:v", ns)
            val = ""
            if v_el is not None and v_el.text:
                val = shared[int(v_el.text)] if t_attr == "s" else v_el.text
            d[_col_letter(ref)] = val
        return d

    # ── Account name (scan first 10 rows) ─────────────────────────────────
    account_name = ""
    for xml_row in xml_rows[:10]:
        rd = _row_dict(xml_row)
        cols = sorted(rd.keys())
        for i, col in enumerate(cols):
            if "account name" in rd[col].lower():
                # value is in the next non-empty column on the same row
                for next_col in cols[i + 1:]:
                    nv = rd[next_col].strip(" -")
                    if nv:
                        account_name = nv
                        break
                break
        if account_name:
            break

    # ── Find header row ───────────────────────────────────────────────────
    hdr_row_idx = None
    date_col = credit_col = narr_col = None
    for i, xml_row in enumerate(xml_rows):
        rd = _row_dict(xml_row)
        inv = {v.lower().strip(): col for col, v in rd.items() if v.strip()}
        if "date" in inv and "credit" in inv and "debit" in inv:
            hdr_row_idx = i
            date_col   = inv["date"]
            credit_col = inv["credit"]
            narr_col   = inv.get("narration") or inv.get("description") or inv.get("details")
            break

    if hdr_row_idx is None:
        raise ValueError("Could not find header row (Date/Credit/Debit) in Excel file.")

    # ── Find debit column from same header row ────────────────────────────
    _hdr_rd2  = _row_dict(xml_rows[hdr_row_idx])
    _hdr_inv2 = {v.lower().strip(): col for col, v in _hdr_rd2.items() if v.strip()}
    debit_col_d = _hdr_inv2.get("debit")

    # ── Parse credit + debit rows ─────────────────────────────────────────
    buckets: dict = {}
    for xml_row in xml_rows[hdr_row_idx + 1:]:
        rd = _row_dict(xml_row)
        date_raw   = rd.get(date_col, "")
        credit_raw = rd.get(credit_col, "")
        if not date_raw:
            continue
        # Date: try serial first, then ISO/DD-MM-YYYY strings
        ym = _serial_to_ym(date_raw)
        if not ym:
            m = re.match(r"^(20\d{2})-(\d{2})", date_raw)
            if m:
                ym = f"{m.group(1)}-{m.group(2)}"
            else:
                m = re.match(r"^(\d{2})/(\d{2})/(20\d{2})", date_raw)
                ym = f"{m.group(3)}-{m.group(2)}" if m else None
        if not ym:
            continue
        narration = rd.get(narr_col, "").strip() if narr_col else ""

        # Log debit for officer visibility
        if debit_col_d:
            try:
                debit = float(str(rd.get(debit_col_d, "") or "").replace(",", ""))
                if debit > 0:
                    add_debit(ym, narration, debit)
            except (ValueError, TypeError):
                pass

        if not credit_raw:
            continue
        try:
            credit = float(str(credit_raw).replace(",", ""))
        except ValueError:
            continue
        if credit <= 0:
            continue
        add_credit(buckets, ym, credit, narration, account_name)

    return buckets, account_name


def parse_excel(file_bytes: bytes) -> tuple[dict, str]:
    """
    Parse an Excel bank statement.
    Primary path: openpyxl (handles most files).
    Fallback path: direct zip+lxml parsing for files where openpyxl fails
    due to an invalid/corrupt stylesheet XML (e.g. Moniepoint Business exports).
    """
    # ── Try openpyxl first ────────────────────────────────────────────────
    if OPENPYXL_AVAILABLE:
        try:
            import io
            wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
            ws = wb.active
            rows = [[cell.value for cell in row] for row in ws.iter_rows()]
            wb.close()
        except Exception:
            # Stylesheet or other XML error — fall through to direct parser
            return _parse_excel_direct(file_bytes)
    else:
        return _parse_excel_direct(file_bytes)

    fmt = _detect_excel_format(rows)
    if not fmt:
        # openpyxl loaded OK but format unrecognised — try direct parser
        return _parse_excel_direct(file_bytes)

    account_name = ""
    for row in rows[:fmt["hdr_idx"]]:
        for j, cell in enumerate(row):
            cv = str(cell or "").lower().strip()
            if cv in ("account name:", "account name", "name"):
                for nc in row[j+1:]:
                    nv = str(nc or "").strip()
                    if nv:
                        account_name = nv.rstrip("-").strip()
                        break
                if not account_name and j + 1 < len(row):
                    account_name = str(rows[rows.index(row)][j+1] or "").strip()
            if account_name:
                break
        if account_name:
            break

    buckets: dict = {}
    hdr_idx      = fmt["hdr_idx"]
    date_col     = fmt["date_col"]
    narr_col     = fmt["narration_col"]
    credit_col   = fmt["credit_col"]
    debit_col    = fmt.get("debit_col")
    fmt_type     = fmt["type"]

    for row in rows[hdr_idx + 1:]:
        if not row or row[date_col] is None:
            continue

        if fmt_type == "moniepoint_excel":
            ym = _excel_serial_to_ym(row[date_col])
        else:
            date_str = str(row[date_col] or "").strip()
            m = re.match(r"^(20\d{2})-(\d{2})", date_str)
            if m:
                ym = f"{m.group(1)}-{m.group(2)}"
            else:
                m = re.match(r"^(\d{2})/(\d{2})/(20\d{2})", date_str)
                ym = f"{m.group(3)}-{m.group(2)}" if m else None

        if not ym:
            continue

        narration = str(row[narr_col] or "").strip() if narr_col is not None else ""

        try:
            credit_raw = str(row[credit_col] or "").replace(",", "").strip()
            credit = float(credit_raw) if credit_raw else 0.0
        except ValueError:
            credit = 0.0

        # Log debit transactions for officer visibility
        if debit_col is not None:
            try:
                debit_raw = str(row[debit_col] or "").replace(",", "").strip()
                debit = float(debit_raw) if debit_raw else 0.0
                if debit > 0:
                    add_debit(ym, narration, debit)
            except (ValueError, IndexError):
                pass

        if credit <= 0:
            continue

        add_credit(buckets, ym, credit, narration, account_name)

    return buckets, account_name


def extract_account_no_excel(file_bytes: bytes) -> str:
    """Extract the account identifier from an Excel (Mono) bank statement.

    Mono exports carry label/value metadata in the top rows. Priority:
      1. 'Nuban'        — the real 10-digit account number (often blank)
      2. 'Account Number' / 'Account No'
      3. 'Client ID'    — Mono's per-customer identifier (BVN-style),
                          used as the account reference when Nuban is blank.
    """
    rows = []
    if OPENPYXL_AVAILABLE:
        try:
            import io
            wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
            ws = wb.active
            for ri, row in enumerate(ws.iter_rows(values_only=True)):
                rows.append(list(row))
                if ri >= 30:        # metadata always sits in the top rows
                    break
            wb.close()
        except Exception:
            return ""

    def _find(labels: list[str]) -> str:
        for row in rows:
            if not row:
                continue
            label = str(row[0] or "").strip().lower().rstrip(":")
            if label in labels:
                for cell in row[1:]:
                    val = str(cell or "").strip()
                    if val:
                        return re.sub(r"\D", "", val) or val
        return ""

    return (
        _find(["nuban"])
        or _find(["account number", "account no", "account no."])
        or _find(["client id"])
    )


# ════════════════════════════════════════════════════════════════════════════
# JAIZ BANK PARSER
# ════════════════════════════════════════════════════════════════════════════
def parse_jaiz(full_text: str) -> tuple[dict, str]:
    """
    Jaiz Bank statement parser.

    Column layout: TRANS DATE | NARRATION | VALUE DATE | DEBIT | CREDIT | BALANCE
    Date format: DD-Mon-YYYY (4-digit year, e.g. 01-Nov-2025 → 2025-11).

    PyPDF2 splits each date across two lines:
      Line A:  "DD-Mon-"          (trailing dash, year on next line)
      Line B:  "YYYY narration"   (year glued to narration start)
    Pre-processing joins these before parsing.

    Row end pattern (value date glued to debit, no space):
      "Ref:DD-Mon-YYYYDEBIT CREDIT BALANCE"
    e.g. "Ref:01-Nov-202550.00 0.00 1,265,379.95"
    """
    buckets: dict = {}

    # ── 1. Account name ───────────────────────────────────────────────────
    account_name = "Unknown"
    m_self = re.search(r'\bIFO\s+([A-Z][A-Z ]{3,}?)\s+Self\b', full_text)
    if m_self:
        account_name = m_self.group(1).strip()
    else:
        m_name = re.search(
            r'CUSTOMER NAME:\s+(?!ACCOUNT)([A-Z][A-Z ]{3,})', full_text
        )
        if m_name:
            account_name = m_name.group(1).strip()

    # ── 2. Pre-process: rejoin split dates ────────────────────────────────
    # PyPDF2 renders "DD-Mon-\n2025" — join into "DD-Mon-2025"
    full_text = re.sub(r'(\d{2}-[A-Za-z]{3}-)\n(\d{4})', r'\1\2', full_text)

    # ── 3. Regexes ────────────────────────────────────────────────────────
    # Transaction start: line beginning with DD-Mon-YYYY (2 or 4 digit year)
    TX_START = re.compile(r'^(\d{2}-[A-Za-z]{3}-\d{2,4})\s+(.*)')

    # Row end: value date glued to debit (no space), then credit, then balance
    # e.g. "...Ref:01-Nov-202550.00 0.00 1,265,379.95"
    ROW_END = re.compile(
        r'\d{2}-[A-Za-z]{3}-\d{2,4}'   # value date (glued to debit)
        r'([\d,]+\.\d{2})'              # debit (no space before)
        r'\s+([\d,]+\.\d{2})'           # credit
        r'\s+([\d,]+\.\d{2})\s*$'       # balance
    )

    # Page-header boilerplate lines to skip
    JAIZ_HDR = re.compile(
        r'^(?:CUSTOMER NAME:|SERVICES LTD|ADDRESS:|UNCLEARED|BALANCE:|'
        r'OPENING|CLOSING|BRANCH:|ACCOUNT TYPE:|CURRENCY:|'
        r'TRANSACTI|ON DATE|DATENARRATION|DATEDEBIT|Page\s+\d|'
        r'JAIZ BANK|JC\d{4,}|KANO)',
        re.I,
    )

    def _to_ym(date_str: str) -> str:
        """'DD-Mon-YYYY' or 'DD-Mon-YY' → 'YYYY-MM'"""
        parts = date_str.split('-')
        mm = MONTH_NUM.get(parts[1].lower(), "00")
        yr = parts[2] if len(parts[2]) == 4 else f"20{parts[2]}"
        return f"{yr}-{mm}"

    # ── 4. Parse ──────────────────────────────────────────────────────────
    lines = full_text.splitlines()
    pending_ym: str | None = None
    pending_acc: list[str] = []

    def _flush() -> None:
        nonlocal pending_ym, pending_acc
        if pending_ym and pending_acc:
            combined = ' '.join(pending_acc)
            m = ROW_END.search(combined)
            if m:
                debit  = float(m.group(1).replace(',', ''))
                credit = float(m.group(2).replace(',', ''))
                narration = combined[:m.start()].strip()
                if credit > 0:
                    add_credit(buckets, pending_ym, credit, narration, account_name)
                elif debit > 0:
                    add_debit(pending_ym, narration, debit)
        pending_ym  = None
        pending_acc = []

    for raw_line in lines:
        line = raw_line.strip()
        if not line:
            continue
        if JAIZ_HDR.match(line):
            _flush()
            continue

        dm = TX_START.match(line)
        if dm:
            _flush()
            pending_ym  = _to_ym(dm.group(1))
            pending_acc = [dm.group(2).strip()]
        elif pending_ym is not None:
            pending_acc.append(line)

    _flush()
    return buckets, account_name


# ════════════════════════════════════════════════════════════════════════════
# ZENITH BANK — NEW e-STATEMENT FORMAT (2025+)
# ════════════════════════════════════════════════════════════════════════════
def parse_zenith_new(full_text: str) -> tuple[dict, str]:
    """
    Zenith Bank e-Statement (new 2025+ format).

    Column layout: DATE | DESCRIPTION | DEBIT | CREDIT | VALUE DATE | BALANCE
    Date format:   DD/MM/YYYY

    PyPDF2 quirks:
    - Narration wraps onto continuation lines.
    - The last continuation line has the reference number glued directly to
      the debit amount with no space (e.g. "000585986050" + "0.00" →
      "0005859860500.00"). This makes debit extraction unreliable.

    Solution: Extract CREDIT as the number immediately preceding the VALUE
    DATE (which is always CREDIT then VALUE_DATE with 0 or 1 spaces between).
    Credit > 0 → income; credit == 0.00 → debit/fee (skip).
    """
    buckets: dict = {}

    # ── Account name ──────────────────────────────────────────────────────
    account_name = "Unknown"
    m_name = re.search(
        r'ACCOUNT NAME:\s+([A-Z][A-Z0-9 &\-\.]{2,}?)(?:\s+Account Statement)',
        full_text
    )
    if m_name:
        account_name = m_name.group(1).strip()
    if account_name == "Unknown":
        # Fallback: first ALL-CAPS word(s) after "ACCOUNT NAME:"
        m2 = re.search(r'ACCOUNT NAME:\s+([A-Z][A-Z ]{3,})', full_text)
        if m2:
            account_name = m2.group(1).strip()

    # ── Regexes ───────────────────────────────────────────────────────────
    # Transaction start: line begins with DD/MM/YYYY
    TX_START = re.compile(r'^(\d{2}/\d{2}/\d{4})\s+(.*)')

    # Extract CREDIT + VALUE_DATE + BALANCE from end of accumulated text.
    # KEY: credit column is the number IMMEDIATELY before the value date
    # (may be glued with 0 spaces, or separated by 1 space).
    # This correctly handles "0.0001/12/2025" → credit=0.00, and
    # "234,000.00 02/12/2025" → credit=234,000.00.
    ROW_END = re.compile(
        r'([\d,]+\.\d{2})'             # credit column (just before value date)
        r'\s*(\d{2}/\d{2}/\d{4})\s+'  # value date (opt space before)
        r'([\d,]+\.\d{2})\s*$'          # balance
    )

    # Header/boilerplate lines that should not be treated as transactions
    HDR = re.compile(
        r'^(?:ZENITH BANK|DOPEMU|ACCOUNT NAME|CURRENCY:|ACCOUNT No|'
        r'LAGOS|DATE DESCRIPTION|Opening Balance|Period:|'
        r'\d\s+[A-Z]{2,}\s+STR|Page\s+\d)',
        re.I
    )

    def _to_ym(date_str: str) -> str:
        parts = date_str.split('/')
        return f"{parts[2]}-{parts[1]}"

    # ── Opening balance seeds the delta chain ─────────────────────────────
    ob_m = re.search(r'Opening\s+Balance\s+([\d,]+\.\d{2})', full_text, re.I)

    # ── Parse ─────────────────────────────────────────────────────────────
    pending_ym: str | None = None
    pending_acc: list[str] = []
    prev_balance: float | None = float(ob_m.group(1).replace(',', '')) if ob_m else None

    def _flush() -> None:
        nonlocal pending_ym, pending_acc, prev_balance
        if pending_ym and pending_acc:
            combined = ' '.join(pending_acc)
            m = ROW_END.search(combined)
            if m:
                credit  = float(m.group(1).replace(',', ''))
                balance = float(m.group(3).replace(',', ''))
                narration = combined[:m.start()].strip()
                if credit > 0:
                    add_credit(buckets, pending_ym, credit, narration, account_name)
                elif prev_balance is not None:
                    # Debit: use balance delta (ref# glued to debit makes direct
                    # extraction unreliable — delta is always correct)
                    delta = prev_balance - balance
                    if delta > 0.005:
                        # Clean narration: strip trailing long digit string
                        # (reference number concatenated by PyPDF2)
                        clean_narr = re.sub(r'\s*\d{8,}\s*$', '', narration).strip()
                        add_debit(pending_ym, clean_narr or narration, round(delta, 2))
                prev_balance = balance
        pending_ym  = None
        pending_acc = []

    for raw_line in full_text.splitlines():
        line = raw_line.strip()
        if not line:
            continue
        if HDR.match(line):
            _flush()
            continue

        dm = TX_START.match(line)
        if dm:
            _flush()
            pending_ym  = _to_ym(dm.group(1))
            pending_acc = [dm.group(2).strip()]
        elif pending_ym is not None:
            pending_acc.append(line)

    _flush()
    return buckets, account_name


# ════════════════════════════════════════════════════════════════════════════
# PARALLEX BANK PARSER
# ════════════════════════════════════════════════════════════════════════════
def parse_parallex(full_text: str) -> tuple[dict, str]:
    """
    Parallex Bank native statement parser.

    Column layout:  Trans Date | Value Date | Debit | Credit | Balance | Narration
    Date format:    DD/MM/YYYY
    Credit rows:    Debit = 0.00, Credit > 0

    Account name is the first Title-Case proper-name line at the top of the
    document (e.g. "Rita Onyeka Ute"), appearing before the statement header.
    """
    buckets: dict = {}
    account_name = "Unknown"

    # Account name: first line that looks like a proper name —
    # 2–4 Title-Case words, no digits, no punctuation.
    # Appears before "This is your Account statement" on page 1.
    # PyPDF2 uses non-breaking spaces (\xa0) inside the name — normalise first.
    for raw in full_text.splitlines():
        line = re.sub(r'[\xa0 ]', ' ', raw).strip()
        if re.match(r'^[A-Z][a-z]+(?:\s+[A-Z][a-z]+){1,3}$', line):
            account_name = line
            break

    # Transaction row:
    #   DD/MM/YYYY  DD/MM/YYYY  <debit>  <credit>  <balance>[narration...]
    # PyPDF2 sometimes omits the space between balance and narration, so
    # \s* (zero or more spaces) is used instead of \s+ after the balance.
    TX_ROW = re.compile(
        r'^(\d{2}/\d{2}/\d{4})\s+\d{2}/\d{2}/\d{4}\s+'
        r'([\d,]+\.\d{2})\s+([\d,]+\.\d{2})\s+[\d,]+\.\d{2}'
        r'\s*(.*)?$'
    )

    def _to_ym(date_str: str) -> str:
        # DD/MM/YYYY → YYYY-MM
        parts = date_str.split('/')
        return f"{parts[2]}-{parts[1]}"

    # PyPDF2 repeats every transaction line once per PDF page for Parallex
    # statements; deduplicate by exact line content (the running balance is
    # included in the line so genuine same-day same-amount credits from
    # different senders will have different balances and thus unique lines).
    seen_lines: set[str] = set()

    for raw in full_text.splitlines():
        line = raw.strip()
        if not line or line in seen_lines:
            continue
        seen_lines.add(line)

        m = TX_ROW.match(line)
        if not m:
            continue
        trans_date = m.group(1)
        debit     = float(m.group(2).replace(',', ''))
        credit    = float(m.group(3).replace(',', ''))
        narration = (m.group(4) or '').strip()

        ym = _to_ym(trans_date)
        if credit > 0 and debit == 0:
            add_credit(buckets, ym, credit, narration, account_name)
        elif debit > 0 and credit == 0:
            add_debit(ym, narration, debit)

    return buckets, account_name


# ════════════════════════════════════════════════════════════════════════════
# RENMONEY MFB PARSER
# ════════════════════════════════════════════════════════════════════════════
def parse_renmoney(full_text: str) -> tuple[dict, str]:
    """
    Renmoney MFB statement parser — 4-pass design.

    PyPDF2 output characteristics:
    - Dates split across two lines: "2026-\\nMM-DD" → normalised to "2026-MM-DD|"
    - ₦ may render as ASCII "?" (0x3F)

    Credit vs debit distinction:
        Credit inline: narration - ₦amount ₦balance   (dash BEFORE ₦; TWO amounts after dash)
        Debit  inline: narration₦amount - ₦balance    (₦ BEFORE dash; ONE amount after dash)
        Credit split:  line ends with "-₦" → next line "<amount>₦"
        Debit  split:  line ends with "₦"  → next line "<amount>- ₦<balance>"

    Passes:
        1  — CREDIT | prefix entries (DOTALL handles any multi-line split)
        2  — General inline credits (non-CREDIT|/INTEREST_APPLIED; two amounts after "- ₦")
        3  — General split credits (line ends with "-₦"; next non-empty line = amount₦)
        4  — INTEREST_APPLIED entries

    Passes 1 & 4 are mutually exclusive with Passes 2 & 3 via the
    (?!CREDIT\\s*\\||INTEREST_APPLIED) negative lookahead and startswith guards,
    preventing any double-counting.

    Account name: from the date-header line, e.g. "16 May, 2026  Olusola Micheal Kehinde".
    """
    buckets: dict = {}

    # Normalise split dates: "2026-\nMM-DD" → "2026-MM-DD|"
    text_n = re.sub(r'(20\d{2})-\n(\d{2}-\d{2})', r'\1-\2|', full_text)

    # Account name — [ ]+ (spaces only) prevents crossing the newline into
    # the following "Product Name: Renmoney Account" line.
    account_name = "Unknown"
    m_name = re.search(
        r'\d{1,2}\s+\w+,?\s+\d{4}\s+([A-Z][a-z]+(?:[ ]+[A-Z][a-z]+)+)',
        full_text,
    )
    if m_name:
        account_name = m_name.group(1).strip()

    # ── Pass 1: CREDIT | entries (DOTALL handles any multi-line split) ────────
    # Format: YYYY-MM-DD|CREDIT | <narration> - <₦/?><amount>
    _CREDIT_RE = re.compile(
        r'(20\d{2}-\d{2}-\d{2})\|CREDIT\s*\|(.+?)-[^\d\r\n]{0,5}[\r\n]*([\d,]+\.\d{2})',
        re.DOTALL,
    )
    for m in _CREDIT_RE.finditer(text_n):
        date_str  = m.group(1)
        narration = m.group(2).strip()
        amount    = float(m.group(3).replace(',', ''))
        ym = f"{date_str[:4]}-{date_str[5:7]}"
        add_credit(buckets, ym, amount, narration, account_name)

    # ── Pass 2: General inline credits ───────────────────────────────────────
    # Matches: DATE|narration - ₦amount₁ ₦amount₂  (all on one line)
    # TWO amounts after "- [₦?]" = credit; debit lines have only ONE amount
    # after the dash (just the balance), so they never match.
    _INLINE_CR = re.compile(
        r'^(20\d{2})-(\d{2})-\d{2}\|'
        r'(?!CREDIT\s*\||INTEREST_APPLIED)'
        r'(.+?)'
        r'\s*-\s*[₦?]\s*([\d,]+\.\d{2})'
        r'\s+[₦?]?\s*[\d,]+\.\d{2}\s*$',
        re.MULTILINE,
    )
    for m in _INLINE_CR.finditer(text_n):
        ym     = f"{m.group(1)}-{m.group(2)}"
        narr   = m.group(3).strip()
        amount = float(m.group(4).replace(',', ''))
        add_credit(buckets, ym, amount, narr, account_name)

    # ── Pass 3: General split credits ────────────────────────────────────────
    # Credit split: date-anchored line ends with "-[₦?]" (dash then currency)
    # Next non-empty line: "<amount>[₦?]"
    # Debit split ends with just "[₦?]" (no preceding dash) → no match here.
    _DATE_START = re.compile(r'^(20\d{2})-(\d{2})-\d{2}\|')
    _SPLIT_END  = re.compile(r'-\s*[₦?]\s*$')
    _SPLIT_AMT  = re.compile(r'^([\d,]+\.\d{2})[₦?]')

    lines = text_n.splitlines()
    i = 0
    while i < len(lines):
        line = lines[i].strip()
        dm = _DATE_START.match(line)
        if dm:
            rest = line[dm.end():]
            if (not rest.startswith('CREDIT ')
                    and not rest.startswith('INTEREST_APPLIED')):
                se = _SPLIT_END.search(rest)
                if se:
                    # Find next non-empty line
                    j = i + 1
                    while j < len(lines) and not lines[j].strip():
                        j += 1
                    if j < len(lines):
                        am = _SPLIT_AMT.match(lines[j].strip())
                        if am:
                            ym     = f"{dm.group(1)}-{dm.group(2)}"
                            amount = float(am.group(1).replace(',', ''))
                            narr   = rest[:se.start()].strip()
                            add_credit(buckets, ym, amount, narr, account_name)
                            i = j + 1
                            continue
        i += 1

    # ── Pass 4: INTEREST_APPLIED entries ──────────────────────────────────────
    # Format: YYYY-MM-DD|INTEREST_APPLIED - <₦/?><amount>
    _INTEREST_RE = re.compile(
        r'(20\d{2}-\d{2}-\d{2})\|INTEREST_APPLIED\s*-\s*[^\d\s]{0,3}\s*([\d,]+\.\d{2})'
    )
    for m in _INTEREST_RE.finditer(text_n):
        date_str = m.group(1)
        amount   = float(m.group(2).replace(',', ''))
        ym = f"{date_str[:4]}-{date_str[5:7]}"
        add_credit(buckets, ym, amount, "INTEREST_APPLIED", account_name)

    # ── Pass 5: DEBIT entries — block-based ───────────────────────────────────
    # A "block" is everything from one date anchor to the next.  Narrations can
    # wrap over 1..N lines, so line-anchored regexes miss any debit whose amount
    # lands on a continuation line.  Within each block:
    #   Debit signature:  ₦<amount> - ₦<balance>   (amount BEFORE the dash;
    #                     newline may fall between ₦ and the amount)
    #   Credit signature: - ₦<amount> ... ₦<balance> (dash BEFORE the amount)
    # Whichever signature appears FIRST decides the row type.  Credit blocks are
    # skipped here — Passes 1–4 already booked them.
    _anchor_idx = [i for i, l in enumerate(lines)
                   if _DATE_START.match(l.strip())]
    _BLK_DR = re.compile(r'[₦?]\s*\n?([\d,]+\.\d{2})\s*-\s*[₦?]\s*[\d,]+\.\d{2}')
    _BLK_CR = re.compile(r'-\s*[₦?]\s*\n?([\d,]+\.\d{2})\s*\n?\s*[₦?]')
    for k, a in enumerate(_anchor_idx):
        end   = _anchor_idx[k + 1] if k + 1 < len(_anchor_idx) else len(lines)
        block = '\n'.join(lines[a:end])
        dm    = _DATE_START.match(block.strip())
        if not dm:
            continue
        rest = block.strip()[dm.end():]
        if rest.startswith('CREDIT ') or rest.startswith('INTEREST_APPLIED'):
            continue
        md = _BLK_DR.search(block)
        mc = _BLK_CR.search(block)
        if md and (not mc or md.start() < mc.start()):
            ym     = f"{dm.group(1)}-{dm.group(2)}"
            amount = float(md.group(1).replace(',', ''))
            narr   = ' '.join(rest[:max(md.start() - dm.end() - 1, 0)].split())
            if amount > 0:
                add_debit(ym, narr, amount)

    return buckets, account_name


# ════════════════════════════════════════════════════════════════════════════
# FIDELITY DIRECT PARSER
# ════════════════════════════════════════════════════════════════════════════
def parse_fidelity_direct(full_text: str) -> tuple[dict, str]:
    """
    Fidelity Bank native (non-mybankStatement) statement parser.

    Column layout:  TRANSACTION DATE | VALUE DATE | REFERENCE |
                    CHANNEL | DESCRIPTION | PAY IN | PAY OUT | BALANCE

    Date format: DD-Mon-YY (2-digit year, e.g. 17-Feb-26 → 2026-02).

    Credit detection: balance-delta method — a transaction is a credit
    when curr_balance > prev_balance.  Opening Balance seeds prev_balance.

    Channel "Online Banking" is split by PyPDF2:
      Line 1:  DD-Mon-YY  DD-Mon-YY[Online]   (no space before "Online")
      Line 2+: Banking<description>[amounts]
    Single-word channels (Others, POS, ATM …) appear on the date line itself.
    """
    buckets: dict = {}

    # ── 1. Account name ───────────────────────────────────────────────────
    # After the internal YYN code line, the next consecutive ALL-CAPS
    # letter-only lines are the account name (e.g. "CNM ENTERTAINMENT\nLIMITED").
    account_name = "Unknown"
    m_name = re.search(
        r'YYN[^\n]*\n([A-Z][A-Z ]+(?:\n[A-Z][A-Z ]+)*)',
        full_text,
    )
    if m_name:
        account_name = " ".join(m_name.group(1).split())
    if len(account_name) < 3 or account_name == "Unknown":
        account_name = _extract_account_name(full_text) or "Unknown"

    # ── 2. Opening balance ────────────────────────────────────────────────
    ob_m = re.search(r'Opening Balance\s+([\d,]+\.\d{2})', full_text)
    prev_bal = float(ob_m.group(1).replace(',', '')) if ob_m else 0.0

    # ── 3. Regexes ────────────────────────────────────────────────────────
    # Date line: DD-Mon-YY  DD-Mon-YY  [optional "Online" concatenated]  [rest]
    FID_DATE = re.compile(
        r'^(\d{2}-[A-Za-z]{3}-\d{2})\s+(\d{2}-[A-Za-z]{3}-\d{2})(Online)?(.*)',
        re.I,
    )
    # Last two decimal amounts on a line = (transaction_amount, closing_balance)
    AMTS_END = re.compile(r'([\d,]+\.\d{2})\s+([\d,]+\.\d{2})\s*$')
    # Channel prefixes to strip from the start of narration text
    CHANNEL_RE = re.compile(
        r'^(Banking|Online\s+Banking|POS|Others|ATM|Internet|Cheque|Transfer|'
        r'Cash|Interest|Bank\s+Charges|Instant\s+Banking)\s*',
        re.I,
    )

    def _to_ym(date_str: str) -> str:
        """'DD-Mon-YY' → 'YYYY-MM'  (century always 20xx)"""
        parts = date_str.split('-')
        mm = MONTH_NUM.get(parts[1].lower(), "00")
        return f"20{parts[2]}-{mm}"

    # ── 4. Find start of transactions section ─────────────────────────────
    lines = full_text.splitlines()
    start_idx = 0
    for idx, ln in enumerate(lines):
        if 'opening balance' in ln.lower() and re.search(r'[\d,]+\.\d{2}', ln):
            start_idx = idx + 1
            break

    # ── 5. Main parse loop ────────────────────────────────────────────────
    i = start_idx
    while i < len(lines):
        line = lines[i].strip()
        if not line:
            i += 1
            continue

        dm = FID_DATE.match(line)
        if not dm:
            i += 1
            continue

        tx_date = dm.group(1)                    # e.g. "17-Feb-26"
        ym      = _to_ym(tx_date)
        rest    = dm.group(4).strip()            # text after value-date (and any "Online")

        # ── Case A: all amounts on this date line (Others / POS / ATM …) ─
        am = AMTS_END.search(rest) if rest else None
        if am:
            curr_bal  = float(am.group(2).replace(',', ''))
            narr_raw  = rest[:am.start()].strip()
            narration = CHANNEL_RE.sub('', narr_raw).strip()
            delta     = curr_bal - prev_bal
            if delta > 0:
                add_credit(buckets, ym, round(delta, 2), narration, account_name)
            elif delta < 0:
                add_debit(ym, narration, round(abs(delta), 2), tx_date)
            prev_bal = curr_bal
            i += 1
            continue

        # ── Case B: multi-line (Online Banking / page-split) ─────────────
        i += 1
        acc = rest        # may be empty when line ended with "Online"

        for _ in range(8):          # guard — max 8 continuation lines
            if i >= len(lines):
                break
            nxt = lines[i].strip()

            # New date line → abandon this incomplete tx (page-split)
            if FID_DATE.match(nxt):
                break

            # Strip channel prefix ("Banking" / "Others" / etc.)
            nxt_clean = CHANNEL_RE.sub('', nxt, count=1).strip()
            if not nxt_clean:       # "Banking" alone → page-boundary artefact
                i += 1
                continue

            acc = (acc + ' ' + nxt_clean).strip()

            am = AMTS_END.search(acc)
            if am:
                curr_bal  = float(am.group(2).replace(',', ''))
                narr_raw  = acc[:am.start()].strip()
                narration = CHANNEL_RE.sub('', narr_raw).strip()
                delta     = curr_bal - prev_bal
                if delta > 0:
                    add_credit(buckets, ym, round(delta, 2), narration, account_name)
                elif delta < 0:
                    add_debit(ym, narration, round(abs(delta), 2), tx_date)
                prev_bal = curr_bal
                i += 1
                break

            i += 1
        # If the guard exhausted without finding amounts, i has already
        # advanced past the continuation lines; next iteration resumes normally.

    return buckets, account_name


# ════════════════════════════════════════════════════════════════════════════
# MAIN ENTRY POINT
# ════════════════════════════════════════════════════════════════════════════
def parse_transactions(file_bytes: bytes, password: str = "",
                       filename: str = "") -> tuple[dict, dict, str, str, list]:
    """
    Auto-detects PDF vs Excel and routes accordingly.
    Returns (buckets, summary_credits, bank_name, account_name, transactions)
    buckets: {ym: {gross, count, self_transfer, reversal, non_business, loan_disbursal, real_credit}}
    transactions: list of {ym, narration, amount, category} for keyword search
    """
    global _txn_log, _debit_log, _last_full_text
    _txn_log   = []       # Reset for each new parse
    _debit_log = []       # Reset debit log
    _last_full_text = ""  # Clear cached text from any prior call

    # ── Excel detection ───────────────────────────────────────────────────
    is_excel = (filename.lower().endswith((".xlsx", ".xls")) or
                file_bytes[:4] in (b"PK\x03\x04", b"\xd0\xcf\x11\xe0"))
    if is_excel:
        buckets, account_name = parse_excel(file_bytes)
        bank = "Mono Excel"
        summary: dict = {}
        return buckets, summary, bank, account_name, list(_txn_log), list(_debit_log)

    # ── PDF ───────────────────────────────────────────────────────────────
    full_text = extract_pdf_text(pdf_bytes=file_bytes, password=password)
    bank = detect_bank(full_text)
    summary = parse_summary_credits(full_text)

    # mybankStatement-engine banks all share the same explicit Debit/Credit
    # column format and are handled by parse_gtbank regardless of bank name.
    _MYBANKSTATEMENT_BANKS = {
        "GTBank", "Access", "FirstBank", "UBA",
        "Fidelity", "Union", "Stanbic", "FCMB", "Wema", "Sterling",
        "Ecobank",          # logo-only bank — uses same mybankStatement portal format
        "mybankStatement",  # unidentified logo-only portal bank (officer overrides)
    }

    if bank == "Carbon":
        buckets, account_name = parse_carbon(full_text)
    elif bank == "Providus":
        buckets, account_name = parse_providus(full_text)
    elif bank == "FairMoney":
        buckets, account_name = parse_fairmoney(full_text)
    elif bank == "OPay":
        buckets, account_name = parse_opay(full_text)
    elif bank == "OPay_v2":
        buckets, account_name = parse_opay_v2(file_bytes, full_text)
    elif bank == "OPay_Business":
        buckets, account_name = parse_opay_business(file_bytes)
    elif bank == "Moniepoint_Business_v2":
        buckets, account_name = parse_moniepoint_business_v2(full_text)
    elif bank == "Moniepoint_Business":
        buckets, account_name = parse_moniepoint_business(full_text)
    elif bank == "Kuda":
        buckets, account_name = parse_kuda(full_text)
    elif bank == "PalmPay_New":
        buckets, account_name = parse_palmpay_new(full_text)
    elif bank == "PalmPay_Business":
        buckets, account_name = parse_palmpay_business(full_text)
    elif bank == "Access_Oracle":
        buckets, account_name = parse_access_oracle(full_text)
    elif bank == "Fidelity_Direct":
        buckets, account_name = parse_fidelity_direct(full_text)
    elif bank == "Jaiz":
        buckets, account_name = parse_jaiz(full_text)
    elif bank == "Parallex":
        buckets, account_name = parse_parallex(full_text)
    elif bank == "Renmoney":
        buckets, account_name = parse_renmoney(full_text)
    elif bank in _MYBANKSTATEMENT_BANKS:
        # Sterling Bank: PyPDF2 silently drops pages with certain encodings.
        # pdfplumber captures all pages and matches the stated total exactly.
        _gtb_text = (
            extract_pdf_text_pdfplumber(file_bytes, password)
            if bank == "Sterling"
            else full_text
        )
        buckets, account_name = parse_gtbank(_gtb_text)
    elif bank == "Zenith":
        buckets, account_name = parse_zenith(full_text)
    elif bank == "Zenith_New":
        buckets, account_name = parse_zenith_new(full_text)
    elif bank == "Zenith_Corporate":
        buckets, account_name = parse_zenith_corporate(full_text)
    else:
        buckets, account_name = parse_generic(full_text)

    # Cache full_text so caller can reuse it for account-number extraction and
    # accuracy verification — eliminates 2–3 extra PDF re-parses in app.py.
    _last_full_text = full_text
    del full_text
    gc.collect()
    return buckets, summary, bank, account_name, list(_txn_log), list(_debit_log)


def monthly_analysis(buckets: dict, summary: dict | None = None) -> list[dict]:
    """
    Compute per-month eligible income.
    Deductions: self_transfer + reversal + non_business + loan_disbursal.
    self_transfer covers OWealth, Renflex, Renvault, RenSavings, savings
    platforms, and own-name round-trips — these are NOT business income.
    """
    rows = []
    for ym in sorted(set(buckets) | set(summary or {})):
        b = buckets.get(ym, _empty_bucket())
        gross = (summary or {}).get(ym, b["gross"])
        # Self-transfers (savings round-trips, own-name) are deducted as they
        # are not genuine business income.
        deductions = (b.get("self_transfer", 0) + b.get("reversal", 0) +
                      b.get("non_business", 0) + b.get("loan_disbursal", 0))
        rows.append({
            "ym": ym,
            "label": ym_label(ym),
            "gross": gross,
            "parsed_gross": b["gross"],
            "self_transfer": b.get("self_transfer", 0),
            "reversal": b.get("reversal", 0),
            "non_business": b.get("non_business", 0),
            "loan_disbursal": b.get("loan_disbursal", 0),
            "deductions": deductions,
            "eligible_income": max(gross - deductions, 0),
            "count": b.get("count", 0),
        })
    return rows


# ════════════════════════════════════════════════════════════════════════════
# FIRSTCENTRAL CREDIT BUREAU PARSER
# ════════════════════════════════════════════════════════════════════════════
def parse_firstcentral(pdf_bytes: bytes, password: str = "") -> dict:
    full_text = extract_pdf_text(pdf_bytes, password)

    STATUS_VALS = r"Closed|Open|Written Off|Lost Written Off|LostWritten Off|WrittenOff"
    CLASS_VALS  = r"Performing|Lost|Derogatory|Delinquent|Non Performing|Non-Performing"

    heading_re = re.compile(
        r'Details of Credit Agreement with "([^"]+)" for Account Number:\s*'
        r'([^\n\r]+(?:-\r?\n\s*[^\n\r]+)?)'
    )
    matches    = list(heading_re.finditer(full_text))
    detail_map: dict[str, dict] = {}

    def _fv(text: str, pattern: str) -> str:
        m = re.search(pattern, text, re.I)
        return m.group(1).strip() if m else ""

    for i, match in enumerate(matches):
        subscriber = match.group(1).strip()
        acct_no    = re.sub(r"[\s\r\n]+", "", match.group(2)).strip()
        if not acct_no:
            continue
        block_end  = matches[i+1].index if i < len(matches)-1 else min(len(full_text), match.index+5000)
        window     = re.sub(r"\s+", " ", full_text[match.index:block_end]).strip()

        acct_sts = (_fv(window, f"Account Status.{{0,60}}?({STATUS_VALS})") or
                    _fv(window, f"({STATUS_VALS})\\s+Account Status"))
        fac_cls  = (_fv(window, f"Facility Classification.{{0,60}}?({CLASS_VALS})") or
                    _fv(window, f"({CLASS_VALS})\\s+Facility Classification"))
        instl    = _parse_currency(
            _fv(window, r"Instalment Amount\s+([\d,]+\.\d{2})") or
            _fv(window, r"([\d,]+\.\d{2})\s+Instalment Amount"))
        outst    = _parse_currency(
            _fv(window, r"Current Balance\s+([\d,]+\.\d{2})") or
            _fv(window, r"([\d,]+\.\d{2})\s+Current Balance"))
        dur_raw  = (_fv(window, r"Loan Duration\s+((?:\d+)|(?:Not Available))\s+Day\(s\)") or
                    _fv(window, r"((?:\d+)|(?:Not Available))\s+Day\(s\)\s+Loan Duration") or
                    "Not Available")

        detail_map[acct_no] = {
            "subscriber": subscriber,
            "account_status": acct_sts or "Unknown",
            "facility_classification": fac_cls or "Unknown",
            "instalment_amount": instl,
            "outstanding_balance": outst,
            "loan_duration_days": dur_raw,
            "tenor_months": _get_tenor_months(dur_raw),
        }

    summary_map: dict[str, dict] = {}
    sum_start = full_text.find("Credit Agreements Summary")
    det_start = full_text.find("Details of Credit Agreement with")
    if sum_start >= 0 and det_start > sum_start:
        sumtext = re.sub(r"\s+", " ", full_text[sum_start:det_start]).strip()
        for acct_no, detail in detail_map.items():
            idx = sumtext.find(acct_no)
            if idx < 0:
                continue
            cs  = max(0, idx-250)
            ce  = min(len(sumtext), idx+250)
            ctx = sumtext[cs:ce]
            ap  = idx - cs
            figs = [float(m.group().replace(",","")) for m in re.finditer(r"[\d,]+\.\d{2}", ctx)]
            pc  = [m[0] for m in re.finditer(f"({CLASS_VALS})", ctx, re.I) if m.start() > ap]
            ps  = [m[0] for m in re.finditer(f"({STATUS_VALS})", ctx, re.I) if m.start() > ap]
            ac  = [m[0] for m in re.finditer(f"({CLASS_VALS})", ctx, re.I)]
            as_ = [m[0] for m in re.finditer(f"({STATUS_VALS})", ctx, re.I)]
            arr, ins, out, avl = 0, 0, 0, 0
            if len(figs) >= 4: arr, ins, out, avl = figs[:4]
            elif len(figs) == 3: arr, ins, out = figs[:3]
            elif len(figs) == 2: arr, ins = figs[:2]
            elif len(figs) == 1: arr = figs[0]
            summary_map[acct_no] = {
                "facility_classification": ((pc or ac or ["Unknown"])[0]).strip(),
                "account_status": ((ps or as_ or ["Unknown"])[0]).strip(),
                "instalment_amount": ins, "outstanding_balance": out,
                "arrear_amount": arr, "availed_limit": avl,
                "figure_count": len(figs),
            }

    records: list[CreditAccount] = []
    seen: set[str] = set()
    for acct_no, det in detail_map.items():
        if acct_no in seen:
            continue
        seen.add(acct_no)
        sm = summary_map.get(acct_no, {})
        fac_cls = det["facility_classification"] if det["facility_classification"] != "Unknown" else sm.get("facility_classification","Unknown")
        acct_sts= det["account_status"] if det["account_status"] != "Unknown" else sm.get("account_status","Unknown")
        fc      = sm.get("figure_count", 0)
        instl   = det["instalment_amount"] if det["instalment_amount"] > 0 else (sm.get("instalment_amount",0) if fc >= 3 else 0)
        outst   = det["outstanding_balance"] if det["outstanding_balance"] > 0 else (sm.get("outstanding_balance",0) if fc >= 3 else 0)
        arrear  = sm.get("arrear_amount", 0)
        tenor   = det["tenor_months"] or _get_tenor_months(det["loan_duration_days"])

        sn = re.sub(r"[^a-z]", "", acct_sts.lower())
        cn = re.sub(r"[^a-z]", "", fac_cls.lower())
        is_closed = sn == "closed"
        is_bad = (sn in ("open","writtenoff","lostwrittenoff") and
                  cn in ("lost","derogatory","delinquent","nonperforming") and
                  (outst > 10000 or arrear > 10000))
        is_perf = sn == "open" and cn == "performing"
        mo, drv = 0.0, False
        if is_perf:
            if instl > 0: mo = instl
            elif outst > 0 and tenor: mo = outst / tenor; drv = True

        records.append(CreditAccount(
            subscriber_name=det["subscriber"], account_number=acct_no,
            account_status=acct_sts, facility_classification=fac_cls,
            instalment_amount=instl, outstanding_balance=outst,
            loan_duration_days=det["loan_duration_days"], tenor_months=tenor,
            monthly_obligation=mo, derived_from_balance=drv,
            include_in_obligation=mo > 0, is_closed=is_closed, is_bad_credit=is_bad,
        ))

    visible = [r for r in records if not r.is_closed]
    return {
        "records": visible,
        "total_monthly_obligation": sum(r.monthly_obligation for r in visible if r.include_in_obligation),
        "bad_credit_accounts": [r for r in visible if r.is_bad_credit],
    }
